"""
Dev & Inspection Creator
Builds the Development Package section of the workbook using a template.
"""

from __future__ import annotations

import math
import os
import re
import urllib.parse
from copy import copy
from io import BytesIO

import streamlit as st
import requests

from google_sheets_uploader import (
    GoogleSheetsUploadError,
    upload_workbook_to_google_sheet,
)

# Google API imports for PDF export
try:
    from google.oauth2.service_account import Credentials as SACredentials
    from googleapiclient.discovery import build
    from google.auth.transport.requests import Request as GoogleRequest
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
    from openpyxl.utils import column_index_from_string, get_column_letter
except Exception:  # pragma: no cover - fallback if dependency missing at runtime
    load_workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None
    NamedStyle = None
    column_index_from_string = None
    get_column_letter = None


# Pricing constants
BASE_PRICE_STANDARD = 2790.00  # Standard pricing is now fixed at $2,790
ACTIVEWEAR_PRICE_LESS_THAN_5 = 3560.00
ACTIVEWEAR_PRICE_5_OR_MORE = 2965.00

OPTIONAL_PRICES = {
    "wash_dye": 1330.00,
    "design": 1330.00,
    "source": 1330.00,
    "treatment": 760.00,
}

# A La Carte hourly rates
A_LA_CARTE_RATE_STANDARD = 190.00  # $190/hr for INTAKE, PATTERN, SAMPLE, FITTING, ADJUSTMENT
A_LA_CARTE_RATE_SAMPLES = 90.00    # $90/hr for FINAL SAMPLES and DUPLICATES
SUMMARY_LABEL_COL = 14  # Column N
SUMMARY_VALUE_COL = 16  # Column P
SUMMARY_DEV_ROW = 10
SUMMARY_OPT_ROW = 12
SUMMARY_DISCOUNT_ROW = 14
SUMMARY_SUM_END_ROW = 13  # Row before discount row
SUMMARY_TOTAL_DUE_BASE_ROW = 20
DELIVERABLE_BLOCK_START = 22
DELIVERABLE_BLOCK_END = 34
DELIVERABLE_BLOCK_HEIGHT = DELIVERABLE_BLOCK_END - DELIVERABLE_BLOCK_START + 1
DELIVERABLE_COL_START = 2  # Column B
DELIVERABLE_COL_END = 16   # Column P
TEMPLATE_FILENAME = "Copy of TEG 2025 WORKBOOK TEMPLATES.xlsx"
TARGET_SHEET = "DEV & INSPECTION"
ROW_INDICES = [10, 12, 14, 16, 18]  # Rows reserved for style entries


@st.cache_data(show_spinner=False)
def get_template_path() -> str:
    """Return the absolute path to the Excel template."""
    base_dir = os.path.dirname(os.path.dirname(__file__))
    template_path = os.path.join(base_dir, "inputs", TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Template '{TEMPLATE_FILENAME}' was not found in the inputs folder."
        )
    return template_path


def calculate_base_price(num_styles: int, is_activewear: bool) -> float:
    """Calculate base price based on number of styles and activewear flag.
    
    Standard pricing is fixed at $2,790.
    Activewear pricing is based on total styles:
    - Less than 5 total styles: $3,560
    - 5 or more total styles: $2,965
    """
    if is_activewear:
        if num_styles < 5:
            return ACTIVEWEAR_PRICE_LESS_THAN_5
        else:
            return ACTIVEWEAR_PRICE_5_OR_MORE
    else:
        # Standard pricing is fixed at $2,790
        return BASE_PRICE_STANDARD


def copy_cell_formatting(source_cell, target_cell) -> None:
    """Copy formatting (fill, border, alignment) from source to target cell."""
    if PatternFill is None or Border is None or Alignment is None:
        return
    
    try:
        # Copy fill (background color)
        if source_cell.fill and source_cell.fill.start_color:
            target_cell.fill = PatternFill(
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color,
                fill_type=source_cell.fill.fill_type
            )
        
        # Copy border
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
    except Exception:
        pass  # Skip if copying fails


def apply_arial_20_font(cell) -> None:
    """Apply Arial font size 20 to a cell."""
    if Font is not None:
        try:
            cell.font = Font(name="Arial", size=20)
        except Exception:
            pass


def apply_full_border(cell) -> None:
    """Apply full thin borders around a cell."""
    if Border is None or Side is None:
        return
    try:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    except Exception:
        pass


def apply_full_border_pair(ws, column, start_row: int, end_row: int) -> None:
    """Apply borders to both cells in a vertical pair (before merging)."""
    if column_index_from_string is None:
        return
    col_idx = column if isinstance(column, int) else column_index_from_string(column)
    top_cell = ws.cell(row=start_row, column=col_idx)
    bottom_cell = ws.cell(row=end_row, column=col_idx)
    apply_full_border(top_cell)
    apply_full_border(bottom_cell)


def capture_deliverables_block(ws):
    """Capture the deliverables block (values + formatting) from the template."""
    block_rows = []
    for row in range(DELIVERABLE_BLOCK_START, DELIVERABLE_BLOCK_END + 1):
        row_data = []
        for col in range(DELIVERABLE_COL_START, DELIVERABLE_COL_END + 1):
            cell = ws.cell(row=row, column=col)
            row_data.append(
                {
                    "value": cell.value,
                    "number_format": cell.number_format,
                    "font": copy(cell.font) if cell.font else None,
                    "fill": copy(cell.fill) if cell.fill else None,
                    "border": copy(cell.border) if cell.border else None,
                    "alignment": copy(cell.alignment) if cell.alignment else None,
                }
            )
        block_rows.append(row_data)
    
    merged_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row >= DELIVERABLE_BLOCK_START
            and merged_range.max_row <= DELIVERABLE_BLOCK_END
            and merged_range.min_col >= DELIVERABLE_COL_START
            and merged_range.max_col <= DELIVERABLE_COL_END
        ):
            merged_ranges.append(
                (
                    merged_range.min_row - DELIVERABLE_BLOCK_START,
                    merged_range.max_row - DELIVERABLE_BLOCK_START,
                    merged_range.min_col,
                    merged_range.max_col,
                )
            )
    
    return {"rows": block_rows, "merges": merged_ranges}


def restore_deliverables_block(ws, template_block: dict, target_start_row: int) -> None:
    """Restore the deliverables block at the specified start row."""
    if not template_block:
        return
    
    rows_data = template_block.get("rows", [])
    block_height = len(rows_data)
    target_end_row = target_start_row + block_height - 1
    
    # Clear existing merges in target area
    to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.max_row < target_start_row
            or merged_range.min_row > target_end_row
            or merged_range.max_col < DELIVERABLE_COL_START
            or merged_range.min_col > DELIVERABLE_COL_END
        ):
            continue
        to_unmerge.append(merged_range)
    
    for merged_range in to_unmerge:
        try:
            ws.unmerge_cells(range_string=str(merged_range))
        except Exception:
            pass
    
    # Write cell data and formatting
    for row_offset, row_cells in enumerate(rows_data):
        target_row = target_start_row + row_offset
        for col_offset, cell_data in enumerate(row_cells):
            target_col = DELIVERABLE_COL_START + col_offset
            coord = None
            if get_column_letter is not None:
                coord = f"{get_column_letter(target_col)}{target_row}"
            value = cell_data.get("value") if cell_data else None
            if coord:
                safe_set_cell_value(ws, coord, value)
            else:
                # Fallback: unmerge first, then set value
                try:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= target_row <= merged_range.max_row and
                            merged_range.min_col <= target_col <= merged_range.max_col):
                            ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
                ws.cell(row=target_row, column=target_col).value = value
            cell = ws.cell(row=target_row, column=target_col)
            cell.number_format = cell_data.get("number_format") or cell.number_format
            if cell_data.get("font"):
                cell.font = copy(cell_data["font"])
            if cell_data.get("fill"):
                cell.fill = copy(cell_data["fill"])
            if cell_data.get("border"):
                cell.border = copy(cell_data["border"])
            if cell_data.get("alignment"):
                cell.alignment = copy(cell_data["alignment"])
    
    # Recreate merged ranges (adjusted for new start row)
    for min_row_offset, max_row_offset, min_col, max_col in template_block.get("merges", []):
        start_row = target_start_row + min_row_offset
        end_row = target_start_row + max_row_offset
        try:
            ws.merge_cells(
                start_row=start_row,
                start_column=min_col,
                end_row=end_row,
                end_column=max_col,
            )
        except Exception:
            pass


def safe_merge_cells(ws, range_str: str) -> bool:
    """Safely merge cells, checking if they're already merged first."""
    try:
        # Parse the range string (e.g., "B10:B11")
        parts = range_str.split(':')
        if len(parts) != 2:
            return False
        
        from openpyxl.utils import column_index_from_string
        
        # Parse start cell
        start_col_letter = ''.join([c for c in parts[0] if c.isalpha()])
        start_row = int(''.join([c for c in parts[0] if c.isdigit()]))
        start_col = column_index_from_string(start_col_letter)
        
        # Parse end cell
        end_col_letter = ''.join([c for c in parts[1] if c.isalpha()])
        end_row = int(''.join([c for c in parts[1] if c.isdigit()]))
        end_col = column_index_from_string(end_col_letter)
        
        # Check if any cell in this range is already part of a merged range
        # Do multiple passes to catch all overlapping merges
        max_passes = 3
        for pass_num in range(max_passes):
            ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if ranges overlap (any cell in our range overlaps with merged range)
                if not (merged_range.max_row < start_row or merged_range.min_row > end_row or
                        merged_range.max_col < start_col or merged_range.min_col > end_col):
                    # Ranges overlap, mark for unmerging
                    ranges_to_unmerge.append(merged_range)
            
            if not ranges_to_unmerge:
                break  # No more overlapping merges
            
            # Unmerge overlapping ranges (in reverse to avoid index issues)
            for merged_range in reversed(ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass
        
        # Now merge the cells vertically
        ws.merge_cells(range_str)
        return True
    except Exception:
        return False  # Return False if merge fails


def safe_get_cell_value(ws, row: int, col: int):
    """Safely get cell value, handling merged cells."""
    try:
        # Check if cell is part of a merged range
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                # Get value from top-left cell of merged range
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left.value
        # Not merged, get value directly
        cell = ws.cell(row=row, column=col)
        return cell.value
    except Exception:
        return None


def safe_set_cell_value(ws, cell_ref: str, value) -> None:
    """Safely set cell value, handling merged cells by writing to top-left cell."""
    try:
        # Parse cell reference to get row and column
        from openpyxl.utils import column_index_from_string
        col_letter = ''.join([c for c in cell_ref if c.isalpha()])
        row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
        col_num = column_index_from_string(col_letter)
        
        # Check if cell is part of a merged range and unmerge ALL ranges that contain this cell
        merged_ranges_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_num <= merged_range.max_row and
                merged_range.min_col <= col_num <= merged_range.max_col):
                merged_ranges_to_unmerge.append(merged_range)
        
        # Unmerge all ranges that contain this cell
        for merged_range in merged_ranges_to_unmerge:
            try:
                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                range_str = f"{min_cell.coordinate}:{max_cell.coordinate}"
                ws.unmerge_cells(range_str)
            except Exception:
                pass  # If unmerge fails, try to write anyway
        
        # Write to the cell using cell() method which always returns a writable cell
        # Always use cell() method, not ws[cell_ref], to avoid MergedCell issues
        # After unmerging, we need to get a fresh cell reference
        target_cell = ws.cell(row=row_num, column=col_num)
        # Double-check: if somehow we still have a MergedCell, get the top-left cell
        if not hasattr(target_cell, 'coordinate'):
            # It's a MergedCell, find the top-left cell of any remaining merged range
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col_num <= merged_range.max_col):
                    target_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        target_cell.value = value
        
    except Exception as e:
        # If all else fails, try one more time with cell() method directly
        try:
            from openpyxl.utils import column_index_from_string
            col_letter = ''.join([c for c in cell_ref if c.isalpha()])
            row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
            col_num = column_index_from_string(col_letter)
            # Always use cell() method - never use ws[cell_ref] as it can return MergedCell
            # Force unmerge any ranges containing this cell first
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col_num <= merged_range.max_col):
                    try:
                        min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                        ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                    except:
                        pass
            target_cell = ws.cell(row=row_num, column=col_num)
            target_cell.value = value
        except Exception:
            pass  # Skip if all methods fail


def update_header_labels(ws, client_name: str) -> None:
    """Ensure headers and client info match the spec."""
    header_map = {
        "H9": "WASH/DYE",
        "I9": "DESIGN",
        "J9": "SOURCE",
        "K9": "TREATMENT",
        "L9": "TOTAL",
    }
    for cell, label in header_map.items():
        safe_set_cell_value(ws, cell, label)
        # Center align row 9 headers (H9, I9, J9, K9, L9)
        if Alignment is not None:
            from openpyxl.utils import column_index_from_string
            col_letter = ''.join([c for c in cell if c.isalpha()])
            row_num = int(''.join([c for c in cell if c.isdigit()]))
            col_num = column_index_from_string(col_letter)
            header_cell = ws.cell(row=row_num, column=col_num)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # TEGMADE, JUST FOR - merged and centered in B3:O3
    # Unmerge first if needed, set value, then merge
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row == 3 and merged_range.max_row == 3 and
            merged_range.min_col <= 2 and merged_range.max_col >= 15):  # B=2, O=15
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                pass
    # Unmerge B3:O3 if it exists, then set value
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row == 3 and merged_range.max_row == 3 and
            merged_range.min_col <= 2 and merged_range.max_col >= 15):  # B=2, O=15
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                pass
    
    safe_set_cell_value(ws, "B3", "TEGMADE, JUST FOR")
    safe_merge_cells(ws, "B3:O3")
    if Alignment is not None:
        # Re-get cell after merging to ensure we have the top-left cell
        cell_b3 = ws.cell(row=3, column=2)
        # Force right alignment - create new Alignment object to override template
        cell_b3.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    
    # CLIENT NAME - merged and centered in P3:AA3
    client_display = (client_name or "").strip().upper()
    # Unmerge P3:AA3 if it exists
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row == 3 and merged_range.max_row == 3 and
            merged_range.min_col <= 16 and merged_range.max_col >= 27):  # P=16, AA=27
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                pass
    safe_set_cell_value(ws, "P3", client_display)
    safe_merge_cells(ws, "P3:AA3")
    if Font is not None:
        # Re-get cell after merging to ensure we have the top-left cell
        cell_p3 = ws.cell(row=3, column=16)
        cell_p3.font = Font(
            color="00C9A57A",
            name="Schibsted Grotesk",
            size=48,
            bold=True,
        )
    if Alignment is not None:
        # Re-get cell after merging to ensure we have the top-left cell
        cell_p3 = ws.cell(row=3, column=16)
        # Force left alignment - create new Alignment object to override template
        cell_p3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    if Alignment is not None:
        center_cells = ["L20", "P10", "P11", "P12", "P13", "P20"]
        for ref in center_cells:
            # Use cell() method to avoid MergedCell issues
            col_letter = ''.join([c for c in ref if c.isalpha()])
            row_num = int(''.join([c for c in ref if c.isdigit()]))
            from openpyxl.utils import column_index_from_string
            col_num = column_index_from_string(col_letter)
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")


def clear_style_rows(ws, num_styles: int = 0) -> None:
    """Blank out style rows (B–L) and the totals row, preserving format for <= 5 styles."""
    if num_styles <= 5:
        # For 5 or fewer styles, clear ALL template style rows (10-18) so past values don't remain
        # Use safe_set_cell_value to preserve formatting/merges
        style_rows = [10, 12, 14, 16, 18]
        for row_idx in style_rows:
            for col_letter in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L']:
                safe_set_cell_value(ws, f"{col_letter}{row_idx}", None)
        
        # Clear totals row (row 20)
        safe_set_cell_value(ws, "B20", None)
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "H20", None)
        safe_set_cell_value(ws, "L20", None)
        safe_set_cell_value(ws, f"N{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, "N21", None)
        safe_set_cell_value(ws, "P21", None)
    else:
        # For more than 5 styles, clear all style rows and totals row
        max_style_rows = num_styles
        for i in range(max_style_rows):
            row_idx = 10 + (i * 2)
            # Only clear if this is a style row (not the totals row)
            totals_row = 20 + (num_styles - 5) * 2
            if row_idx == totals_row:
                continue
            for col_idx in range(2, 13):  # Columns B through L (1-based)
                cell = ws.cell(row=row_idx, column=col_idx)
                # Use safe_set_cell_value to handle merged cells
                if get_column_letter:
                    safe_set_cell_value(ws, f"{get_column_letter(col_idx)}{row_idx}", None)
                else:
                    # Fallback: check if cell is part of a merged range
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Get the top-left cell of the merged range
                            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            safe_set_cell_value(ws, top_left.coordinate, None)
                            is_merged = True
                            break
                    if not is_merged:
                        safe_set_cell_value(ws, cell.coordinate, None)
        
        # Clear the totals row
        totals_row = 20 + (num_styles - 5) * 2
        safe_set_cell_value(ws, f"B{totals_row}", None)
        safe_set_cell_value(ws, f"F{totals_row}", None)
        safe_set_cell_value(ws, f"H{totals_row}", None)
        safe_set_cell_value(ws, f"L{totals_row}", None)
        safe_set_cell_value(ws, f"N{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, "N21", None)
        safe_set_cell_value(ws, "P21", None)


def apply_development_package(
    ws,
    *,
    client_name: str,
    client_email: str,
    representative: str,
    phone_number: str,
    style_entries: list[dict],
    custom_styles: list[dict],
    a_la_carte_items: list[dict],
    discount_percentage: float,
) -> tuple[float, float]:
    """Write the inputs into the workbook and return totals."""
    # Header metadata - merge and center cells
    # D6:F6 for client email
    safe_set_cell_value(ws, "D6", client_email.strip())
    safe_merge_cells(ws, "D6:F6")
    if Alignment is not None:
        cell_d6 = ws.cell(row=6, column=4)  # D=4
        cell_d6.alignment = Alignment(horizontal="center", vertical="center")
    
    # J6:L6 for representative
    safe_set_cell_value(ws, "J6", (representative or "").strip().upper())
    safe_merge_cells(ws, "J6:L6")
    if Alignment is not None:
        cell_j6 = ws.cell(row=6, column=10)  # J=10
        cell_j6.alignment = Alignment(horizontal="center", vertical="center")
    
    # P6:Q6 for phone number
    if phone_number and phone_number.strip():
        safe_set_cell_value(ws, "P6", phone_number.strip())
        safe_merge_cells(ws, "P6:Q6")
        if Alignment is not None:
            cell_p6 = ws.cell(row=6, column=16)  # P=16
            cell_p6.alignment = Alignment(horizontal="center", vertical="center")
    
    # B8:F8 for "DEVELOPMENT PACKAGE"
    # Set value first, then merge, then apply formatting (safe_set_cell_value unmerges, so we set directly)
    cell_b8 = ws.cell(row=8, column=2)  # B=2
    cell_b8.value = "DEVELOPMENT PACKAGE"
    if safe_merge_cells(ws, "B8:F8"):
        # Re-get cell after merging to ensure we have the top-left cell of merged range
        cell_b8 = ws.cell(row=8, column=2)
    if Alignment is not None:
        cell_b8.alignment = Alignment(horizontal="center", vertical="center")
    
    # H8:L8 for "OPTIONAL ADD-ONS"
    # Set value first, then merge, then apply formatting
    cell_h8 = ws.cell(row=8, column=8)  # H=8
    cell_h8.value = "OPTIONAL ADD-ONS"
    if safe_merge_cells(ws, "H8:L8"):
        # Re-get cell after merging to ensure we have the top-left cell of merged range
        cell_h8 = ws.cell(row=8, column=8)
    if Alignment is not None:
        cell_h8.alignment = Alignment(horizontal="center", vertical="center")

    optional_cells = {
        "H": "wash_dye",
        "I": "design",
        "J": "source",
        "K": "treatment",
    }

    total_development = 0.0
    total_optional = 0.0
    num_styles = len(style_entries)
    num_custom_styles = len(custom_styles)
    total_styles_count = num_styles + num_custom_styles  # Total for pricing tier calculation
    discount_percentage = max(0.0, float(discount_percentage or 0))
    deliverables_template = capture_deliverables_block(ws)

    base_capacity = len(ROW_INDICES)
    extra_styles = max(num_styles - base_capacity, 0)
    rows_to_insert_regular = extra_styles * 2
    
    # Calculate how many rows we'll need for Custom Items too (if any)
    # This ensures we insert all rows upfront before writing any styles
    rows_to_insert_custom = 0
    if num_custom_styles > 0:
        # Calculate where totals row will be after regular insertions
        if num_styles > 5:
            totals_row_after_regular = 20 + (num_styles - 5) * 2
        else:
            totals_row_after_regular = 20
        
        # Calculate where Custom Items would start (after all regular styles, before totals)
        if num_styles > 5:
            custom_start_row = 20 + (num_styles - 5) * 2
        else:
            # For <= 5 styles, custom items start after last regular style (row 18) at row 20
            # But row 20 is the totals row, so we need to insert rows BEFORE totals
            custom_start_row = 10 + num_styles * 2  # This will be 20 for 5 styles
        
        # Calculate custom row indices
        custom_row_indices_precalc = []
        for i in range(num_custom_styles):
            custom_row_indices_precalc.append(custom_start_row + (i * 2))
        
        # Check if we need more rows for Custom Items
        # If custom items would overlap with totals row, we need to insert rows BEFORE totals
        if custom_row_indices_precalc:
            last_custom_row = custom_row_indices_precalc[-1]
            if last_custom_row >= totals_row_after_regular - 2:
                # We need to insert rows before the totals row
                required_totals_row = last_custom_row + 2
                rows_to_insert_custom = required_totals_row - totals_row_after_regular
                if rows_to_insert_custom < 0:
                    rows_to_insert_custom = 0
            elif num_styles <= 5 and custom_start_row >= totals_row_after_regular:
                # Special case: when <= 5 styles and custom items would start at or after totals row
                # We need to insert rows before totals row (row 20)
                rows_to_insert_custom = num_custom_styles * 2

    # Insert all rows upfront (regular + custom) before writing any styles
    total_rows_to_insert = rows_to_insert_regular + rows_to_insert_custom
    if total_rows_to_insert > 0:
        ws.insert_rows(SUMMARY_TOTAL_DUE_BASE_ROW, amount=total_rows_to_insert)

        # Copy formatting from template row 18 to new rows (preserve colors, borders, alignment)
        # New rows start at 20, 22, 24, etc.
        template_row = 18  # Use row 18 as template for formatting
        for i in range(total_rows_to_insert):
            new_row = 20 + i
            for col_idx in range(2, 13):  # Columns B through L
                source_cell = ws.cell(row=template_row, column=col_idx)
                target_cell = ws.cell(row=new_row, column=col_idx)
                copy_cell_formatting(source_cell, target_cell)

        # Unmerge ALL cells in newly inserted rows to avoid MergedCell errors
        # Excel automatically adjusts merged ranges when rows are inserted, which can cause conflicts
        first_new_row = SUMMARY_TOTAL_DUE_BASE_ROW
        last_new_row = SUMMARY_TOTAL_DUE_BASE_ROW + total_rows_to_insert - 1
        
        # Collect all merged ranges that intersect with the newly inserted rows
        # We need to unmerge both horizontal and vertical merges in the new rows
        # Do this multiple times to catch all merges (some might be created after unmerging others)
        max_iterations = 3
        for iteration in range(max_iterations):
            merged_ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if merged range intersects with newly inserted rows
                # (min_row <= last_new_row and max_row >= first_new_row)
                if (merged_range.min_row <= last_new_row and merged_range.max_row >= first_new_row):
                    merged_ranges_to_unmerge.append(merged_range)
            
            if not merged_ranges_to_unmerge:
                break  # No more merges to unmerge
            
            # Unmerge the identified ranges (do this in reverse to avoid index issues)
            for merged_range in reversed(merged_ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    range_str = f"{min_cell.coordinate}:{max_cell.coordinate}"
                    ws.unmerge_cells(range_str)
                except Exception:
                    pass  # Skip if unmerge fails

    # Clear style rows and totals row (after insertion so row numbers are correct)
    clear_style_rows(ws, num_styles=num_styles)

    # Generate row indices dynamically based on actual number of styles
    # Each style uses 2 rows (merged): 10-11, 12-13, 14-15, 16-17, 18-19, 20-21, etc.
    # We write to the first row of each pair (10, 12, 14, 16, 18, 20, 22, etc.)
    dynamic_row_indices = []
    start_row = 10
    for i in range(num_styles):
        dynamic_row_indices.append(start_row + (i * 2))

    for idx, row_idx in enumerate(dynamic_row_indices):
        entry = style_entries[idx]
        style_name = entry.get("name", "").strip() or "STYLE"
        complexity_pct = float(entry.get("complexity", 0.0))
        is_activewear = entry.get("activewear", False)
        row_options = entry.get("options", {})

        # Calculate base price based on tiered pricing and activewear
        # Use total_styles_count (regular + custom) for pricing tier
        line_base_price = calculate_base_price(total_styles_count, is_activewear)

        # Check if this is a new row (row_idx > 18) that needs Arial 20 font
        is_new_row = num_styles > 5 and row_idx > 18
        
        # Each style row spans 2 rows (merged cells)
        row_second = row_idx + 1

        # For new rows, explicitly unmerge all cells in this row pair before writing
        # This ensures we start with clean, unmerged cells (prevents horizontal merges)
        if is_new_row:
            ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if merged range intersects with this row pair
                if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                    merged_range.min_col <= 12 and merged_range.max_col >= 2):  # Columns B-L
                    ranges_to_unmerge.append(merged_range)
            
            # Unmerge all overlapping ranges
            for merged_range in reversed(ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass

        # Write style number (#) - use style_number from entry (101, 102, 103, etc.)
        style_number = entry.get("style_number", 101 + idx)  # Default to 101, 102, 103... if not set
        safe_set_cell_value(ws, f"B{row_idx}", style_number)
        cell_b = ws.cell(row=row_idx, column=2)
        # Always merge B10:B11, B12:B13, etc. (not just for new rows)
        apply_full_border_pair(ws, 2, row_idx, row_second)
        safe_merge_cells(ws, f"B{row_idx}:B{row_second}")
        if is_new_row:
            apply_arial_20_font(cell_b)
        if Alignment is not None:
            cell_b.alignment = Alignment(horizontal="left", vertical="center")
        
        # Write style name (merged across 2 rows, left-aligned)
        safe_set_cell_value(ws, f"C{row_idx}", style_name.upper())
        cell_c = ws.cell(row=row_idx, column=3)
        # Always merge C10:C11, C12:C13, etc. (not just for new rows)
        apply_full_border_pair(ws, 3, row_idx, row_second)
        safe_merge_cells(ws, f"C{row_idx}:C{row_second}")
        if is_new_row:
            apply_arial_20_font(cell_c)
        if Alignment is not None:
            cell_c.alignment = Alignment(horizontal="left", vertical="center")
        
        # Set complexity - show percentage when provided
        cell_e = ws.cell(row=row_idx, column=5)
        if complexity_pct == 0:
            safe_set_cell_value(ws, f"E{row_idx}", None)
            # Even when complexity is 0, merge and center for all rows
            apply_full_border_pair(ws, 5, row_idx, row_second)
            safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_e)
            if Alignment is not None:
                cell_e.alignment = Alignment(horizontal="center", vertical="center")
        else:
            safe_set_cell_value(ws, f"E{row_idx}", complexity_pct / 100.0)
            cell_e.number_format = '0%'  # Percentage format
            # Always merge E10:E11, E12:E13, etc. (not just for new rows)
            apply_full_border_pair(ws, 5, row_idx, row_second)
            safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_e)
            if Alignment is not None:
                cell_e.alignment = Alignment(horizontal="center", vertical="center")

        # Write base price (currency format, integer)
        safe_set_cell_value(ws, f"D{row_idx}", int(line_base_price))
        cell_d = ws.cell(row=row_idx, column=4)
        cell_d.number_format = '$#,##0'  # Currency format
        # Always merge D10:D11, D12:D13, etc. (not just for new rows)
        apply_full_border_pair(ws, 4, row_idx, row_second)
        safe_merge_cells(ws, f"D{row_idx}:D{row_second}")
        if is_new_row:
            apply_arial_20_font(cell_d)
        if Alignment is not None:
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write total formula (column F)
        if complexity_pct == 0:
            safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}")
        else:
            safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}*(1+E{row_idx})")
        cell_f = ws.cell(row=row_idx, column=6)
        cell_f.number_format = '$#,##0'  # Currency format
        # Always merge F10:F11, F12:F13, etc. (not just for new rows)
        apply_full_border_pair(ws, 6, row_idx, row_second)
        safe_merge_cells(ws, f"F{row_idx}:F{row_second}")
        if is_new_row:
            apply_arial_20_font(cell_f)
        if Alignment is not None:
            cell_f.alignment = Alignment(horizontal="center", vertical="center")

        # Optional add-ons per row (columns H, I, J, K)
        row_optional_sum = 0.0
        for col_letter, key in optional_cells.items():
            col_num = ord(col_letter) - 64  # Convert letter to column number
            if row_options.get(key):
                price = int(OPTIONAL_PRICES[key])  # Ensure integer
                safe_set_cell_value(ws, f"{col_letter}{row_idx}", price)
                cell_opt = ws.cell(row=row_idx, column=col_num)
                cell_opt.number_format = '$#,##0'  # Currency format
                # Always merge H10:H11, I10:I11, etc. (not just for new rows)
                apply_full_border_pair(ws, col_letter, row_idx, row_second)
                safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                if Alignment is not None:
                    cell_opt.alignment = Alignment(horizontal="center", vertical="center")
                row_optional_sum += price
            else:
                safe_set_cell_value(ws, f"{col_letter}{row_idx}", None)
                cell_opt = ws.cell(row=row_idx, column=col_num)
                # Always merge even if empty (not just for new rows)
                apply_full_border_pair(ws, col_letter, row_idx, row_second)
                safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                if Alignment is not None:
                    cell_opt.alignment = Alignment(horizontal="center", vertical="center")
        
        safe_set_cell_value(ws, f"L{row_idx}", f"=SUM(H{row_idx}:K{row_idx})")
        cell_l = ws.cell(row=row_idx, column=12)
        cell_l.number_format = '$#,##0'  # Currency format
        # Always merge L10:L11, L12:L13, etc. (not just for new rows)
        apply_full_border_pair(ws, 12, row_idx, row_second)
        safe_merge_cells(ws, f"L{row_idx}:L{row_second}")
        if is_new_row:
            apply_arial_20_font(cell_l)
        if Alignment is not None:
            cell_l.alignment = Alignment(horizontal="center", vertical="center")

        total_development += line_base_price * (1 + complexity_pct / 100.0)
        total_optional += row_optional_sum

    # Process Custom Items (user-defined price, complexity, no add-ons)
    if num_custom_styles > 0:
        # Calculate starting row for Custom Items (after all regular styles)
        # After row insertion, if we have <= 5 styles, custom items start at row 20 (which was inserted)
        # If we have > 5 styles, custom items start after the last regular style
        if num_styles <= 5:
            # For <= 5 styles, custom items start at row 20 (after row 18, the last regular style row)
            # Row 20 was the totals row, but we inserted rows before it, so now custom items use row 20
            custom_start_row = 20
        else:
            # For > 5 styles, custom items start after the last regular style
            custom_start_row = dynamic_row_indices[-1] + 2 if dynamic_row_indices else 10
        
        custom_row_indices = []
        for i in range(num_custom_styles):
            custom_row_indices.append(custom_start_row + (i * 2))
        
        # Custom row indices are already calculated correctly since we inserted all rows upfront
        # No need to insert more rows here - they were already inserted above
        
        for idx, row_idx in enumerate(custom_row_indices):
            entry = custom_styles[idx]
            style_name = entry.get("name", "").strip() or "STYLE"
            custom_price = float(entry.get("price", 0.0))
            complexity_pct = float(entry.get("complexity", 0.0))
            
            is_new_row = row_idx > 18
            row_second = row_idx + 1
            
            # For new rows, explicitly unmerge all cells in this row pair before writing
            # This ensures we start with clean, unmerged cells (prevents horizontal merges)
            if is_new_row:
                ranges_to_unmerge = []
                for merged_range in list(ws.merged_cells.ranges):
                    # Check if merged range intersects with this row pair
                    if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                        merged_range.min_col <= 12 and merged_range.max_col >= 2):  # Columns B-L
                        ranges_to_unmerge.append(merged_range)
                
                # Unmerge all overlapping ranges
                for merged_range in reversed(ranges_to_unmerge):
                    try:
                        min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                        ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                    except Exception:
                        pass
            
            # Write style number (#) - use the style_number from entry if available, otherwise default
            # Use style_number from entry if set, otherwise default to 101 + num_styles + idx
            style_number = entry.get("style_number", 101 + num_styles + idx)
            safe_set_cell_value(ws, f"B{row_idx}", style_number)
            cell_b = ws.cell(row=row_idx, column=2)
            # Always merge B10:B11, B12:B13, etc. (not just for new rows)
            apply_full_border_pair(ws, 2, row_idx, row_second)
            safe_merge_cells(ws, f"B{row_idx}:B{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_b)
            if Alignment is not None:
                cell_b.alignment = Alignment(horizontal="left", vertical="center")
            
            # Write style name
            safe_set_cell_value(ws, f"C{row_idx}", style_name.upper())
            cell_c = ws.cell(row=row_idx, column=3)
            # Always merge C10:C11, C12:C13, etc. (not just for new rows)
            apply_full_border_pair(ws, 3, row_idx, row_second)
            safe_merge_cells(ws, f"C{row_idx}:C{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_c)
            if Alignment is not None:
                cell_c.alignment = Alignment(horizontal="left", vertical="center")
            
            # Write custom price
            safe_set_cell_value(ws, f"D{row_idx}", int(custom_price))
            cell_d = ws.cell(row=row_idx, column=4)
            cell_d.number_format = '$#,##0'
            # Always merge D10:D11, D12:D13, etc. (not just for new rows)
            apply_full_border_pair(ws, 4, row_idx, row_second)
            safe_merge_cells(ws, f"D{row_idx}:D{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_d)
            if Alignment is not None:
                cell_d.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write complexity
            if complexity_pct == 0:
                safe_set_cell_value(ws, f"E{row_idx}", None)
                cell_e = ws.cell(row=row_idx, column=5)
            else:
                safe_set_cell_value(ws, f"E{row_idx}", complexity_pct / 100.0)
                cell_e = ws.cell(row=row_idx, column=5)
                cell_e.number_format = '0%'
            # Always merge E10:E11, E12:E13, etc. (not just for new rows)
            apply_full_border_pair(ws, 5, row_idx, row_second)
            safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_e)
            if Alignment is not None:
                cell_e.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write total formula
            if complexity_pct == 0:
                safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}")
            else:
                safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}*(1+E{row_idx})")
            cell_f = ws.cell(row=row_idx, column=6)
            cell_f.number_format = '$#,##0'
            # Always merge F10:F11, F12:F13, etc. (not just for new rows)
            apply_full_border_pair(ws, 6, row_idx, row_second)
            safe_merge_cells(ws, f"F{row_idx}:F{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_f)
            if Alignment is not None:
                cell_f.alignment = Alignment(horizontal="center", vertical="center")
            
            # Clear add-ons (Custom Items don't have add-ons)
            for col_letter in ["H", "I", "J", "K"]:
                safe_set_cell_value(ws, f"{col_letter}{row_idx}", None)
                col_num = ord(col_letter) - 64
                cell_opt = ws.cell(row=row_idx, column=col_num)
                # Always merge even if empty (not just for new rows)
                apply_full_border_pair(ws, col_letter, row_idx, row_second)
                safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                if Alignment is not None:
                    cell_opt.alignment = Alignment(horizontal="center", vertical="center")
            
            # Clear total optional add-on
            safe_set_cell_value(ws, f"L{row_idx}", None)
            cell_l = ws.cell(row=row_idx, column=12)
            # Always merge L10:L11, L12:L13, etc. (not just for new rows)
            apply_full_border_pair(ws, 12, row_idx, row_second)
            safe_merge_cells(ws, f"L{row_idx}:L{row_second}")
            if is_new_row:
                apply_arial_20_font(cell_l)
            if Alignment is not None:
                cell_l.alignment = Alignment(horizontal="center", vertical="center")
            
            total_development += custom_price * (1 + complexity_pct / 100.0)
        
        # Update last_style_row to include Custom Items (for totals calculations)
        if custom_row_indices:
            last_style_row = custom_row_indices[-1]
    
    # Determine last_style_row if no Custom Items (for totals calculations)
    if num_custom_styles == 0:
        if dynamic_row_indices:
            last_style_row = dynamic_row_indices[-1]
        else:
            last_style_row = 10

    # Determine last_regular_style_row (only regular + activewear, excluding Custom Items)
    # This is used for deliverables counts (Patterns, First Samples, Final Samples)
    if dynamic_row_indices:
        last_regular_style_row = dynamic_row_indices[-1]
    else:
        last_regular_style_row = 10

    # Count activewear and regular styles
    num_activewear = sum(1 for entry in style_entries if entry.get("activewear", False))
    num_regular = sum(1 for entry in style_entries if not entry.get("activewear", False))
    
    total_extra_rows = max(total_styles_count - len(ROW_INDICES), 0) * 2
    deliverables_block_start = DELIVERABLE_BLOCK_START + total_extra_rows
    deliverables_block_end = deliverables_block_start + DELIVERABLE_BLOCK_HEIGHT - 1
    restore_deliverables_block(ws, deliverables_template, deliverables_block_start)
    
    # If there are activewear styles, modify deliverables section
    if num_activewear > 0 and column_index_from_string is not None:
        label_column_idx = column_index_from_string("B")
        col_c_idx = column_index_from_string("C")
        col_d_idx = column_index_from_string("D")
        final_samples_row = None
        row_final_samples_new = None  # Store the row where new FINAL SAMPLES is created
        
        # Find the "FINAL SAMPLES" row (should be around row 31-32)
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=label_column_idx).value
            if isinstance(value, str):
                value_lower = value.strip().lower()
                if "final samples" in value_lower and final_samples_row is None:
                    final_samples_row = scan_row
                    break
        
        if final_samples_row:
            # Step 1: Replace FINAL SAMPLES (rows 31-32) with SECOND SAMPLES (only columns B-D)
            # Unmerge cells in columns B-D for these rows first
            for row in [final_samples_row, final_samples_row + 1]:
                for col in [label_column_idx, col_c_idx, col_d_idx]:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass

            # Clear values in row 32 (second row of pair) for columns B-D
            for col in [label_column_idx, col_c_idx, col_d_idx]:
                if get_column_letter:
                    safe_set_cell_value(ws, f"{get_column_letter(col)}{final_samples_row + 1}", None)
                else:
                    # Fallback: unmerge first, then set value
                    try:
                        for merged_range in list(ws.merged_cells.ranges):
                            if (merged_range.min_row <= final_samples_row + 1 <= merged_range.max_row and
                                merged_range.min_col <= col <= merged_range.max_col):
                                ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
                    safe_set_cell_value(ws, f"{get_column_letter(col) if get_column_letter else 'A'}{final_samples_row + 1}", None)

            # Clear column C in row 31 (to ensure no leftover values)
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{final_samples_row}", None)
            else:
                # Fallback: unmerge first, then set value
                try:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                            merged_range.min_col <= col_c_idx <= merged_range.max_col):
                            ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx) if get_column_letter else 'C'}{final_samples_row}", None)

            # Set SECOND SAMPLES in row 31, column B, count in column D
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{final_samples_row}", "SECOND SAMPLES")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{final_samples_row}", num_activewear)
            else:
                # Fallback: use safe_set_cell_value which handles unmerging
                safe_set_cell_value(ws, f"B{final_samples_row}", "SECOND SAMPLES")
                safe_set_cell_value(ws, f"D{final_samples_row}", num_activewear)
            cell_d = ws.cell(row=final_samples_row, column=col_d_idx)
            cell_d.number_format = "0"
            
            # Use row 29 (ROUND OF REVISIONS) as reference - it should be around row 29
            # Find ROUND OF REVISIONS row (should be row 29)
            reference_row = None
            for scan_row in range(deliverables_block_start, final_samples_row):
                value = ws.cell(row=scan_row, column=label_column_idx).value
                if isinstance(value, str) and "round of revisions" in value.lower():
                    reference_row = scan_row
                    break
            
            # If not found, assume it's row 29 (typical position)
            if reference_row is None:
                reference_row = 29
            
            # Check what merges exist for columns B-D in rows 29-30 (reference_row to reference_row+1)
            b_merged = False
            c_merged = False
            d_merged = False
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row == reference_row and merged_range.max_row == reference_row + 1):
                    if merged_range.min_col == 2 and merged_range.max_col == 2:  # Column B
                        b_merged = True
                    elif merged_range.min_col == 3 and merged_range.max_col == 3:  # Column C
                        c_merged = True
                    elif merged_range.min_col == 4 and merged_range.max_col == 4:  # Column D
                        d_merged = True
            
            # Copy exact formatting from B29-D30 to B31-D32 (SECOND SAMPLES)
            # First, copy cell formatting (but not alignment for column B - we'll center it after merge)
            for col in [label_column_idx, col_c_idx, col_d_idx]:
                source_cell_29 = ws.cell(row=reference_row, column=col)
                source_cell_30 = ws.cell(row=reference_row + 1, column=col)
                target_cell_31 = ws.cell(row=final_samples_row, column=col)
                target_cell_32 = ws.cell(row=final_samples_row + 1, column=col)
                
                # Copy formatting from row 29 to row 31
                if source_cell_29.font:
                    target_cell_31.font = copy(source_cell_29.font)
                if source_cell_29.fill:
                    target_cell_31.fill = copy(source_cell_29.fill)
                if source_cell_29.border:
                    target_cell_31.border = copy(source_cell_29.border)
                # Don't copy alignment for column B - we'll set it to center after merge
                if col != label_column_idx and source_cell_29.alignment:
                    target_cell_31.alignment = copy(source_cell_29.alignment)
                target_cell_31.number_format = source_cell_29.number_format
                
                # Copy formatting from row 30 to row 32
                if source_cell_30.font:
                    target_cell_32.font = copy(source_cell_30.font)
                if source_cell_30.fill:
                    target_cell_32.fill = copy(source_cell_30.fill)
                if source_cell_30.border:
                    target_cell_32.border = copy(source_cell_30.border)
                # Don't copy alignment for column B - we'll set it to center after merge
                if col != label_column_idx and source_cell_30.alignment:
                    target_cell_32.alignment = copy(source_cell_30.alignment)
                target_cell_32.number_format = source_cell_30.number_format
            
            # Merge B:C together and D separately, with center alignment
            if safe_merge_cells:
                # Merge B:C together (columns B and C merged across 2 rows)
                safe_merge_cells(ws, f"B{final_samples_row}:C{final_samples_row + 1}")
                # Set center alignment for merged B:C
                if Alignment:
                    cell_bc = ws.cell(row=final_samples_row, column=label_column_idx)
                    cell_bc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                # Merge column D
                safe_merge_cells(ws, f"D{final_samples_row}:D{final_samples_row + 1}")
                # Set center alignment for merged column D
                if Alignment:
                    cell_d = ws.cell(row=final_samples_row, column=col_d_idx)
                    cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Step 2: Insert 6 rows after row 32 (3 new deliverable entries, each with 2 rows)
            insert_row = final_samples_row + 2  # After row 32
            rows_to_insert = 6
            
            # Insert rows
            ws.insert_rows(insert_row, amount=rows_to_insert)
            
            # Unmerge any cells in the newly inserted rows
            for i in range(rows_to_insert):
                target_row = insert_row + i
                for col in range(DELIVERABLE_COL_START, DELIVERABLE_COL_END + 1):
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= target_row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass
            
            # Step 3: Set values for the new rows and define row variables
            # Row 33-34: 2ND ROUND OF FITTINGS (count in column D)
            row_2nd_fittings = insert_row
            # Clear column C to ensure no leftover values
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_2nd_fittings}", None)
            else:
                # Fallback: unmerge first, then set value
                try:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= row_2nd_fittings <= merged_range.max_row and
                            merged_range.min_col <= col_c_idx <= merged_range.max_col):
                            ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx) if get_column_letter else 'C'}{row_2nd_fittings}", None)
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_2nd_fittings}", "2ND ROUND OF FITTINGS")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{row_2nd_fittings}", 1)
            else:
                # Fallback: use safe_set_cell_value which handles unmerging
                safe_set_cell_value(ws, f"B{row_2nd_fittings}", "2ND ROUND OF FITTINGS")
                safe_set_cell_value(ws, f"D{row_2nd_fittings}", 1)
            cell_d = ws.cell(row=row_2nd_fittings, column=col_d_idx)
            cell_d.number_format = "0"
            row_2nd_fittings_2 = insert_row + 1
            
            # Row 35-36: 2ND ROUND OF REVISIONS (count in column D)
            row_2nd_revisions = insert_row + 2
            # Clear column C to ensure no leftover values
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_2nd_revisions}", None)
            else:
                # Fallback: unmerge first, then set value
                try:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= row_2nd_revisions <= merged_range.max_row and
                            merged_range.min_col <= col_c_idx <= merged_range.max_col):
                            ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx) if get_column_letter else 'C'}{row_2nd_revisions}", None)
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_2nd_revisions}", "2ND ROUND OF REVISIONS")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{row_2nd_revisions}", 1)
            else:
                # Fallback: use safe_set_cell_value which handles unmerging
                safe_set_cell_value(ws, f"B{row_2nd_revisions}", "2ND ROUND OF REVISIONS")
                safe_set_cell_value(ws, f"D{row_2nd_revisions}", 1)
            cell_d = ws.cell(row=row_2nd_revisions, column=col_d_idx)
            cell_d.number_format = "0"
            row_2nd_revisions_2 = insert_row + 3
            
            # Row 37-38: FINAL SAMPLES (count of all styles)
            row_final_samples = insert_row + 4
            row_final_samples_new = row_final_samples  # Store for later use
            # Clear column C to ensure no leftover values
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_final_samples}", None)
            else:
                # Fallback: unmerge first, then set value
                try:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= row_final_samples <= merged_range.max_row and
                            merged_range.min_col <= col_c_idx <= merged_range.max_col):
                            ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx) if get_column_letter else 'C'}{row_final_samples}", None)
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_final_samples}", "FINAL SAMPLES")
                # Will be set by formulas section below
            else:
                # Fallback: use safe_set_cell_value which handles unmerging
                safe_set_cell_value(ws, f"B{row_final_samples}", "FINAL SAMPLES")
            row_final_samples_2 = insert_row + 5
            
            # Step 4: Copy exact formatting from B29-D30 to each new row pair (B33-D34, B35-D36, B37-D38)
            # Use row 29 as source for first row of each pair, row 30 for second row
            # Don't copy alignment for column B - we'll center it after merge
            for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                row_second = base_row + 1
                # Copy formatting from row 29 to first row of pair
                for col in [label_column_idx, col_c_idx, col_d_idx]:
                    source_cell_29 = ws.cell(row=reference_row, column=col)
                    source_cell_30 = ws.cell(row=reference_row + 1, column=col)
                    target_cell_first = ws.cell(row=base_row, column=col)
                    target_cell_second = ws.cell(row=row_second, column=col)
                    
                    # Copy formatting from row 29 to first row
                    if source_cell_29.font:
                        target_cell_first.font = copy(source_cell_29.font)
                    if source_cell_29.fill:
                        target_cell_first.fill = copy(source_cell_29.fill)
                    if source_cell_29.border:
                        target_cell_first.border = copy(source_cell_29.border)
                    # Don't copy alignment for column B - we'll set it to center after merge
                    if col != label_column_idx and source_cell_29.alignment:
                        target_cell_first.alignment = copy(source_cell_29.alignment)
                    target_cell_first.number_format = source_cell_29.number_format
                    
                    # Copy formatting from row 30 to second row
                    if source_cell_30.font:
                        target_cell_second.font = copy(source_cell_30.font)
                    if source_cell_30.fill:
                        target_cell_second.fill = copy(source_cell_30.fill)
                    if source_cell_30.border:
                        target_cell_second.border = copy(source_cell_30.border)
                    # Don't copy alignment for column B - we'll set it to center after merge
                    if col != label_column_idx and source_cell_30.alignment:
                        target_cell_second.alignment = copy(source_cell_30.alignment)
                    target_cell_second.number_format = source_cell_30.number_format
            
            # Step 5: Merge B:C together and D separately, with center alignment for each new row pair
            if safe_merge_cells:
                for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                    row_second = base_row + 1
                    # Merge B:C together (columns B and C merged across 2 rows)
                    safe_merge_cells(ws, f"B{base_row}:C{row_second}")
                    # Set center alignment for merged B:C
                    if Alignment:
                        cell_bc = ws.cell(row=base_row, column=label_column_idx)
                        cell_bc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    # Merge column D
                    safe_merge_cells(ws, f"D{base_row}:D{row_second}")
                    # Set center alignment for merged column D
                    if Alignment:
                        cell_d = ws.cell(row=base_row, column=col_d_idx)
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    
                    # Set FINAL SAMPLES count immediately after merging
                    if base_row == row_final_samples:
                        # Unmerge column D temporarily to set the value
                        try:
                            ws.unmerge_cells(f"D{base_row}:D{row_second}")
                        except Exception:
                            pass
                        # Set the count value (total styles: regular + activewear, excluding custom line items)
                        count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                        # Clear the cell completely (including any formulas) before setting value
                        safe_set_cell_value(ws, f"D{base_row}", None)
                        # Set direct numeric value (not a formula)
                        safe_set_cell_value(ws, f"D{base_row}", count_value)
                        count_cell = ws.cell(row=base_row, column=col_d_idx)
                        count_cell.number_format = "0"
                        # Re-merge column D
                        safe_merge_cells(ws, f"D{base_row}:D{row_second}")
                        # Restore center alignment
                        if Alignment:
                            cell_d = ws.cell(row=base_row, column=col_d_idx)
                            cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Restore merged cells for right side columns (H-P) using template merge patterns
            if deliverables_template and "merges" in deliverables_template:
                right_side_merges = []
                for min_row_offset, max_row_offset, min_col, max_col in deliverables_template.get("merges", []):
                    if (min_col >= 8 and max_col <= 16 and 
                        max_row_offset - min_row_offset == 1):
                        right_side_merges.append((min_col, max_col))
                
                # Apply merge patterns to new rows and center align
                for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                    row_second = base_row + 1
                    for min_col, max_col in right_side_merges:
                        try:
                            if get_column_letter:
                                start_col_letter = get_column_letter(min_col)
                                end_col_letter = get_column_letter(max_col)
                                range_str = f"{start_col_letter}{base_row}:{end_col_letter}{row_second}"
                                safe_merge_cells(ws, range_str)
                                # Center align the merged cell
                                if Alignment:
                                    cell = ws.cell(row=base_row, column=min_col)
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        except Exception:
                            pass
            
            # Center align all right side columns (H-P) in the entire deliverables section
            if Alignment:
                for row in range(deliverables_block_start, deliverables_block_end + 1):
                    for col in range(8, 17):  # Columns H (8) through P (16)
                        cell = ws.cell(row=row, column=col)
                        if cell.alignment:
                            cell.alignment = Alignment(
                                horizontal='center',
                                vertical=cell.alignment.vertical,
                                wrap_text=cell.alignment.wrap_text,
                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                indent=cell.alignment.indent
                            )
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Update deliverables_block_end to include new rows
            deliverables_block_end += rows_to_insert
        
        # Center align all right side columns (H-P) in the entire deliverables section (including original rows)
        if Alignment:
            for row in range(deliverables_block_start, deliverables_block_end + 1):
                for col in range(8, 17):  # Columns H (8) through P (16)
                    cell = ws.cell(row=row, column=col)
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    else:
        # Even if no activewear, center align right side columns (H-P) in deliverables section
        if Alignment:
            for row in range(deliverables_block_start, deliverables_block_end + 1):
                for col in range(8, 17):  # Columns H (8) through P (16)
                    cell = ws.cell(row=row, column=col)
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Update deliverables table counts (column J) for add-on selections
    if column_index_from_string is not None and total_styles_count > 0:
        label_column_idx = column_index_from_string("H")
        target_col_j = column_index_from_string("J")
        deliverable_addon_map = [
            ("WASH/TREATMENT", "H"),
            ("DESIGN", "I"),
            ("SOURCING", "J"),
            ("TREATMENT", "K"),
        ]

        def find_label_row(label_text: str) -> int | None:
            """Locate the row index for a given deliverable label."""
            lowered = label_text.strip().lower()
            partial_match_row = None
            for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
                value = ws.cell(row=scan_row, column=label_column_idx).value
                if not isinstance(value, str):
                    continue
                value_clean = value.strip().lower()
                if value_clean == lowered:
                    return scan_row
                if lowered in value_clean and partial_match_row is None:
                    partial_match_row = scan_row
            return partial_match_row

        for label_text, addon_col_letter in deliverable_addon_map:
            row_idx = find_label_row(label_text)
            if row_idx is None:
                continue
            addon_range = f"{addon_col_letter}10:{addon_col_letter}{last_style_row}"
            cell = ws.cell(row=row_idx, column=target_col_j)
            cell.value = f"=COUNT({addon_range})"
            cell.number_format = "0"

        # Round of Fittings: always 1
        fittings_row = find_label_row("ROUND OF FITTINGS")
        if fittings_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= fittings_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=fittings_row, column=col_d_idx)
            count_cell.value = None
            # Set direct numeric value (not a formula)
            count_cell.value = 1
            count_cell.number_format = "0"
        
        # Round of Revisions: 
        # - 1 if there's ONLY Regular styles (non-Activewear) OR ONLY Activewear styles
        # - 2 if there's BOTH Regular AND Activewear styles
        # Note: ROUND OF REVISIONS is in column B, not column H, so we need to search column B
        label_col_b = column_index_from_string("B")
        revisions_row = None
        revisions_label_lower = "round of revisions"
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=label_col_b).value
            if isinstance(value, str) and revisions_label_lower in value.strip().lower():
                revisions_row = scan_row
                break
        if revisions_row:
            col_d_idx = column_index_from_string("D")
            # Check if this cell is part of a merged range and remember the merge pattern
            was_merged = False
            merge_pattern = None
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= revisions_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    was_merged = True
                    merge_pattern = (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
                    break
            # SIMPLEST APPROACH: Use num_regular and num_activewear already calculated from style_entries
            # These are calculated above from the actual style_entries data - 100% reliable
            # No need to read from Excel cells which might have text/number issues
            
            # Calculate revisions count: 2 if both regular AND activewear exist, 1 otherwise
            # num_regular and num_activewear are calculated at lines 884-885 from style_entries
            revisions_count = 2 if (num_regular > 0 and num_activewear > 0) else 1
            
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=revisions_row, column=col_d_idx)
            count_cell.value = None
            
            # Set the calculated value directly (not a formula) - this will ALWAYS work
            count_cell.value = revisions_count
            count_cell.number_format = "0"
            if Alignment is not None:
                count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Re-merge AFTER setting the value
            if was_merged and merge_pattern and safe_merge_cells:
                min_row, max_row, min_col, max_col = merge_pattern
                if get_column_letter:
                    start_col_letter = get_column_letter(min_col)
                    end_col_letter = get_column_letter(max_col)
                    range_str = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"
                    safe_merge_cells(ws, range_str)
                    # Re-apply alignment after merging
                    if Alignment is not None:
                        merged_cell = ws.cell(row=revisions_row, column=col_d_idx)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        # If activewear exists, handle SECOND SAMPLES and the new rows
        if num_activewear > 0:
            # SECOND SAMPLES: count of activewear styles (replaced FINAL SAMPLES at rows 31-32)
            # Note: count is already set in column D above, but we ensure it's correct here
            second_sample_row = find_label_row("SECOND SAMPLES")
            if second_sample_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_sample_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=second_sample_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = num_activewear
                count_cell.number_format = "0"
            
            # 2nd Round of Fittings: always 1 (already set in column D above)
            second_fittings_row = find_label_row("2ND ROUND OF FITTINGS")
            if second_fittings_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_fittings_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=second_fittings_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = 1
                count_cell.number_format = "0"
            
            # 2nd Round of Revisions: always 1 for Active category (already set in column D above)
            second_revisions_row = find_label_row("2ND ROUND OF REVISIONS")
            if second_revisions_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_revisions_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=second_revisions_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = 1
                count_cell.number_format = "0"
            
            # Final Samples: count is already set in the merge section above, but verify it's correct here
            # (This is a backup check - the count should already be set when merging FINAL SAMPLES)
            # Use row_final_samples_new if available (the newly created FINAL SAMPLES row), otherwise search for it
            # Always prefer row_final_samples_new since it's the newly created row
            final_samples_row_to_use = row_final_samples_new if row_final_samples_new else find_label_row("FINAL SAMPLES")
            if final_samples_row_to_use:
                final_samples_row = final_samples_row_to_use
                # Use direct count instead of formula
                count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                col_d_idx = column_index_from_string("D")
                # Unmerge temporarily to set value and clear any formulas
                was_merged = False
                merge_pattern = None
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        was_merged = True
                        merge_pattern = (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                        break
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=final_samples_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = count_value
                count_cell.number_format = "0"
                # Re-merge if it was merged before
                if was_merged and merge_pattern and safe_merge_cells:
                    min_row, max_row, min_col, max_col = merge_pattern
                    if get_column_letter:
                        start_col_letter = get_column_letter(min_col)
                        end_col_letter = get_column_letter(max_col)
                        range_str = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"
                        safe_merge_cells(ws, range_str)
                    if Alignment:
                        cell_d = ws.cell(row=final_samples_row, column=col_d_idx)
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        # Update PATTERNS and FIRST SAMPLES to have the same value as FINAL SAMPLES (num_styles)
        # This applies to both activewear and non-activewear cases, after find_label_row is defined
        # Note: find_label_row searches column H, but PATTERNS/FIRST SAMPLES are in column B
        # So we need to search column B instead
        patterns_row = None
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=column_index_from_string("B")).value
            if isinstance(value, str) and "pattern" in value.lower():
                patterns_row = scan_row
                break
        
        if patterns_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= patterns_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=patterns_row, column=col_d_idx)
            count_cell.value = None
            # Set direct numeric value (same as FINAL SAMPLES: num_styles)
            count_cell.value = num_styles
            count_cell.number_format = "0"
            # Merge and center (PATTERNS typically spans 2 rows)
            patterns_row_second = patterns_row + 1
            if safe_merge_cells:
                safe_merge_cells(ws, f"D{patterns_row}:D{patterns_row_second}")
            if Alignment:
                count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        first_samples_row = None
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=column_index_from_string("B")).value
            if isinstance(value, str) and "first" in value.lower() and "sample" in value.lower():
                first_samples_row = scan_row
                break
        
        if first_samples_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= first_samples_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=first_samples_row, column=col_d_idx)
            count_cell.value = None
            # Set direct numeric value (same as FINAL SAMPLES: num_styles)
            count_cell.value = num_styles
            count_cell.number_format = "0"
            # Merge and center (FIRST SAMPLES typically spans 2 rows)
            first_samples_row_second = first_samples_row + 1
            if safe_merge_cells:
                safe_merge_cells(ws, f"D{first_samples_row}:D{first_samples_row_second}")
            if Alignment:
                count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
        else:
            # Final Samples: all styles (regular only, no activewear, excluding custom line items)
            final_samples_row = find_label_row("FINAL SAMPLES")
            if final_samples_row:
                # Use direct count instead of formula
                count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=final_samples_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = count_value
                count_cell.number_format = "0"
            
            # Now that FINAL SAMPLES is set, update PATTERNS and FIRST SAMPLES to have the same value
            # Patterns: all styles (regular only, no activewear, excluding custom line items)
            # Set to the same value as FINAL SAMPLES (num_styles) - direct value, not a formula
            patterns_row = find_label_row("PATTERNS")
            if patterns_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= patterns_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=patterns_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (same as FINAL SAMPLES: num_styles)
                count_cell.value = num_styles
                count_cell.number_format = "0"
            
            # First Samples: all styles (regular only, no activewear, excluding custom line items)
            # Set to the same value as FINAL SAMPLES (num_styles) - direct value, not a formula
            first_samples_row = find_label_row("FIRST SAMPLES")
            if first_samples_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= first_samples_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=first_samples_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (same as FINAL SAMPLES: num_styles)
                count_cell.value = num_styles
                count_cell.number_format = "0"
    
    # Move all cells from N39:AA42 up 6 rows to N33:AA36, then apply formatting like ADJUSTMENT on row 31
    if column_index_from_string is not None and get_column_letter is not None:
        # Source range: N39:AA42 (columns N=14 through AA=27, rows 39-42)
        # Target range: N33:AA36 (moved up 6 rows)
        source_start_row = 39
        source_end_row = 42
        target_start_row = 33
        target_end_row = 36
        source_start_col = column_index_from_string("N")  # 14
        source_end_col = column_index_from_string("AA")  # 27
        
        # Copy all cell data from source to target
        cell_data = []
        for row in range(source_start_row, source_end_row + 1):
            row_data = []
            for col in range(source_start_col, source_end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append({
                    "value": cell.value,
                    "number_format": cell.number_format,
                    "font": copy(cell.font) if cell.font else None,
                    "fill": copy(cell.fill) if cell.fill else None,
                    "border": copy(cell.border) if cell.border else None,
                    "alignment": copy(cell.alignment) if cell.alignment else None,
                })
            cell_data.append(row_data)
        
        # Clear source cells and unmerge
        for row in range(source_start_row, source_end_row + 1):
            for col in range(source_start_col, source_end_col + 1):
                # Unmerge if needed
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        try:
                            min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                            ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                        except Exception:
                            pass
                safe_set_cell_value(ws, f"{get_column_letter(col)}{row}", None)
        
        # Write data to target cells
        for row_offset, row_data in enumerate(cell_data):
            target_row = target_start_row + row_offset
            for col_offset, cell_info in enumerate(row_data):
                target_col = source_start_col + col_offset
                col_letter = get_column_letter(target_col)
                safe_set_cell_value(ws, f"{col_letter}{target_row}", cell_info["value"])
                cell = ws.cell(row=target_row, column=target_col)
                if cell_info.get("number_format"):
                    cell.number_format = cell_info["number_format"]
                if cell_info.get("font"):
                    cell.font = cell_info["font"]
                if cell_info.get("fill"):
                    cell.fill = cell_info["fill"]
                if cell_info.get("border"):
                    cell.border = cell_info["border"]
        
        # Apply formatting like ADJUSTMENT on row 31: merge across 2 rows and center align
        # Each row pair should be merged: N33:N34, N35:N36, etc.
        for row in range(target_start_row, target_end_row + 1, 2):  # Process pairs: 33-34, 35-36
            row_second = row + 1
            if row_second > target_end_row:
                break
            for col in range(source_start_col, source_end_col + 1):
                col_letter = get_column_letter(col)
                # Merge the cell pair vertically (like ADJUSTMENT)
                if safe_merge_cells:
                    safe_merge_cells(ws, f"{col_letter}{row}:{col_letter}{row_second}")
                # Center align
                if Alignment is not None:
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # Apply borders like ADJUSTMENT
                apply_full_border_pair(ws, col, row, row_second)

        # Additional required merges/centering for the moved block
        merge_targets = [
            "N33:P34", "N35:P36",
            "R33:S34", "R35:S36",
            "T27:X36",
            "Y33:AA34", "Y35:AA36",
        ]
        for rng in merge_targets:
            if safe_merge_cells:
                safe_merge_cells(ws, rng)
            if Alignment is not None:
                try:
                    start_col_letter = ''.join([c for c in rng.split(':')[0] if c.isalpha()])
                    start_row_num = int(''.join([c for c in rng.split(':')[0] if c.isdigit()]))
                    start_col_idx = column_index_from_string(start_col_letter)
                    cell = ws.cell(row=start_row_num, column=start_col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                except Exception:
                    pass

        # Ensure source area N39:AA42 is fully cleared (values and borders)
        for row in range(source_start_row, source_end_row + 1):
            for col in range(source_start_col, source_end_col + 1):
                safe_set_cell_value(ws, f"{get_column_letter(col)}{row}", None)
                try:
                    ws.cell(row=row, column=col).border = Border()
                except Exception:
                    pass

    # Totals section - dynamically calculate totals row and range based on number of styles
    # For 5 or fewer styles: totals at row 20 (original position), but if custom items exist, they push it down
    # For more than 5 styles: totals row shifts down by (total_styles_count - 5) * 2 rows
    if total_styles_count > 0:
        first_style_row = dynamic_row_indices[0] if dynamic_row_indices else 10
        if num_styles <= 5:
            # Original totals row position is 20, but if we inserted rows for custom items, it moved
            if num_custom_styles > 0:
                # Custom items were inserted before row 20, so totals row moved down
                totals_row = 20 + (num_custom_styles * 2)
            else:
                totals_row = 20  # Original totals row position
        else:
            # Totals row shifts down by the number of rows we inserted
            totals_row = 20 + (num_styles - 5) * 2
            if num_custom_styles > 0:
                # Add space for custom items too
                totals_row += num_custom_styles * 2
        
        # Unmerge any merged cells in the totals row to avoid issues
        merged_ranges_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row):
                merged_ranges_to_unmerge.append(merged_range)
        
        for merged_range in merged_ranges_to_unmerge:
            try:
                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
            except Exception:
                pass
        
        # Set totals row labels - merge and center like "DEVELOPMENT PACKAGE"
        # "TOTAL DEVELOPMENT" spans B:F (like "DEVELOPMENT PACKAGE" at B8:F8)
        # Set value first, then merge, then apply formatting (safe_set_cell_value unmerges, so we set directly)
        cell_b_totals = ws.cell(row=totals_row, column=2)
        cell_b_totals.value = "TOTAL DEVELOPMENT"
        if safe_merge_cells(ws, f"B{totals_row}:F{totals_row}"):
            # Re-get cell after merging to ensure we have the top-left cell of merged range
            cell_b_totals = ws.cell(row=totals_row, column=2)
        if Font is not None:
            cell_b_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_b_totals.alignment = Alignment(horizontal="center", vertical="center")
        
        # "TOTAL OPTIONAL ADD-ONS" spans H:L (columns H, I, J, K, L)
        # Set value first, then merge, then apply formatting
        cell_h_totals = ws.cell(row=totals_row, column=8)
        cell_h_totals.value = "TOTAL OPTIONAL ADD-ONS"
        if safe_merge_cells(ws, f"H{totals_row}:L{totals_row}"):
            # Re-get cell after merging to ensure we have the top-left cell of merged range
            cell_h_totals = ws.cell(row=totals_row, column=8)
        if Font is not None:
            cell_h_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_h_totals.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set totals formulas - sum all style rows (dynamic based on actual style rows)
        safe_set_cell_value(ws, f"F{totals_row}", f"=SUM(F10:F{last_style_row})")
        cell_f_totals = ws.cell(row=totals_row, column=6)
        cell_f_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_f_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_f_totals.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #709171 to TOTAL DEVELOPMENT
        if PatternFill is not None:
            cell_f_totals.fill = PatternFill(start_color="709171", end_color="709171", fill_type="solid")
        
        safe_set_cell_value(ws, f"L{totals_row}", f"=SUM(L10:L{last_style_row})")
        cell_l_totals = ws.cell(row=totals_row, column=12)
        cell_l_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_l_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_l_totals.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #f0cfbb to TOTAL OPTIONAL ADD-ONS
        if PatternFill is not None:
            cell_l_totals.fill = PatternFill(start_color="F0CFBB", end_color="F0CFBB", fill_type="solid")

        # Discount row (fixed location: Y16:Z17 for label, AA16:AA17 for value)
        # Only show discount if discount_percentage > 0, otherwise clear the cells
        discount_row = 16
        discount_decimal = discount_percentage / 100.0 if discount_percentage else 0.0
        
        if discount_percentage > 0:
            # Set discount label in Y16:Z17 (merged across 2 rows, 2 columns)
            safe_set_cell_value(ws, "Y16", f"DISCOUNT ({discount_percentage:.0f}%)")
            safe_merge_cells(ws, "Y16:Z17")
            cell_y16 = ws.cell(row=16, column=25)  # Y = column 25
            if Font is not None:
                cell_y16.font = Font(name="Arial", size=20, bold=True, color=cell_y16.font.color if cell_y16.font else None)
            if Alignment is not None:
                cell_y16.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set discount value in AA16:AA17 (merged across 2 rows, 1 column)
            safe_set_cell_value(ws, "AA16", f"=SUM(AA10:AA15)*{discount_decimal}")
            safe_merge_cells(ws, "AA16:AA17")
            cell_aa16 = ws.cell(row=16, column=27)  # AA = column 27
            cell_aa16.number_format = '$#,##0'
            if Font is not None:
                cell_aa16.font = Font(name="Arial", size=20, bold=True, color=cell_aa16.font.color if cell_aa16.font else None)
            if Alignment is not None:
                cell_aa16.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Clear discount row if discount is 0% - remove/clean Y16:Z17 and AA16:AA17
            safe_set_cell_value(ws, "Y16", None)
            safe_set_cell_value(ws, "Z16", None)
            safe_set_cell_value(ws, "Y17", None)
            safe_set_cell_value(ws, "Z17", None)
            safe_set_cell_value(ws, "AA16", None)
            safe_set_cell_value(ws, "AA17", None)
            
            # Merge and center Y16:Z17 and AA16:AA17 after clearing
            if safe_merge_cells:
                safe_merge_cells(ws, "Y16:Z17")
                safe_merge_cells(ws, "AA16:AA17")
                # Center align the merged cells
                if Alignment is not None:
                    merged_cell_y = ws.cell(row=16, column=25)
                    merged_cell_y.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    merged_cell_aa = ws.cell(row=16, column=27)
                    merged_cell_aa.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Ensure N12 and P12 have inferior (bottom) borders
            if Border is not None and Side is not None:
                bottom_side = Side(style="thin")
                # Get existing borders for N12 and P12, then add bottom border
                row_12 = discount_row - 2  # Row 12 (discount_row is 14)
                cell_n12 = ws.cell(row=row_12, column=SUMMARY_LABEL_COL)
                cell_p12 = ws.cell(row=row_12, column=SUMMARY_VALUE_COL)
                
                # Preserve existing borders and add bottom border
                existing_n12_border = cell_n12.border if cell_n12.border else Border()
                existing_p12_border = cell_p12.border if cell_p12.border else Border()
                
                cell_n12.border = Border(
                    left=existing_n12_border.left,
                    right=existing_n12_border.right,
                    top=existing_n12_border.top,
                    bottom=bottom_side
                )
                cell_p12.border = Border(
                    left=existing_p12_border.left,
                    right=existing_p12_border.right,
                    top=existing_p12_border.top,
                    bottom=bottom_side
                )
        
        # Clear N23 if it contains discount percentage (remove duplicate)
        # N23 should remain empty or contain "TOTAL DUE AT SIGNING" if that's what the template has
        cell_n23_check = ws.cell(row=23, column=SUMMARY_LABEL_COL)
        if cell_n23_check.value and "%" in str(cell_n23_check.value):
            # Clear the duplicate discount percentage from N23
            safe_set_cell_value(ws, "N23", None)
        
        # Make all cells in totals row bold
        if Font is not None:
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
                else:
                    cell.font = Font(bold=True)
        
        # Apply Arial 20 font to totals row if it's a new row (num_styles > 5)
        if num_styles > 5:
            for col_idx in [2, 6, 8, 12, 14, 16]:  # Columns B, F, H, L, N, P
                cell = ws.cell(row=totals_row, column=col_idx)
                apply_arial_20_font(cell)
                # Ensure bold is maintained
                if Font is not None:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
        else:
            # For <=5 styles (totals_row==20), ensure Arial size 20 bold for key cells
            for col_idx in [2, 6, 8, 12, 14, 16]:
                cell = ws.cell(row=totals_row, column=col_idx)
                if Font is not None:
                    cell.font = Font(name="Arial", size=20, bold=True, color=cell.font.color if cell.font else None)
        # Ensure P (column 16) is also Arial size 20 for <=5 styles
        if num_styles <= 5:
            cell_p_totals = ws.cell(row=totals_row, column=16)
            if Font is not None:
                cell_p_totals.font = Font(name="Arial", size=20, bold=True, color=cell_p_totals.font.color if cell_p_totals.font else None)
        # Apply inferior (bottom) border to entire totals row
        if Border is not None and Side is not None:
            bottom_side = Side(style="thin")
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                # Skip columns A, G, M for bottom borders
                if col_idx in [1, 7, 13]:
                    new_bottom = cell.border.bottom
                else:
                    new_bottom = bottom_side
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=new_bottom,
                )

        # Apply full borders (top/bottom/left/right) to key totals cells (B, F, H, L, N, P)
        if Border is not None and Side is not None:
            full_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for col_idx in [2, 6, 8, 12, 14, 16]:
                cell = ws.cell(row=totals_row, column=col_idx)
                cell.border = full_border
        
    else:
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "L20", None)
    
    # Write A La Carte items based on template analysis
    # Correct mapping: N=Style Number (#), O=Style Name (STYLE), P=INTAKE, Q=1ST PATTERN, R=1ST SAMPLE, S=FITTING, T=ADJUSTMENT, U=FINAL SAMPLES, V=DUPLICATES
    # W has TOTAL formula =SUM(N10:V10), so we write to N-V (not W)
    if a_la_carte_items and column_index_from_string is not None:
        start_row = 10
        col_n = column_index_from_string("N")  # 14 - Style Number (#)
        col_o = column_index_from_string("O")  # 15 - Style Name (STYLE)
        col_p = column_index_from_string("P")  # 16 - INTAKE
        col_q = column_index_from_string("Q")  # 17 - 1ST PATTERN
        col_r = column_index_from_string("R")  # 18 - 1ST SAMPLE
        col_s = column_index_from_string("S")  # 19 - FITTING
        col_t = column_index_from_string("T")  # 20 - ADJUSTMENT
        col_u = column_index_from_string("U")  # 21 - FINAL SAMPLES
        col_v = column_index_from_string("V")  # 22 - DUPLICATES
        
        # First, set headers in row 9 if they don't exist
        # N9 = #, O9 = STYLE, P9 = INTAKE, Q9 = 1ST PATTERN, etc.
        # Use safe_get_cell_value to avoid MergedCell read-only errors
        n9_value = safe_get_cell_value(ws, 9, col_n)
        if n9_value is None or str(n9_value).strip() == "":
            safe_set_cell_value(ws, f"{get_column_letter(col_n)}9", "#")
        
        o9_value = safe_get_cell_value(ws, 9, col_o)
        if o9_value is None or str(o9_value).strip() == "":
            safe_set_cell_value(ws, f"{get_column_letter(col_o)}9", "STYLE")
        
        p9_value = safe_get_cell_value(ws, 9, col_p)
        if p9_value is None or "INTAKE" not in str(p9_value).upper():
            safe_set_cell_value(ws, f"{get_column_letter(col_p)}9", "INTAKE")
        
        # Set up A La Carte header row (row 20) - merge and align left B20:E20 and H20:K20
        # B20:E20 for "A LA CARTE PACKAGE" header (left-aligned)
        # H20:K20 for another header (left-aligned)
        header_row = 20
        if safe_merge_cells(ws, f"B{header_row}:E{header_row}"):
            cell_b20 = ws.cell(row=header_row, column=2)
            if Alignment is not None:
                cell_b20.alignment = Alignment(horizontal="left", vertical="center")
        if safe_merge_cells(ws, f"H{header_row}:K{header_row}"):
            cell_h20 = ws.cell(row=header_row, column=8)
            if Alignment is not None:
                cell_h20.alignment = Alignment(horizontal="left", vertical="center")
        
        # Merge, center, and align left N20:V20
        if column_index_from_string is not None and get_column_letter is not None:
            col_n = column_index_from_string("N")  # 14
            col_v = column_index_from_string("V")  # 22
            if safe_merge_cells:
                safe_merge_cells(ws, f"N{header_row}:V{header_row}")
            if Alignment is not None:
                cell_n20 = ws.cell(row=header_row, column=col_n)
                cell_n20.alignment = Alignment(horizontal="left", vertical="center")
            # Merge and center Y20:Z20
            col_y = column_index_from_string("Y")
            if safe_merge_cells:
                safe_merge_cells(ws, f"Y{header_row}:Z{header_row}")
            if Alignment is not None:
                cell_y20 = ws.cell(row=header_row, column=col_y)
                cell_y20.alignment = Alignment(horizontal="center", vertical="center")

        # Clean P10 and P12 (no summary formulas here because A La Carte occupies this section)
        safe_set_cell_value(ws, "P10", None)
        safe_set_cell_value(ws, "P12", None)
        
        for i, item in enumerate(a_la_carte_items):
            # Each entry uses two rows, similar to style rows (10-11, 12-13, ...)
            row = start_row + (i * 2)
            row_second = row + 1
            
            # Style Number - column N
            safe_set_cell_value(ws, f"{get_column_letter(col_n)}{row}", item.get("style_number", ""))
            cell_n = ws.cell(row=row, column=col_n)
            cell_n.number_format = "0"  # Integer format
            apply_full_border_pair(ws, col_n, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_n)}{row}:{get_column_letter(col_n)}{row_second}")
            
            # Style Name - column O
            safe_set_cell_value(ws, f"{get_column_letter(col_o)}{row}", item.get("name", ""))
            cell_o = ws.cell(row=row, column=col_o)
            apply_full_border_pair(ws, col_o, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_o)}{row}:{get_column_letter(col_o)}{row_second}")
            
            # INTAKE - column P (hours * $190)
            # Write INTAKE to all A La Carte rows (summary formulas will overwrite P10, P12, P14 later)
            intake_hours = item.get("intake_session", 0)
            intake_price = int(intake_hours * A_LA_CARTE_RATE_STANDARD)
            safe_set_cell_value(ws, f"{get_column_letter(col_p)}{row}", intake_price)
            cell_p = ws.cell(row=row, column=col_p)
            cell_p.number_format = '$#,##0'  # Currency format
            apply_full_border_pair(ws, col_p, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_p)}{row}:{get_column_letter(col_p)}{row_second}")
            
            # 1ST PATTERN - column Q (hours * $190)
            pattern_hours = item.get("first_pattern", 0)
            pattern_price = int(pattern_hours * A_LA_CARTE_RATE_STANDARD)
            safe_set_cell_value(ws, f"{get_column_letter(col_q)}{row}", pattern_price)
            cell_q = ws.cell(row=row, column=col_q)
            cell_q.number_format = '$#,##0'  # Currency format
            apply_full_border_pair(ws, col_q, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_q)}{row}:{get_column_letter(col_q)}{row_second}")
            
            # 1ST SAMPLE - column R (hours * $190)
            sample_hours = item.get("first_sample", 0)
            sample_price = int(sample_hours * A_LA_CARTE_RATE_STANDARD)
            safe_set_cell_value(ws, f"{get_column_letter(col_r)}{row}", sample_price)
            cell_r = ws.cell(row=row, column=col_r)
            cell_r.number_format = '$#,##0'  # Currency format
            apply_full_border_pair(ws, col_r, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_r)}{row}:{get_column_letter(col_r)}{row_second}")
            
            # FITTING - column S (hours * $190)
            fitting_hours = item.get("fitting", 0)
            fitting_price = int(fitting_hours * A_LA_CARTE_RATE_STANDARD)
            safe_set_cell_value(ws, f"{get_column_letter(col_s)}{row}", fitting_price)
            cell_s = ws.cell(row=row, column=col_s)
            cell_s.number_format = '$#,##0'  # Currency format
            apply_full_border_pair(ws, col_s, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_s)}{row}:{get_column_letter(col_s)}{row_second}")
            
            # ADJUSTMENT - column T (hours * $190)
            adjustment_hours = item.get("adjustment", 0)
            adjustment_price = int(adjustment_hours * A_LA_CARTE_RATE_STANDARD)
            safe_set_cell_value(ws, f"{get_column_letter(col_t)}{row}", adjustment_price)
            cell_t = ws.cell(row=row, column=col_t)
            cell_t.number_format = '$#,##0'  # Currency format
            apply_full_border_pair(ws, col_t, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_t)}{row}:{get_column_letter(col_t)}{row_second}")
            
            # FINAL SAMPLES - column U (hours * $90)
            final_samples_hours = item.get("final_samples", 0)
            final_samples_price = int(final_samples_hours * A_LA_CARTE_RATE_SAMPLES)
            safe_set_cell_value(ws, f"{get_column_letter(col_u)}{row}", final_samples_price)
            cell_u = ws.cell(row=row, column=col_u)
            cell_u.number_format = '$#,##0'  # Currency format
            apply_full_border_pair(ws, col_u, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_u)}{row}:{get_column_letter(col_u)}{row_second}")
            
            # DUPLICATES - column V (hours * $90)
            duplicates_hours = item.get("duplicates", 0)
            duplicates_price = int(duplicates_hours * A_LA_CARTE_RATE_SAMPLES)
            safe_set_cell_value(ws, f"{get_column_letter(col_v)}{row}", duplicates_price)
            cell_v = ws.cell(row=row, column=col_v)
            cell_v.number_format = '$#,##0'  # Currency format
            apply_full_border_pair(ws, col_v, row, row_second)
            safe_merge_cells(ws, f"{get_column_letter(col_v)}{row}:{get_column_letter(col_v)}{row_second}")
            
            # Apply center alignment to all numeric cells
            if Alignment is not None:
                # Align all cells including INTAKE (summary formulas will overwrite P10, P12, P14 later)
                cells_to_align = [cell_n, cell_o, cell_p, cell_q, cell_r, cell_s, cell_t, cell_u, cell_v]
                for cell in cells_to_align:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # Style name is already centered
            
            # Note: Total is calculated via formula =SUM(N10:V10) in column W, we don't write it directly
        
        # Update Q23 with total INTAKE hours from A La Carte items
        # Q23 is column Q (17), row 23 - should show total INTAKE hours, merged with Q24
        if len(a_la_carte_items) > 0:
            # INTAKE SESSION total (already correct)
            total_intake_hours = sum(item.get("intake_session", 0) for item in a_la_carte_items)
            col_q_idx = column_index_from_string("Q")
            safe_set_cell_value(ws, "Q23", total_intake_hours)
            cell_q23 = ws.cell(row=23, column=col_q_idx)
            cell_q23.number_format = "0"  # Integer format
            if safe_merge_cells:
                safe_merge_cells(ws, "Q23:Q24")
            if Alignment is not None:
                cell_q23 = ws.cell(row=23, column=col_q_idx)
                cell_q23.alignment = Alignment(horizontal="center", vertical="center")

            # Mirror the same direct totals for the remaining A La Carte columns (Q–V)
            ala_totals_map = [
                ("Q", "intake_session"),
                ("Q", "first_pattern"),
                ("Q", "first_sample"),
                ("Q", "fitting"),
                ("Q", "adjustment"),
                ("Q", "final_samples"),
                ("Q", "duplicates"),
            ]
            start_row_totals = 23
            for idx, (_, key) in enumerate(ala_totals_map):
                target_row = start_row_totals + (idx * 2)
                col_letter = "Q"
                col_idx = column_index_from_string(col_letter)
                total_hours = sum(item.get(key, 0) for item in a_la_carte_items)
                safe_set_cell_value(ws, f"{col_letter}{target_row}", total_hours)
                cell_total = ws.cell(row=target_row, column=col_idx)
                cell_total.number_format = "0"
                if safe_merge_cells:
                    safe_merge_cells(ws, f"{col_letter}{target_row}:{col_letter}{target_row + 1}")
                if Alignment is not None:
                    cell_total = ws.cell(row=target_row, column=col_idx)
                    cell_total.alignment = Alignment(horizontal="center", vertical="center")
        
        # Clear extra A La Carte rows beyond the actual items (template has 5 rows, we may only use 2)
        # After clearing, merge and center the cells to maintain formatting
        num_a_la_carte = len(a_la_carte_items)
        if num_a_la_carte < 5:
            # Clear rows for items we don't have (each item uses 2 rows)
            for i in range(num_a_la_carte, 5):
                clear_row = start_row + (i * 2)
                clear_row_second = clear_row + 1
                # Clear all columns N through V
                for col_letter in ["N", "O", "P", "Q", "R", "S", "T", "U", "V"]:
                    col_idx = column_index_from_string(col_letter)
                    # Unmerge if needed
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= clear_row_second and merged_range.max_row >= clear_row and
                            merged_range.min_col <= col_idx <= merged_range.max_col):
                            try:
                                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                            except Exception:
                                pass
                    # Clear the cells
                    safe_set_cell_value(ws, f"{col_letter}{clear_row}", None)
                    safe_set_cell_value(ws, f"{col_letter}{clear_row_second}", None)
                    # Merge and center the cleared cells (maintain formatting like other A La Carte rows)
                    if safe_merge_cells:
                        safe_merge_cells(ws, f"{col_letter}{clear_row}:{col_letter}{clear_row_second}")
                    if Alignment is not None:
                        cell = ws.cell(row=clear_row, column=col_idx)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Update summary formulas AFTER A La Carte items (so they overwrite any A La Carte values in P10, P12, P14)
    # This must be done after A La Carte to ensure summary formulas are preserved
    if total_styles_count > 0:
        # Find and update "TOTAL DUE AT SIGNING" formula
        # The label is in column N (14) and the formula is in column P (16) of the totals row
        cell_n_totals = ws.cell(row=totals_row, column=14)  # Column N
        if cell_n_totals.value and "TOTAL DUE AT SIGNING" in str(cell_n_totals.value).upper():
            # Update the formula in column P to reference the dynamic totals row
            # If discount is 0, don't subtract discount; otherwise subtract AA16 (discount value)
            if discount_percentage > 0:
                safe_set_cell_value(ws, f"P{totals_row}", f"=SUM(P10:P13)-AA16")
            else:
                safe_set_cell_value(ws, f"P{totals_row}", f"=SUM(P10:P13)")
            cell_p_totals = ws.cell(row=totals_row, column=16)  # Re-get after safe_set
            cell_p_totals.number_format = '$#,##0'  # Currency format
            # Apply font size 20 and bold to TOTAL DUE AT SIGNING formula
            if Font is not None:
                cell_p_totals.font = Font(name="Arial", size=20, bold=True, color=cell_p_totals.font.color if cell_p_totals.font else None)
            if Alignment is not None:
                cell_p_totals.alignment = Alignment(horizontal="center", vertical="center")
            # Apply cell color #ffff00 to TOTAL DUE AT SIGNING
            if PatternFill is not None:
                cell_p_totals.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Also make the label bold with size 20
            if Font is not None:
                cell_n_totals.font = Font(name="Arial", size=20, bold=True, color=cell_n_totals.font.color if cell_n_totals.font else None)
        
        # Also update any other formulas in column P that reference F20 or L20 statically
        # Check a few rows around the totals row (but skip rows 10, 12, 14 which we already handled)
        for check_row in range(totals_row - 2, totals_row + 3):
            if check_row in [10, 12, 14]:  # Skip rows we already handled
                continue
            cell_p = ws.cell(row=check_row, column=16)  # Column P
            if cell_p.value and isinstance(cell_p.value, str) and cell_p.value.startswith("="):
                # Check if it references F20 or L20 (static references)
                if "F20" in cell_p.value or "L20" in cell_p.value:
                    # Replace with dynamic references
                    formula = cell_p.value.replace("F20", f"F{totals_row}").replace("L20", f"L{totals_row}")
                    # Use safe_set_cell_value to handle merged cells
                    safe_set_cell_value(ws, f"P{check_row}", formula)
                    cell_p = ws.cell(row=check_row, column=16)  # Re-get cell after safe_set
                    if Font is not None:
                        existing_font = cell_p.font
                        cell_p.font = Font(
                            name=existing_font.name if existing_font and existing_font.name else "Arial",
                            size=existing_font.size if existing_font and existing_font.size else 20,
                            bold=True,
                            color=existing_font.color if existing_font else None
                        )
                    if Alignment is not None:
                        cell_p.alignment = Alignment(horizontal="center", vertical="center")
    
    return total_development, total_optional


def build_workbook_bytes(
    *,
    client_name: str,
    client_email: str,
    representative: str,
    phone_number: str,
    style_entries: list[dict],
    custom_styles: list[dict],
    a_la_carte_items: list[dict],
    discount_percentage: float,
    notes: list[str] = None,
) -> tuple[bytes, float, float]:
    """Load the template, update it, and return bytes plus totals."""
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. Please add it to the environment first."
        )

    template_path = get_template_path()
    wb = load_workbook(template_path)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Worksheet '{TARGET_SHEET}' is missing from the template."
        )
    ws = wb[TARGET_SHEET]

    # Rename worksheet to "Workbook"
    ws.title = "Workbook"

    # Remove all other worksheets, keeping only "Workbook"
    sheets_to_remove = [name for name in wb.sheetnames if name != "Workbook"]
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])

    # Get reference to the Workbook worksheet after cleanup
    ws = wb["Workbook"]

    update_header_labels(ws, client_name)
    total_dev, total_optional = apply_development_package(
        ws,
        client_name=client_name,
        client_email=client_email,
        representative=representative,
        phone_number=phone_number,
        style_entries=style_entries,
        custom_styles=custom_styles,
        a_la_carte_items=a_la_carte_items,
        discount_percentage=discount_percentage,
    )
    
    # Write notes to column N starting below "PROJECT NOTES"
    # Find "PROJECT NOTES" dynamically (it moves when more styles are added)
    if notes:
        project_notes_row = None
        project_notes_col = None
        # Search for "PROJECT NOTES" in column N (column 14) and column B (column 2)
        search_columns = []
        if column_index_from_string is not None:
            search_columns.append(column_index_from_string("Y"))  # Preferred location
        search_columns.extend([SUMMARY_LABEL_COL, 2])  # Fallbacks: N and B

        for col in search_columns:
            for search_row in range(20, 60):  # Search a bit further down in case rows shifted
                cell_value = ws.cell(row=search_row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if "PROJECT" in cell_value.upper() and "NOTES" in cell_value.upper():
                        project_notes_row = search_row
                        project_notes_col = col
                        break
            if project_notes_row:
                break
        
        if project_notes_row and project_notes_col:
            # Place notes starting one row below the header; ensure we start below Y22 as requested
            notes_start_row = max(project_notes_row + 1, 23)
            note_index = 0
            for note in notes:
                if note and note.strip():
                    # Place notes at every other row (skip one row between notes)
                    cell_row = notes_start_row + (note_index * 2)
                    cell_row_second = cell_row + 1
                    col_letter = get_column_letter(project_notes_col) if get_column_letter else None
                    note_text = note.strip().upper()
                    target_col = project_notes_col or SUMMARY_LABEL_COL
                    if col_letter:
                        safe_set_cell_value(ws, f"{col_letter}{cell_row}", note_text)
                        # Merge Y:AA across 2 rows (Y23:AA24, Y25:AA26, etc.)
                        end_col = target_col + 2
                        try:
                            end_col_letter = get_column_letter(end_col)
                            # Merge horizontally first (Y23:AA23), then merge vertically (Y23:AA24)
                            safe_merge_cells(ws, f"{col_letter}{cell_row}:{end_col_letter}{cell_row_second}")
                        except Exception:
                            pass
                    else:
                        safe_set_cell_value(ws, f"N{cell_row}", note_text)
                    cell = ws.cell(row=cell_row, column=target_col)
                    if Font is not None:
                        cell.font = Font(name="Arial", size=20)
                    if Alignment is not None:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    note_index += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), total_dev, total_optional


@st.cache_data(ttl=300)  # Cache for 5 minutes
def get_sales_records():
    """Get sales records from monday.com for dropdown selection."""
    try:
        from database_utils import get_sales_data
        
        sales_data = get_sales_data()
        items = sales_data.get("data", {}).get("boards", [{}])[0].get("items_page", {}).get("items", [])
        
        # Return list of (item_id, item_name) tuples
        records = [(item.get("id"), item.get("name", "")) for item in items if item.get("name")]
        return sorted(records, key=lambda x: x[1])  # Sort by name
    except Exception as e:
        st.warning(f"Could not load sales records: {e}")
        return []


def update_monday_item_workbook_url(item_id: str, workbook_url: str) -> bool:
    """Update a monday.com item with the workbook URL in a 'Workbook Link' field."""
    try:
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            st.error("Monday.com API token not found in secrets.")
            return False
        
        # First, we need to find the column ID for "Workbook Link" or create it
        # For now, we'll use a URL column type. The column ID needs to be found or created in monday.com
        # This is a placeholder - the actual column ID needs to be configured
        
        # Query to get board columns to find the "Workbook Link" column
        query = f"""
        query {{
            items(ids: [{item_id}]) {{
                board {{
                    id
                    columns {{
                        id
                        title
                        type
                    }}
                }}
            }}
        }}
        """
        
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            st.error(f"Error fetching monday.com columns: {data['errors']}")
            return False
        
        items = data.get("data", {}).get("items", [])
        if not items:
            st.error("Item not found in monday.com")
            return False
        
        board = items[0].get("board", {})
        board_id = board.get("id")
        columns = board.get("columns", [])
        
        if not board_id:
            st.error("Could not determine board ID from monday.com item")
            return False
        
        # Find "Workbook Link" or "Workbook URL Link" column
        workbook_column = None
        for col in columns:
            title_lower = col.get("title", "").lower()
            if "workbook" in title_lower and ("link" in title_lower or "url" in title_lower):
                workbook_column = col
                break
        
        if not workbook_column:
            st.warning("⚠️ 'Workbook Link' column not found in monday.com. Please create a URL column named 'Workbook Link' or 'Workbook URL Link' in the Sales board.")
            return False
        
        column_id = workbook_column.get("id")
        column_type = workbook_column.get("type")
        
        # Update the item with the workbook URL
        # For URL columns, the value format is: {"url": "https://...", "text": "Link Text"}
        if column_type == "link":
            mutation = f"""
            mutation {{
                change_column_value(
                    board_id: {board_id},
                    item_id: {item_id},
                    column_id: "{column_id}",
                    value: "{{\\"url\\": \\"{workbook_url}\\", \\"text\\": \\"View Workbook\\"}}"
                ) {{
                    id
                }}
            }}
            """
        else:
            # For text columns, just use the URL as text
            mutation = f"""
            mutation {{
                change_column_value(
                    board_id: {board_id},
                    item_id: {item_id},
                    column_id: "{column_id}",
                    value: "{workbook_url}"
                ) {{
                    id
                }}
            }}
            """
        
        response = requests.post(url, json={"query": mutation}, headers=headers, timeout=30)
        result = response.json()
        
        if "errors" in result:
            st.error(f"Error updating monday.com: {result['errors']}")
            return False
        
        return True
        
    except Exception as e:
        st.error(f"Failed to update monday.com: {e}")
        return False


def extract_spreadsheet_id_from_url(url: str) -> str | None:
    """Extract spreadsheet ID from Google Sheets or Google Drive URL.
    
    Args:
        url: Google Sheets URL (e.g., https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit)
             or Google Drive URL (e.g., https://drive.google.com/file/d/FILE_ID/view)
    
    Returns:
        Spreadsheet/file ID or None if not found
    """
    import re
    
    # Pattern for Google Sheets URL: /spreadsheets/d/{id}/
    sheets_pattern = r'/spreadsheets/d/([a-zA-Z0-9-_]+)'
    match = re.search(sheets_pattern, url)
    if match:
        return match.group(1)
    
    # Pattern for Google Drive URL: /file/d/{id}/
    drive_pattern = r'/file/d/([a-zA-Z0-9-_]+)'
    match = re.search(drive_pattern, url)
    if match:
        return match.group(1)
    
    return None


def export_google_sheet_as_pdf(sheet_url: str) -> bytes:
    """Export Google Sheet as PDF using Google Sheets export API.
    
    Args:
        sheet_url: The Google Sheet URL
        
    Returns:
        PDF file as bytes
        
    Raises:
        RuntimeError: If spreadsheet ID cannot be extracted or export fails
    """
    # Extract spreadsheet ID from URL
    spreadsheet_id = extract_spreadsheet_id_from_url(sheet_url)
    if not spreadsheet_id:
        raise RuntimeError(f"Could not extract spreadsheet ID from URL: {sheet_url}")
    
    # Build export URL with PDF format
    # Using landscape orientation and removing gridlines for better readability
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=pdf&portrait=false&gridlines=false"
    
    # Get Google credentials for authentication
    if not GOOGLE_API_AVAILABLE:
        raise RuntimeError("Google API libraries not available. Please install google-auth and google-api-python-client.")
    
    try:
        info = st.secrets.get("google_service_account")
        if not info:
            raise RuntimeError("Google service account credentials not found in secrets")
        
        credentials = SACredentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        
        # Refresh credentials to get access token
        if not credentials.valid:
            credentials.refresh(GoogleRequest())
        
        # Fetch the PDF using authenticated request
        import urllib.request
        
        req = urllib.request.Request(export_url)
        req.add_header('Authorization', f'Bearer {credentials.token}')
        
        with urllib.request.urlopen(req) as response:
            pdf_bytes = response.read()
            return pdf_bytes
            
    except Exception as e:
        raise RuntimeError(f"Failed to export Google Sheet as PDF: {e}")


def get_board_id_from_item(item_id: str) -> str | None:
    """Get board ID from a Monday.com item ID.
    
    Args:
        item_id: The Monday.com item ID
        
    Returns:
        Board ID or None if not found
    """
    try:
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            return None
        
        query = f"""
        query {{
            items(ids: [{item_id}]) {{
                board {{
                    id
                }}
            }}
        }}
        """
        
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            return None
        
        items = data.get("data", {}).get("items", [])
        if not items:
            return None
        
        board = items[0].get("board", {})
        return board.get("id")
        
    except Exception:
        return None


def upload_file_to_monday_item(item_id: str, board_id: str, file_bytes: bytes, filename: str) -> bool:
    """Upload a file to a monday.com item using the GraphQL file upload API."""
    try:
        import json
        
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            st.error("Monday.com API token not found in secrets.")
            return False
        
        # Step 1: Get the files column ID for this board
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        # Query to find the files column
        query = f"""
        query {{
            boards(ids: [{board_id}]) {{
                columns {{
                    id
                    title
                    type
                }}
            }}
        }}
        """
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            st.error(f"Error fetching board columns: {data['errors']}")
            return False
        
        boards = data.get("data", {}).get("boards", [])
        if not boards:
            st.error("Board not found")
            return False
        
        columns = boards[0].get("columns", [])
        
        # Find files column (type is "file" or title contains "file")
        files_column = None
        for col in columns:
            if col.get("type") == "file" or "file" in col.get("title", "").lower():
                files_column = col
                break
        
        if not files_column:
            st.warning("⚠️ No files column found on the board. File upload requires a files column.")
            return False
        
        column_id = files_column.get("id")
        
        # Step 2: Upload file using GraphQL file upload API
        # Use the /v2/file endpoint for file uploads with multipart/form-data
        file_url = "https://api.monday.com/v2/file"
        
        # GraphQL mutation for adding file to column
        mutation = """
        mutation addFile($file: File!) {
            add_file_to_column(
                file: $file,
                item_id: %s,
                column_id: "%s"
            ) {
                id
            }
        }
        """ % (item_id, column_id)
        
        # Prepare multipart form data for GraphQL file upload
        # The format is: query, variables, map, and file
        variables = {
            "file": None
        }
        
        file_map = {
            "file": ["variables.file"]
        }
        
        # Determine MIME type based on file extension
        if filename.lower().endswith('.pdf'):
            mime_type = 'application/pdf'
        elif filename.lower().endswith('.xlsx'):
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mime_type = 'application/octet-stream'
        
        # Create multipart form data
        files_data = {
            'query': (None, mutation),
            'variables': (None, json.dumps(variables)),
            'map': (None, json.dumps(file_map)),
            'file': (filename, file_bytes, mime_type)
        }
        
        # Upload the file
        upload_response = requests.post(
            file_url,
            headers={"Authorization": api_token},
            files=files_data,
            timeout=60
        )
        
        if upload_response.status_code not in [200, 201]:
            error_text = upload_response.text
            try:
                error_json = upload_response.json()
                if "errors" in error_json:
                    st.error(f"Error uploading file: {error_json['errors']}")
                else:
                    st.error(f"Error uploading file: {error_text}")
            except:
                st.error(f"Error uploading file: {upload_response.status_code} - {error_text}")
            return False
        
        result = upload_response.json()
        
        if "errors" in result:
            st.error(f"Error uploading file: {result['errors']}")
            return False
        
        # Check if file was added successfully
        if result.get("data", {}).get("add_file_to_column", {}).get("id"):
            return True
        else:
            st.warning("File upload may have succeeded but no file ID returned")
            return True  # Assume success if no errors
        
    except Exception as e:
        st.error(f"Failed to upload file to monday.com: {e}")
        return False


def main() -> None:
    st.set_page_config(
        page_title="Dev & Inspection Creator",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.title("📊 Dev & Inspection Creator")

    st.markdown(
        """
<style>

    /* Enlarge checkbox */
    [data-testid="stCheckbox"] label span {
        height: 1.8rem;
        width: 1.8rem;
        transform: translateX(35px);
    }

    /* Hide normal label text (since you use empty label) */
    [data-testid="stCheckbox"] label p {
        font-size: 0;
        margin: 0;
    }

    /* Move right the entire checkbox inside its Streamlit column */
    [data-testid="stCheckbox"] {
        display: flex !important;
        justify-content: right !important;  /* Center horizontally */
        align-items: right !important;       /* Center vertically */
        height: 100%;                          /* Align with text inputs */
        padding: 0 !important;
        margin: 0 !important;
    }

    /* Also move right the internal label wrapper */
    [data-testid="stCheckbox"] > label {
        display: flex !important;
        justify-content: right !important;
        align-items: right !important;
        width: 100%;
        height: 100%;
        padding: 0 !important;
        margin: 0 !important;
    }

    /* Hide ALL sidebar list items by default */
    [data-testid="stSidebarNav"] li {
        display: none !important;
    }

    /* Show list items that contain allowed tool pages using :has() selector */
    [data-testid="stSidebarNav"] li:has(a[href*="quickbooks"]),
    [data-testid="stSidebarNav"] li:has(a[href*="signnow"]),
    [data-testid="stSidebarNav"] li:has(a[href*="/tools"]),
    [data-testid="stSidebarNav"] li:has(a[href*="workbook"]),
    [data-testid="stSidebarNav"] li:has(a[href*="deck_creator"]) {
        display: block !important;
    }

</style>
<script>
// JavaScript to show only tool pages and hide everything else
(function() {
    function showToolPagesOnly() {
        const navItems = document.querySelectorAll('[data-testid="stSidebarNav"] li');
        const allowedPages = ['quickbooks', 'signnow', 'tools', 'workbook', 'deck_creator'];
        
        navItems.forEach(item => {
            const link = item.querySelector('a');
            if (!link) {
                item.style.setProperty('display', 'none', 'important');
                return;
            }
            
            const href = (link.getAttribute('href') || '').toLowerCase();
            const text = link.textContent.trim().toLowerCase();
            
            // Check if this is an allowed tool page
            const isToolPage = allowedPages.some(page => {
                return href.includes(page) || text.includes(page.toLowerCase());
            });
            
            // Make sure it's not ads dashboard or other dashboards
            const isDashboard = (text.includes('ads') && text.includes('dashboard')) || 
                              (href.includes('ads') && href.includes('dashboard'));
            
            if (isToolPage && !isDashboard) {
                item.style.setProperty('display', 'block', 'important');
                link.style.setProperty('display', 'block', 'important');
            } else {
                item.style.setProperty('display', 'none', 'important');
            }
        });
    }
    
    // Run immediately and on load
    showToolPagesOnly();
    window.addEventListener('load', function() {
        setTimeout(showToolPagesOnly, 50);
        setTimeout(showToolPagesOnly, 200);
        setTimeout(showToolPagesOnly, 500);
    });
    
    // Watch for DOM changes
    const observer = new MutationObserver(function() {
        showToolPagesOnly();
    });
    
    setTimeout(function() {
        const sidebar = document.querySelector('[data-testid="stSidebarNav"]');
        if (sidebar) {
            observer.observe(sidebar, { 
                childList: true, 
                subtree: true,
                attributes: true
            });
        }
    }, 100);
})();
</script>
""",
        unsafe_allow_html=True,
    )

    st.caption(
        "Fill in the Development Package inputs and download a formatted workbook "
        "based on the official template."
    )

    # Get query parameters from Monday.com link
    query_params = st.query_params
    first_name = query_params.get("first_name", "").strip()
    last_name = query_params.get("last_name", "").strip()
    email = query_params.get("email", "").strip()
    representative = query_params.get("representative", "").strip()
    num_styles_param = query_params.get("num_styles", "").strip()
    item_id = query_params.get("item_id", "").strip()
    
    # Clean up representative (remove "$" prefix if present from Monday.com formula)
    if representative.startswith("$"):
        representative = representative[1:].strip()
    
    # Combine first_name and last_name for client_name
    client_name_default = ""
    if first_name or last_name:
        client_name_default = f"{first_name} {last_name}".strip()
    
    # Use query params as default values if available
    client_name = st.text_input(
        "Client Name", 
        value=client_name_default,
        placeholder="Enter client name"
    )

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        client_email = st.text_input(
            "Client Email", 
            value=email,
            placeholder="client@email.com"
        )
    with col_b:
        phone_number = st.text_input(
            "Phone Number",
            value=st.session_state.get("phone_number", ""),
            placeholder="(555) 123-4567"
        )
        st.session_state["phone_number"] = phone_number
    with col_c:
        representative = st.text_input(
            "Representative", 
            value=representative,
            placeholder="Enter representative"
        )

    # Initialize session state for style entries
    if "style_entries" not in st.session_state:
        st.session_state["style_entries"] = []
    
    # Initialize Custom Items
    if "custom_styles" not in st.session_state:
        st.session_state["custom_styles"] = []
    
    # Initialize A La Carte Items
    if "a_la_carte_items" not in st.session_state:
        st.session_state["a_la_carte_items"] = []
    
    st.subheader("**Styles**")
    
    # Number of Styles field (auto-fills from query param, but editable) - in first column, half width
    col_num_styles, col_spacer = st.columns([1, 3])
    with col_num_styles:
        num_styles_default = len(st.session_state["style_entries"])
        if num_styles_param:
            try:
                num_styles_default = int(num_styles_param)
            except ValueError:
                num_styles_default = len(st.session_state["style_entries"])
        
        num_styles_input = st.number_input(
            "Number of Styles",
            min_value=0,
            value=num_styles_default,
            step=1,
            key="num_styles_input",
            help="Number of styles/items. Auto-filled from Monday.com if available."
        )
        
        # Update style entries list to match num_styles_input
        current_count = len(st.session_state["style_entries"])
        if num_styles_input > current_count:
            # Add new styles with blank names
            for i in range(current_count, num_styles_input):
                st.session_state["style_entries"].append({
                    "name": "",  # Leave blank, don't default to "Style 1", etc.
                    "activewear": False,
                    "complexity": 0.0,
                    "style_number": 101 + i,  # Default style numbers: 101, 102, 103...
                    "options": {
                        "wash_dye": False,
                        "design": False,
                        "source": False,
                        "treatment": False,
                    },
                })
        elif num_styles_input < current_count:
            # Remove excess styles
            st.session_state["style_entries"] = st.session_state["style_entries"][:num_styles_input]
    
    # Column headers
    header_cols = st.columns([1.1, 1.8, 0.8, 1.2, 1.2, 1, 1, 1.1])
    with header_cols[0]:
        st.markdown("**Style Number**")
    with header_cols[1]:
        st.markdown("**Style Name**")
    with header_cols[2]:
        st.markdown("**Activewear?**")
    with header_cols[3]:
        st.markdown("**Complexity (%)**")
    with header_cols[4]:
        st.markdown("**Wash/Dye ($1,330)**")
    with header_cols[5]:
        st.markdown("**Design ($1,330)**")
    with header_cols[6]:
        st.markdown("**Source ($1,330)**")
    with header_cols[7]:
        st.markdown("**Treatment ($760)**")
    
    # Display existing style entries in horizontal rows
    if st.session_state["style_entries"]:
        for i, entry in enumerate(st.session_state["style_entries"]):
            with st.container():
                cols = st.columns([1.2, 1.8, 1, 1.2, 1.2, 1, 1, 1])
                with cols[0]:
                    # Style Number field with default value (101, 102, 103...)
                    default_style_number = entry.get("style_number", 101 + i)
                    style_number = st.number_input(
                        "Style Number",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with cols[1]:
                    style_name = st.text_input(
                        "Custom Item Name",
                        value=entry.get("name", ""),
                        key=f"style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Dress, Winter Coat",
                    )
                    entry["name"] = style_name
                with cols[2]:
                    activewear = st.checkbox(
                        "",
                        value=entry.get("activewear", False),
                        key=f"activewear_{i}",
                        label_visibility="visible",
                    )
                    entry["activewear"] = activewear
                with cols[3]:
                    complexity = st.number_input(
                        "Complexity (%)",
                        min_value=0,
                        max_value=200,
                        value=int(entry.get("complexity", 100)),
                        step=5,
                        format="%d",
                        key=f"complexity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["complexity"] = float(complexity)
                with cols[4]:
                    wash_dye = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("wash_dye", False),
                        key=f"wash_dye_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["wash_dye"] = wash_dye
                with cols[5]:
                    design = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("design", False),
                        key=f"design_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["design"] = design
                with cols[6]:
                    source = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("source", False),
                        key=f"source_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["source"] = source
                with cols[7]:
                    treatment = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("treatment", False),
                        key=f"treatment_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["treatment"] = treatment

    # Custom Item section
    st.markdown("---")
    st.subheader("**Custom Item**")
    
    # Calculate number of regular styles for custom item numbering
    num_regular_styles = len(st.session_state.get("style_entries", []))
    
    # Column headers for Custom Items
    custom_header_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
    with custom_header_cols[0]:
        st.markdown("**Style Number**")
    with custom_header_cols[1]:
        st.markdown("**Custom Item Name**")
    with custom_header_cols[2]:
        st.markdown("**Price ($)**")
    with custom_header_cols[3]:
        st.markdown("**Complexity (%)**")
    with custom_header_cols[4]:
        st.markdown("**Remove**")
    
    # Display existing Custom Items
    if st.session_state["custom_styles"]:
        for i, entry in enumerate(st.session_state["custom_styles"]):
            with st.container():
                custom_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
                with custom_cols[0]:
                    # Style Number field with default value (101 + num_regular_styles + i)
                    default_style_number = entry.get("style_number", 101 + num_regular_styles + i)
                    style_number = st.number_input(
                        "Style Number",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"custom_style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with custom_cols[1]:
                    custom_style_name = st.text_input(
                        "Custom Item Name",
                        value=entry.get("name", ""),
                        key=f"custom_style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Custom Item",
                    )
                    entry["name"] = custom_style_name
                with custom_cols[2]:
                    custom_price = st.number_input(
                        "Price",
                        min_value=0.0,
                        value=float(entry.get("price", 0.0)),
                        step=100.0,
                        format="%.2f",
                        key=f"custom_price_{i}",
                        label_visibility="collapsed",
                    )
                    entry["price"] = float(custom_price)
                with custom_cols[3]:
                    custom_complexity = st.number_input(
                        "Complexity (%)",
                        min_value=0,
                        max_value=200,
                        value=int(entry.get("complexity", 0)),
                        step=5,
                        format="%d",
                        key=f"custom_complexity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["complexity"] = float(custom_complexity)
                with custom_cols[4]:
                    if st.button("❌", key=f"remove_custom_{i}", help="Remove this Custom Item"):
                        st.session_state["custom_styles"].pop(i)
                        st.rerun()
    
    # Add new Custom Item interface
    st.markdown("**Add New Custom Item**")
    add_custom_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
    default_new_custom_name = st.session_state.get("new_custom_style_name", "")
    default_new_custom_price = st.session_state.get("new_custom_price", 0.0)
    default_new_custom_complexity = st.session_state.get("new_custom_complexity", 0)
    default_new_custom_style_number = st.session_state.get("new_custom_style_number", 101 + num_regular_styles + len(st.session_state.get("custom_styles", [])))
    
    with add_custom_cols[0]:
        new_custom_style_number = st.number_input(
            "Style Number",
            min_value=1,
            value=int(default_new_custom_style_number),
            step=1,
            key="new_custom_style_number",
            label_visibility="collapsed",
        )
    with add_custom_cols[1]:
        new_custom_style_name = st.text_input(
            "Custom Item Name",
            value=default_new_custom_name,
            key="new_custom_style_name",
            label_visibility="collapsed",
            placeholder="e.g., Custom Item",
        )
    with add_custom_cols[2]:
        new_custom_price = st.number_input(
            "Price",
            min_value=0.0,
            value=default_new_custom_price,
            step=100.0,
            format="%.2f",
            key="new_custom_price",
            label_visibility="collapsed",
        )
    with add_custom_cols[3]:
        new_custom_complexity = st.number_input(
            "Complexity (%)",
            min_value=0,
            max_value=200,
            value=default_new_custom_complexity,
            step=5,
            format="%d",
            key="new_custom_complexity",
            label_visibility="collapsed",
        )
    with add_custom_cols[4]:
        if st.button("➕ Add", key="add_custom_style", help="Add this Custom Item"):
            if new_custom_style_name.strip() and new_custom_price > 0:
                st.session_state["custom_styles"].append({
                    "name": new_custom_style_name.strip(),
                    "price": float(new_custom_price),
                    "complexity": float(new_custom_complexity),
                    "style_number": int(new_custom_style_number),
                })
                # Reset add-new-custom-style inputs
                for key in [
                    "new_custom_style_name",
                    "new_custom_price",
                    "new_custom_complexity",
                    "new_custom_style_number",
                ]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
            else:
                st.warning("Please enter a Custom Item name and price before adding.")

    # A La Carte Package section
    st.markdown("---")
    st.subheader("**A La Carte Package**")
    
    # Calculate number for a la carte item numbering (continues from custom items)
    num_regular_styles = len(st.session_state.get("style_entries", []))
    num_custom_styles = len(st.session_state.get("custom_styles", []))
    a_la_carte_start_number = 101 + num_regular_styles + num_custom_styles
    
    # Column headers for A La Carte Items
    a_la_carte_header_cols = st.columns([1.2, 2, 1, 1, 1, 1, 1, 1, 1, 0.8])
    with a_la_carte_header_cols[0]:
        st.markdown("**Style Number**")
    with a_la_carte_header_cols[1]:
        st.markdown("**Style Name**")
    with a_la_carte_header_cols[2]:
        st.markdown("**INTAKE**")
    with a_la_carte_header_cols[3]:
        st.markdown("**1ST PATTERN**")
    with a_la_carte_header_cols[4]:
        st.markdown("**1ST SAMPLE**")
    with a_la_carte_header_cols[5]:
        st.markdown("**FITTING**")
    with a_la_carte_header_cols[6]:
        st.markdown("**ADJUSTMENT**")
    with a_la_carte_header_cols[7]:
        st.markdown("**FINAL SAMPLES**")
    with a_la_carte_header_cols[8]:
        st.markdown("**DUPLICATES**")
    with a_la_carte_header_cols[9]:
        st.markdown("**Remove**")
    
    # Display existing A La Carte Items
    if st.session_state["a_la_carte_items"]:
        for i, entry in enumerate(st.session_state["a_la_carte_items"]):
            with st.container():
                a_la_carte_cols = st.columns([1.2, 2, 1, 1, 1, 1, 1, 1, 1, 0.8])
                with a_la_carte_cols[0]:
                    # Style Number field
                    default_style_number = entry.get("style_number", a_la_carte_start_number + i)
                    style_number = st.number_input(
                        "Style Number",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"a_la_carte_style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with a_la_carte_cols[1]:
                    style_name = st.text_input(
                        "Style Name",
                        value=entry.get("name", ""),
                        key=f"a_la_carte_style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Style Name",
                    )
                    entry["name"] = style_name
                with a_la_carte_cols[2]:
                    intake_hours = st.number_input(
                        "INTAKE",
                        min_value=0,
                        value=int(entry.get("intake_session", 0)),
                        step=1,
                        format="%d",
                        key=f"a_la_carte_intake_{i}",
                        label_visibility="collapsed",
                    )
                    entry["intake_session"] = int(intake_hours)
                with a_la_carte_cols[3]:
                    first_pattern_hours = st.number_input(
                        "1ST PATTERN",
                        min_value=0,
                        value=int(entry.get("first_pattern", 0)),
                        step=1,
                        format="%d",
                        key=f"a_la_carte_first_pattern_{i}",
                        label_visibility="collapsed",
                    )
                    entry["first_pattern"] = int(first_pattern_hours)
                with a_la_carte_cols[4]:
                    first_sample_hours = st.number_input(
                        "1ST SAMPLE",
                        min_value=0,
                        value=int(entry.get("first_sample", 0)),
                        step=1,
                        format="%d",
                        key=f"a_la_carte_first_sample_{i}",
                        label_visibility="collapsed",
                    )
                    entry["first_sample"] = int(first_sample_hours)
                with a_la_carte_cols[5]:
                    fitting_hours = st.number_input(
                        "FITTING",
                        min_value=0,
                        value=int(entry.get("fitting", 0)),
                        step=1,
                        format="%d",
                        key=f"a_la_carte_fitting_{i}",
                        label_visibility="collapsed",
                    )
                    entry["fitting"] = int(fitting_hours)
                with a_la_carte_cols[6]:
                    adjustment_hours = st.number_input(
                        "ADJUSTMENT",
                        min_value=0,
                        value=int(entry.get("adjustment", 0)),
                        step=1,
                        format="%d",
                        key=f"a_la_carte_adjustment_{i}",
                        label_visibility="collapsed",
                    )
                    entry["adjustment"] = int(adjustment_hours)
                with a_la_carte_cols[7]:
                    final_samples_hours = st.number_input(
                        "FINAL SAMPLES",
                        min_value=0,
                        value=int(entry.get("final_samples", 0)),
                        step=1,
                        format="%d",
                        key=f"a_la_carte_final_samples_{i}",
                        label_visibility="collapsed",
                    )
                    entry["final_samples"] = int(final_samples_hours)
                with a_la_carte_cols[8]:
                    duplicates_hours = st.number_input(
                        "DUPLICATES",
                        min_value=0,
                        value=int(entry.get("duplicates", 0)),
                        step=1,
                        format="%d",
                        key=f"a_la_carte_duplicates_{i}",
                        label_visibility="collapsed",
                    )
                    entry["duplicates"] = int(duplicates_hours)
                    # Calculate total (sum of all hours) - kept for Excel export, not displayed
                    total_hours = (
                        entry.get("intake_session", 0) +
                        entry.get("first_pattern", 0) +
                        entry.get("first_sample", 0) +
                        entry.get("fitting", 0) +
                        entry.get("adjustment", 0) +
                        entry.get("final_samples", 0) +
                        entry.get("duplicates", 0)
                    )
                    entry["total_hours"] = total_hours
                with a_la_carte_cols[9]:
                    if st.button("❌", key=f"remove_a_la_carte_{i}", help="Remove this A La Carte Item"):
                        st.session_state["a_la_carte_items"].pop(i)
                        st.rerun()
    
    # Add new A La Carte Item interface
    st.markdown("**Add New A La Carte Item**")
    add_a_la_carte_cols = st.columns([1.2, 2, 1, 1, 1, 1, 1, 1, 1, 0.8])
    default_new_a_la_carte_style_name = st.session_state.get("new_a_la_carte_style_name", "")
    default_new_a_la_carte_style_number = st.session_state.get("new_a_la_carte_style_number", a_la_carte_start_number + len(st.session_state.get("a_la_carte_items", [])))
    
    with add_a_la_carte_cols[0]:
        new_a_la_carte_style_number = st.number_input(
            "Style Number",
            min_value=1,
            value=int(default_new_a_la_carte_style_number),
            step=1,
            key="new_a_la_carte_style_number",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[1]:
        new_a_la_carte_style_name = st.text_input(
            "Style Name",
            value=default_new_a_la_carte_style_name,
            key="new_a_la_carte_style_name",
            label_visibility="collapsed",
            placeholder="e.g., Style Name",
        )
    with add_a_la_carte_cols[2]:
        new_intake = st.number_input(
            "INTAKE",
            min_value=0,
            value=0,
            step=1,
            format="%d",
            key="new_a_la_carte_intake",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[3]:
        new_first_pattern = st.number_input(
            "1ST PATTERN",
            min_value=0,
            value=0,
            step=1,
            format="%d",
            key="new_a_la_carte_first_pattern",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[4]:
        new_first_sample = st.number_input(
            "1ST SAMPLE",
            min_value=0,
            value=0,
            step=1,
            format="%d",
            key="new_a_la_carte_first_sample",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[5]:
        new_fitting = st.number_input(
            "FITTING",
            min_value=0,
            value=0,
            step=1,
            format="%d",
            key="new_a_la_carte_fitting",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[6]:
        new_adjustment = st.number_input(
            "ADJUSTMENT",
            min_value=0,
            value=0,
            step=1,
            format="%d",
            key="new_a_la_carte_adjustment",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[7]:
        new_final_samples = st.number_input(
            "FINAL SAMPLES",
            min_value=0,
            value=0,
            step=1,
            format="%d",
            key="new_a_la_carte_final_samples",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[8]:
        new_duplicates = st.number_input(
            "DUPLICATES",
            min_value=0,
            value=0,
            step=1,
            format="%d",
            key="new_a_la_carte_duplicates",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[9]:
        if st.button("➕ Add", key="add_a_la_carte_style", help="Add this A La Carte Item"):
            if new_a_la_carte_style_name.strip():
                st.session_state["a_la_carte_items"].append({
                    "name": new_a_la_carte_style_name.strip(),
                    "style_number": int(new_a_la_carte_style_number),
                    "intake_session": int(new_intake),
                    "first_pattern": int(new_first_pattern),
                    "first_sample": int(new_first_sample),
                    "fitting": int(new_fitting),
                    "adjustment": int(new_adjustment),
                    "final_samples": int(new_final_samples),
                    "duplicates": int(new_duplicates),
                    "total_hours": int(new_intake + new_first_pattern + new_first_sample + new_fitting + new_adjustment + new_final_samples + new_duplicates),
                })
                # Reset add-new-a-la-carte inputs
                for key in [
                    "new_a_la_carte_style_name",
                    "new_a_la_carte_style_number",
                ]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
            else:
                st.warning("Please enter a Style Name before adding.")

    # Check if at least one item exists after all sections
    if not st.session_state["style_entries"] and not st.session_state["custom_styles"] and not st.session_state.get("a_la_carte_items", []):
        st.info("Add at least one style, Custom Item, or A La Carte Item to enable the generator.")
        return

    st.markdown("---")
    st.subheader("Discount")
    discount_cols = st.columns([1.8, 1, 1, 1.2, 1, 0.8, 0.8, 1, 0.8])
    with discount_cols[0]:
        discount_percentage = st.number_input(
            "Discount (%)",
            min_value=0,
            max_value=100,
            value=st.session_state.get("discount_percentage_value", 0),
            step=1,
        )
    st.session_state["discount_percentage_value"] = discount_percentage
    
    # Notes Section
    st.markdown("---")
    st.subheader("Notes")
    
    # Initialize notes list in session state
    if "notes_list" not in st.session_state:
        st.session_state["notes_list"] = [""]
    
    # Display existing notes
    for i, note in enumerate(st.session_state["notes_list"]):
        note_cols = st.columns([10, 1])
        with note_cols[0]:
            updated_note = st.text_input(
                "Note",
                value=note,
                key=f"note_{i}",
                placeholder="Enter a note...",
                label_visibility="collapsed"
            )
            st.session_state["notes_list"][i] = updated_note
        with note_cols[1]:
            if st.button("❌", key=f"remove_note_{i}", help="Remove this note"):
                st.session_state["notes_list"].pop(i)
                st.rerun()
    
    # Add new note button
    if st.button("➕ Add Note", key="add_note"):
        st.session_state["notes_list"].append("")
        st.rerun()
    
    # Filter out empty notes for workbook generation
    notes = [note for note in st.session_state["notes_list"] if note and note.strip()]

    try:
        excel_bytes, _, _ = build_workbook_bytes(
            client_name=client_name,
            client_email=client_email,
            representative=representative,
            phone_number=phone_number,
            style_entries=st.session_state["style_entries"],
            custom_styles=st.session_state.get("custom_styles", []),
            a_la_carte_items=st.session_state.get("a_la_carte_items", []),
            discount_percentage=discount_percentage,
            notes=notes if notes else [],
        )
    except FileNotFoundError as exc:
        st.error(str(exc))
        return
    except Exception as exc:  # pragma: no cover - streamlit runtime feedback
        st.error(f"Failed to build workbook: {exc}")
        return

    safe_client = re.sub(r"[^A-Za-z0-9_-]+", "_", (client_name or "").strip()) or "client"
    download_name = f"dev_inspection_{safe_client.lower()}.xlsx"

    st.success("Workbook is ready.")
    st.download_button(
        label="Generate Workbook",
        data=excel_bytes,
        file_name=download_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    # Google Drive upload section (service account only)
    st.markdown("---")
    st.subheader("Google Sheets --> Monday.com Upload")

    sheet_title = (client_name or "Workbook").strip() or "Workbook"
    sheet_title = f"{sheet_title} - Development Package"

    st.caption("Uploads will use the shared Google Drive folder configured for the service account.")

    if st.button("Upload to Monday.com", type="primary"):
        with st.spinner("Uploading workbook to Google Sheets and updating Monday.com..."):
            try:
                # Upload to Google Sheets
                sheet_url, converted = upload_workbook_to_google_sheet(excel_bytes, sheet_title)
                st.session_state["google_sheet_url"] = sheet_url
                
                if converted:
                    st.success(f"✅ Google Sheet created: [Open Sheet]({sheet_url})")
                else:
                    st.success(
                        f"✅ Workbook uploaded as XLSX: [Download / Open]({sheet_url})  \n"
                        "Google denied automatic conversion because service accounts report zero Drive quota. "
                        "Open the file in Google Drive and use **File → Save as Google Sheets** if you need an editable sheet."
                    )
                
                # Update Monday.com with the Google Sheet URL if item_id is provided
                if item_id:
                    # Update Monday.com item with the Google Sheet link
                    if update_monday_item_workbook_url(item_id, sheet_url):
                        st.success(f"✅ Monday.com item updated with Google Sheet link!")
                    else:
                        st.warning("⚠️ Google Sheet uploaded, but failed to update Monday.com item. Please update manually.")
                    
                    # Also upload PDF version to Files column (only if converted to Google Sheet)
                    if converted:
                        try:
                            # Get board_id from item_id
                            board_id = get_board_id_from_item(item_id)
                            if board_id:
                                # Export Google Sheet as PDF using Google Sheets export API
                                pdf_bytes = export_google_sheet_as_pdf(sheet_url)
                                pdf_filename = f"{safe_client}_workbook.pdf"
                                
                                # Upload PDF to Monday.com Files column
                                if upload_file_to_monday_item(item_id, board_id, pdf_bytes, pdf_filename):
                                    st.success(f"✅ PDF version uploaded to Monday.com Files column!")
                                else:
                                    st.warning("⚠️ PDF upload to Monday.com Files column failed. Link was updated successfully.")
                            else:
                                st.warning("⚠️ Could not retrieve board ID. PDF upload skipped.")
                        except Exception as pdf_error:
                            st.warning(f"⚠️ PDF export/upload failed: {pdf_error}. Link was updated successfully.")
                    else:
                        st.info("ℹ️ PDF export skipped (workbook not converted to Google Sheet format).")
                else:
                    st.info("ℹ️ No Monday.com item ID provided. Workbook uploaded to Google Sheets only.")

            except GoogleSheetsUploadError as exc:
                message = str(exc)
                st.error(f"❌ Google Sheets upload failed: {message}")
                if "storage quota" in message.lower():
                    _show_drive_quota_help()
            except Exception as exc:  # pragma: no cover - runtime failures
                st.error(f"❌ Unexpected error: {exc}")


if __name__ == "__main__":
    main()
