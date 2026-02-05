"""
A La Carte Creator
Builds the Development Package section of the workbook using a template.
"""

from __future__ import annotations

import io
import math
import os
import re
import urllib.parse
from copy import copy
from io import BytesIO

import streamlit as st
import requests

from google_sheets_uploader import (
    GoogleSheetsUploadError,
    upload_workbook_to_google_sheet,
)

# Google API imports for PDF export
try:
    from google.oauth2.service_account import Credentials as SACredentials
    from googleapiclient.discovery import build
    from google.auth.transport.requests import Request as GoogleRequest
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
    from openpyxl.utils import column_index_from_string, get_column_letter
except Exception:  # pragma: no cover - fallback if dependency missing at runtime
    load_workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None
    NamedStyle = None
    column_index_from_string = None
    get_column_letter = None


# Pricing constants - Package rates by # OF STYLES (style rows only)
# 2-4 STYLES, 5-9 STYLES, 10-14 STYLES, +15 STYLES. Activewear includes *2 FITTINGS.
PRICING_TABLE = {
    "Regular": {
        "< 2": 3360.00,   # 1 style → same as 2-4
        "< 5": 3360.00,   # 2-4 STYLES
        "< 10": 2856.00,  # 5-9 STYLES
        "< 15": 2688.00,  # 10-14 STYLES
        "15+": 2520.00,   # +15 STYLES
    },
    "Activewear/Lingerie/Swim": {
        "< 2": 3800.00,
        "< 5": 3800.00,
        "< 10": 3230.00,
        "< 15": 3040.00,
        "15+": 2850.00,
    },
    "Pattern Blocks": {
        "< 2": 2500.00,
        "< 5": 2500.00,
        "< 10": 2125.00,
        "< 15": 2000.00,
        "15+": 1875.00,
    }
}

OPTIONAL_PRICES = {
    "wash_dye": 1500.00,  # For Development section
    "dye_testing": 1500.00,  # For A La Carte section
    "planning": 1500.00,  # For A La Carte section
    "design": 1500.00,
    "treatment": 860.00,  # For Development section
    # Note: "source" removed - no longer an option
}
SUMMARY_LABEL_COL = 14  # Column N
SUMMARY_VALUE_COL = 16  # Column P
SUMMARY_DEV_ROW = 10
SUMMARY_OPT_ROW = 12
SUMMARY_SUBTOTAL_ROW = 14
SUMMARY_DISCOUNT_ROW = 16
SUMMARY_SUM_END_ROW = 13  # Row before subtotal row
SUMMARY_TOTAL_DUE_BASE_ROW = 20
DELIVERABLE_BLOCK_START = 22
DELIVERABLE_BLOCK_END = 34
DELIVERABLE_BLOCK_HEIGHT = DELIVERABLE_BLOCK_END - DELIVERABLE_BLOCK_START + 1
DELIVERABLE_COL_START = 2  # Column B
DELIVERABLE_COL_END = 16   # Column P
TEMPLATE_FILENAME = "Copy of TEG 2025 WORKBOOK TEMPLATES.xlsx"
TARGET_SHEET = "DEVELOPMENT ONLY"
ROW_INDICES = [10, 12, 14, 16, 18]  # Rows reserved for style entries


@st.cache_data(show_spinner=False)
def get_template_path() -> str:
    """Return the absolute path to the Excel template."""
    base_dir = os.path.dirname(os.path.dirname(__file__))
    template_path = os.path.join(base_dir, "inputs", TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Template '{TEMPLATE_FILENAME}' was not found in the inputs folder."
        )
    return template_path


def calculate_base_price(num_styles: int, style_type: str) -> float:
    """Calculate base price based on number of styles and style type.
    
    Pricing is based on style count brackets and style type:
    - 1 style / 2-4 STYLES: Regular=$3,360, Activewear=$3,800, Pattern Blocks=$2,500
    - 5-9 STYLES: Regular=$2,856, Activewear=$3,230, Pattern Blocks=$2,125
    - 10-14 STYLES: Regular=$2,688, Activewear=$3,040, Pattern Blocks=$2,000
    - +15 STYLES: Regular=$2,520, Activewear=$2,850, Pattern Blocks=$1,875
    
    Args:
        num_styles: Number of style rows (used to determine bracket; custom items are not counted)
        style_type: One of "Regular", "Activewear/Lingerie/Swim", or "Pattern Blocks"
    """
    if style_type not in PRICING_TABLE:
        style_type = "Regular"
    if num_styles < 2:
        bracket = "< 2"
    elif num_styles < 5:
        bracket = "< 5"
    elif num_styles < 10:
        bracket = "< 10"
    elif num_styles < 15:
        bracket = "< 15"
    else:
        bracket = "15+"
    return PRICING_TABLE[style_type][bracket]


def copy_cell_formatting(source_cell, target_cell) -> None:
    """Copy formatting (fill, border, alignment) from source to target cell."""
    if PatternFill is None or Border is None or Alignment is None:
        return
    
    try:
        # Copy fill (background color)
        if source_cell.fill and source_cell.fill.start_color:
            target_cell.fill = PatternFill(
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color,
                fill_type=source_cell.fill.fill_type
            )
        
        # Copy border
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
    except Exception:
        pass  # Skip if copying fails


def apply_arial_20_font(cell) -> None:
    """Apply Arial font size 20 to a cell."""
    if Font is not None:
        try:
            cell.font = Font(name="Arial", size=20)
        except Exception:
            pass


def apply_full_border(cell) -> None:
    """Apply full thin borders around a cell."""
    if Border is None or Side is None:
        return
    try:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    except Exception:
        pass


def apply_full_border_pair(ws, column, start_row: int, end_row: int) -> None:
    """Apply borders to both cells in a vertical pair (before merging)."""
    if column_index_from_string is None:
        return
    col_idx = column if isinstance(column, int) else column_index_from_string(column)
    top_cell = ws.cell(row=start_row, column=col_idx)
    bottom_cell = ws.cell(row=end_row, column=col_idx)
    apply_full_border(top_cell)
    apply_full_border(bottom_cell)


def capture_deliverables_block(ws):
    """Capture the deliverables block (values + formatting) from the template."""
    block_rows = []
    for row in range(DELIVERABLE_BLOCK_START, DELIVERABLE_BLOCK_END + 1):
        row_data = []
        for col in range(DELIVERABLE_COL_START, DELIVERABLE_COL_END + 1):
            cell = ws.cell(row=row, column=col)
            row_data.append(
                {
                    "value": cell.value,
                    "number_format": cell.number_format,
                    "font": copy(cell.font) if cell.font else None,
                    "fill": copy(cell.fill) if cell.fill else None,
                    "border": copy(cell.border) if cell.border else None,
                    "alignment": copy(cell.alignment) if cell.alignment else None,
                }
            )
        block_rows.append(row_data)
    
    merged_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row >= DELIVERABLE_BLOCK_START
            and merged_range.max_row <= DELIVERABLE_BLOCK_END
            and merged_range.min_col >= DELIVERABLE_COL_START
            and merged_range.max_col <= DELIVERABLE_COL_END
        ):
            merged_ranges.append(
                (
                    merged_range.min_row - DELIVERABLE_BLOCK_START,
                    merged_range.max_row - DELIVERABLE_BLOCK_START,
                    merged_range.min_col,
                    merged_range.max_col,
                )
            )
    
    return {"rows": block_rows, "merges": merged_ranges}


def restore_deliverables_block(ws, template_block: dict, target_start_row: int) -> None:
    """Restore the deliverables block at the specified start row."""
    if not template_block:
        return
    
    rows_data = template_block.get("rows", [])
    block_height = len(rows_data)
    target_end_row = target_start_row + block_height - 1
    
    # Clear existing merges in target area
    to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.max_row < target_start_row
            or merged_range.min_row > target_end_row
            or merged_range.max_col < DELIVERABLE_COL_START
            or merged_range.min_col > DELIVERABLE_COL_END
        ):
            continue
        to_unmerge.append(merged_range)
    
    for merged_range in to_unmerge:
        try:
            ws.unmerge_cells(range_string=str(merged_range))
        except Exception:
            pass
    
    # Write cell data and formatting
    for row_offset, row_cells in enumerate(rows_data):
        target_row = target_start_row + row_offset
        for col_offset, cell_data in enumerate(row_cells):
            target_col = DELIVERABLE_COL_START + col_offset
            coord = None
            if get_column_letter is not None:
                coord = f"{get_column_letter(target_col)}{target_row}"
            value = cell_data.get("value") if cell_data else None
            if coord:
                safe_set_cell_value(ws, coord, value)
            else:
                ws.cell(row=target_row, column=target_col).value = value
            cell = ws.cell(row=target_row, column=target_col)
            cell.number_format = cell_data.get("number_format") or cell.number_format
            if cell_data.get("font"):
                cell.font = copy(cell_data["font"])
            if cell_data.get("fill"):
                cell.fill = copy(cell_data["fill"])
            if cell_data.get("border"):
                cell.border = copy(cell_data["border"])
            if cell_data.get("alignment"):
                cell.alignment = copy(cell_data["alignment"])
    
    # Recreate merged ranges (adjusted for new start row)
    for min_row_offset, max_row_offset, min_col, max_col in template_block.get("merges", []):
        start_row = target_start_row + min_row_offset
        end_row = target_start_row + max_row_offset
        try:
            ws.merge_cells(
                start_row=start_row,
                start_column=min_col,
                end_row=end_row,
                end_column=max_col,
            )
        except Exception:
            pass


def safe_merge_cells(ws, range_str: str) -> bool:
    """Safely merge cells, checking if they're already merged first."""
    try:
        # Parse the range string (e.g., "B10:B11")
        parts = range_str.split(':')
        if len(parts) != 2:
            return False
        
        from openpyxl.utils import column_index_from_string
        
        # Parse start cell
        start_col_letter = ''.join([c for c in parts[0] if c.isalpha()])
        start_row = int(''.join([c for c in parts[0] if c.isdigit()]))
        start_col = column_index_from_string(start_col_letter)
        
        # Parse end cell
        end_col_letter = ''.join([c for c in parts[1] if c.isalpha()])
        end_row = int(''.join([c for c in parts[1] if c.isdigit()]))
        end_col = column_index_from_string(end_col_letter)
        
        # Check if any cell in this range is already part of a merged range
        # Do multiple passes to catch all overlapping merges
        max_passes = 3
        for pass_num in range(max_passes):
            ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if ranges overlap (any cell in our range overlaps with merged range)
                if not (merged_range.max_row < start_row or merged_range.min_row > end_row or
                        merged_range.max_col < start_col or merged_range.min_col > end_col):
                    # Ranges overlap, mark for unmerging
                    ranges_to_unmerge.append(merged_range)
            
            if not ranges_to_unmerge:
                break  # No more overlapping merges
            
            # Unmerge overlapping ranges (in reverse to avoid index issues)
            for merged_range in reversed(ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass
        
        # Now merge the cells vertically
        ws.merge_cells(range_str)
        return True
    except Exception:
        return False  # Return False if merge fails


def safe_set_cell_value(ws, cell_ref: str, value) -> None:
    """Safely set cell value, handling merged cells by writing to top-left cell."""
    try:
        # Parse cell reference to get row and column
        from openpyxl.utils import coordinate_from_string, column_index_from_string
        col_letter = ''.join([c for c in cell_ref if c.isalpha()])
        row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
        col_num = column_index_from_string(col_letter)
        
        # Check if cell is part of a merged range
        cell_in_merged = False
        target_range = None
        
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_num <= merged_range.max_row and
                merged_range.min_col <= col_num <= merged_range.max_col):
                cell_in_merged = True
                target_range = merged_range
                break
        
        if cell_in_merged and target_range:
            # Unmerge the range first, then write to the original cell
            try:
                min_cell = ws.cell(row=target_range.min_row, column=target_range.min_col)
                max_cell = ws.cell(row=target_range.max_row, column=target_range.max_col)
                range_str = f"{min_cell.coordinate}:{max_cell.coordinate}"
                ws.unmerge_cells(range_str)
            except Exception:
                pass  # If unmerge fails, try to write anyway
        
        # Write to the cell using cell() method which always returns a writable cell
        target_cell = ws.cell(row=row_num, column=col_num)
        target_cell.value = value
        
    except Exception:
        # Fallback: try using the cell reference directly
        try:
            cell = ws[cell_ref]
            # If it's a MergedCell, we need to unmerge first
            if hasattr(cell, 'value'):
                try:
                    cell.value = value
                except (AttributeError, TypeError):
                    # It's a MergedCell, find and unmerge
                    cell_row = cell.row
                    cell_col = cell.column
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= cell_row <= merged_range.max_row and
                            merged_range.min_col <= cell_col <= merged_range.max_col):
                            try:
                                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                                ws.cell(row=cell_row, column=cell_col).value = value
                                break
                            except Exception:
                                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                                break
        except Exception:
            pass  # Skip if all methods fail


def safe_get_writable_cell(ws, row: int, column: int):
    """Get a writable cell, handling MergedCell by returning top-left cell of merged range."""
    # First check if the cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= column <= merged_range.max_col):
            # It's part of a merged range, return top-left cell (always writable)
            # But verify it's actually writable
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            try:
                # Test if writable by trying to read value
                _ = top_left.value
                return top_left
            except (AttributeError, TypeError):
                # Even top-left might be MergedCell in edge cases, find the real top-left
                # This shouldn't happen, but handle it
                return top_left
    
    # Not part of a merged range, try to get the cell
    # But even if not in merged_cells.ranges, ws.cell() might still return MergedCell
    # So we need to test if it's writable
    cell = ws.cell(row=row, column=column)
    # Check if it's a MergedCell by trying to access value (MergedCell raises AttributeError)
    try:
        # Try to read value - if it works, it's a regular cell
        _ = cell.value
        return cell
    except (AttributeError, TypeError):
        # It's a MergedCell even though it's not in merged_cells.ranges
        # This can happen if the merge was just created or there's a timing issue
        # Find the top-left cell by checking all merged ranges again
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= column <= merged_range.max_col):
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        # If we can't find it, the cell might be in a newly merged range
        # Try to get it anyway - caller should handle exceptions
        return cell


def safe_get_cell_value(ws, row: int, column: int):
    """Safely get cell value, handling MergedCell by returning value from top-left cell."""
    # First check if the cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= column <= merged_range.max_col):
            # It's part of a merged range, get value from top-left cell
            try:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                try:
                    return top_left_cell.value
                except (AttributeError, TypeError):
                    # Even top-left might be MergedCell, try to get value anyway
                    return top_left_cell.value
            except Exception:
                return None
    
    # Not part of a merged range, try to get the cell value
    try:
        cell = ws.cell(row=row, column=column)
        try:
            return cell.value
        except (AttributeError, TypeError):
            # It's a MergedCell, find the top-left cell
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= column <= merged_range.max_col):
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    try:
                        return top_left_cell.value
                    except Exception:
                        return None
            return None
    except Exception:
        return None


def update_header_labels(ws, client_name: str) -> None:
    """Ensure headers and client info match the spec."""
    header_map = {
        "H9": "WASH/DYE",
        "I9": "DESIGN",
        "J9": "TREATMENT",
        "K9": "TOTAL",
    }
    for cell, label in header_map.items():
        safe_set_cell_value(ws, cell, label)

    safe_set_cell_value(ws, "B3", "TEGMADE, JUST FOR")
    client_display = (client_name or "").strip().upper()
    safe_set_cell_value(ws, "J3", client_display)
    if Font is not None:
        ws["J3"].font = Font(
            color="00C9A57A",
            name="Schibsted Grotesk",
            size=48,
            bold=True,
        )
    if Alignment is not None:
        ws["J3"].alignment = Alignment(horizontal="left", vertical="center")

    if Alignment is not None:
        center_cells = ["L20", "P10", "P11", "P12", "P13", "P20"]
        for ref in center_cells:
            ws[ref].alignment = Alignment(horizontal="center", vertical="center")


def clear_style_rows(ws, num_styles: int = 0) -> None:
    """Blank out style rows (B–L) and the totals row, preserving format for <= 5 styles."""
    if num_styles <= 5:
        # For 5 or fewer styles, clear ALL template style rows (10-18) so past values don't remain
        # Use safe_set_cell_value to preserve formatting/merges
        style_rows = [10, 12, 14, 16, 18]
        for row_idx in style_rows:
            for col_letter in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L']:
                safe_set_cell_value(ws, f"{col_letter}{row_idx}", None)
        
        # Clear totals row (row 20)
        safe_set_cell_value(ws, "B20", None)
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "H20", None)
        safe_set_cell_value(ws, "L20", None)
        safe_set_cell_value(ws, f"N{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"N{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, "N19", None)
        safe_set_cell_value(ws, "P19", None)
    else:
        # For more than 5 styles, clear all style rows and totals row
        max_style_rows = num_styles
        totals_row_clear = 20 + (num_styles - 5) * 2
        for i in range(max_style_rows):
            row_idx = 10 + (i * 2)
            # Only clear if this is a style row (not the totals row)
            if row_idx == totals_row_clear:
                continue
            for col_idx in range(2, 13):  # Columns B through L (1-based)
                cell = ws.cell(row=row_idx, column=col_idx)
                # Check if cell is part of a merged range
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left.value = None
                        is_merged = True
                        break
                if not is_merged:
                    cell.value = None
        
        # Clear the totals row and TOTAL DUE AT SIGNING row (dynamic position)
        safe_set_cell_value(ws, f"B{totals_row_clear}", None)
        safe_set_cell_value(ws, f"F{totals_row_clear}", None)
        safe_set_cell_value(ws, f"H{totals_row_clear}", None)
        safe_set_cell_value(ws, f"L{totals_row_clear}", None)
        safe_set_cell_value(ws, f"N{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"N{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"N{totals_row_clear - 1}", None)
        safe_set_cell_value(ws, f"P{totals_row_clear - 1}", None)


def apply_development_package(
    ws,
    *,
    client_name: str,
    client_email: str,
    representative: str,
    style_entries: list[dict],
    custom_styles: list[dict],
    discount_percentage: float,
) -> tuple[float, float]:
    """Write the inputs into the workbook and return totals."""
    # Header metadata
    safe_set_cell_value(ws, "D6", client_email.strip())
    safe_set_cell_value(ws, "J6", (representative or "").strip().upper())
    safe_set_cell_value(ws, "B8", "DEVELOPMENT PACKAGE")

    optional_cells = {
        "H": "wash_dye",
        "I": "design",
        "J": "treatment",
        # Note: "source" removed - no longer an option
    }

    total_development = 0.0
    total_optional = 0.0
    num_styles = len(style_entries)
    num_custom_styles = len(custom_styles)
    total_styles_count = num_styles + num_custom_styles  # Total for pricing tier calculation
    discount_percentage = max(0.0, float(discount_percentage or 0))
    deliverables_template = capture_deliverables_block(ws)

    base_capacity = len(ROW_INDICES)
    extra_styles = max(num_styles - base_capacity, 0)
    rows_to_insert_regular = extra_styles * 2
    
    # Calculate how many rows we'll need for Custom Items too (if any)
    # This ensures we insert all rows upfront before writing any styles
    rows_to_insert_custom = 0
    if num_custom_styles > 0:
        # Calculate where Custom Items would start (after all regular styles)
        if num_styles > 5:
            custom_start_row = 20 + (num_styles - 5) * 2
        else:
            custom_start_row = 10 + num_styles * 2
        # Calculate where totals row will be after regular insertions
        if num_styles > 5:
            totals_row_after_regular = 20 + (num_styles - 5) * 2
        else:
            totals_row_after_regular = 20
        # Calculate custom row indices
        custom_row_indices_precalc = []
        for i in range(num_custom_styles):
            custom_row_indices_precalc.append(custom_start_row + (i * 2))
        # Check if we need more rows for Custom Items
        if custom_row_indices_precalc:
            last_custom_row = custom_row_indices_precalc[-1]
            if last_custom_row >= totals_row_after_regular - 2:
                required_totals_row = last_custom_row + 2
                rows_to_insert_custom = required_totals_row - totals_row_after_regular
                if rows_to_insert_custom < 0:
                    rows_to_insert_custom = 0

    # Insert all rows upfront (regular + custom) before writing any styles
    total_rows_to_insert = rows_to_insert_regular + rows_to_insert_custom
    if total_rows_to_insert > 0:
        ws.insert_rows(SUMMARY_TOTAL_DUE_BASE_ROW, amount=total_rows_to_insert)

        # Copy formatting from template row 18 to new rows (preserve colors, borders, alignment)
        # New rows start at 20, 22, 24, etc.
        template_row = 18  # Use row 18 as template for formatting
        for i in range(total_rows_to_insert):
            new_row = 20 + i
            for col_idx in range(2, 13):  # Columns B through L
                source_cell = ws.cell(row=template_row, column=col_idx)
                target_cell = ws.cell(row=new_row, column=col_idx)
                copy_cell_formatting(source_cell, target_cell)

        # Unmerge ALL cells in newly inserted rows to avoid MergedCell errors
        # Excel automatically adjusts merged ranges when rows are inserted, which can cause conflicts
        first_new_row = SUMMARY_TOTAL_DUE_BASE_ROW
        last_new_row = SUMMARY_TOTAL_DUE_BASE_ROW + total_rows_to_insert - 1
        
        # Collect all merged ranges that intersect with the newly inserted rows
        # We need to unmerge both horizontal and vertical merges in the new rows
        # Do this multiple times to catch all merges (some might be created after unmerging others)
        max_iterations = 3
        for iteration in range(max_iterations):
            merged_ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if merged range intersects with newly inserted rows
                # (min_row <= last_new_row and max_row >= first_new_row)
                if (merged_range.min_row <= last_new_row and merged_range.max_row >= first_new_row):
                    merged_ranges_to_unmerge.append(merged_range)
            
            if not merged_ranges_to_unmerge:
                break  # No more merges to unmerge
            
            # Unmerge the identified ranges (do this in reverse to avoid index issues)
            for merged_range in reversed(merged_ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    range_str = f"{min_cell.coordinate}:{max_cell.coordinate}"
                    ws.unmerge_cells(range_str)
                except Exception:
                    pass  # Skip if unmerge fails

    # Clear style rows and totals row (after insertion so row numbers are correct)
    clear_style_rows(ws, num_styles=num_styles)

    # Generate row indices dynamically based on actual number of styles
    # Each style uses 2 rows (merged): 10-11, 12-13, 14-15, 16-17, 18-19, 20-21, etc.
    # We write to the first row of each pair (10, 12, 14, 16, 18, 20, 22, etc.)
    dynamic_row_indices = []
    start_row = 10
    for i in range(num_styles):
        dynamic_row_indices.append(start_row + (i * 2))

    for idx, row_idx in enumerate(dynamic_row_indices):
        entry = style_entries[idx]
        style_name = entry.get("name", "").strip() or "STYLE"
        complexity_pct = float(entry.get("complexity", 0.0))
        # Handle migration from old "activewear" boolean to "style_type"
        style_type = entry.get("style_type", "Regular")
        if "style_type" not in entry and entry.get("activewear", False):
            style_type = "Activewear/Lingerie/Swim"
        row_options = entry.get("options", {})

        # Calculate base price based on tiered pricing and style type
        # Package rate bracket is based on number of styles only (not custom items)
        line_base_price = calculate_base_price(num_styles, style_type)

        # Check if this is a new row (row_idx > 18) that needs Arial 20 font
        is_new_row = num_styles > 5 and row_idx > 18
        
        # Each style row spans 2 rows (merged cells)
        row_second = row_idx + 1

        # For new rows, explicitly unmerge all cells in this row pair before writing
        # This ensures we start with clean, unmerged cells (prevents horizontal merges)
        if is_new_row:
            ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if merged range intersects with this row pair
                if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                    merged_range.min_col <= 12 and merged_range.max_col >= 2):  # Columns B-L
                    ranges_to_unmerge.append(merged_range)
            
            # Unmerge all overlapping ranges
            for merged_range in reversed(ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass

        # Write style number (#) - use style_number from entry (101, 102, 103, etc.)
        cell_b = ws.cell(row=row_idx, column=2)
        style_number = entry.get("style_number", 101 + idx)  # Default to 101, 102, 103... if not set
        cell_b.value = style_number
        if is_new_row:
            apply_full_border_pair(ws, 2, row_idx, row_second)
            safe_merge_cells(ws, f"B{row_idx}:B{row_second}")
            apply_arial_20_font(cell_b)
            if Alignment is not None:
                cell_b.alignment = Alignment(horizontal="left", vertical="center")
        
        # Write style name (merged across 2 rows, left-aligned)
        cell_c = ws.cell(row=row_idx, column=3)
        cell_c.value = style_name.upper()
        if is_new_row:
            apply_full_border_pair(ws, 3, row_idx, row_second)
            safe_merge_cells(ws, f"C{row_idx}:C{row_second}")
            apply_arial_20_font(cell_c)
            if Alignment is not None:
                cell_c.alignment = Alignment(horizontal="left", vertical="center")
        # For existing rows (10-18), template already has cells merged, don't touch
        
        # Set complexity - show percentage when provided
        cell_e = ws.cell(row=row_idx, column=5)
        if complexity_pct == 0:
            cell_e.value = None
            # Even when complexity is 0, merge and center for new rows
            if is_new_row:
                apply_full_border_pair(ws, 5, row_idx, row_second)
                safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
                apply_arial_20_font(cell_e)
                if Alignment is not None:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell_e.value = complexity_pct / 100.0
            cell_e.number_format = '0%'  # Percentage format
            if is_new_row:
                apply_full_border_pair(ws, 5, row_idx, row_second)
                safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
                apply_arial_20_font(cell_e)
                if Alignment is not None:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")
                elif Alignment is not None and cell_e.alignment:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")

        # Write base price (currency format, integer)
        cell_d = ws.cell(row=row_idx, column=4)
        cell_d.value = int(line_base_price)
        cell_d.number_format = '$#,##0'  # Currency format
        if is_new_row:
            apply_full_border_pair(ws, 4, row_idx, row_second)
            safe_merge_cells(ws, f"D{row_idx}:D{row_second}")
            apply_arial_20_font(cell_d)
            if Alignment is not None:
                cell_d.alignment = Alignment(horizontal="center", vertical="center")
        elif Alignment is not None and cell_d.alignment:
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write total formula (column F)
        cell_f = ws.cell(row=row_idx, column=6)
        if complexity_pct == 0:
            cell_f.value = f"=D{row_idx}"
        else:
            cell_f.value = f"=D{row_idx}*(1+E{row_idx})"
        cell_f.number_format = '$#,##0'  # Currency format
        if is_new_row:
            apply_full_border_pair(ws, 6, row_idx, row_second)
            safe_merge_cells(ws, f"F{row_idx}:F{row_second}")
            apply_arial_20_font(cell_f)
        if Alignment is not None:
            cell_f.alignment = Alignment(horizontal="center", vertical="center")

        # Optional add-ons per row (columns H, I, J, K)
        row_optional_sum = 0.0
        for col_letter, key in optional_cells.items():
            col_num = ord(col_letter) - 64  # Convert letter to column number
            cell_opt = ws.cell(row=row_idx, column=col_num)
            if row_options.get(key):
                price = int(OPTIONAL_PRICES[key])  # Ensure integer
                cell_opt.value = price
                cell_opt.number_format = '$#,##0'  # Currency format
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                    # Merge and center columns H, I, J
                    apply_full_border_pair(ws, col_letter, row_idx, row_second)
                    safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                    if Alignment is not None:
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
                row_optional_sum += price
            else:
                cell_opt.value = None
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                    # Merge and center even if empty
                    apply_full_border_pair(ws, col_letter, row_idx, row_second)
                    safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                    if Alignment is not None:
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
        
        # TOTAL OPTIONAL ADD-ONS now uses columns K and L (merged K10-L11, K12-L13, etc.)
        # Follow the pattern from workbook_creator.py: set value FIRST, then merge
        cell_k = ws.cell(row=row_idx, column=11)
        cell_k.value = f"=SUM(H{row_idx}:J{row_idx})"
        cell_k.number_format = '$#,##0'  # Currency format
        
        # Always merge and center K-L for each style row (K10-L11, K12-L13, etc.)
        # Unmerge first if needed
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_second <= merged_range.max_row and
                merged_range.min_row >= row_idx and
                merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        
        # Re-set value after unmerge (unmerge might have cleared it)
        cell_k = ws.cell(row=row_idx, column=11)
        cell_k.value = f"=SUM(H{row_idx}:J{row_idx})"
        cell_k.number_format = '$#,##0'
        
        # Apply borders to individual cells BEFORE merging (for all rows, not just is_new_row)
        apply_full_border_pair(ws, 11, row_idx, row_second)
        apply_full_border_pair(ws, 12, row_idx, row_second)
        
        # Now merge (value should be preserved in top-left cell K)
        # Always merge K-L for ALL rows (including template rows <= 18), not just is_new_row
        safe_merge_cells(ws, f"K{row_idx}:L{row_second}")
        
        # Get the writable merged cell and apply formatting (use safe_get_writable_cell to handle MergedCell)
        cell_k = safe_get_writable_cell(ws, row_idx, 11)
        
        # Always ensure full borders on merged cell (borders were applied before merge, but ensure they persist)
        if Border is not None and Side is not None:
            thin = Side(style="thin")
            full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            try:
                cell_k.border = full_border
            except Exception:
                pass
        
        # Always apply font and alignment (matching workbook_creator.py - font only for is_new_row, alignment always)
        if is_new_row:
            apply_arial_20_font(cell_k)
        if Alignment is not None:
            cell_k.alignment = Alignment(horizontal="center", vertical="center")

        total_development += line_base_price * (1 + complexity_pct / 100.0)
        total_optional += row_optional_sum

    # Process Custom Items (user-defined price, complexity, no add-ons)
    if num_custom_styles > 0:
        # Calculate starting row for Custom Items (after all regular styles)
        custom_start_row = dynamic_row_indices[-1] + 2 if dynamic_row_indices else 10
        custom_row_indices = []
        for i in range(num_custom_styles):
            custom_row_indices.append(custom_start_row + (i * 2))
        
        # Custom row indices are already calculated correctly since we inserted all rows upfront
        # No need to insert more rows here - they were already inserted above
        
        for idx, row_idx in enumerate(custom_row_indices):
            entry = custom_styles[idx]
            style_name = entry.get("name", "").strip() or "STYLE"
            custom_price = float(entry.get("price", 0.0))
            complexity_pct = float(entry.get("complexity", 0.0))
            
            is_new_row = row_idx > 18
            row_second = row_idx + 1
            
            # For new rows, explicitly unmerge all cells in this row pair before writing
            # This ensures we start with clean, unmerged cells (prevents horizontal merges)
            if is_new_row:
                ranges_to_unmerge = []
                for merged_range in list(ws.merged_cells.ranges):
                    # Check if merged range intersects with this row pair
                    if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                        merged_range.min_col <= 12 and merged_range.max_col >= 2):  # Columns B-L
                        ranges_to_unmerge.append(merged_range)
                
                # Unmerge all overlapping ranges
                for merged_range in reversed(ranges_to_unmerge):
                    try:
                        min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                        ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                    except Exception:
                        pass
            
            # Write style number (#) - use the style_number from entry if available, otherwise default
            cell_b = ws.cell(row=row_idx, column=2)
            # Use style_number from entry if set, otherwise default to 101 + num_styles + idx
            style_number = entry.get("style_number", 101 + num_styles + idx)
            cell_b.value = style_number
            if is_new_row:
                apply_full_border_pair(ws, 2, row_idx, row_second)
                safe_merge_cells(ws, f"B{row_idx}:B{row_second}")
                apply_arial_20_font(cell_b)
                if Alignment is not None:
                    cell_b.alignment = Alignment(horizontal="left", vertical="center")
            
            # Write style name
            cell_c = ws.cell(row=row_idx, column=3)
            cell_c.value = style_name.upper()
            if is_new_row:
                apply_full_border_pair(ws, 3, row_idx, row_second)
                safe_merge_cells(ws, f"C{row_idx}:C{row_second}")
                apply_arial_20_font(cell_c)
                if Alignment is not None:
                    cell_c.alignment = Alignment(horizontal="left", vertical="center")
            
            # Write custom price
            cell_d = ws.cell(row=row_idx, column=4)
            cell_d.value = int(custom_price)
            cell_d.number_format = '$#,##0'
            if is_new_row:
                apply_full_border_pair(ws, 4, row_idx, row_second)
                safe_merge_cells(ws, f"D{row_idx}:D{row_second}")
                apply_arial_20_font(cell_d)
                if Alignment is not None:
                    cell_d.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write complexity
            cell_e = ws.cell(row=row_idx, column=5)
            if complexity_pct == 0:
                cell_e.value = None
            else:
                cell_e.value = complexity_pct / 100.0
                cell_e.number_format = '0%'
            if is_new_row:
                apply_full_border_pair(ws, 5, row_idx, row_second)
                safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
                apply_arial_20_font(cell_e)
                if Alignment is not None:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write total formula
            cell_f = ws.cell(row=row_idx, column=6)
            if complexity_pct == 0:
                cell_f.value = f"=D{row_idx}"
            else:
                cell_f.value = f"=D{row_idx}*(1+E{row_idx})"
            cell_f.number_format = '$#,##0'
            if is_new_row:
                apply_full_border_pair(ws, 6, row_idx, row_second)
                safe_merge_cells(ws, f"F{row_idx}:F{row_second}")
                apply_arial_20_font(cell_f)
                if Alignment is not None:
                    cell_f.alignment = Alignment(horizontal="center", vertical="center")
            
            # Clear add-ons (Custom Items don't have add-ons)
            for col_letter in ["H", "I", "J", "K"]:
                col_num = ord(col_letter) - 64
                cell_opt = ws.cell(row=row_idx, column=col_num)
                cell_opt.value = None
                if is_new_row:
                    apply_full_border_pair(ws, col_letter, row_idx, row_second)
                    safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                    if Alignment is not None:
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
            
            # Clear total optional add-on - merge K-L for custom items when total_styles_count > 5 or is_new_row
            cell_k = ws.cell(row=row_idx, column=11)
            cell_k.value = None
            cell_l = ws.cell(row=row_idx, column=12)
            cell_l.value = None
            
            # Merge K-L for custom items if total_styles_count > 5 or is_new_row (matching workbook_creator.py)
            if is_new_row or total_styles_count > 5:
                # Unmerge first if needed
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= row_second <= merged_range.max_row and
                        merged_range.min_row >= row_idx and
                        merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                
                # Apply borders to individual cells BEFORE merging
                apply_full_border_pair(ws, 11, row_idx, row_second)
                apply_full_border_pair(ws, 12, row_idx, row_second)
                
                # Merge K-L
                safe_merge_cells(ws, f"K{row_idx}:L{row_second}")
                
                # Get writable cell and apply formatting
                cell_k = safe_get_writable_cell(ws, row_idx, 11)
                if total_styles_count > 5 and Border is not None and Side is not None:
                    thin = Side(style="thin")
                    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    try:
                        cell_k.border = full_border
                    except Exception:
                        pass
                if Alignment is not None:
                    cell_k.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # For template rows (<= 18), just merge L individually
                if is_new_row:
                    apply_full_border_pair(ws, 12, row_idx, row_second)
                    safe_merge_cells(ws, f"L{row_idx}:L{row_second}")
                    if Alignment is not None:
                        cell_l.alignment = Alignment(horizontal="center", vertical="center")
            
            total_development += custom_price * (1 + complexity_pct / 100.0)
        
        # For less than 5 styles, merge, center and apply all borders to empty OPTIONAL ADD-ONS totals
        if total_styles_count <= 5 and Border is not None and Side is not None:
            # Find which style rows are empty (not used)
            used_rows = set()
            if dynamic_row_indices:
                for row_idx in dynamic_row_indices:
                    used_rows.add(row_idx)
            
            # Process each row in ROW_INDICES that wasn't used
            for row_idx in ROW_INDICES:
                if row_idx not in used_rows:
                    row_second = row_idx + 1
                    # Check if K-L cells are empty (no formula/value)
                    cell_k_value = safe_get_cell_value(ws, row_idx, 11)
                    if cell_k_value is None or (isinstance(cell_k_value, str) and cell_k_value.strip() == ""):
                        # Unmerge any existing merges in K-L for this row pair
                        for merged_range in list(ws.merged_cells.ranges):
                            if (merged_range.min_row <= row_idx <= merged_range.max_row <= row_second and
                                merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                                try:
                                    ws.unmerge_cells(range_string=str(merged_range))
                                except Exception:
                                    pass
                        
                        # Apply full borders to K and L columns
                        apply_full_border_pair(ws, 11, row_idx, row_second)
                        apply_full_border_pair(ws, 12, row_idx, row_second)
                        
                        # Merge and center K-L
                        safe_merge_cells(ws, f"K{row_idx}:L{row_second}")
                        if Alignment is not None:
                            # Get writable cell for alignment
                            cell_k = safe_get_writable_cell(ws, row_idx, 11)
                            try:
                                cell_k.alignment = Alignment(horizontal="center", vertical="center")
                            except Exception:
                                pass
                        # Clear L cell since it's merged with K (handle MergedCell)
                        try:
                            cell_l = safe_get_writable_cell(ws, row_idx, 12)
                            cell_l.value = None
                        except Exception:
                            # If it's a MergedCell, the value is already None or handled by merge
                            pass
        
        # Update last_style_row to include Custom Items (for totals calculations)
        if custom_row_indices:
            last_style_row = custom_row_indices[-1]
    
    # For less than 5 styles, merge, center and apply all borders to empty OPTIONAL ADD-ONS totals
    if total_styles_count <= 5 and Border is not None and Side is not None:
        # Find which style rows are empty (not used)
        used_rows = set()
        if dynamic_row_indices:
            for row_idx in dynamic_row_indices:
                used_rows.add(row_idx)
        
        # Process each row in ROW_INDICES that wasn't used
        for row_idx in ROW_INDICES:
            if row_idx not in used_rows:
                row_second = row_idx + 1
                # Check if K-L cells are empty (no formula/value)
                cell_k_value = safe_get_cell_value(ws, row_idx, 11)
                if cell_k_value is None or (isinstance(cell_k_value, str) and cell_k_value.strip() == ""):
                    # Unmerge any existing merges in K-L for this row pair
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= row_idx <= merged_range.max_row <= row_second and
                            merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass
                    
                    # Apply full borders to K and L columns
                    apply_full_border_pair(ws, 11, row_idx, row_second)
                    apply_full_border_pair(ws, 12, row_idx, row_second)
                    
                    # Merge and center K-L
                    safe_merge_cells(ws, f"K{row_idx}:L{row_second}")
                    if Alignment is not None:
                        # Get writable cell for alignment
                        cell_k = safe_get_writable_cell(ws, row_idx, 11)
                        try:
                            cell_k.alignment = Alignment(horizontal="center", vertical="center")
                        except Exception:
                            pass
                    # Clear L cell since it's merged with K (handle MergedCell)
                    try:
                        cell_l = safe_get_writable_cell(ws, row_idx, 12)
                        cell_l.value = None
                    except Exception:
                        # If it's a MergedCell, the value is already None or handled by merge
                        pass

    # Determine last_style_row if no Custom Items (for totals calculations)
    if num_custom_styles == 0:
        if dynamic_row_indices:
            last_style_row = dynamic_row_indices[-1]
        else:
            last_style_row = 10

    # Determine last_regular_style_row (only regular + activewear, excluding Custom Items)
    # This is used for deliverables counts (Patterns, First Samples, Final Samples)
    if dynamic_row_indices:
        last_regular_style_row = dynamic_row_indices[-1]
    else:
        last_regular_style_row = 10

    # Count activewear and regular styles (checking style_type instead of activewear boolean)
    num_activewear = sum(1 for entry in style_entries if entry.get("style_type", "Regular") == "Activewear/Lingerie/Swim" or (entry.get("style_type") is None and entry.get("activewear", False)))
    num_regular = sum(1 for entry in style_entries if entry.get("style_type", "Regular") != "Activewear/Lingerie/Swim" and not (entry.get("style_type") is None and entry.get("activewear", False)))
    
    total_extra_rows = max(total_styles_count - len(ROW_INDICES), 0) * 2
    deliverables_block_start = DELIVERABLE_BLOCK_START + total_extra_rows
    deliverables_block_end = deliverables_block_start + DELIVERABLE_BLOCK_HEIGHT - 1
    restore_deliverables_block(ws, deliverables_template, deliverables_block_start)
    
    # If there are activewear styles, modify deliverables section
    if num_activewear > 0 and column_index_from_string is not None:
        label_column_idx = column_index_from_string("B")
        col_c_idx = column_index_from_string("C")
        col_d_idx = column_index_from_string("D")
        final_samples_row = None
        row_final_samples_new = None  # Store the row where new FINAL SAMPLES is created
        
        # Find the "FINAL SAMPLES" row (should be around row 31-32)
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=label_column_idx).value
            if isinstance(value, str):
                value_lower = value.strip().lower()
                if "final samples" in value_lower and final_samples_row is None:
                    final_samples_row = scan_row
                    break
        
        if final_samples_row:
            # Step 1: Replace FINAL SAMPLES (rows 31-32) with SECOND SAMPLES (only columns B-D)
            # Unmerge cells in columns B-D for these rows first
            for row in [final_samples_row, final_samples_row + 1]:
                for col in [label_column_idx, col_c_idx, col_d_idx]:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass

            # Clear values in row 32 (second row of pair) for columns B-D
            for col in [label_column_idx, col_c_idx, col_d_idx]:
                if get_column_letter:
                    safe_set_cell_value(ws, f"{get_column_letter(col)}{final_samples_row + 1}", None)
                else:
                    ws.cell(row=final_samples_row + 1, column=col).value = None

            # Clear column C in row 31 (to ensure no leftover values)
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{final_samples_row}", None)
            else:
                ws.cell(row=final_samples_row, column=col_c_idx).value = None

            # Set SECOND SAMPLES in row 31, column B, count in column D
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{final_samples_row}", "SECOND SAMPLES")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{final_samples_row}", num_activewear)
            else:
                ws.cell(row=final_samples_row, column=label_column_idx).value = "SECOND SAMPLES"
                ws.cell(row=final_samples_row, column=col_d_idx).value = num_activewear
            ws.cell(row=final_samples_row, column=col_d_idx).number_format = "0"
            
            # Use row 29 (ROUND OF REVISIONS) as reference - it should be around row 29
            # Find ROUND OF REVISIONS row (should be row 29)
            reference_row = None
            for scan_row in range(deliverables_block_start, final_samples_row):
                value = ws.cell(row=scan_row, column=label_column_idx).value
                if isinstance(value, str) and "round of revisions" in value.lower():
                    reference_row = scan_row
                    break
            
            # If not found, assume it's row 29 (typical position)
            if reference_row is None:
                reference_row = 29
            
            # Check what merges exist for columns B-D in rows 29-30 (reference_row to reference_row+1)
            b_merged = False
            c_merged = False
            d_merged = False
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row == reference_row and merged_range.max_row == reference_row + 1):
                    if merged_range.min_col == 2 and merged_range.max_col == 2:  # Column B
                        b_merged = True
                    elif merged_range.min_col == 3 and merged_range.max_col == 3:  # Column C
                        c_merged = True
                    elif merged_range.min_col == 4 and merged_range.max_col == 4:  # Column D
                        d_merged = True
            
            # Copy exact formatting from B29-D30 to B31-D32 (SECOND SAMPLES)
            # First, copy cell formatting (but not alignment for column B - we'll center it after merge)
            for col in [label_column_idx, col_c_idx, col_d_idx]:
                source_cell_29 = ws.cell(row=reference_row, column=col)
                source_cell_30 = ws.cell(row=reference_row + 1, column=col)
                target_cell_31 = ws.cell(row=final_samples_row, column=col)
                target_cell_32 = ws.cell(row=final_samples_row + 1, column=col)
                
                # Copy formatting from row 29 to row 31
                if source_cell_29.font:
                    target_cell_31.font = copy(source_cell_29.font)
                if source_cell_29.fill:
                    target_cell_31.fill = copy(source_cell_29.fill)
                if source_cell_29.border:
                    target_cell_31.border = copy(source_cell_29.border)
                # Don't copy alignment for column B - we'll set it to center after merge
                if col != label_column_idx and source_cell_29.alignment:
                    target_cell_31.alignment = copy(source_cell_29.alignment)
                target_cell_31.number_format = source_cell_29.number_format
                
                # Copy formatting from row 30 to row 32
                if source_cell_30.font:
                    target_cell_32.font = copy(source_cell_30.font)
                if source_cell_30.fill:
                    target_cell_32.fill = copy(source_cell_30.fill)
                if source_cell_30.border:
                    target_cell_32.border = copy(source_cell_30.border)
                # Don't copy alignment for column B - we'll set it to center after merge
                if col != label_column_idx and source_cell_30.alignment:
                    target_cell_32.alignment = copy(source_cell_30.alignment)
                target_cell_32.number_format = source_cell_30.number_format
            
            # Merge B:C together and D separately, with center alignment
            if safe_merge_cells:
                # Merge B:C together (columns B and C merged across 2 rows)
                safe_merge_cells(ws, f"B{final_samples_row}:C{final_samples_row + 1}")
                # Set center alignment for merged B:C
                if Alignment:
                    cell_bc = ws.cell(row=final_samples_row, column=label_column_idx)
                    cell_bc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                # Merge column D
                safe_merge_cells(ws, f"D{final_samples_row}:D{final_samples_row + 1}")
                # Set center alignment for merged column D
                if Alignment:
                    cell_d = ws.cell(row=final_samples_row, column=col_d_idx)
                    cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Step 2: Insert 6 rows after row 32 (3 new deliverable entries, each with 2 rows)
            insert_row = final_samples_row + 2  # After row 32
            rows_to_insert = 6
            
            # Insert rows
            ws.insert_rows(insert_row, amount=rows_to_insert)
            
            # Unmerge any cells in the newly inserted rows
            for i in range(rows_to_insert):
                target_row = insert_row + i
                for col in range(DELIVERABLE_COL_START, DELIVERABLE_COL_END + 1):
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= target_row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass
            
            # Step 3: Set values for the new rows and define row variables
            # Row 33-34: 2ND ROUND OF FITTINGS (count in column D)
            row_2nd_fittings = insert_row
            # Clear column C to ensure no leftover values
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_2nd_fittings}", None)
            else:
                ws.cell(row=row_2nd_fittings, column=col_c_idx).value = None
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_2nd_fittings}", "2ND ROUND OF FITTINGS")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{row_2nd_fittings}", 1)
            else:
                ws.cell(row=row_2nd_fittings, column=label_column_idx).value = "2ND ROUND OF FITTINGS"
                ws.cell(row=row_2nd_fittings, column=col_d_idx).value = 1
            ws.cell(row=row_2nd_fittings, column=col_d_idx).number_format = "0"
            row_2nd_fittings_2 = insert_row + 1
            
            # Row 35-36: 2ND ROUND OF REVISIONS (count in column D)
            row_2nd_revisions = insert_row + 2
            # Clear column C to ensure no leftover values
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_2nd_revisions}", None)
            else:
                ws.cell(row=row_2nd_revisions, column=col_c_idx).value = None
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_2nd_revisions}", "2ND ROUND OF REVISIONS")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{row_2nd_revisions}", 1)
            else:
                ws.cell(row=row_2nd_revisions, column=label_column_idx).value = "2ND ROUND OF REVISIONS"
                ws.cell(row=row_2nd_revisions, column=col_d_idx).value = 1
            ws.cell(row=row_2nd_revisions, column=col_d_idx).number_format = "0"
            row_2nd_revisions_2 = insert_row + 3
            
            # Row 37-38: FINAL SAMPLES (count of all styles)
            row_final_samples = insert_row + 4
            row_final_samples_new = row_final_samples  # Store for later use
            # Clear column C to ensure no leftover values
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_final_samples}", None)
            else:
                ws.cell(row=row_final_samples, column=col_c_idx).value = None
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_final_samples}", "FINAL SAMPLES")
                # Will be set by formulas section below
            else:
                ws.cell(row=row_final_samples, column=label_column_idx).value = "FINAL SAMPLES"
            row_final_samples_2 = insert_row + 5
            
            # Step 4: Copy exact formatting from B29-D30 to each new row pair (B33-D34, B35-D36, B37-D38)
            # Use row 29 as source for first row of each pair, row 30 for second row
            # Don't copy alignment for column B - we'll center it after merge
            for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                row_second = base_row + 1
                # Copy formatting from row 29 to first row of pair
                for col in [label_column_idx, col_c_idx, col_d_idx]:
                    source_cell_29 = ws.cell(row=reference_row, column=col)
                    source_cell_30 = ws.cell(row=reference_row + 1, column=col)
                    target_cell_first = ws.cell(row=base_row, column=col)
                    target_cell_second = ws.cell(row=row_second, column=col)
                    
                    # Copy formatting from row 29 to first row
                    if source_cell_29.font:
                        target_cell_first.font = copy(source_cell_29.font)
                    if source_cell_29.fill:
                        target_cell_first.fill = copy(source_cell_29.fill)
                    if source_cell_29.border:
                        target_cell_first.border = copy(source_cell_29.border)
                    # Don't copy alignment for column B - we'll set it to center after merge
                    if col != label_column_idx and source_cell_29.alignment:
                        target_cell_first.alignment = copy(source_cell_29.alignment)
                    target_cell_first.number_format = source_cell_29.number_format
                    
                    # Copy formatting from row 30 to second row
                    if source_cell_30.font:
                        target_cell_second.font = copy(source_cell_30.font)
                    if source_cell_30.fill:
                        target_cell_second.fill = copy(source_cell_30.fill)
                    if source_cell_30.border:
                        target_cell_second.border = copy(source_cell_30.border)
                    # Don't copy alignment for column B - we'll set it to center after merge
                    if col != label_column_idx and source_cell_30.alignment:
                        target_cell_second.alignment = copy(source_cell_30.alignment)
                    target_cell_second.number_format = source_cell_30.number_format
            
            # Step 5: Merge B:C together and D separately, with center alignment for each new row pair
            if safe_merge_cells:
                for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                    row_second = base_row + 1
                    # Merge B:C together (columns B and C merged across 2 rows)
                    safe_merge_cells(ws, f"B{base_row}:C{row_second}")
                    # Set center alignment for merged B:C
                    if Alignment:
                        cell_bc = ws.cell(row=base_row, column=label_column_idx)
                        cell_bc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    # Merge column D
                    safe_merge_cells(ws, f"D{base_row}:D{row_second}")
                    # Set center alignment for merged column D
                    if Alignment:
                        cell_d = ws.cell(row=base_row, column=col_d_idx)
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    
                    # Set FINAL SAMPLES count immediately after merging
                    if base_row == row_final_samples:
                        # Unmerge column D temporarily to set the value
                        try:
                            ws.unmerge_cells(f"D{base_row}:D{row_second}")
                        except Exception:
                            pass
                        # Set the count value (total styles: regular + activewear, excluding custom line items)
                        count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                        # Clear the cell completely (including any formulas) before setting value
                        count_cell = ws.cell(row=base_row, column=col_d_idx)
                        count_cell.value = None
                        # Set direct numeric value (not a formula)
                        count_cell.value = count_value
                        count_cell.number_format = "0"
                        # Re-merge column D
                        safe_merge_cells(ws, f"D{base_row}:D{row_second}")
                        # Restore center alignment
                        if Alignment:
                            cell_d = ws.cell(row=base_row, column=col_d_idx)
                            cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Restore merged cells for right side columns (H-P) using template merge patterns
            if deliverables_template and "merges" in deliverables_template:
                right_side_merges = []
                for min_row_offset, max_row_offset, min_col, max_col in deliverables_template.get("merges", []):
                    if (min_col >= 8 and max_col <= 16 and 
                        max_row_offset - min_row_offset == 1):
                        right_side_merges.append((min_col, max_col))
                
                # Apply merge patterns to new rows and center align
                for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                    row_second = base_row + 1
                    for min_col, max_col in right_side_merges:
                        try:
                            if get_column_letter:
                                start_col_letter = get_column_letter(min_col)
                                end_col_letter = get_column_letter(max_col)
                                range_str = f"{start_col_letter}{base_row}:{end_col_letter}{row_second}"
                                safe_merge_cells(ws, range_str)
                                # Center align the merged cell
                                if Alignment:
                                    cell = ws.cell(row=base_row, column=min_col)
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        except Exception:
                            pass
            
            # Center align all right side columns (H-P) in the entire deliverables section
            if Alignment:
                for row in range(deliverables_block_start, deliverables_block_end + 1):
                    for col in range(8, 17):  # Columns H (8) through P (16)
                        cell = ws.cell(row=row, column=col)
                        if cell.alignment:
                            cell.alignment = Alignment(
                                horizontal='center',
                                vertical=cell.alignment.vertical,
                                wrap_text=cell.alignment.wrap_text,
                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                indent=cell.alignment.indent
                            )
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Update deliverables_block_end to include new rows
            deliverables_block_end += rows_to_insert
        
        # Center align all right side columns (H-P) in the entire deliverables section (including original rows)
        if Alignment:
            for row in range(deliverables_block_start, deliverables_block_end + 1):
                for col in range(8, 17):  # Columns H (8) through P (16)
                    cell = ws.cell(row=row, column=col)
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    else:
        # Even if no activewear, center align right side columns (H-P) in deliverables section
        if Alignment:
            for row in range(deliverables_block_start, deliverables_block_end + 1):
                for col in range(8, 17):  # Columns H (8) through P (16)
                    cell = ws.cell(row=row, column=col)
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Update deliverables table counts (column J) for add-on selections
    if column_index_from_string is not None and total_styles_count > 0:
        label_column_idx = column_index_from_string("H")
        target_col_j = column_index_from_string("J")
        deliverable_addon_map = [
            ("WASH/DYE", "H"),
            ("DESIGN", "I"),
            ("TREATMENT", "J"),
        ]

        def find_label_row(label_text: str) -> int | None:
            """Locate the row index for a given deliverable label."""
            lowered = label_text.strip().lower()
            partial_match_row = None
            for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
                value = ws.cell(row=scan_row, column=label_column_idx).value
                if not isinstance(value, str):
                    continue
                value_clean = value.strip().lower()
                if value_clean == lowered:
                    return scan_row
                if lowered in value_clean and partial_match_row is None:
                    partial_match_row = scan_row
            return partial_match_row

        for label_text, addon_col_letter in deliverable_addon_map:
            row_idx = find_label_row(label_text)
            if row_idx is None:
                continue
            addon_range = f"{addon_col_letter}10:{addon_col_letter}{last_style_row}"
            cell = ws.cell(row=row_idx, column=target_col_j)
            cell.value = f"=COUNT({addon_range})"
            cell.number_format = "0"

        # Round of Fittings: always 1
        fittings_row = find_label_row("ROUND OF FITTINGS")
        if fittings_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= fittings_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=fittings_row, column=col_d_idx)
            count_cell.value = None
            # Set direct numeric value (not a formula)
            count_cell.value = 1
            count_cell.number_format = "0"
        
        # Round of Revisions: 
        # - 1 if there's ONLY Regular styles (non-Activewear) OR ONLY Activewear styles
        # - 2 if there's BOTH Regular AND Activewear styles
        # Note: ROUND OF REVISIONS is in column B, not column H, so we need to search column B
        label_col_b = column_index_from_string("B")
        revisions_row = None
        revisions_label_lower = "round of revisions"
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=label_col_b).value
            if isinstance(value, str) and revisions_label_lower in value.strip().lower():
                revisions_row = scan_row
                break
        if revisions_row:
            col_d_idx = column_index_from_string("D")
            # Check if this cell is part of a merged range and remember the merge pattern
            was_merged = False
            merge_pattern = None
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= revisions_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    was_merged = True
                    merge_pattern = (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
                    break
            # SIMPLEST APPROACH: Use num_regular and num_activewear already calculated from style_entries
            # These are calculated above from the actual style_entries data - 100% reliable
            # No need to read from Excel cells which might have text/number issues
            
            # Calculate revisions count: 2 if both regular AND activewear exist, 1 otherwise
            # num_regular and num_activewear are calculated at lines 884-885 from style_entries
            revisions_count = 2 if (num_regular > 0 and num_activewear > 0) else 1
            
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=revisions_row, column=col_d_idx)
            count_cell.value = None
            
            # Set the calculated value directly (not a formula) - this will ALWAYS work
            count_cell.value = revisions_count
            count_cell.number_format = "0"
            if Alignment is not None:
                count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Re-merge AFTER setting the value
            if was_merged and merge_pattern and safe_merge_cells:
                min_row, max_row, min_col, max_col = merge_pattern
                if get_column_letter:
                    start_col_letter = get_column_letter(min_col)
                    end_col_letter = get_column_letter(max_col)
                    range_str = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"
                    safe_merge_cells(ws, range_str)
                    # Re-apply alignment after merging
                    if Alignment is not None:
                        merged_cell = ws.cell(row=revisions_row, column=col_d_idx)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        # If activewear exists, handle SECOND SAMPLES and the new rows
        if num_activewear > 0:
            # SECOND SAMPLES: count of activewear styles (replaced FINAL SAMPLES at rows 31-32)
            # Note: count is already set in column D above, but we ensure it's correct here
            second_sample_row = find_label_row("SECOND SAMPLES")
            if second_sample_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_sample_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=second_sample_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = num_activewear
                count_cell.number_format = "0"
            
            # 2nd Round of Fittings: always 1 (already set in column D above)
            second_fittings_row = find_label_row("2ND ROUND OF FITTINGS")
            if second_fittings_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_fittings_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=second_fittings_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = 1
                count_cell.number_format = "0"
            
            # 2nd Round of Revisions: always 1 for Active category (already set in column D above)
            second_revisions_row = find_label_row("2ND ROUND OF REVISIONS")
            if second_revisions_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_revisions_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=second_revisions_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = 1
                count_cell.number_format = "0"
            
            # Final Samples: count is already set in the merge section above, but verify it's correct here
            # (This is a backup check - the count should already be set when merging FINAL SAMPLES)
            # Use row_final_samples_new if available (the newly created FINAL SAMPLES row), otherwise search for it
            # Always prefer row_final_samples_new since it's the newly created row
            final_samples_row_to_use = row_final_samples_new if row_final_samples_new else find_label_row("FINAL SAMPLES")
            if final_samples_row_to_use:
                final_samples_row = final_samples_row_to_use
                # Use direct count instead of formula
                count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                col_d_idx = column_index_from_string("D")
                # Unmerge temporarily to set value and clear any formulas
                was_merged = False
                merge_pattern = None
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        was_merged = True
                        merge_pattern = (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                        break
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=final_samples_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = count_value
                count_cell.number_format = "0"
                # Re-merge if it was merged before
                if was_merged and merge_pattern and safe_merge_cells:
                    min_row, max_row, min_col, max_col = merge_pattern
                    if get_column_letter:
                        start_col_letter = get_column_letter(min_col)
                        end_col_letter = get_column_letter(max_col)
                        range_str = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"
                        safe_merge_cells(ws, range_str)
                    if Alignment:
                        cell_d = ws.cell(row=final_samples_row, column=col_d_idx)
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        else:
            # When there's no activewear, set FINAL SAMPLES here
            # Final Samples: all styles (regular only, no activewear, excluding custom line items)
            # Note: FINAL SAMPLES is in column B, not column H, so we need to search column B
            final_samples_row = None
            label_col_b = column_index_from_string("B")
            for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
                value = ws.cell(row=scan_row, column=label_col_b).value
                if isinstance(value, str) and "final" in value.lower() and "sample" in value.lower():
                    final_samples_row = scan_row
                    break
            if final_samples_row:
                # Use direct count instead of formula
                count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                col_d_idx = column_index_from_string("D")
                # Unmerge temporarily to set value and clear any formulas
                was_merged = False
                merge_pattern = None
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        was_merged = True
                        merge_pattern = (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                        break
                # Clear the cell completely (including any formulas)
                count_cell = ws.cell(row=final_samples_row, column=col_d_idx)
                count_cell.value = None
                # Set direct numeric value (not a formula)
                count_cell.value = count_value
                count_cell.number_format = "0"
                # Re-merge if it was merged before
                if was_merged and merge_pattern and safe_merge_cells:
                    min_row, max_row, min_col, max_col = merge_pattern
                    if get_column_letter:
                        start_col_letter = get_column_letter(min_col)
                        end_col_letter = get_column_letter(max_col)
                        range_str = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"
                        safe_merge_cells(ws, range_str)
                    if Alignment:
                        cell_d = ws.cell(row=final_samples_row, column=col_d_idx)
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        # Update PATTERNS and FIRST SAMPLES to have the same value as FINAL SAMPLES (num_styles)
        # This applies to both activewear and non-activewear cases, after find_label_row is defined
        # Note: find_label_row searches column H, but PATTERNS/FIRST SAMPLES are in column B
        # So we need to search column B instead
        patterns_row = None
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=column_index_from_string("B")).value
            if isinstance(value, str) and "pattern" in value.lower():
                patterns_row = scan_row
                break
        
        if patterns_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= patterns_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=patterns_row, column=col_d_idx)
            count_cell.value = None
            # Set direct numeric value (same as FINAL SAMPLES: num_styles)
            count_cell.value = num_styles
            count_cell.number_format = "0"
            # Merge and center (PATTERNS typically spans 2 rows)
            patterns_row_second = patterns_row + 1
            if safe_merge_cells:
                safe_merge_cells(ws, f"D{patterns_row}:D{patterns_row_second}")
            if Alignment:
                count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        first_samples_row = None
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = ws.cell(row=scan_row, column=column_index_from_string("B")).value
            if isinstance(value, str) and "first" in value.lower() and "sample" in value.lower():
                first_samples_row = scan_row
                break
        
        if first_samples_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= first_samples_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear the cell completely (including any formulas)
            count_cell = ws.cell(row=first_samples_row, column=col_d_idx)
            count_cell.value = None
            # Set direct numeric value (same as FINAL SAMPLES: num_styles)
            count_cell.value = num_styles
            count_cell.number_format = "0"
            # Merge and center (FIRST SAMPLES typically spans 2 rows)
            first_samples_row_second = first_samples_row + 1
            if safe_merge_cells:
                safe_merge_cells(ws, f"D{first_samples_row}:D{first_samples_row_second}")
            if Alignment:
                count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Totals section - dynamically calculate totals row and range based on number of styles
    # For 5 or fewer styles: totals at row 20 (original position)
    # For more than 5 styles: totals row shifts down by (total_styles_count - 5) * 2 rows
    if total_styles_count > 0:
        first_style_row = dynamic_row_indices[0] if dynamic_row_indices else 10
        if total_styles_count <= 5:
            totals_row = 20  # Original totals row position
        else:
            # Totals row shifts down by the number of rows we inserted
            totals_row = 20 + (total_styles_count - 5) * 2
        
        # Unmerge any merged cells in the totals row to avoid issues
        merged_ranges_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row):
                merged_ranges_to_unmerge.append(merged_range)
        
        for merged_range in merged_ranges_to_unmerge:
            try:
                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
            except Exception:
                pass
        
        # Set totals row labels
        cell_b_totals = ws.cell(row=totals_row, column=2)
        cell_b_totals.value = "TOTAL DEVELOPMENT"
        if Font is not None:
            cell_b_totals.font = Font(bold=True)
        
        # Merge and center "TOTAL OPTIONAL ADD-ONS" H20-J20 (matching workbook_creator.py)
        cell_h_totals = ws.cell(row=totals_row, column=8)
        cell_h_totals.value = "TOTAL OPTIONAL ADD-ONS"
        if Font is not None:
            cell_h_totals.font = Font(bold=True)
        # Unmerge any existing merges in H-J for totals row
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row and
                merged_range.min_col <= 8 <= merged_range.max_col <= 10):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        # Merge H-J
        safe_merge_cells(ws, f"H{totals_row}:J{totals_row}")
        # Get writable cell after merge for alignment and borders
        cell_h_totals = safe_get_writable_cell(ws, totals_row, 8)
        if Alignment is not None:
            try:
                cell_h_totals.alignment = Alignment(horizontal="left", vertical="center")
            except Exception:
                pass
        # Apply full borders to H-J merged cell (for more than 5 styles)
        if total_styles_count > 5 and Border is not None and Side is not None:
            thin = Side(style="thin")
            try:
                cell_h_totals.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            except Exception:
                pass
        
        # Set totals formulas - sum all style rows (dynamic based on actual style rows)
        cell_f_totals = ws.cell(row=totals_row, column=6)
        cell_f_totals.value = f"=SUM(F10:F{last_style_row})"
        cell_f_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_f_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_f_totals.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #709171 to TOTAL DEVELOPMENT
        if PatternFill is not None:
            cell_f_totals.fill = PatternFill(start_color="709171", end_color="709171", fill_type="solid")
        
        # TOTAL OPTIONAL ADD-ONS now uses columns K-L (merged K20-L20)
        cell_k_totals = ws.cell(row=totals_row, column=11)
        cell_l_totals = ws.cell(row=totals_row, column=12)
        # Formula now sums K column (which contains SUM(H:J) for each row)
        cell_k_totals.value = f"=SUM(K10:K{last_style_row})"
        cell_k_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_k_totals.font = Font(bold=True)
        # Unmerge any existing merges in K-L for totals row
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row and
                merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        # Merge and center K-L
        safe_merge_cells(ws, f"K{totals_row}:L{totals_row}")
        # Get writable cell after merge for alignment, font, fill, and borders
        cell_k_totals = safe_get_writable_cell(ws, totals_row, 11)
        if Alignment is not None:
            try:
                cell_k_totals.alignment = Alignment(horizontal="center", vertical="center")
            except Exception:
                pass
        # Apply font size 20 to TOTAL OPTIONAL ADD-ONS
        if Font is not None:
            try:
                cell_k_totals.font = Font(name="Arial", size=20, bold=True)
            except Exception:
                pass
        # Apply cell color #f0cfbb to TOTAL OPTIONAL ADD-ONS
        if PatternFill is not None:
            try:
                cell_k_totals.fill = PatternFill(start_color="F0CFBB", end_color="F0CFBB", fill_type="solid")
            except Exception:
                pass
        # Clear L cell since it's merged with K
        try:
            cell_l_totals.value = None
        except Exception:
            pass
        
        # Apply full borders to K-L merged cell (ALWAYS, not just for more than 5 styles)
        # For merged cells, we need to ensure the border is applied correctly
        # Get the merged cell and apply borders
        cell_k_totals = safe_get_writable_cell(ws, totals_row, 11)
        if Border is not None and Side is not None:
            thin = Side(style="thin")
            full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            try:
                # Apply border to the merged cell (top-left cell K)
                cell_k_totals.border = full_border
            except Exception:
                try:
                    # Also apply border to column L (rightmost of merged range)
                    cell_l_totals = ws.cell(row=totals_row, column=12)
                    cell_l_totals.border = full_border
                except Exception:
                    pass

        # SUB-TOTAL row (new) and Discount row (moved down)
        subtotal_row = SUMMARY_SUBTOTAL_ROW
        discount_row = SUMMARY_DISCOUNT_ROW
        discount_decimal = discount_percentage / 100.0 if discount_percentage else 0.0
        cell_n_subtotal = ws.cell(row=subtotal_row, column=SUMMARY_LABEL_COL)
        cell_p_subtotal = ws.cell(row=subtotal_row, column=SUMMARY_VALUE_COL)
        cell_n_discount = ws.cell(row=discount_row, column=SUMMARY_LABEL_COL)
        cell_p_discount = ws.cell(row=discount_row, column=SUMMARY_VALUE_COL)
        
        if discount_percentage > 0:
            # SUB-TOTAL = TOTAL DEVELOPMENT + TOTAL OPTIONAL ADD-ONS
            cell_n_subtotal.value = "SUB-TOTAL"
            if Font is not None:
                cell_n_subtotal.font = Font(name="Arial", size=20, bold=True, color=cell_n_subtotal.font.color if cell_n_subtotal.font else None)
            if Alignment is not None:
                cell_n_subtotal.alignment = Alignment(horizontal="center", vertical="center")
            cell_p_subtotal.value = "=SUM(P10:P13)"
            cell_p_subtotal.number_format = '$#,##0'
            if Font is not None:
                cell_p_subtotal.font = Font(name="Arial", size=20, bold=True, color=cell_p_subtotal.font.color if cell_p_subtotal.font else None)
            if Alignment is not None:
                cell_p_subtotal.alignment = Alignment(horizontal="center", vertical="center")
            
            # DISCOUNT uses SUB-TOTAL as base
            cell_n_discount.value = f"DISCOUNT ({discount_percentage:.0f}%)"
            if Font is not None:
                cell_n_discount.font = Font(name="Arial", size=20, bold=True, color=cell_n_discount.font.color if cell_n_discount.font else None)
            if Alignment is not None:
                cell_n_discount.alignment = Alignment(horizontal="center", vertical="center")
            
            cell_p_discount.value = f"=P{subtotal_row}*{discount_decimal}"
            cell_p_discount.number_format = '$#,##0'
            if Font is not None:
                cell_p_discount.font = Font(name="Arial", size=20, bold=True, color=cell_p_discount.font.color if cell_p_discount.font else None)
            if Alignment is not None:
                cell_p_discount.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Clear SUB-TOTAL and DISCOUNT rows when no discount
            for clear_row in [subtotal_row, subtotal_row + 1, discount_row, discount_row + 1]:
                safe_set_cell_value(ws, f"N{clear_row}", None)
                safe_set_cell_value(ws, f"P{clear_row}", None)
            
            # Merge and center cleared ranges for SUB-TOTAL and DISCOUNT
            if safe_merge_cells:
                safe_merge_cells(ws, f"N{subtotal_row}:P{subtotal_row + 1}")
                safe_merge_cells(ws, f"N{discount_row}:P{discount_row + 1}")
                if Alignment is not None:
                    merged_cell_subtotal = ws.cell(row=subtotal_row, column=SUMMARY_LABEL_COL)
                    merged_cell_subtotal.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    merged_cell_discount = ws.cell(row=discount_row, column=SUMMARY_LABEL_COL)
                    merged_cell_discount.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Ensure N12 and P12 have inferior (bottom) borders
            if Border is not None and Side is not None:
                bottom_side = Side(style="thin")
                row_12 = SUMMARY_OPT_ROW  # Row 12
                cell_n12 = ws.cell(row=row_12, column=SUMMARY_LABEL_COL)
                cell_p12 = ws.cell(row=row_12, column=SUMMARY_VALUE_COL)
                
                existing_n12_border = cell_n12.border if cell_n12.border else Border()
                existing_p12_border = cell_p12.border if cell_p12.border else Border()
                
                cell_n12.border = Border(
                    left=existing_n12_border.left,
                    right=existing_n12_border.right,
                    top=existing_n12_border.top,
                    bottom=bottom_side
                )
                cell_p12.border = Border(
                    left=existing_p12_border.left,
                    right=existing_p12_border.right,
                    top=existing_p12_border.top,
                    bottom=bottom_side
                )
        
        # Clear N23 if it contains discount percentage (remove duplicate)
        # N23 should remain empty or contain "TOTAL DUE AT SIGNING" if that's what the template has
        cell_n23_check = ws.cell(row=23, column=SUMMARY_LABEL_COL)
        if cell_n23_check.value and "%" in str(cell_n23_check.value):
            # Clear the duplicate discount percentage from N23
            cell_n23_check.value = None
        
        # Make all cells in totals row bold
        if Font is not None:
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
                else:
                    cell.font = Font(bold=True)
        
        # Apply Arial 20 font to totals row if it's a new row (num_styles > 5)
        if num_styles > 5:
            for col_idx in [2, 6, 8, 12, 14, 16]:  # Columns B, F, H, L, N, P
                cell = ws.cell(row=totals_row, column=col_idx)
                apply_arial_20_font(cell)
                # Ensure bold is maintained
                if Font is not None:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
        else:
            # For <=5 styles (totals_row==20), ensure Arial size 20 bold for key cells
            for col_idx in [2, 6, 8, 12, 14, 16]:
                cell = ws.cell(row=totals_row, column=col_idx)
                if Font is not None:
                    cell.font = Font(name="Arial", size=20, bold=True, color=cell.font.color if cell.font else None)
        # Ensure P (column 16) is also Arial size 20 for <=5 styles
        if num_styles <= 5:
            cell_p_totals = ws.cell(row=totals_row, column=16)
            if Font is not None:
                cell_p_totals.font = Font(name="Arial", size=20, bold=True, color=cell_p_totals.font.color if cell_p_totals.font else None)
        # Apply inferior (bottom) border to entire totals row
        if Border is not None and Side is not None:
            bottom_side = Side(style="thin")
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                # Skip columns A, G, M for bottom borders
                if col_idx in [1, 7, 13]:
                    new_bottom = cell.border.bottom
                else:
                    new_bottom = bottom_side
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=new_bottom,
                )

        # Apply full borders (top/bottom/left/right) to key totals cells (B, F, H, L, N, P)
        if Border is not None and Side is not None:
            full_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for col_idx in [2, 6, 8, 12, 14, 16]:
                cell = ws.cell(row=totals_row, column=col_idx)
                cell.border = full_border
        
        # Note: P10 and P12 are updated in build_workbook_bytes after both tabs are created
        # to combine totals from both DEVELOPMENT ONLY and A LA CARTE tabs
        
        # Format notes in DELIVERABLES section (matching workbook_creator.py)
        try:
            first_note_row = None
            second_note_row = None

            # Search a reasonable band of rows where the notes live
            for row in range(20, 90):
                for col in (5, 6):  # Columns E (5) and F (6)
                    value = safe_get_cell_value(ws, row, col)
                    if not isinstance(value, str):
                        continue
                    upper = value.upper()
                    if ("DESIGNS ARE REVIEWED" in upper and
                        "SURCHARGE APPLIES" in upper and
                        first_note_row is None):
                        first_note_row = row
                    if ("DEVELOPMENT DOES NOT INCLUDE" in upper and
                        "BULK PRODUCTION INVENTORY" in upper and
                        second_note_row is None):
                        second_note_row = row
                if first_note_row is not None and second_note_row is not None:
                    break

            # Helper to (re)merge and fully border a note block in E-F
            def _format_note_block(start_row: int, row_span: int) -> None:
                end_row = start_row + row_span - 1
                # Unmerge any existing merges in E-F intersecting this vertical span
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_col <= 6 and merged_range.max_col >= 5 and
                        merged_range.min_row <= end_row and merged_range.max_row >= start_row):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass

                # Merge E-F across the span
                safe_merge_cells(ws, f"E{start_row}:F{end_row}")

                # Center-align the merged cell
                if Alignment is not None:
                    top_left = safe_get_writable_cell(ws, start_row, 5)
                    try:
                        top_left.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=True,
                        )
                    except Exception:
                        pass

                # Apply full borders to all cells in the E-F block
                if Border is not None and Side is not None:
                    thin = Side(style="thin")
                    for r in range(start_row, end_row + 1):
                        for c in (5, 6):
                            cell = ws.cell(row=r, column=c)
                            try:
                                cell.border = Border(
                                    left=thin,
                                    right=thin,
                                    top=thin,
                                    bottom=thin,
                                )
                            except Exception:
                                pass

            # First note: 6-row span
            if first_note_row is not None:
                _format_note_block(first_note_row, 6)

            # Second note: 8-row span (position will already include any extra rows
            # introduced by Activewear logic, so we just span 8 rows from its text row)
            if second_note_row is not None:
                _format_note_block(second_note_row, 8)

                # After we format the main 8-row block, remove any *other* instances
                for row in range(20, 90):
                    # Skip inside the formatted 8-row block
                    if second_note_row <= row <= second_note_row + 7:
                        continue
                    for col in (5, 6):  # E/F
                        value = safe_get_cell_value(ws, row, col)
                        if not isinstance(value, str):
                            continue
                        upper = value.upper()
                        if ("DEVELOPMENT DOES NOT INCLUDE" in upper and
                            "BULK PRODUCTION INVENTORY" in upper):
                            try:
                                safe_set_cell_value(
                                    ws,
                                    f"{'E' if col == 5 else 'F'}{row}",
                                    None,
                                )
                            except Exception:
                                pass

        except Exception:
            # Notes formatting should never break workbook creation
            pass
        
        # Format TEG TECH PACK and COSTING WORKBOOK (matching workbook_creator.py)
        try:
            if column_index_from_string is not None:
                col_b_idx = column_index_from_string("B")
                col_c_idx = column_index_from_string("C")
                col_d_idx = column_index_from_string("D")
                teg_tech_pack_row = None
                costing_workbook_row = None
                
                # Scan for TEG TECH PACK and COSTING WORKBOOK
                scan_end_row = min(deliverables_block_end + 5, ws.max_row + 1)
                for scan_row in range(deliverables_block_start, scan_end_row + 1):
                    value = safe_get_cell_value(ws, scan_row, col_b_idx)
                    if isinstance(value, str):
                        value_lower = value.lower().strip()
                        if "teg" in value_lower and "tech" in value_lower and "pack" in value_lower:
                            if teg_tech_pack_row is None:
                                teg_tech_pack_row = scan_row
                        elif "costing" in value_lower and "workbook" in value_lower:
                            if costing_workbook_row is None:
                                costing_workbook_row = scan_row
                
                # Set TEG TECH PACK - merge and center like other deliverables
                if teg_tech_pack_row:
                    # Ensure TEG TECH PACK label exists in column B
                    teg_label = safe_get_cell_value(ws, teg_tech_pack_row, col_b_idx)
                    if not teg_label or not isinstance(teg_label, str) or "teg" not in teg_label.lower():
                        safe_set_cell_value(ws, f"B{teg_tech_pack_row}", "TEG TECH PACK")
                    
                    # Unmerge B:C and D first to set value
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= teg_tech_pack_row <= merged_range.max_row and
                            (merged_range.min_col <= col_b_idx <= merged_range.max_col or
                             merged_range.min_col <= col_d_idx <= merged_range.max_col)):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass
                    # Clear and set value to num_styles
                    safe_set_cell_value(ws, f"D{teg_tech_pack_row}", None)
                    safe_set_cell_value(ws, f"D{teg_tech_pack_row}", num_styles)
                    # Merge B:C and D, then center
                    teg_tech_pack_row_2 = teg_tech_pack_row + 1
                    if safe_merge_cells:
                        safe_merge_cells(ws, f"B{teg_tech_pack_row}:C{teg_tech_pack_row_2}")
                        safe_merge_cells(ws, f"D{teg_tech_pack_row}:D{teg_tech_pack_row_2}")
                    # Set number format and alignment
                    count_cell = safe_get_writable_cell(ws, teg_tech_pack_row, col_d_idx)
                    label_cell = safe_get_writable_cell(ws, teg_tech_pack_row, col_b_idx)
                    try:
                        count_cell.number_format = "0"
                        if Alignment:
                            count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                            label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    except Exception:
                        pass
                
                if costing_workbook_row and teg_tech_pack_row:
                    # Ensure COSTING WORKBOOK label exists in column B
                    costing_label = safe_get_cell_value(ws, costing_workbook_row, col_b_idx)
                    if not costing_label or not isinstance(costing_label, str) or "costing" not in costing_label.lower():
                        safe_set_cell_value(ws, f"B{costing_workbook_row}", "COSTING WORKBOOK")
                    
                    # Read the actual value from TEG TECH PACK cell to ensure they match
                    teg_tech_pack_value = safe_get_cell_value(ws, teg_tech_pack_row, col_d_idx)
                    # If TEG TECH PACK value is wrong, use num_styles directly
                    if teg_tech_pack_value != num_styles:
                        teg_tech_pack_value = num_styles
                    
                    # Unmerge B:C and D first to set value
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= costing_workbook_row <= merged_range.max_row and
                            (merged_range.min_col <= col_b_idx <= merged_range.max_col or
                             merged_range.min_col <= col_d_idx <= merged_range.max_col)):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass
                    # Set COSTING WORKBOOK value
                    if get_column_letter:
                        costing_workbook_ref = f"{get_column_letter(col_d_idx)}{costing_workbook_row}"
                        safe_set_cell_value(ws, costing_workbook_ref, teg_tech_pack_value)
                    else:
                        safe_set_cell_value(ws, f"D{costing_workbook_row}", teg_tech_pack_value)
                    # Merge B:C and D, then center
                    costing_workbook_row_2 = costing_workbook_row + 1
                    if safe_merge_cells:
                        safe_merge_cells(ws, f"B{costing_workbook_row}:C{costing_workbook_row_2}")
                        safe_merge_cells(ws, f"D{costing_workbook_row}:D{costing_workbook_row_2}")
                    # Set number format and alignment
                    count_cell = safe_get_writable_cell(ws, costing_workbook_row, col_d_idx)
                    label_cell = safe_get_writable_cell(ws, costing_workbook_row, col_b_idx)
                    try:
                        count_cell.number_format = "0"
                        if Alignment:
                            count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                            label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    except Exception:
                        pass
                
                # Add right borders to column D (values) for all deliverables including TEG TECH PACK and COSTING WORKBOOK
                if Border is not None and Side is not None:
                    thin = Side(style="thin")
                    # Scan for all deliverables and add right border to their D column
                    for scan_row in range(deliverables_block_start, min(deliverables_block_end + 5, ws.max_row + 1)):
                        value = safe_get_cell_value(ws, scan_row, col_b_idx)
                        if isinstance(value, str) and value.strip():
                            value_lower = value.lower().strip()
                            # Check if it's a deliverable label
                            deliverable_keywords = ["pattern", "sample", "fitting", "revision", "tech pack", "costing", "workbook"]
                            if any(keyword in value_lower for keyword in deliverable_keywords):
                                # Found a deliverable, add right border to its D column
                                count_cell = safe_get_writable_cell(ws, scan_row, col_d_idx)
                                try:
                                    existing_border = count_cell.border
                                    if existing_border:
                                        count_cell.border = Border(
                                            left=existing_border.left,
                                            right=thin,
                                            top=existing_border.top,
                                            bottom=existing_border.bottom
                                        )
                                    else:
                                        count_cell.border = Border(right=thin)
                                except Exception:
                                    pass
        except Exception:
            # TEG TECH PACK/COSTING WORKBOOK formatting should not break workbook creation
            pass
        
        # Activewear box: Only build the tall merged E–F "box" when there is at least one Activewear style
        try:
            if (
                num_activewear > 0
                and Border is not None
                and Side is not None
                and column_index_from_string is not None
            ):
                label_col_b = column_index_from_string("B")
                final_samples_row_scan = None
                costing_workbook_row_scan = None

                # Scan a reasonable band where these labels live
                for row in range(deliverables_block_start, deliverables_block_end + 15):
                    value = safe_get_cell_value(ws, row, label_col_b)
                    if not isinstance(value, str):
                        continue
                    lower = value.strip().lower()
                    if "final samples" in lower and final_samples_row_scan is None:
                        final_samples_row_scan = row
                    elif ("costing" in lower and "workbook" in lower and
                          costing_workbook_row_scan is None):
                        costing_workbook_row_scan = row

                if final_samples_row_scan is not None and costing_workbook_row_scan is not None:
                    start_row = final_samples_row_scan
                    end_row = costing_workbook_row_scan
                    # Include one extra row below COSTING WORKBOOK to match the visual box
                    box_end_row = end_row + 1

                    thin = Side(style="thin")
                    # First, apply full borders to E/F for the whole vertical span
                    for r in range(start_row, box_end_row + 1):
                        for c in (5, 6):  # E and F
                            cell = ws.cell(row=r, column=c)
                            try:
                                cell.border = Border(
                                    left=thin,
                                    right=thin,
                                    top=thin,
                                    bottom=thin,
                                )
                            except Exception:
                                pass

                    # Then merge this entire E–F block into a single tall cell and center its content
                    for merged_range in list(ws.merged_cells.ranges):
                        if (
                            merged_range.min_col <= 6
                            and merged_range.max_col >= 5
                            and merged_range.min_row <= box_end_row
                            and merged_range.max_row >= start_row
                        ):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass

                    # Merge E-F from FINAL SAMPLES row down through COSTING WORKBOOK box
                    safe_merge_cells(ws, f"E{start_row}:F{box_end_row}")

                    if Alignment is not None:
                        top_left = safe_get_writable_cell(ws, start_row, 5)
                        try:
                            top_left.alignment = Alignment(
                                horizontal="center",
                                vertical="center",
                                wrap_text=True,
                            )
                        except Exception:
                            pass
        except Exception:
            # Border "box" improvement should not break workbook creation
            pass
        
        # Place "TOTAL DUE AT SIGNING" in the 2 rows just above totals row: N(totals_row-1):O(totals_row) and P(totals_row-1):P(totals_row).
        # For <=5 styles: totals_row=20 → N19:O20, P19:P20. For 6 styles: totals_row=22 → N21:O22, P21:P22. Always dynamic.
        total_due_start = totals_row - 1
        total_due_end = totals_row
        cell_n_label = ws.cell(row=total_due_start, column=14)
        # Always place in dynamic position when we have a totals row (so 6+ styles get N21:O22, not N19:O20)
        # When totals row > 20, unmerge and clear old position (N19:O20, P19:P20) to avoid duplicate
        if total_due_start > 19:
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= 20 and merged_range.max_row >= 19 and
                    merged_range.min_col >= 14 and merged_range.max_col <= 16):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear value, borders, and fill (yellow) from original N19:O20, P19:P20 so no leftover formatting
            for r in (19, 20):
                for c in (14, 15, 16):
                    try:
                        cell = ws.cell(row=r, column=c)
                        cell.value = None
                        if Border is not None:
                            cell.border = Border()
                        if PatternFill is not None:
                            cell.fill = PatternFill(fill_type="none")
                    except Exception:
                        pass
        # Unmerge any existing merge in N-P for the target two rows (total_due_start:total_due_end)
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= total_due_end and merged_range.max_row >= total_due_start and
                merged_range.min_col >= 14 and merged_range.max_col <= 16):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        # Apply borders to every cell in both ranges BEFORE merging so the full perimeter is bordered
        if Border is not None and Side is not None:
            thin_side = Side(style="thin")
            no_side = Side(style=None)
            for r in range(total_due_start, total_due_end + 1):
                top_side = thin_side if r == total_due_start else no_side
                bottom_side = thin_side if r == total_due_end else no_side
                for c in range(14, 16 + 1):  # N=14, O=15, P=16
                    try:
                        cell = ws.cell(row=r, column=c)
                        if c == 14:
                            cell.border = Border(left=thin_side, right=no_side, top=top_side, bottom=bottom_side)
                        elif c == 15:
                            cell.border = Border(left=no_side, right=thin_side, top=top_side, bottom=bottom_side)
                        else:
                            cell.border = Border(left=thin_side, right=thin_side, top=top_side, bottom=bottom_side)
                    except Exception:
                        pass
        # Merge N(total_due_start):O(total_due_end) (label spanning 2 rows)
        cell_n_label.value = "TOTAL DUE AT SIGNING"
        safe_merge_cells(ws, f"N{total_due_start}:O{total_due_end}")
        cell_n_label = safe_get_writable_cell(ws, total_due_start, 14)
        # Merge P(total_due_start):P(total_due_end) (value spanning 2 rows)
        cell_p_val = ws.cell(row=total_due_start, column=16)  # Column P
        if discount_percentage > 0:
            cell_p_val.value = f"=P{SUMMARY_SUBTOTAL_ROW}-P{SUMMARY_DISCOUNT_ROW}"
        else:
            cell_p_val.value = f"=SUM(P10:P13)"
        safe_merge_cells(ws, f"P{total_due_start}:P{total_due_end}")
        cell_p_val = safe_get_writable_cell(ws, total_due_start, 16)
        cell_p_val.number_format = '$#,##0'  # Currency format
        # Apply font size 20 and bold to TOTAL DUE AT SIGNING formula
        if Font is not None:
            cell_p_val.font = Font(name="Arial", size=20, bold=True, color=cell_p_val.font.color if cell_p_val.font else None)
        if Alignment is not None:
            cell_p_val.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #ffff00 to TOTAL DUE AT SIGNING
        if PatternFill is not None:
            cell_p_val.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Label (merged N:O) bold size 20
        if Font is not None:
            cell_n_label.font = Font(name="Arial", size=20, bold=True, color=cell_n_label.font.color if cell_n_label.font else None)
        if Alignment is not None:
            cell_n_label.alignment = Alignment(horizontal="left", vertical="center")
        
        # Also update any other formulas in column P that reference F20 or L20 statically
        # Check a few rows around the totals row
        for check_row in range(totals_row - 2, totals_row + 3):
            cell_p = safe_get_writable_cell(ws, check_row, 16)  # Column P
            if cell_p.value and isinstance(cell_p.value, str) and cell_p.value.startswith("="):
                # Check if it references F20 or L20 (static references)
                if "F20" in cell_p.value or "L20" in cell_p.value:
                    # Replace with dynamic references
                    formula = cell_p.value.replace("F20", f"F{totals_row}").replace("L20", f"L{totals_row}")
                    cell_p.value = formula
                    if Font is not None:
                        existing_font = cell_p.font
                        cell_p.font = Font(
                            name=existing_font.name if existing_font and existing_font.name else "Arial",
                            size=existing_font.size if existing_font and existing_font.size else 20,
                            bold=True,
                            color=existing_font.color if existing_font else None
                        )
                    if Alignment is not None:
                        cell_p.alignment = Alignment(horizontal="center", vertical="center")
    else:
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "L20", None)
    return total_development, total_optional


# A La Carte hourly rates
A_LA_CARTE_RATE_STANDARD = 215.00  # $215/hr for INTAKE, PATTERN, SAMPLE, FITTING, ADJUSTMENT
A_LA_CARTE_RATE_SAMPLES = 110.00    # $110/hr for FINAL SAMPLES and DUPLICATES


def apply_ala_carte_package(
    ws,
    *,
    client_name: str,
    client_email: str,
    representative: str,
    a_la_carte_items: list[dict],
) -> tuple[float, float]:
    """Write A La Carte items to the A LA CARTE tab and return totals.
    
    Args:
        ws: The worksheet (A LA CARTE tab)
        client_name: Client name
        client_email: Client email
        representative: Representative name
        a_la_carte_items: List of A La Carte items with hours and optional add-ons
    
    Returns:
        Tuple of (total_ala_carte, total_optional_ala_carte)
    """
    # Populate header cells (same positions as DEVELOPMENT ONLY tab): TEGMADE, client name, email, representative, date
    safe_set_cell_value(ws, "B3", "TEGMADE, JUST FOR")
    client_display = (client_name or "").strip().upper()
    safe_set_cell_value(ws, "J3", client_display)
    if Font is not None:
        try:
            ws["J3"].font = Font(
                color="00C9A57A",
                name="Schibsted Grotesk",
                size=48,
                bold=True,
            )
        except Exception:
            pass
    if Alignment is not None:
        try:
            ws["J3"].alignment = Alignment(horizontal="left", vertical="center")
        except Exception:
            pass
    safe_set_cell_value(ws, "D6", (client_email or "").strip())
    safe_set_cell_value(ws, "J6", (representative or "").strip().upper())
    safe_set_cell_value(ws, "B8", "A LA CARTE PACKAGE")
    # Date next to "DATE" label is in the template in the right position; do not write to L6 (would duplicate)

    # Set headers for OPTIONAL ADD-ONS (A LA CARTE) in row 9: S=DYE TESTING, T=PLANNING, U=DESIGN, V=TOTAL
    if get_column_letter:
        safe_set_cell_value(ws, "S9", "DYE TESTING")
        safe_set_cell_value(ws, "T9", "PLANNING")
        safe_set_cell_value(ws, "U9", "DESIGN")
        safe_set_cell_value(ws, "V9", "TOTAL")
    
    # Helper function to apply font size 20 to a cell
    def apply_font_20(cell):
        if Font is not None:
            existing_font = cell.font
            cell.font = Font(
                name=existing_font.name if existing_font and existing_font.name else "Arial",
                size=20,
                bold=existing_font.bold if existing_font else False,
                color=existing_font.color if existing_font else None
            )
    
    # Apply formatting to OPTIONAL ADD-ONS headers in row 9
    if column_index_from_string:
        for col_letter in ["S", "T", "U", "V"]:
            col_idx = column_index_from_string(col_letter)
            if col_idx:
                header_cell = ws.cell(row=9, column=col_idx)
                apply_font_20(header_cell)
                if Alignment is not None:
                    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    total_ala_carte = 0.0
    total_optional_ala_carte = 0.0
    
    if not a_la_carte_items:
        return total_ala_carte, total_optional_ala_carte
    
    num_items = len(a_la_carte_items)
    
    # If more than 5 items, insert rows after row 19 and move everything below down
    rows_to_insert = 0
    if num_items > 5:
        rows_to_insert = (num_items - 5) * 2  # Each item uses 2 rows
    
    # Capture the B20:V40 block before inserting rows (totals + deliverables + rates)
    # This preserves all merges and formatting
    ala_carte_block_start = 20
    ala_carte_block_end = 40  # Extended to include rates section
    ala_carte_block_col_start = 2  # Column B
    ala_carte_block_col_end = 22   # Column V (was U)
    
    ala_carte_block_template = None
    if rows_to_insert > 0:
        # Capture the block before insertion
        block_rows = []
        for row in range(ala_carte_block_start, ala_carte_block_end + 1):
            row_data = []
            for col in range(ala_carte_block_col_start, ala_carte_block_col_end + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(
                    {
                        "value": cell.value,
                        "number_format": cell.number_format,
                        "font": copy(cell.font) if cell.font else None,
                        "fill": copy(cell.fill) if cell.fill else None,
                        "border": copy(cell.border) if cell.border else None,
                        "alignment": copy(cell.alignment) if cell.alignment else None,
                    }
                )
            block_rows.append(row_data)
        
        merged_ranges = []
        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row >= ala_carte_block_start
                and merged_range.max_row <= ala_carte_block_end
                and merged_range.min_col >= ala_carte_block_col_start
                and merged_range.max_col <= ala_carte_block_col_end
            ):
                merged_ranges.append(
                    (
                        merged_range.min_row - ala_carte_block_start,
                        merged_range.max_row - ala_carte_block_start,
                        merged_range.min_col,
                        merged_range.max_col,
                    )
                )
        
        ala_carte_block_template = {"rows": block_rows, "merges": merged_ranges}
        
        # Insert rows at row 20 (after the last template row 19)
        # This will automatically shift everything from row 20 onwards down
        ws.insert_rows(20, amount=rows_to_insert)
        
        # Copy formatting from row 19 to new rows (for columns B-Q)
        template_row = 19
        for i in range(rows_to_insert):
            new_row = 20 + i
            for col in range(2, 18):  # Columns B through Q
                source_cell = ws.cell(row=template_row, column=col)
                target_cell = ws.cell(row=new_row, column=col)
                copy_cell_formatting(source_cell, target_cell)
                # Ensure white backgrounds for style columns
                if PatternFill is not None:
                    target_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # A La Carte items start at row 10, each item uses 2 rows (merged)
    start_row = 10
    
    # Column indices - match template "Copy of TEG 2025 WORKBOOK TEMPLATES.xlsx" A LA CARTE tab:
    # D: INTAKE HOURS, E: INTAKE TOTAL
    # F: 1ST PATTERN HOURS, G: 1ST PATTERN TOTAL
    # H: 1ST SAMPLE HOURS, I: 1ST SAMPLE TOTAL
    # J: FITTING HOURS, K: FITTING TOTAL
    # L: ADJUSTMENT HOURS, M: ADJUSTMENT TOTAL
    # N: FINAL SAMPLE HOURS, O: FINAL SAMPLE QUANT, P: FINAL SAMPLE TOTAL
    # Q: TOTAL (row total)
    # R-U: OPTIONAL ADD-ONS (DYE TESTING, PLANNING, DESIGN, TOTAL)
    col_b = column_index_from_string("B") if column_index_from_string else 2
    col_c = column_index_from_string("C") if column_index_from_string else 3
    col_d = column_index_from_string("D") if column_index_from_string else 4  # INTAKE HOURS
    col_e = column_index_from_string("E") if column_index_from_string else 5  # INTAKE TOTAL
    col_f = column_index_from_string("F") if column_index_from_string else 6  # 1ST PATTERN HOURS
    col_g = column_index_from_string("G") if column_index_from_string else 7  # 1ST PATTERN TOTAL
    col_h = column_index_from_string("H") if column_index_from_string else 8  # 1ST SAMPLE HOURS
    col_i = column_index_from_string("I") if column_index_from_string else 9  # 1ST SAMPLE TOTAL
    col_j = column_index_from_string("J") if column_index_from_string else 10  # FITTING HOURS
    col_k = column_index_from_string("K") if column_index_from_string else 11  # FITTING TOTAL
    col_l = column_index_from_string("L") if column_index_from_string else 12  # ADJUSTMENT HOURS
    col_m = column_index_from_string("M") if column_index_from_string else 13  # ADJUSTMENT TOTAL
    col_n = column_index_from_string("N") if column_index_from_string else 14  # FINAL SAMPLE HOURS
    col_o = column_index_from_string("O") if column_index_from_string else 15  # FINAL SAMPLE QUANT
    col_p = column_index_from_string("P") if column_index_from_string else 16  # FINAL SAMPLE TOTAL
    col_q = column_index_from_string("Q") if column_index_from_string else 17  # TOTAL (row total)
    col_r = column_index_from_string("S") if column_index_from_string else 19  # DYE TESTING (S)
    col_s = column_index_from_string("T") if column_index_from_string else 20  # PLANNING (T)
    col_t = column_index_from_string("U") if column_index_from_string else 21  # DESIGN (U)
    col_u = column_index_from_string("V") if column_index_from_string else 22  # OPTIONAL TOTAL (V)
    
    # Rate cells in deliverables section (dynamic based on rows_to_insert)
    # Note: "STANDARD RATE ($/HR):" label is in B35:E36, value is in F35:F36 (base position)
    # "DUPLICATES RATE ($/HR):" label is in B37:E38, value is in F37:F38 (base position)
    # When rows are inserted, these move down by rows_to_insert
    rate_standard_row = 35 + rows_to_insert  # Dynamic: moves down when rows are inserted
    rate_duplicates_row = 37 + rows_to_insert  # Dynamic: moves down when rows are inserted
    rate_col = column_index_from_string("F") if column_index_from_string else 6  # Rates are now in column F
    
    quantity_sum = 0  # Sum of all quantities for deliverables section
    
    for i, item in enumerate(a_la_carte_items):
        # Each entry uses two rows, similar to style rows (10-11, 12-13, ...)
        row = start_row + (i * 2)
        row_second = row + 1
        
        # Style Number - column B
        safe_set_cell_value(ws, f"{get_column_letter(col_b)}{row}", item.get("style_number", ""))
        cell_b = ws.cell(row=row, column=col_b)
        cell_b.number_format = "0"  # Integer format
        apply_font_20(cell_b)
        apply_full_border_pair(ws, col_b, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_b)}{row}:{get_column_letter(col_b)}{row_second}")
        
        # Style Name - column C
        safe_set_cell_value(ws, f"{get_column_letter(col_c)}{row}", item.get("name", ""))
        cell_c = ws.cell(row=row, column=col_c)
        apply_font_20(cell_c)
        apply_full_border_pair(ws, col_c, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_c)}{row}:{get_column_letter(col_c)}{row_second}")
        
        # INTAKE HOURS - column D
        intake_hours = float(item.get("intake_session", 0))
        safe_set_cell_value(ws, f"{get_column_letter(col_d)}{row}", intake_hours if intake_hours > 0 else None)
        cell_d = ws.cell(row=row, column=col_d)
        cell_d.number_format = "0.00" if intake_hours > 0 and intake_hours != int(intake_hours) else "0"
        apply_font_20(cell_d)
        apply_full_border_pair(ws, col_d, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_d)}{row}:{get_column_letter(col_d)}{row_second}")
        
        # INTAKE TOTAL - column E (formula: hours * rate)
        if intake_hours > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_e)}{row}", f"=D{row}*${get_column_letter(rate_col)}${rate_standard_row}")
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_e)}{row}", None)
        cell_e = ws.cell(row=row, column=col_e)
        cell_e.number_format = '$#,##0'
        apply_font_20(cell_e)
        apply_full_border_pair(ws, col_e, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_e)}{row}:{get_column_letter(col_e)}{row_second}")
        
        # 1ST PATTERN HOURS - column F
        pattern_hours = float(item.get("first_pattern", 0))
        safe_set_cell_value(ws, f"{get_column_letter(col_f)}{row}", pattern_hours if pattern_hours > 0 else None)
        cell_f = ws.cell(row=row, column=col_f)
        cell_f.number_format = "0.00" if pattern_hours > 0 and pattern_hours != int(pattern_hours) else "0"
        apply_font_20(cell_f)
        apply_full_border_pair(ws, col_f, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_f)}{row}:{get_column_letter(col_f)}{row_second}")
        
        # 1ST PATTERN TOTAL - column G (formula: hours * rate)
        if pattern_hours > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_g)}{row}", f"=F{row}*${get_column_letter(rate_col)}${rate_standard_row}")
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_g)}{row}", None)
        cell_g = ws.cell(row=row, column=col_g)
        cell_g.number_format = '$#,##0'
        apply_font_20(cell_g)
        apply_full_border_pair(ws, col_g, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_g)}{row}:{get_column_letter(col_g)}{row_second}")
        
        # 1ST SAMPLE HOURS - column H
        sample_hours = float(item.get("first_sample", 0))
        safe_set_cell_value(ws, f"{get_column_letter(col_h)}{row}", sample_hours if sample_hours > 0 else None)
        cell_h = ws.cell(row=row, column=col_h)
        cell_h.number_format = "0.00" if sample_hours > 0 and sample_hours != int(sample_hours) else "0"
        apply_font_20(cell_h)
        apply_full_border_pair(ws, col_h, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_h)}{row}:{get_column_letter(col_h)}{row_second}")
        
        # 1ST SAMPLE TOTAL - column I (formula: hours * rate)
        if sample_hours > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_i)}{row}", f"=H{row}*${get_column_letter(rate_col)}${rate_standard_row}")
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_i)}{row}", None)
        cell_i = ws.cell(row=row, column=col_i)
        cell_i.number_format = '$#,##0'
        apply_font_20(cell_i)
        apply_full_border_pair(ws, col_i, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_i)}{row}:{get_column_letter(col_i)}{row_second}")
        
        # FITTING HOURS - column J
        fitting_hours = float(item.get("fitting", 0))
        safe_set_cell_value(ws, f"{get_column_letter(col_j)}{row}", fitting_hours if fitting_hours > 0 else None)
        cell_j = ws.cell(row=row, column=col_j)
        cell_j.number_format = "0.00" if fitting_hours > 0 and fitting_hours != int(fitting_hours) else "0"
        apply_font_20(cell_j)
        apply_full_border_pair(ws, col_j, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_j)}{row}:{get_column_letter(col_j)}{row_second}")
        
        # FITTING TOTAL - column K (formula: hours * rate)
        if fitting_hours > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_k)}{row}", f"=J{row}*${get_column_letter(rate_col)}${rate_standard_row}")
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_k)}{row}", None)
        cell_k = ws.cell(row=row, column=col_k)
        cell_k.number_format = '$#,##0'
        apply_font_20(cell_k)
        apply_full_border_pair(ws, col_k, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_k)}{row}:{get_column_letter(col_k)}{row_second}")
        
        # ADJUSTMENT HOURS - column L
        adjustment_hours = float(item.get("adjustment", 0))
        safe_set_cell_value(ws, f"{get_column_letter(col_l)}{row}", adjustment_hours if adjustment_hours > 0 else None)
        cell_l = ws.cell(row=row, column=col_l)
        cell_l.number_format = "0.00" if adjustment_hours > 0 and adjustment_hours != int(adjustment_hours) else "0"
        apply_font_20(cell_l)
        apply_full_border_pair(ws, col_l, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_l)}{row}:{get_column_letter(col_l)}{row_second}")
        
        # ADJUSTMENT TOTAL - column M (formula: hours * standard rate)
        if adjustment_hours > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_m)}{row}", f"=L{row}*${get_column_letter(rate_col)}${rate_standard_row}")
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_m)}{row}", None)
        cell_m = ws.cell(row=row, column=col_m)
        cell_m.number_format = '$#,##0'
        apply_font_20(cell_m)
        apply_full_border_pair(ws, col_m, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_m)}{row}:{get_column_letter(col_m)}{row_second}")
        
        # FINAL SAMPLE HOURS - column N (per-unit hours; was "duplicates" hours)
        final_sample_hours = float(item.get("duplicates", 0))
        safe_set_cell_value(ws, f"{get_column_letter(col_n)}{row}", final_sample_hours if final_sample_hours > 0 else None)
        cell_n = ws.cell(row=row, column=col_n)
        cell_n.number_format = "0.00" if final_sample_hours > 0 and final_sample_hours != int(final_sample_hours) else "0"
        apply_font_20(cell_n)
        apply_full_border_pair(ws, col_n, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_n)}{row}:{get_column_letter(col_n)}{row_second}")
        
        # FINAL SAMPLE QUANT - column O (was "quantity")
        final_sample_quant = int(item.get("quantity", 1))
        safe_set_cell_value(ws, f"{get_column_letter(col_o)}{row}", final_sample_quant if final_sample_quant > 0 else None)
        cell_o_qty = ws.cell(row=row, column=col_o)
        cell_o_qty.number_format = "0"
        apply_font_20(cell_o_qty)
        apply_full_border_pair(ws, col_o, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_o)}{row}:{get_column_letter(col_o)}{row_second}")
        quantity_sum += final_sample_quant
        
        # FINAL SAMPLE TOTAL - column P (formula: hours * quantity * final sample rate)
        if final_sample_hours > 0 and final_sample_quant > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_p)}{row}", f"=N{row}*O{row}*${get_column_letter(rate_col)}${rate_duplicates_row}")
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_p)}{row}", None)
        cell_p = ws.cell(row=row, column=col_p)
        cell_p.number_format = '$#,##0'
        apply_font_20(cell_p)
        apply_full_border_pair(ws, col_p, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_p)}{row}:{get_column_letter(col_p)}{row_second}")
        
        # TOTAL - column Q (formula: sum of all totals = E+G+I+K+M+P)
        safe_set_cell_value(ws, f"{get_column_letter(col_q)}{row}", f"=E{row}+G{row}+I{row}+K{row}+M{row}+P{row}")
        cell_q = ws.cell(row=row, column=col_q)
        cell_q.number_format = '$#,##0'
        apply_font_20(cell_q)
        apply_full_border_pair(ws, col_q, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_q)}{row}:{get_column_letter(col_q)}{row_second}")
        
        # Calculate total for tracking (for grand total calculation)
        intake_price = intake_hours * 215 if intake_hours > 0 else 0
        pattern_price = pattern_hours * 215 if pattern_hours > 0 else 0
        sample_price = sample_hours * 215 if sample_hours > 0 else 0
        fitting_price = fitting_hours * 215 if fitting_hours > 0 else 0
        adjustment_price = adjustment_hours * 215 if adjustment_hours > 0 else 0
        final_sample_price = final_sample_hours * final_sample_quant * 110 if (final_sample_hours > 0 and final_sample_quant > 0) else 0
        item_total = intake_price + pattern_price + sample_price + fitting_price + adjustment_price + final_sample_price
        total_ala_carte += item_total
        
        # Optional Add-ons for A La Carte (columns R-U)
        # R: DYE TESTING, S: PLANNING, T: DESIGN, U: TOTAL
        row_options = item.get("options", {})
        
        # DYE TESTING - column R (leave blank if $0)
        dye_testing_price = OPTIONAL_PRICES["dye_testing"] if row_options.get("dye_testing", False) else 0
        if dye_testing_price > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_r)}{row}", dye_testing_price)
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_r)}{row}", None)  # Leave blank if $0
        cell_r_opt = ws.cell(row=row, column=col_r)
        cell_r_opt.number_format = '$#,##0'
        apply_font_20(cell_r_opt)
        apply_full_border_pair(ws, col_r, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_r)}{row}:{get_column_letter(col_r)}{row_second}")
        
        # PLANNING - column S (leave blank if $0)
        planning_price = OPTIONAL_PRICES["planning"] if row_options.get("planning", False) else 0
        if planning_price > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_s)}{row}", planning_price)
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_s)}{row}", None)  # Leave blank if $0
        cell_s_planning = ws.cell(row=row, column=col_s)
        cell_s_planning.number_format = '$#,##0'
        apply_font_20(cell_s_planning)
        apply_full_border_pair(ws, col_s, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_s)}{row}:{get_column_letter(col_s)}{row_second}")
        
        # DESIGN - column T (leave blank if $0)
        design_price = OPTIONAL_PRICES["design"] if row_options.get("design", False) else 0
        if design_price > 0:
            safe_set_cell_value(ws, f"{get_column_letter(col_t)}{row}", design_price)
        else:
            safe_set_cell_value(ws, f"{get_column_letter(col_t)}{row}", None)  # Leave blank if $0
        cell_t_design = ws.cell(row=row, column=col_t)
        cell_t_design.number_format = '$#,##0'
        apply_font_20(cell_t_design)
        apply_full_border_pair(ws, col_t, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_t)}{row}:{get_column_letter(col_t)}{row_second}")
        
        # TOTAL - column U (formula: sum of optional add-ons R+S+T)
        safe_set_cell_value(ws, f"{get_column_letter(col_u)}{row}", f"=S{row}+T{row}+U{row}")
        cell_u_total = ws.cell(row=row, column=col_u)
        cell_u_total.number_format = '$#,##0'
        apply_font_20(cell_u_total)
        apply_full_border_pair(ws, col_u, row, row_second)
        safe_merge_cells(ws, f"{get_column_letter(col_u)}{row}:{get_column_letter(col_u)}{row_second}")
        total_optional_ala_carte += dye_testing_price + planning_price + design_price
        
        # Apply center alignment to all cells
        if Alignment is not None:
            cells_to_align = [cell_b, cell_c, cell_d, cell_e, cell_f, cell_g, cell_h, cell_i, cell_j, cell_k, cell_l,
                             cell_m, cell_n, cell_o_qty, cell_p, cell_q, cell_r_opt, cell_s_planning, cell_t_design, cell_u_total]
            for cell in cells_to_align:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # No derived calculations needed - all totals are now formulas in Excel
    
    # Clear unused rows (template has 5 rows: 10, 12, 14, 16, 18)
    # Only clear if we have 5 or fewer items (if more than 5, we've already inserted rows)
    rows_to_clear = []
    if num_items == 1:
        rows_to_clear = [12, 14, 16, 18]  # Clear items 2, 3, 4, 5
    elif num_items == 2:
        rows_to_clear = [14, 16, 18]  # Clear items 3, 4, 5
    elif num_items == 3:
        rows_to_clear = [16, 18]  # Clear items 4, 5
    elif num_items == 4:
        rows_to_clear = [18]  # Clear item 5
    
    for clear_row in rows_to_clear:
        clear_row_second = clear_row + 1
        # Clear all columns B through V
        for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V"]:
            col_idx = column_index_from_string(col_letter) if column_index_from_string else None
            if col_idx:
                # Unmerge if needed
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= clear_row_second and merged_range.max_row >= clear_row and
                        merged_range.min_col <= col_idx <= merged_range.max_col):
                        try:
                            min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                            ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                        except Exception:
                            pass
                # Clear the cells
                safe_set_cell_value(ws, f"{col_letter}{clear_row}", None)
                safe_set_cell_value(ws, f"{col_letter}{clear_row_second}", None)
                # Merge and center the cleared cells (maintain formatting)
                if safe_merge_cells:
                    safe_merge_cells(ws, f"{col_letter}{clear_row}:{col_letter}{clear_row_second}")
                if Alignment is not None:
                    cell = ws.cell(row=clear_row, column=col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    apply_font_20(cell)
    
    # Restore the B20:Q36 block at the new position (shifted down by rows_to_insert)
    if rows_to_insert > 0 and ala_carte_block_template:
        target_start_row = ala_carte_block_start + rows_to_insert
        rows_data = ala_carte_block_template.get("rows", [])
        block_height = len(rows_data)
        target_end_row = target_start_row + block_height - 1
        
        # Clear existing merges in target area
        to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            if (
                merged_range.max_row < target_start_row
                or merged_range.min_row > target_end_row
                or merged_range.max_col < ala_carte_block_col_start
                or merged_range.min_col > ala_carte_block_col_end
            ):
                continue
            to_unmerge.append(merged_range)
        
        for merged_range in to_unmerge:
            try:
                ws.unmerge_cells(range_string=str(merged_range))
            except Exception:
                pass
        
        # Write cell data and formatting
        # Skip writing formulas for totals row (row 20) - we'll update them after restore
        totals_row_offset = 0  # Row 20 is at offset 0 in the block (20 - 20 = 0)
        for row_offset, row_cells in enumerate(rows_data):
            target_row = target_start_row + row_offset
            is_totals_row = (row_offset == totals_row_offset)
            
            for col_offset, cell_data in enumerate(row_cells):
                target_col = ala_carte_block_col_start + col_offset
                coord = None
                if get_column_letter is not None:
                    coord = f"{get_column_letter(target_col)}{target_row}"
                value = cell_data.get("value") if cell_data else None
                
                # Skip formulas in totals row (columns P and V) - we'll update them after
                # Also skip formulas in column U (21) for deliverables section (OPTIONAL ADD-ONS counts) - we'll update them after
                col_p_check = 16
                col_u_check = 22   # V = optional add-ons total column
                col_t_deliverables = 21  # U = OPTIONAL ADD-ONS (A LA CARTE) deliverables counts
                is_deliverables_section = (target_row >= 22 + rows_to_insert and target_row <= 40 + rows_to_insert)
                
                if is_totals_row and target_col == col_p_check:
                    # Don't restore the formula value for column P, just formatting
                    # The formula will be set after restore
                    pass
                elif is_totals_row and target_col == col_u_check:
                    # Don't restore the formula value for column U, just formatting
                    # The formula will be set after restore
                    pass
                elif is_deliverables_section and target_col == col_t_deliverables and isinstance(value, str) and (value.startswith("=") or value.startswith("=@")):
                    # Don't restore formulas in column U for deliverables section (OPTIONAL ADD-ONS counts), we'll update them after
                    # Skip both regular formulas (=) and formulas with @ symbol (=@)
                    pass
                else:
                    if coord:
                        safe_set_cell_value(ws, coord, value)
                    else:
                        ws.cell(row=target_row, column=target_col).value = value
                
                cell = ws.cell(row=target_row, column=target_col)
                cell.number_format = cell_data.get("number_format") or cell.number_format
                if cell_data.get("font"):
                    cell.font = copy(cell_data["font"])
                if cell_data.get("fill"):
                    cell.fill = copy(cell_data["fill"])
                if cell_data.get("border"):
                    cell.border = copy(cell_data["border"])
                if cell_data.get("alignment"):
                    cell.alignment = copy(cell_data["alignment"])
                # Apply font size 20
                apply_font_20(cell)
        
        # Recreate merged ranges (adjusted for new start row)
        for min_row_offset, max_row_offset, min_col, max_col in ala_carte_block_template.get("merges", []):
            start_row = target_start_row + min_row_offset
            end_row = target_start_row + max_row_offset
            try:
                if safe_merge_cells and get_column_letter:
                    start_col_letter = get_column_letter(min_col)
                    end_col_letter = get_column_letter(max_col)
                    range_str = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
                    safe_merge_cells(ws, range_str)
            except Exception:
                pass
    
    # Calculate and display totals in row 20 (shifted down if rows were inserted)
    # Update formulas AFTER restoring the block to ensure they reference correct rows
    if len(a_la_carte_items) > 0:
        row_20 = 20 + rows_to_insert  # Shift down if rows were inserted
        # Formulas always start at K10/Q10 and end at the row before the totals row
        first_ala_row = 10  # Always start at row 10
        last_ala_row = row_20 - 2  # Last item row is 2 rows before totals row (items use 2 rows each)
        
        # Unmerge and update TOTAL A LA CARTE formula in column Q (row total sum)
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_20 <= merged_range.max_row and
                merged_range.min_col <= col_q <= merged_range.max_col):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass
        
        # Set the correct formula: sum of row totals Q10 to last item row
        safe_set_cell_value(ws, f"Q{row_20}", f"=SUM(Q{first_ala_row}:Q{last_ala_row})")
        cell_q_total = ws.cell(row=row_20, column=col_q)
        cell_q_total.number_format = '$#,##0'
        apply_font_20(cell_q_total)
        if Font is not None:
            cell_q_total.font = Font(name="Arial", size=20, bold=True, color=cell_q_total.font.color if cell_q_total.font else None)
        if Alignment is not None:
            cell_q_total.alignment = Alignment(horizontal="center", vertical="center")
        
        # Unmerge and update TOTAL OPTIONAL ADD-ONS (A LA CARTE) formula in column V
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_20 <= merged_range.max_row and
                merged_range.min_col <= col_u <= merged_range.max_col):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass
        
        # Set the correct formula: sum of optional add-ons total column V
        safe_set_cell_value(ws, f"V{row_20}", f"=SUM(V{first_ala_row}:V{last_ala_row})")
        cell_u_total = ws.cell(row=row_20, column=col_u)
        cell_u_total.number_format = '$#,##0'
        apply_font_20(cell_u_total)
        if Font is not None:
            cell_u_total.font = Font(name="Arial", size=20, bold=True, color=cell_u_total.font.color if cell_u_total.font else None)
        if Alignment is not None:
            cell_u_total.alignment = Alignment(horizontal="center", vertical="center")
        
        # Re-merge the totals row: B20:P20 for "TOTAL A LA CARTE" (value in Q20); S20:U20 for "TOTAL OPTIONAL ADD-ONS (A LA CARTE)" (value in V20)
        if safe_merge_cells and get_column_letter:
            # Re-merge B20:P20 so Q20 is free for TOTAL A LA CARTE value
            safe_merge_cells(ws, f"B{row_20}:P{row_20}")
            # Re-merge S20:U20 for the optional add-ons label (value is in V20)
            safe_merge_cells(ws, f"S{row_20}:U{row_20}")
    
    # Update deliverables counts in column F (A LA CARTE BREAKDOWN)
    # Count how many items have non-zero hours for each service
    if len(a_la_carte_items) > 0 and column_index_from_string is not None:
        col_b = column_index_from_string("B")
        col_f_deliverables = column_index_from_string("F")  # Deliverables go in column F (was E)
        
        # Map of deliverable labels to their corresponding hour field names or special handling
        # Fitting and Adjustment: item count, hard-code 1. Final Sample: count (not time).
        deliverables_map = [
            ("INTAKE SESSION", "intake_session"),
            ("1ST PATTERN", "first_pattern"),
            ("1ST SAMPLE", "first_sample"),
            ("FITTING", None),    # Item count, hard-code 1
            ("ADJUSTMENT", None),  # Item count, hard-code 1
            ("FINAL SAMPLE", "final_sample_quant"),  # Sum of Final Sample Quant (not "duplicates")
        ]
        
        def find_deliverable_row(label_text: str) -> int | None:
            """Find the row index for a deliverable label in column B."""
            lowered = label_text.strip().lower()
            # Deliverables section starts at row 22, shifted down by rows_to_insert
            start_scan = 22 + rows_to_insert
            end_scan = 40 + rows_to_insert  # Extended to include rates section
            for scan_row in range(start_scan, end_scan + 1):
                value = ws.cell(row=scan_row, column=col_b).value
                if isinstance(value, str):
                    value_clean = value.strip().lower()
                    if lowered in value_clean:
                        return scan_row
            return None
        
        # Calculate total hours/counts for each deliverable
        count_labels = {"intake session", "1st pattern", "1st sample"}
        for label_text, hour_field in deliverables_map:
            row_idx = find_deliverable_row(label_text)
            if row_idx is None:
                continue
            
            label_key = label_text.strip().lower()
            
            # FITTING and ADJUSTMENT: item count, hard-code 1 (not time)
            if label_key == "fitting" or label_key == "adjustment":
                total_value = 1
            # FINAL SAMPLE: sum of Final Sample Quant (quantity_sum)
            elif label_key == "final sample":
                total_value = quantity_sum
            elif hour_field is None:
                total_value = 0
            elif label_key in count_labels:
                total_value = sum(1 for item in a_la_carte_items if float(item.get(hour_field, 0)) > 0)
            else:
                total_value = sum(float(item.get(hour_field, 0)) for item in a_la_carte_items)
            
            # Unmerge column F if needed
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= row_idx <= merged_range.max_row and
                    merged_range.min_col <= col_f_deliverables <= merged_range.max_col):
                    try:
                        min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                        ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                    except Exception:
                        pass
            
            # Set the total value in column F
            safe_set_cell_value(ws, f"F{row_idx}", total_value)
            cell_f_count = ws.cell(row=row_idx, column=col_f_deliverables)
            # Use integer format for counts and quantities (Fitting, Adjustment, Final Sample, Intake, Pattern, Sample)
            if label_key in ("fitting", "adjustment", "final sample") or label_key in count_labels or total_value == int(total_value):
                cell_f_count.number_format = "0"
            else:
                cell_f_count.number_format = "0.00"
            apply_font_20(cell_f_count)
            if Alignment is not None:
                cell_f_count.alignment = Alignment(horizontal="center", vertical="center")
            
            # Re-merge if it was merged (typically F23:F24, F25:F26, etc.)
            row_second = row_idx + 1
            if safe_merge_cells:
                safe_merge_cells(ws, f"F{row_idx}:F{row_second}")
    
    # Update deliverables counts (OPTIONAL ADD-ONS: DYE TESTING S, PLANNING T, DESIGN U)
    # These formulas count the optional add-ons selected for A La Carte items
    if len(a_la_carte_items) > 0 and column_index_from_string is not None:
        label_column_idx = column_index_from_string("S")  # Labels in column S for A LA CARTE tab (DYE TESTING)
        target_col_t = column_index_from_string("U")      # Counts go in column U for OPTIONAL ADD-ONS (A LA CARTE)
        deliverable_addon_map = [
            ("DYE TESTING", "S", "S"),    # DYE TESTING counts column S
            ("PLANNING", "T", "T"),       # PLANNING counts column T
            ("DESIGN", "U", "U"),         # DESIGN counts column U
        ]
        
        # Deliverables section starts at row 22, shifted down by rows_to_insert
        deliverables_block_start = 22 + rows_to_insert
        deliverables_block_end = 36 + rows_to_insert
        
        def find_label_row(label_text: str) -> int | None:
            """Locate the row index for a given deliverable label (same as DEVELOPMENT ONLY)."""
            lowered = label_text.strip().lower()
            # Also create a key word version (e.g., "wash/treatment" -> "wash")
            key_words = lowered.replace("/", " ").split()
            primary_key = key_words[0] if key_words else lowered
            
            partial_match_row = None
            for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
                value = ws.cell(row=scan_row, column=label_column_idx).value
                if not isinstance(value, str):
                    continue
                value_clean = value.strip().lower()
                # Exact match first
                if value_clean == lowered:
                    return scan_row
                # Partial match (label text contained in cell value)
                if lowered in value_clean and partial_match_row is None:
                    partial_match_row = scan_row
                # Also check reverse (cell value contained in label text)
                if value_clean in lowered and partial_match_row is None:
                    partial_match_row = scan_row
                # Check if primary key word matches (e.g., "design" matches "DESIGN")
                if primary_key in value_clean and partial_match_row is None:
                    partial_match_row = scan_row
            return partial_match_row
        
        # Calculate the last row with an A La Carte item
        first_ala_row = 10
        last_ala_row = 10 + ((len(a_la_carte_items) - 1) * 2)
        
        # Update formulas for each optional add-on (same approach as DEVELOPMENT ONLY)
        for label_text, start_col_letter, end_col_letter in deliverable_addon_map:
            row_idx = find_label_row(label_text)
            
            # If not found, try searching all columns in the deliverables section
            if row_idx is None:
                # Search in columns H, N, and other potential label columns
                search_columns = [label_column_idx]  # Column H
                if column_index_from_string:
                    # Also try column N (14) which might have labels
                    search_columns.append(column_index_from_string("N"))
                
                for search_col in search_columns:
                    for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
                        value = ws.cell(row=scan_row, column=search_col).value
                        if isinstance(value, str):
                            value_lower = value.strip().lower()
                            label_lower = label_text.strip().lower()
                            # Check if label is in value or value is in label
                            if label_lower in value_lower or value_lower in label_lower:
                                # Also check key words
                                key_words = label_lower.replace("/", " ").split()
                                for key in key_words:
                                    if key in value_lower and len(key) > 2:  # Only if key word is meaningful
                                        row_idx = scan_row
                                        break
                                if row_idx:
                                    break
                    if row_idx:
                        break
            
            if row_idx is None:
                # Still not found - skip this one (but continue with others)
                continue
            
            row_second = row_idx + 1
            col_s = 19
            col_t = 20
            col_u = 21
            
            # Unmerge S, T, U for these two rows so we can re-merge and format
            for _ in range(3):
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                        merged_range.min_col <= col_u and merged_range.max_col >= col_s):
                        try:
                            min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                            ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                        except Exception:
                            pass
            
            # Label: ensure text in S (row_idx), then merge S:T over 2 rows, center, borders
            label_cell = ws.cell(row=row_idx, column=col_s)
            if not label_cell.value or str(label_cell.value).strip() == "":
                label_cell.value = label_text
            if Font is not None:
                apply_font_20(label_cell)
            # Apply borders to every cell in S:T (2 rows) before merging
            if Border is not None and Side is not None:
                thin_side = Side(style="thin")
                no_side = Side(style=None)
                for r in (row_idx, row_second):
                    top_side = thin_side if r == row_idx else no_side
                    bottom_side = thin_side if r == row_second else no_side
                    for c in (col_s, col_t):
                        cell = ws.cell(row=r, column=c)
                        if c == col_s:
                            cell.border = Border(left=thin_side, right=no_side, top=top_side, bottom=bottom_side)
                        else:
                            cell.border = Border(left=no_side, right=thin_side, top=top_side, bottom=bottom_side)
            safe_merge_cells(ws, f"S{row_idx}:T{row_second}")
            label_merged = safe_get_writable_cell(ws, row_idx, col_s)
            if Alignment is not None:
                label_merged.alignment = Alignment(horizontal="center", vertical="center")
            
            # Value: formula in U, merge U over 2 rows, center, borders
            addon_range = f"{start_col_letter}{first_ala_row}:{end_col_letter}{last_ala_row}"
            formula_text = f"=COUNT({addon_range})"
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                    merged_range.min_col <= col_u <= merged_range.max_col):
                    try:
                        min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                        ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                    except Exception:
                        pass
            
            # Remove any array formula designation first
            try:
                if hasattr(ws, 'array_formulae'):
                    array_formulae_to_remove = []
                    for array_range in ws.array_formulae:
                        if (array_range.min_row <= row_second and array_range.max_row >= row_idx and
                            array_range.min_col <= col_u <= array_range.max_col):
                            array_formulae_to_remove.append(array_range)
                    for arr in array_formulae_to_remove:
                        ws.array_formulae.remove(arr)
            except Exception:
                pass
            
            # Apply borders to U (2 rows) before merging
            if Border is not None and Side is not None:
                thin_side = Side(style="thin")
                no_side = Side(style=None)
                for r in (row_idx, row_second):
                    top_side = thin_side if r == row_idx else no_side
                    bottom_side = thin_side if r == row_second else no_side
                    cell_u = ws.cell(row=r, column=col_u)
                    cell_u.border = Border(left=thin_side, right=thin_side, top=top_side, bottom=bottom_side)
            
            value_cell = ws.cell(row=row_idx, column=col_u)
            value_cell.value = formula_text
            value_cell.number_format = "0"
            apply_font_20(value_cell)
            if Alignment is not None:
                value_cell.alignment = Alignment(horizontal="center", vertical="center")
            safe_merge_cells(ws, f"U{row_idx}:U{row_second}")
            value_merged = safe_get_writable_cell(ws, row_idx, col_u)
            if Alignment is not None:
                value_merged.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply font size 20 to all cells in the A LA CARTE tab (rows 8-36+rows_to_insert, columns B-Q)
    end_row = 36 + rows_to_insert
    for row in range(8, end_row + 1):
        for col in range(2, 18):  # Columns B through Q
            cell = ws.cell(row=row, column=col)
            apply_font_20(cell)
    
    return total_ala_carte, total_optional_ala_carte


def build_workbook_bytes(
    *,
    client_name: str,
    client_email: str,
    representative: str,
    style_entries: list[dict],
    custom_styles: list[dict],
    discount_percentage: float,
    a_la_carte_items: list[dict] = None,
    notes: list[str] = None,
) -> tuple[bytes, float, float]:
    """Load the template, update it, and return bytes plus totals.
    
    Generates two tabs:
    1. "DEVELOPMENT ONLY" - Development package with styles and custom items
    2. "A LA CARTE" - A La Carte package items
    """
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. Please add it to the environment first."
        )

    template_path = get_template_path()
    wb = load_workbook(template_path)
    
    # Check for required sheets
    if "DEVELOPMENT ONLY" not in wb.sheetnames:
        raise ValueError(
            f"Worksheet 'DEVELOPMENT ONLY' is missing from the template."
        )
    # A LA CARTE sheet is only required if there are items (checked later)
    
    # Process DEVELOPMENT ONLY tab
    ws_dev = wb["DEVELOPMENT ONLY"]
    ws_dev.title = "DEVELOPMENT ONLY"
    
    update_header_labels(ws_dev, client_name)
    total_dev, total_optional = apply_development_package(
        ws_dev,
        client_name=client_name,
        client_email=client_email,
        representative=representative,
        style_entries=style_entries,
        custom_styles=custom_styles,
        discount_percentage=discount_percentage,
    )
    
    # Write notes to column N starting below "PROJECT NOTES"
    # Find "PROJECT NOTES" dynamically (it moves when more styles are added)
    if notes:
        project_notes_row = None
        # Search for "PROJECT NOTES" in column N (column 14)
        for search_row in range(20, 50):  # Search from row 20 to 50
            cell_value = ws_dev.cell(row=search_row, column=SUMMARY_LABEL_COL).value
            if cell_value and isinstance(cell_value, str):
                if "PROJECT" in cell_value.upper() and "NOTES" in cell_value.upper():
                    project_notes_row = search_row
                    break
        
        if project_notes_row:
            # Place notes starting one row below "PROJECT NOTES" (every other row: N27, N29, etc.)
            notes_start_row = project_notes_row + 1
            note_index = 0
            for note in notes:
                if note and note.strip():
                    # Place notes at every other row (skip one row between notes)
                    cell_row = notes_start_row + (note_index * 2)
                    cell = ws_dev.cell(row=cell_row, column=SUMMARY_LABEL_COL)
                    # Uppercase and center-align the notes
                    cell.value = note.strip().upper()
                    if Font is not None:
                        cell.font = Font(name="Arial", size=20)
                    if Alignment is not None:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    note_index += 1
    
    # Process A LA CARTE tab only if there are items
    a_la_carte_items = a_la_carte_items or []
    total_ala_carte = 0.0
    total_optional_ala_carte = 0.0
    
    if len(a_la_carte_items) > 0:
        # Check for A LA CARTE sheet only if we need it
        if "A LA CARTE" not in wb.sheetnames:
            raise ValueError(
                f"Worksheet 'A LA CARTE' is missing from the template."
            )
        
        ws_ala = wb["A LA CARTE"]
        ws_ala.title = "A LA CARTE"
        
        total_ala_carte, total_optional_ala_carte = apply_ala_carte_package(
            ws_ala,
            client_name=client_name,
            client_email=client_email,
            representative=representative,
            a_la_carte_items=a_la_carte_items,
        )
    
    # Update DEVELOPMENT ONLY tab summary section to combine both tabs
    # Change N10 label from "TOTAL DEVELOPMENT" to "TOTAL PACKAGES"
    cell_n10 = ws_dev.cell(row=10, column=SUMMARY_LABEL_COL)
    safe_set_cell_value(ws_dev, "N10", "TOTAL PACKAGES")
    if Font is not None:
        cell_n10.font = Font(name="Arial", size=20, bold=True, color=cell_n10.font.color if cell_n10.font else None)
    if Alignment is not None:
        cell_n10.alignment = Alignment(horizontal="center", vertical="center")
    
    # Find the totals row in DEVELOPMENT ONLY tab
    # Calculate totals row based on number of styles
    num_styles = len(style_entries)
    num_custom_styles = len(custom_styles)
    total_styles_count = num_styles + num_custom_styles
    if total_styles_count <= 5:
        totals_row_dev = 20
    else:
        totals_row_dev = 20 + (total_styles_count - 5) * 2
    
    # Find the totals row in A LA CARTE tab
    num_ala_items = len(a_la_carte_items)
    rows_to_insert_ala = 0
    if num_ala_items > 5:
        rows_to_insert_ala = (num_ala_items - 5) * 2
    totals_row_ala = 20 + rows_to_insert_ala
    
    # Update P10: TOTAL PACKAGES = TOTAL DEVELOPMENT + TOTAL A LA CARTE
    # P10 should sum F{totals_row_dev} (DEVELOPMENT ONLY) + Q{totals_row_ala} (A LA CARTE total, now in column Q)
    cell_p10 = ws_dev.cell(row=10, column=SUMMARY_VALUE_COL)
    if num_ala_items > 0:
        # Sum from both tabs
        cell_p10.value = f"=F{totals_row_dev}+'A LA CARTE'!Q{totals_row_ala}"
    else:
        # Only DEVELOPMENT ONLY
        cell_p10.value = f"=F{totals_row_dev}"
    cell_p10.number_format = '$#,##0'
    if Font is not None:
        cell_p10.font = Font(name="Arial", size=20, bold=True, color=cell_p10.font.color if cell_p10.font else None)
    if Alignment is not None:
        cell_p10.alignment = Alignment(horizontal="center", vertical="center")
    
    # Update P12: TOTAL OPTIONAL ADD-ONS = TOTAL OPTIONAL ADD-ONS (DEV) + TOTAL OPTIONAL ADD-ONS (A LA CARTE)
    # P12 should sum K{totals_row_dev} (DEVELOPMENT ONLY) + V{totals_row_ala} (A LA CARTE optional add-ons)
    cell_p12 = ws_dev.cell(row=12, column=SUMMARY_VALUE_COL)
    if num_ala_items > 0:
        # Sum from both tabs
        cell_p12.value = f"=K{totals_row_dev}+'A LA CARTE'!V{totals_row_ala}"
    else:
        # Only DEVELOPMENT ONLY
        cell_p12.value = f"=K{totals_row_dev}"
    cell_p12.number_format = '$#,##0'
    if Font is not None:
        cell_p12.font = Font(name="Arial", size=20, bold=True, color=cell_p12.font.color if cell_p12.font else None)
    if Alignment is not None:
        cell_p12.alignment = Alignment(horizontal="center", vertical="center")
    
    # Remove any other worksheets (keep only DEVELOPMENT ONLY and A LA CARTE if it exists)
    sheets_to_keep = ["DEVELOPMENT ONLY"]
    if len(a_la_carte_items) > 0:
        sheets_to_keep.append("A LA CARTE")
    sheets_to_remove = [name for name in wb.sheetnames if name not in sheets_to_keep]
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), total_dev, total_optional


@st.cache_data(ttl=300)  # Cache for 5 minutes
def get_sales_records():
    """Get sales records from monday.com for dropdown selection."""
    try:
        from database_utils import get_sales_data
        
        sales_data = get_sales_data()
        items = sales_data.get("data", {}).get("boards", [{}])[0].get("items_page", {}).get("items", [])
        
        # Return list of (item_id, item_name) tuples
        records = [(item.get("id"), item.get("name", "")) for item in items if item.get("name")]
        return sorted(records, key=lambda x: x[1])  # Sort by name
    except Exception as e:
        st.warning(f"Could not load sales records: {e}")
        return []


def update_monday_item_a_la_carte_link(item_id: str, workbook_url: str) -> bool:
    """Update a monday.com item with the workbook URL in a 'A La Carte Link' field."""
    try:
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            st.error("Monday.com API token not found in secrets.")
            return False
        
        # Query to get board columns to find the "A La Carte Link" column
        query = f"""
        query {{
            items(ids: [{item_id}]) {{
                board {{
                    id
                    columns {{
                        id
                        title
                        type
                    }}
                }}
            }}
        }}
        """
        
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            st.error(f"Error fetching monday.com columns: {data['errors']}")
            return False
        
        items = data.get("data", {}).get("items", [])
        if not items:
            st.error("Item not found in monday.com")
            return False
        
        board = items[0].get("board", {})
        board_id = board.get("id")
        columns = board.get("columns", [])
        
        if not board_id:
            st.error("Could not determine board ID from monday.com item")
            return False
        
        # Find "A La Carte Link" column (case-insensitive, flexible matching)
        a_la_carte_column = None
        for col in columns:
            title_lower = col.get("title", "").lower()
            if ("dev" in title_lower and "inspection" in title_lower and "link" in title_lower) or \
               ("dev" in title_lower and "inspection" in title_lower and "url" in title_lower):
                a_la_carte_column = col
                break
        
        if not a_la_carte_column:
            st.warning("⚠️ 'A La Carte Link' column not found in monday.com. Please create a URL column named 'A La Carte Link' in the Sales board.")
            return False
        
        column_id = a_la_carte_column.get("id")
        column_type = a_la_carte_column.get("type")
        
        # Update the item with the workbook URL
        # For URL columns, the value format is: {"url": "https://...", "text": "Link Text"}
        if column_type == "link":
            mutation = f"""
            mutation {{
                change_column_value(
                    board_id: {board_id},
                    item_id: {item_id},
                    column_id: "{column_id}",
                    value: "{{\\"url\\": \\"{workbook_url}\\", \\"text\\": \\"View A La Carte\\"}}"
                ) {{
                    id
                }}
            }}
            """
        else:
            # For text columns, just use the URL as text
            mutation = f"""
            mutation {{
                change_column_value(
                    board_id: {board_id},
                    item_id: {item_id},
                    column_id: "{column_id}",
                    value: "{workbook_url}"
                ) {{
                    id
                }}
            }}
            """
        
        response = requests.post(url, json={"query": mutation}, headers=headers, timeout=30)
        result = response.json()
        
        if "errors" in result:
            st.error(f"Error updating monday.com: {result['errors']}")
            return False
        
        return True
        
    except Exception as e:
        st.error(f"Failed to update monday.com: {e}")
        return False


def upload_workbook_to_google_sheet_a_la_carte(
    workbook_bytes: bytes, sheet_name: str
) -> tuple[str, bool]:
    """
    Upload the XLSX workbook bytes to Google Drive using a_la_carte specific folder.
    
    Returns (web_url, converted_to_google_sheet).
    """
    from google_sheets_uploader import GoogleSheetsUploadError
    
    if not workbook_bytes:
        raise GoogleSheetsUploadError("Workbook data is empty; nothing to upload.")
    
    try:
        from google.oauth2.service_account import Credentials as SACredentials
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError:
        raise GoogleSheetsUploadError(
            "Google API libraries not available. Please install google-auth and google-api-python-client."
        )
    
    # Get a_la_carte specific folder ID from secrets
    cfg = st.secrets.get("google_drive", {}) or {}
    parent_folder_id = cfg.get("parent_folder_id_a_la_carte")
    
    if not parent_folder_id:
        raise GoogleSheetsUploadError(
            "parent_folder_id_a_la_carte not found in secrets. "
            "Add it to your Streamlit secrets under google_drive section."
        )
    
    # Get credentials
    info = st.secrets.get("google_service_account")
    if not info:
        raise GoogleSheetsUploadError(
            "Google Cloud service account credentials missing in secrets."
        )
    
    credentials = SACredentials.from_service_account_info(
        info,
        scopes=[
            "https://www.googleapis.com/auth/drive",
            "https://www.googleapis.com/auth/drive.file",
        ]
    )
    
    drive_service = build("drive", "v3", credentials=credentials)
    
    # Check if folder is in a Shared Drive
    shared_drive_id = None
    try:
        folder_info = drive_service.files().get(
            fileId=parent_folder_id,
            fields="id, name, driveId",
            supportsAllDrives=True
        ).execute()
        folder_drive_id = folder_info.get("driveId")
        if folder_drive_id:
            shared_drive_id = folder_drive_id
    except Exception:
        pass
    
    file_metadata = {
        "name": sheet_name or "A La Carte Workbook",
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "parents": [parent_folder_id],
    }
    
    media = MediaIoBaseUpload(
        io.BytesIO(workbook_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    
    try:
        create_kwargs = {
            "body": file_metadata,
            "media_body": media,
            "fields": "id, webViewLink",
        }
        if shared_drive_id:
            create_kwargs["supportsAllDrives"] = True
        created_file = drive_service.files().create(**create_kwargs).execute()
    except HttpError as exc:
        error_text = str(exc)
        if getattr(exc, "resp", None) and getattr(exc.resp, "status", None) == 403:
            if "storageQuotaExceeded" in error_text:
                raise GoogleSheetsUploadError(
                    "Google Drive storage quota has been exceeded. "
                    "Please delete older files or empty the Drive trash, then try again."
                ) from exc
        raise GoogleSheetsUploadError(f"Google Drive upload failed: {exc}") from exc
    except Exception as exc:
        raise GoogleSheetsUploadError(f"Google Drive upload failed: {exc}") from exc
    
    file_id = created_file.get("id")
    web_view = created_file.get("webViewLink")
    if not file_id:
        raise GoogleSheetsUploadError("Upload succeeded but Google Drive did not return a file ID.")
    
    converted = False
    try:
        copy_body = {
            "name": sheet_name or "A La Carte Workbook",
            "mimeType": "application/vnd.google-apps.spreadsheet",
            "parents": [parent_folder_id],
        }
        copy_kwargs = {
            "fileId": file_id,
            "body": copy_body,
            "fields": "id, webViewLink",
        }
        if shared_drive_id:
            copy_kwargs["supportsAllDrives"] = True
        converted_file = drive_service.files().copy(**copy_kwargs).execute()
        new_id = converted_file.get("id")
        new_link = converted_file.get("webViewLink")
        if new_id:
            try:
                delete_kwargs = {"fileId": file_id}
                if shared_drive_id:
                    delete_kwargs["supportsAllDrives"] = True
                drive_service.files().delete(**delete_kwargs).execute()
            except Exception:
                pass
            file_id = new_id
            web_view = new_link or web_view
            converted = True
    except HttpError as exc:
        if getattr(exc, "resp", None) and getattr(exc.resp, "status", None) == 403:
            if "storageQuotaExceeded" in str(exc):
                st.warning(
                    "Google Sheets conversion failed because the service account has zero Drive quota. "
                    "The XLSX workbook is still uploaded to the shared folder."
                )
            else:
                st.warning(f"Google Sheets conversion failed: {exc}")
        else:
            st.warning(f"Google Sheets conversion failed: {exc}")
    
    final_url = web_view or f"https://drive.google.com/file/d/{file_id}/view"
    return final_url, converted


def extract_spreadsheet_id_from_url(url: str) -> str | None:
    """Extract spreadsheet ID from Google Sheets or Google Drive URL.
    
    Args:
        url: Google Sheets URL (e.g., https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit)
             or Google Drive URL (e.g., https://drive.google.com/file/d/FILE_ID/view)
    
    Returns:
        Spreadsheet/file ID or None if not found
    """
    import re
    
    # Pattern for Google Sheets URL: /spreadsheets/d/{id}/
    sheets_pattern = r'/spreadsheets/d/([a-zA-Z0-9-_]+)'
    match = re.search(sheets_pattern, url)
    if match:
        return match.group(1)
    
    # Pattern for Google Drive URL: /file/d/{id}/
    drive_pattern = r'/file/d/([a-zA-Z0-9-_]+)'
    match = re.search(drive_pattern, url)
    if match:
        return match.group(1)
    
    return None


def export_google_sheet_as_pdf(sheet_url: str) -> bytes:
    """Export Google Sheet as PDF using Google Sheets export API.
    
    Args:
        sheet_url: The Google Sheet URL
        
    Returns:
        PDF file as bytes
        
    Raises:
        RuntimeError: If spreadsheet ID cannot be extracted or export fails
    """
    # Extract spreadsheet ID from URL
    spreadsheet_id = extract_spreadsheet_id_from_url(sheet_url)
    if not spreadsheet_id:
        raise RuntimeError(f"Could not extract spreadsheet ID from URL: {sheet_url}")
    
    # Build export URL with PDF format
    # Using landscape orientation and removing gridlines for better readability
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=pdf&portrait=false&gridlines=false"
    
    # Get Google credentials for authentication
    if not GOOGLE_API_AVAILABLE:
        raise RuntimeError("Google API libraries not available. Please install google-auth and google-api-python-client.")
    
    try:
        info = st.secrets.get("google_service_account")
        if not info:
            raise RuntimeError("Google service account credentials not found in secrets")
        
        credentials = SACredentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        
        # Refresh credentials to get access token
        if not credentials.valid:
            credentials.refresh(GoogleRequest())
        
        # Fetch the PDF using authenticated request
        import urllib.request
        
        req = urllib.request.Request(export_url)
        req.add_header('Authorization', f'Bearer {credentials.token}')
        
        with urllib.request.urlopen(req) as response:
            pdf_bytes = response.read()
            return pdf_bytes
            
    except Exception as e:
        raise RuntimeError(f"Failed to export Google Sheet as PDF: {e}")


def get_board_id_from_item(item_id: str) -> str | None:
    """Get board ID from a Monday.com item ID.
    
    Args:
        item_id: The Monday.com item ID
        
    Returns:
        Board ID or None if not found
    """
    try:
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            return None
        
        query = f"""
        query {{
            items(ids: [{item_id}]) {{
                board {{
                    id
                }}
            }}
        }}
        """
        
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            return None
        
        items = data.get("data", {}).get("items", [])
        if not items:
            return None
        
        board = items[0].get("board", {})
        return board.get("id")
        
    except Exception:
        return None


def upload_file_to_monday_item(item_id: str, board_id: str, file_bytes: bytes, filename: str) -> bool:
    """Upload a file to a monday.com item using the GraphQL file upload API."""
    try:
        import json
        
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            st.error("Monday.com API token not found in secrets.")
            return False
        
        # Step 1: Get the files column ID for this board
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        # Query to find the files column
        query = f"""
        query {{
            boards(ids: [{board_id}]) {{
                columns {{
                    id
                    title
                    type
                }}
            }}
        }}
        """
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            st.error(f"Error fetching board columns: {data['errors']}")
            return False
        
        boards = data.get("data", {}).get("boards", [])
        if not boards:
            st.error("Board not found")
            return False
        
        columns = boards[0].get("columns", [])
        
        # Find files column (type is "file" or title contains "file")
        files_column = None
        for col in columns:
            if col.get("type") == "file" or "file" in col.get("title", "").lower():
                files_column = col
                break
        
        if not files_column:
            st.warning("⚠️ No files column found on the board. File upload requires a files column.")
            return False
        
        column_id = files_column.get("id")
        
        # Step 2: Upload file using GraphQL file upload API
        # Use the /v2/file endpoint for file uploads with multipart/form-data
        file_url = "https://api.monday.com/v2/file"
        
        # GraphQL mutation for adding file to column
        mutation = """
        mutation addFile($file: File!) {
            add_file_to_column(
                file: $file,
                item_id: %s,
                column_id: "%s"
            ) {
                id
            }
        }
        """ % (item_id, column_id)
        
        # Prepare multipart form data for GraphQL file upload
        # The format is: query, variables, map, and file
        variables = {
            "file": None
        }
        
        file_map = {
            "file": ["variables.file"]
        }
        
        # Determine MIME type based on file extension
        if filename.lower().endswith('.pdf'):
            mime_type = 'application/pdf'
        elif filename.lower().endswith('.xlsx'):
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mime_type = 'application/octet-stream'
        
        # Create multipart form data
        files_data = {
            'query': (None, mutation),
            'variables': (None, json.dumps(variables)),
            'map': (None, json.dumps(file_map)),
            'file': (filename, file_bytes, mime_type)
        }
        
        # Upload the file
        upload_response = requests.post(
            file_url,
            headers={"Authorization": api_token},
            files=files_data,
            timeout=60
        )
        
        if upload_response.status_code not in [200, 201]:
            error_text = upload_response.text
            try:
                error_json = upload_response.json()
                if "errors" in error_json:
                    st.error(f"Error uploading file: {error_json['errors']}")
                else:
                    st.error(f"Error uploading file: {error_text}")
            except:
                st.error(f"Error uploading file: {upload_response.status_code} - {error_text}")
            return False
        
        result = upload_response.json()
        
        if "errors" in result:
            st.error(f"Error uploading file: {result['errors']}")
            return False
        
        # Check if file was added successfully
        if result.get("data", {}).get("add_file_to_column", {}).get("id"):
            return True
        else:
            st.warning("File upload may have succeeded but no file ID returned")
            return True  # Assume success if no errors
        
    except Exception as e:
        st.error(f"Failed to upload file to monday.com: {e}")
        return False


def main() -> None:
    st.set_page_config(
        page_title="A La Carte Creator",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.title("📊 A La Carte Creator")

    st.markdown(
        """
<style>

    /* Enlarge checkbox */
    [data-testid="stCheckbox"] label span {
        height: 1.8rem;
        width: 1.8rem;
        transform: translateX(35px);
    }

    /* Hide normal label text (since you use empty label) */
    [data-testid="stCheckbox"] label p {
        font-size: 0;
        margin: 0;
    }

    /* Move right the entire checkbox inside its Streamlit column */
    [data-testid="stCheckbox"] {
        display: flex !important;
        justify-content: right !important;  /* Center horizontally */
        align-items: right !important;       /* Center vertically */
        height: 100%;                          /* Align with text inputs */
        padding: 0 !important;
        margin: 0 !important;
    }

    /* Also move right the internal label wrapper */
    [data-testid="stCheckbox"] > label {
        display: flex !important;
        justify-content: right !important;
        align-items: right !important;
        width: 100%;
        height: 100%;
        padding: 0 !important;
        margin: 0 !important;
    }

    /* Hide ALL sidebar list items by default */
    [data-testid="stSidebarNav"] li {
        display: none !important;
    }

    /* Show list items that contain allowed tool pages using :has() selector */
    [data-testid="stSidebarNav"] li:has(a[href*="signnow"]),
    [data-testid="stSidebarNav"] li:has(a[href*="/tools"]),
    [data-testid="stSidebarNav"] li:has(a[href*="workbook"]),
    [data-testid="stSidebarNav"] li:has(a[href*="deck_creator"]),
    [data-testid="stSidebarNav"] li:has(a[href*="a_la_carte"]) {
        display: block !important;
    }

</style>
<script>
// JavaScript to show only tool pages and hide everything else
(function() {
    function showToolPagesOnly() {
        const navItems = document.querySelectorAll('[data-testid="stSidebarNav"] li');
        const allowedPages = ['signnow', 'tools', 'workbook', 'deck_creator', 'a_la_carte'];
        
        // Check if we're currently on an ads dashboard page
        const currentUrl = window.location.href.toLowerCase();
        const currentPath = window.location.pathname.toLowerCase();
        const isOnAdsDashboard = currentUrl.includes('ads') && currentUrl.includes('dashboard') ||
                                 currentPath.includes('ads') && currentPath.includes('dashboard');
        
        navItems.forEach(item => {
            const link = item.querySelector('a');
            if (!link) {
                item.style.setProperty('display', 'none', 'important');
                return;
            }
            
            const href = (link.getAttribute('href') || '').toLowerCase();
            const text = link.textContent.trim().toLowerCase();
            
            // Check if this is an allowed tool page
            const isToolPage = allowedPages.some(page => {
                return href.includes(page) || text.includes(page.toLowerCase());
            });
            
            // Make sure it's not ads dashboard or other dashboards
            const isDashboard = (text.includes('ads') && text.includes('dashboard')) || 
                              (href.includes('ads') && href.includes('dashboard'));
            
            // Hide a_la_carte if we're on an ads dashboard page
            const isDevInspection = href.includes('a_la_carte') || text.includes('a_la_carte');
            if (isOnAdsDashboard && isDevInspection) {
                item.style.setProperty('display', 'none', 'important');
                return;
            }
            
            if (isToolPage && !isDashboard) {
                item.style.setProperty('display', 'block', 'important');
                link.style.setProperty('display', 'block', 'important');
            } else {
                item.style.setProperty('display', 'none', 'important');
            }
        });
    }
    
    // Run immediately and on load
    showToolPagesOnly();
    window.addEventListener('load', function() {
        setTimeout(showToolPagesOnly, 50);
        setTimeout(showToolPagesOnly, 200);
        setTimeout(showToolPagesOnly, 500);
    });
    
    // Watch for DOM changes
    const observer = new MutationObserver(function() {
        showToolPagesOnly();
    });
    
    setTimeout(function() {
        const sidebar = document.querySelector('[data-testid="stSidebarNav"]');
        if (sidebar) {
            observer.observe(sidebar, { 
                childList: true, 
                subtree: true,
                attributes: true
            });
        }
    }, 100);
})();
</script>
""",
        unsafe_allow_html=True,
    )

    st.caption(
        "Fill in the Development Package inputs and download a formatted workbook "
        "based on the official template."
    )

    # Get query parameters from Monday.com link
    query_params = st.query_params
    first_name = query_params.get("first_name", "").strip()
    last_name = query_params.get("last_name", "").strip()
    email = query_params.get("email", "").strip()
    representative = query_params.get("representative", "").strip()
    num_styles_param = query_params.get("num_styles", "").strip()
    item_id = query_params.get("item_id", "").strip()
    
    # Clean up representative (remove "$" prefix if present from Monday.com formula)
    if representative.startswith("$"):
        representative = representative[1:].strip()
    
    # Combine first_name and last_name for client_name
    client_name_default = ""
    if first_name or last_name:
        client_name_default = f"{first_name} {last_name}".strip()
    
    # Use query params as default values if available
    client_name = st.text_input(
        "Client Name", 
        value=client_name_default,
        placeholder="Enter client name"
    )

    col_a, col_b = st.columns(2)
    with col_a:
        client_email = st.text_input(
            "Client Email", 
            value=email,
            placeholder="client@email.com"
        )
    with col_b:
        representative = st.text_input(
            "Representative", 
            value=representative,
            placeholder="Enter representative"
        )

    # Initialize session state for style entries
    if "style_entries" not in st.session_state:
        st.session_state["style_entries"] = []
    
    # Initialize Custom Items
    if "custom_styles" not in st.session_state:
        st.session_state["custom_styles"] = []
    
    st.subheader("**Styles**")
    
    # Number of Styles field (auto-fills from query param, but editable) - in first column, half width
    col_num_styles, col_spacer = st.columns([1, 3])
    with col_num_styles:
        num_styles_default = len(st.session_state["style_entries"])
        if num_styles_param:
            try:
                num_styles_default = int(num_styles_param)
            except ValueError:
                num_styles_default = len(st.session_state["style_entries"])
        
        num_styles_input = st.number_input(
            "Number of Styles",
            min_value=0,
            value=num_styles_default,
            step=1,
            key="num_styles_input",
            help="Number of styles/items. Auto-filled from Monday.com if available."
        )
        
        # Update style entries list to match num_styles_input
        current_count = len(st.session_state["style_entries"])
        if num_styles_input > current_count:
            # Add new styles with blank names
            for i in range(current_count, num_styles_input):
                st.session_state["style_entries"].append({
                    "name": "",  # Leave blank, don't default to "Style 1", etc.
                    "style_type": "Regular",  # Default to Regular (matching workbook_creator.py)
                    "complexity": 0.0,
                    "style_number": 101 + i,  # Default style numbers: 101, 102, 103...
                    "options": {
                        "wash_dye": False,
                        "design": False,
                        "source": False,
                        "treatment": False,
                    },
                })
        elif num_styles_input < current_count:
            # Remove excess styles
            st.session_state["style_entries"] = st.session_state["style_entries"][:num_styles_input]
    
    # Column headers (removed Source column)
    header_cols = st.columns([1.1, 1.8, 1.2, 1.2, 1.2, 1, 1.1])
    with header_cols[0]:
        st.markdown("**Style Number**")
    with header_cols[1]:
        st.markdown("**Style Name**")
    with header_cols[2]:
        st.markdown("**Style Type**")
    with header_cols[3]:
        st.markdown("**Complexity (%)**")
    with header_cols[4]:
        st.markdown("**Wash/Dye ($1,500)**")
    with header_cols[5]:
        st.markdown("**Design ($1,500)**")
    with header_cols[6]:
        st.markdown("**Treatment ($860)**")
    
    # Display existing style entries in horizontal rows
    if st.session_state["style_entries"]:
        for i, entry in enumerate(st.session_state["style_entries"]):
            with st.container():
                cols = st.columns([1.2, 1.8, 1.2, 1.2, 1.2, 1, 1.1])  # Removed source column, updated Style Type width
                with cols[0]:
                    # Style Number field with default value (101, 102, 103...)
                    default_style_number = entry.get("style_number", 101 + i)
                    style_number = st.number_input(
                        "Style Number",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with cols[1]:
                    style_name = st.text_input(
                        "Custom Item Name",
                        value=entry.get("name", ""),
                        key=f"style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Dress, Winter Coat",
                    )
                    entry["name"] = style_name
                with cols[2]:
                    # Style Type dropdown instead of checkbox (matching workbook_creator.py)
                    style_type_options = ["Regular", "Activewear/Lingerie/Swim", "Pattern Blocks"]
                    current_style_type = entry.get("style_type", "Regular")
                    # Handle migration from old "activewear" boolean
                    if "style_type" not in entry and entry.get("activewear", False):
                        current_style_type = "Activewear/Lingerie/Swim"
                    style_type = st.selectbox(
                        "Style Type",
                        options=style_type_options,
                        index=style_type_options.index(current_style_type) if current_style_type in style_type_options else 0,
                        key=f"style_type_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_type"] = style_type
                    # Remove old activewear key if it exists
                    if "activewear" in entry:
                        del entry["activewear"]
                with cols[3]:
                    complexity = st.number_input(
                        "Complexity (%)",
                        min_value=0,
                        max_value=200,
                        value=int(entry.get("complexity", 100)),
                        step=5,
                        format="%d",
                        key=f"complexity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["complexity"] = float(complexity)
                with cols[4]:
                    wash_dye = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("wash_dye", False),
                        key=f"wash_dye_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["wash_dye"] = wash_dye
                with cols[5]:
                    design = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("design", False),
                        key=f"design_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["design"] = design
                # Removed source column - no longer an option
                with cols[6]:
                    treatment = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("treatment", False),
                        key=f"treatment_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["treatment"] = treatment

    # Custom Item section
    st.markdown("---")
    st.subheader("**Custom Item**")
    
    # Calculate number of regular styles for custom item numbering
    num_regular_styles = len(st.session_state.get("style_entries", []))
    
    # QuickSelect buttons for common custom items
    quick_select_cols = st.columns([1, 1, 2])
    with quick_select_cols[0]:
        if st.button("➕ Sourcing ($2,050)", key="quick_sourcing", help="Add Sourcing — $2,050"):
            next_num = 101 + num_regular_styles + len(st.session_state["custom_styles"])
            st.session_state["custom_styles"].append({
                "name": "Sourcing",
                "price": 2050.0,
                "complexity": 0.0,
                "style_number": next_num,
            })
            st.rerun()
    with quick_select_cols[1]:
        if st.button("➕ Sourcing Consult ($860)", key="quick_sourcing_consult", help="Add Sourcing Consult — $860"):
            next_num = 101 + num_regular_styles + len(st.session_state["custom_styles"])
            st.session_state["custom_styles"].append({
                "name": "Sourcing Consult",
                "price": 860.0,
                "complexity": 0.0,
                "style_number": next_num,
            })
            st.rerun()
    
    # Column headers for Custom Items
    custom_header_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
    with custom_header_cols[0]:
        st.markdown("**Style Number**")
    with custom_header_cols[1]:
        st.markdown("**Custom Item Name**")
    with custom_header_cols[2]:
        st.markdown("**Price ($)**")
    with custom_header_cols[3]:
        st.markdown("**Complexity (%)**")
    with custom_header_cols[4]:
        st.markdown("**Remove**")
    
    # Display existing Custom Items
    if st.session_state["custom_styles"]:
        for i, entry in enumerate(st.session_state["custom_styles"]):
            with st.container():
                custom_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
                with custom_cols[0]:
                    # Style Number field with default value (101 + num_regular_styles + i)
                    default_style_number = entry.get("style_number", 101 + num_regular_styles + i)
                    style_number = st.number_input(
                        "Style Number",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"custom_style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with custom_cols[1]:
                    custom_style_name = st.text_input(
                        "Custom Item Name",
                        value=entry.get("name", ""),
                        key=f"custom_style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Custom Item",
                    )
                    entry["name"] = custom_style_name
                with custom_cols[2]:
                    custom_price = st.number_input(
                        "Price",
                        min_value=0.0,
                        value=float(entry.get("price", 0.0)),
                        step=100.0,
                        format="%.2f",
                        key=f"custom_price_{i}",
                        label_visibility="collapsed",
                    )
                    entry["price"] = float(custom_price)
                with custom_cols[3]:
                    custom_complexity = st.number_input(
                        "Complexity (%)",
                        min_value=0,
                        max_value=200,
                        value=int(entry.get("complexity", 0)),
                        step=5,
                        format="%d",
                        key=f"custom_complexity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["complexity"] = float(custom_complexity)
                with custom_cols[4]:
                    if st.button("❌", key=f"remove_custom_{i}", help="Remove this Custom Item"):
                        st.session_state["custom_styles"].pop(i)
                        st.rerun()
    
    # Add new Custom Item interface
    st.markdown("**Add New Custom Item**")
    add_custom_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
    default_new_custom_name = st.session_state.get("new_custom_style_name", "")
    default_new_custom_price = st.session_state.get("new_custom_price", 0.0)
    default_new_custom_complexity = st.session_state.get("new_custom_complexity", 0)
    default_new_custom_style_number = st.session_state.get("new_custom_style_number", 101 + num_regular_styles + len(st.session_state.get("custom_styles", [])))
    
    with add_custom_cols[0]:
        new_custom_style_number = st.number_input(
            "Style Number",
            min_value=1,
            value=int(default_new_custom_style_number),
            step=1,
            key="new_custom_style_number",
            label_visibility="collapsed",
        )
    with add_custom_cols[1]:
        new_custom_style_name = st.text_input(
            "Custom Item Name",
            value=default_new_custom_name,
            key="new_custom_style_name",
            label_visibility="collapsed",
            placeholder="e.g., Custom Item",
        )
    with add_custom_cols[2]:
        new_custom_price = st.number_input(
            "Price",
            min_value=0.0,
            value=default_new_custom_price,
            step=100.0,
            format="%.2f",
            key="new_custom_price",
            label_visibility="collapsed",
        )
    with add_custom_cols[3]:
        new_custom_complexity = st.number_input(
            "Complexity (%)",
            min_value=0,
            max_value=200,
            value=default_new_custom_complexity,
            step=5,
            format="%d",
            key="new_custom_complexity",
            label_visibility="collapsed",
        )
    with add_custom_cols[4]:
        if st.button("➕ Add", key="add_custom_style", help="Add this Custom Item"):
            if new_custom_style_name.strip() and new_custom_price > 0:
                st.session_state["custom_styles"].append({
                    "name": new_custom_style_name.strip(),
                    "price": float(new_custom_price),
                    "complexity": float(new_custom_complexity),
                    "style_number": int(new_custom_style_number),
                })
                # Reset add-new-custom-style inputs
                for key in [
                    "new_custom_style_name",
                    "new_custom_price",
                    "new_custom_complexity",
                    "new_custom_style_number",
                ]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
            else:
                st.warning("Please enter a Custom Item name and price before adding.")

    # A La Carte Package section
    st.markdown("---")
    st.subheader("**A La Carte Package**")
    
    # Initialize A La Carte items in session state
    if "a_la_carte_items" not in st.session_state:
        st.session_state["a_la_carte_items"] = []
    
    # Calculate number for a la carte item numbering (continues from custom items)
    num_regular_styles = len(st.session_state.get("style_entries", []))
    num_custom_styles = len(st.session_state.get("custom_styles", []))
    a_la_carte_start_number = 101 + num_regular_styles + num_custom_styles
    
    # Column headers for A La Carte Items (match Excel: Intake, 1st Pattern, 1st Sample, Fitting, Adjustment, Final Sample Hrs, Final Sample Qty)
    a_la_carte_header_cols = st.columns([0.6, 1.2, 0.45, 0.45, 0.45, 0.45, 0.45, 0.45, 0.45, 0.5, 0.5, 0.5, 0.4])
    with a_la_carte_header_cols[0]:
        st.markdown("**#**")
    with a_la_carte_header_cols[1]:
        st.markdown("**Style Name**")
    with a_la_carte_header_cols[2]:
        st.markdown("**INTAKE**")
    with a_la_carte_header_cols[3]:
        st.markdown("**1 PATTERN**")
    with a_la_carte_header_cols[4]:
        st.markdown("**1 SAMPLE**")
    with a_la_carte_header_cols[5]:
        st.markdown("**FITTING**")
    with a_la_carte_header_cols[6]:
        st.markdown("**ADJUST.**")
    with a_la_carte_header_cols[7]:
        st.markdown("**FIN SAMP HR**")
    with a_la_carte_header_cols[8]:
        st.markdown("**FIN SAMP QT**")
    with a_la_carte_header_cols[9]:
        st.markdown("**Dye Testing**")
    with a_la_carte_header_cols[10]:
        st.markdown("**Planning**")
    with a_la_carte_header_cols[11]:
        st.markdown("**Design**")
    with a_la_carte_header_cols[12]:
        st.markdown("**Remove**")
    
    # Display existing A La Carte Items
    if st.session_state["a_la_carte_items"]:
        for i, entry in enumerate(st.session_state["a_la_carte_items"]):
            with st.container():
                a_la_carte_cols = st.columns([0.6, 1.2, 0.45, 0.45, 0.45, 0.45, 0.45, 0.45, 0.45, 0.5, 0.5, 0.5, 0.4])
                with a_la_carte_cols[0]:
                    default_style_number = entry.get("style_number", a_la_carte_start_number + i)
                    style_number = st.number_input(
                        "#",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"a_la_carte_style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with a_la_carte_cols[1]:
                    style_name = st.text_input(
                        "Style Name",
                        value=entry.get("name", ""),
                        key=f"a_la_carte_style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Style Name",
                    )
                    entry["name"] = style_name
                with a_la_carte_cols[2]:
                    intake_hours = st.number_input(
                        "INTAKE",
                        min_value=0.0,
                        value=float(entry.get("intake_session", 0.0)),
                        step=0.25,
                        format="%.2f",
                        key=f"a_la_carte_intake_{i}",
                        label_visibility="collapsed",
                    )
                    entry["intake_session"] = float(intake_hours)
                with a_la_carte_cols[3]:
                    first_pattern_hours = st.number_input(
                        "1ST PATTERN",
                        min_value=0.0,
                        value=float(entry.get("first_pattern", 0.0)),
                        step=0.25,
                        format="%.2f",
                        key=f"a_la_carte_first_pattern_{i}",
                        label_visibility="collapsed",
                    )
                    entry["first_pattern"] = float(first_pattern_hours)
                with a_la_carte_cols[4]:
                    first_sample_hours = st.number_input(
                        "1ST SAMPLE",
                        min_value=0.0,
                        value=float(entry.get("first_sample", 0.0)),
                        step=0.25,
                        format="%.2f",
                        key=f"a_la_carte_first_sample_{i}",
                        label_visibility="collapsed",
                    )
                    entry["first_sample"] = float(first_sample_hours)
                with a_la_carte_cols[5]:
                    fitting_hours = st.number_input(
                        "FITTING",
                        min_value=0.0,
                        value=float(entry.get("fitting", 0.0)),
                        step=0.25,
                        format="%.2f",
                        key=f"a_la_carte_fitting_{i}",
                        label_visibility="collapsed",
                    )
                    entry["fitting"] = float(fitting_hours)
                with a_la_carte_cols[6]:
                    adjustment_hours = st.number_input(
                        "ADJUSTMENT",
                        min_value=0.0,
                        value=float(entry.get("adjustment", 0.0)),
                        step=0.25,
                        format="%.2f",
                        key=f"a_la_carte_adjustment_{i}",
                        label_visibility="collapsed",
                    )
                    entry["adjustment"] = float(adjustment_hours)
                with a_la_carte_cols[7]:
                    final_sample_hours = st.number_input(
                        "FINAL SAMPLE HRS",
                        min_value=0.0,
                        value=float(entry.get("duplicates", 0.0)),
                        step=0.25,
                        format="%.2f",
                        key=f"a_la_carte_duplicates_{i}",
                        label_visibility="collapsed",
                    )
                    entry["duplicates"] = float(final_sample_hours)
                with a_la_carte_cols[8]:
                    quantity_val = st.number_input(
                        "FINAL SAMPLE QTY",
                        min_value=0,
                        value=int(entry.get("quantity", 1)),
                        step=1,
                        key=f"a_la_carte_quantity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["quantity"] = int(quantity_val)
                with a_la_carte_cols[9]:
                    dye_testing = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("dye_testing", False),
                        key=f"a_la_carte_dye_testing_{i}",
                        label_visibility="collapsed",
                    )
                    if "options" not in entry:
                        entry["options"] = {}
                    entry["options"]["dye_testing"] = dye_testing
                with a_la_carte_cols[10]:
                    planning = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("planning", False),
                        key=f"a_la_carte_planning_{i}",
                        label_visibility="collapsed",
                    )
                    entry["options"]["planning"] = planning
                with a_la_carte_cols[11]:
                    design = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("design", False),
                        key=f"a_la_carte_design_{i}",
                        label_visibility="collapsed",
                    )
                    entry["options"]["design"] = design
                with a_la_carte_cols[12]:
                    if st.button("❌", key=f"remove_a_la_carte_{i}", help="Remove this A La Carte Item"):
                        st.session_state["a_la_carte_items"].pop(i)
                        st.rerun()
    
    # Add new A La Carte Item interface
    st.subheader("**Add New A La Carte Item**")
    add_a_la_carte_cols = st.columns([0.6, 1.2, 0.45, 0.45, 0.45, 0.45, 0.45, 0.45, 0.45, 0.5, 0.5, 0.5, 0.4])
    default_new_a_la_carte_style_name = st.session_state.get("new_a_la_carte_style_name", "")
    default_new_a_la_carte_style_number = st.session_state.get("new_a_la_carte_style_number", a_la_carte_start_number + len(st.session_state.get("a_la_carte_items", [])))
    
    with add_a_la_carte_cols[0]:
        new_a_la_carte_style_number = st.number_input(
            "#",
            min_value=1,
            value=int(default_new_a_la_carte_style_number),
            step=1,
            key="new_a_la_carte_style_number",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[1]:
        new_a_la_carte_style_name = st.text_input(
            "Style Name",
            value=default_new_a_la_carte_style_name,
            key="new_a_la_carte_style_name",
            label_visibility="collapsed",
            placeholder="e.g., Style Name",
        )
    with add_a_la_carte_cols[2]:
        new_intake = st.number_input(
            "INTAKE",
            min_value=0.0,
            value=0.0,
            step=0.25,
            format="%.2f",
            key="new_a_la_carte_intake",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[3]:
        new_first_pattern = st.number_input(
            "1ST PATTERN",
            min_value=0.0,
            value=0.0,
            step=0.25,
            format="%.2f",
            key="new_a_la_carte_first_pattern",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[4]:
        new_first_sample = st.number_input(
            "1ST SAMPLE",
            min_value=0.0,
            value=0.0,
            step=0.25,
            format="%.2f",
            key="new_a_la_carte_first_sample",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[5]:
        new_fitting = st.number_input(
            "FITTING",
            min_value=0.0,
            value=0.0,
            step=0.25,
            format="%.2f",
            key="new_a_la_carte_fitting",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[6]:
        new_adjustment = st.number_input(
            "ADJUSTMENT",
            min_value=0.0,
            value=0.0,
            step=0.25,
            format="%.2f",
            key="new_a_la_carte_adjustment",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[7]:
        new_final_sample_hrs = st.number_input(
            "FINAL SAMPLE HRS",
            min_value=0.0,
            value=0.0,
            step=0.25,
            format="%.2f",
            key="new_a_la_carte_duplicates",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[8]:
        new_quantity = st.number_input(
            "FINAL SAMPLE QTY",
            min_value=0,
            value=1,
            step=1,
            key="new_a_la_carte_quantity",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[9]:
        new_dye_testing = st.checkbox(
            "",
            value=False,
            key="new_a_la_carte_dye_testing",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[10]:
        new_planning = st.checkbox(
            "",
            value=False,
            key="new_a_la_carte_planning",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[11]:
        new_design = st.checkbox(
            "",
            value=False,
            key="new_a_la_carte_design",
            label_visibility="collapsed",
        )
    with add_a_la_carte_cols[12]:
        if st.button("➕ Add", key="add_a_la_carte_style", help="Add this A La Carte Item"):
            if new_a_la_carte_style_name.strip():
                st.session_state["a_la_carte_items"].append({
                    "name": new_a_la_carte_style_name.strip(),
                    "style_number": int(new_a_la_carte_style_number),
                    "intake_session": float(new_intake),
                    "first_pattern": float(new_first_pattern),
                    "first_sample": float(new_first_sample),
                    "fitting": float(new_fitting),
                    "adjustment": float(new_adjustment),
                    "duplicates": float(new_final_sample_hrs),
                    "quantity": int(new_quantity),
                    "options": {
                        "dye_testing": new_dye_testing,
                        "planning": new_planning,
                        "design": new_design,
                    },
                })
                for key in [
                    "new_a_la_carte_style_name",
                    "new_a_la_carte_style_number",
                ]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
            else:
                st.warning("Please enter a Style Name before adding.")

    if not st.session_state["style_entries"] and not st.session_state["custom_styles"] and not st.session_state.get("a_la_carte_items", []):
        st.info("Add at least one style, Custom Item, or A La Carte Item to enable the generator.")
        return

    st.markdown("---")
    st.subheader("Discount")
    discount_cols = st.columns([1.8, 1, 1, 1.2, 1, 0.8, 0.8, 1, 0.8])
    with discount_cols[0]:
        discount_percentage = st.number_input(
            "Discount (%)",
            min_value=0,
            max_value=100,
            value=st.session_state.get("discount_percentage_value", 0),
            step=1,
        )
    st.session_state["discount_percentage_value"] = discount_percentage
    
    # Notes Section
    st.markdown("---")
    st.subheader("Notes")
    
    # Initialize notes list in session state
    if "notes_list" not in st.session_state:
        st.session_state["notes_list"] = [""]
    
    # Display existing notes
    for i, note in enumerate(st.session_state["notes_list"]):
        note_cols = st.columns([10, 1])
        with note_cols[0]:
            updated_note = st.text_input(
                "Note",
                value=note,
                key=f"note_{i}",
                placeholder="Enter a note...",
                label_visibility="collapsed"
            )
            st.session_state["notes_list"][i] = updated_note
        with note_cols[1]:
            if st.button("❌", key=f"remove_note_{i}", help="Remove this note"):
                st.session_state["notes_list"].pop(i)
                st.rerun()
    
    # Add new note button
    if st.button("➕ Add Note", key="add_note"):
        st.session_state["notes_list"].append("")
        st.rerun()
    
    # Filter out empty notes for workbook generation
    notes = [note for note in st.session_state["notes_list"] if note and note.strip()]

    try:
        excel_bytes, _, _ = build_workbook_bytes(
            client_name=client_name,
            client_email=client_email,
            representative=representative,
            style_entries=st.session_state["style_entries"],
            custom_styles=st.session_state.get("custom_styles", []),
            discount_percentage=discount_percentage,
            a_la_carte_items=st.session_state.get("a_la_carte_items", []),
            notes=notes if notes else [],
        )
    except FileNotFoundError as exc:
        st.error(str(exc))
        return
    except Exception as exc:  # pragma: no cover - streamlit runtime feedback
        st.error(f"Failed to build workbook: {exc}")
        return

    safe_client = re.sub(r"[^A-Za-z0-9_-]+", "_", (client_name or "").strip()) or "client"
    download_name = f"a_la_carte_{safe_client.lower()}.xlsx"

    st.success("Workbook is ready.")
    st.download_button(
        label="Generate Workbook",
        data=excel_bytes,
        file_name=download_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    # Google Drive upload section (service account only)
    st.markdown("---")
    st.subheader("Google Sheets --> Monday.com Upload")

    sheet_title = (client_name or "Workbook").strip() or "Workbook"
    sheet_title = f"{sheet_title} - A La Carte"

    st.caption("Uploads will use the shared Google Drive folder configured for the service account.")

    if st.button("Upload to Monday.com", type="primary"):
        with st.spinner("Uploading workbook to Google Sheets and updating Monday.com..."):
            try:
                # Upload to Google Sheets using a_la_carte specific folder
                sheet_url, converted = upload_workbook_to_google_sheet_a_la_carte(excel_bytes, sheet_title)
                st.session_state["google_sheet_url"] = sheet_url
                
                if converted:
                    st.success(f"✅ Google Sheet created: [Open Sheet]({sheet_url})")
                else:
                    st.success(
                        f"✅ Workbook uploaded as XLSX: [Download / Open]({sheet_url})  \n"
                        "Google denied automatic conversion because service accounts report zero Drive quota. "
                        "Open the file in Google Drive and use **File → Save as Google Sheets** if you need an editable sheet."
                    )
                
                # Update Monday.com with the Google Sheet URL if item_id is provided
                if item_id:
                    # Update Monday.com item with the Google Sheet link
                    if update_monday_item_a_la_carte_link(item_id, sheet_url):
                        st.success(f"✅ Monday.com item updated with Google Sheet link!")
                    else:
                        st.warning("⚠️ Google Sheet uploaded, but failed to update Monday.com item. Please update manually.")
                    
                    # Also upload PDF version to Files column (only if converted to Google Sheet)
                    if converted:
                        try:
                            # Get board_id from item_id
                            board_id = get_board_id_from_item(item_id)
                            if board_id:
                                # Export Google Sheet as PDF using Google Sheets export API
                                pdf_bytes = export_google_sheet_as_pdf(sheet_url)
                                pdf_filename = f"{safe_client}_workbook.pdf"
                                
                                # Upload PDF to Monday.com Files column
                                if upload_file_to_monday_item(item_id, board_id, pdf_bytes, pdf_filename):
                                    st.success(f"✅ PDF version uploaded to Monday.com Files column!")
                                else:
                                    st.warning("⚠️ PDF upload to Monday.com Files column failed. Link was updated successfully.")
                            else:
                                st.warning("⚠️ Could not retrieve board ID. PDF upload skipped.")
                        except Exception as pdf_error:
                            st.warning(f"⚠️ PDF export/upload failed: {pdf_error}. Link was updated successfully.")
                    else:
                        st.info("ℹ️ PDF export skipped (workbook not converted to Google Sheet format).")
                else:
                    st.info("ℹ️ No Monday.com item ID provided. Workbook uploaded to Google Sheets only.")

            except GoogleSheetsUploadError as exc:
                message = str(exc)
                st.error(f"❌ Google Sheets upload failed: {message}")
            except Exception as exc:  # pragma: no cover - runtime failures
                st.error(f"❌ Unexpected error: {exc}")

if __name__ == "__main__":
    main()
