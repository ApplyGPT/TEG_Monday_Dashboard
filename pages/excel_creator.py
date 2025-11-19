"""
Excel Creator
Builds the Development Package section of the workbook using a template.
"""

from __future__ import annotations

import os
import re
from io import BytesIO

import streamlit as st

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
except Exception:  # pragma: no cover - fallback if dependency missing at runtime
    load_workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None
    NamedStyle = None


# Pricing constants
BASE_PRICE_LESS_THAN_5 = 2780.00
BASE_PRICE_5_OR_MORE = 2325.00
ACTIVEWEAR_PRICE = 3560.00

OPTIONAL_PRICES = {
    "wash_dye": 1330.00,
    "design": 1330.00,
    "source": 1330.00,
    "treatment": 760.00,
}
TEMPLATE_FILENAME = "Copy of TEG 2025 WORKBOOK TEMPLATES.xlsx"
TARGET_SHEET = "DEVELOPMENT ONLY"
ROW_INDICES = [10, 12, 14, 16, 18]  # Rows reserved for style entries


@st.cache_data(show_spinner=False)
def get_template_path() -> str:
    """Return the absolute path to the Excel template."""
    base_dir = os.path.dirname(os.path.dirname(__file__))
    template_path = os.path.join(base_dir, "inputs", TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Template '{TEMPLATE_FILENAME}' was not found in the inputs folder."
        )
    return template_path


def calculate_base_price(num_styles: int, is_activewear: bool) -> float:
    """Calculate base price based on number of styles and activewear flag."""
    if is_activewear:
        return ACTIVEWEAR_PRICE
    elif num_styles < 5:
        return BASE_PRICE_LESS_THAN_5
    else:
        return BASE_PRICE_5_OR_MORE


def copy_cell_formatting(source_cell, target_cell) -> None:
    """Copy formatting (fill, border, alignment) from source to target cell."""
    if PatternFill is None or Border is None or Alignment is None:
        return
    
    try:
        # Copy fill (background color)
        if source_cell.fill and source_cell.fill.start_color:
            target_cell.fill = PatternFill(
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color,
                fill_type=source_cell.fill.fill_type
            )
        
        # Copy border
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
    except Exception:
        pass  # Skip if copying fails


def apply_arial_20_font(cell) -> None:
    """Apply Arial font size 20 to a cell."""
    if Font is not None:
        try:
            cell.font = Font(name="Arial", size=20)
        except Exception:
            pass


def safe_set_cell_value(ws, cell_ref: str, value) -> None:
    """Safely set cell value, handling merged cells by writing to top-left cell."""
    try:
        # Parse cell reference to get row and column
        from openpyxl.utils import coordinate_from_string, column_index_from_string
        col_letter = ''.join([c for c in cell_ref if c.isalpha()])
        row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
        col_num = column_index_from_string(col_letter)
        
        # Check if cell is part of a merged range
        cell_in_merged = False
        target_range = None
        
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_num <= merged_range.max_row and
                merged_range.min_col <= col_num <= merged_range.max_col):
                cell_in_merged = True
                target_range = merged_range
                break
        
        if cell_in_merged and target_range:
            # Unmerge the range first, then write to the original cell
            try:
                min_cell = ws.cell(row=target_range.min_row, column=target_range.min_col)
                max_cell = ws.cell(row=target_range.max_row, column=target_range.max_col)
                range_str = f"{min_cell.coordinate}:{max_cell.coordinate}"
                ws.unmerge_cells(range_str)
            except Exception:
                pass  # If unmerge fails, try to write anyway
        
        # Write to the cell using cell() method which always returns a writable cell
        target_cell = ws.cell(row=row_num, column=col_num)
        target_cell.value = value
        
    except Exception:
        # Fallback: try using the cell reference directly
        try:
            cell = ws[cell_ref]
            # If it's a MergedCell, we need to unmerge first
            if hasattr(cell, 'value'):
                try:
                    cell.value = value
                except (AttributeError, TypeError):
                    # It's a MergedCell, find and unmerge
                    cell_row = cell.row
                    cell_col = cell.column
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= cell_row <= merged_range.max_row and
                            merged_range.min_col <= cell_col <= merged_range.max_col):
                            try:
                                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                                ws.cell(row=cell_row, column=cell_col).value = value
                                break
                            except Exception:
                                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                                break
        except Exception:
            pass  # Skip if all methods fail


def update_header_labels(ws, client_name: str) -> None:
    """Ensure headers and client info match the spec."""
    header_map = {
        "H9": "WASH/DYE",
        "I9": "DESIGN",
        "J9": "SOURCE",
        "K9": "TREATMENT",
        "L9": "TOTAL",
    }
    for cell, label in header_map.items():
        safe_set_cell_value(ws, cell, label)

    safe_set_cell_value(ws, "B3", "TEGMADE, JUST FOR")
    client_display = (client_name or "").strip().upper()
    safe_set_cell_value(ws, "J3", client_display)
    if Font is not None:
        ws["J3"].font = Font(
            color="00C9A57A",
            name="Schibsted Grotesk",
            size=48,
            bold=True,
        )
    if Alignment is not None:
        ws["J3"].alignment = Alignment(horizontal="left", vertical="center")

    if Alignment is not None:
        center_cells = ["L20", "P10", "P11", "P12", "P13", "P20"]
        for ref in center_cells:
            ws[ref].alignment = Alignment(horizontal="center", vertical="center")


def clear_style_rows(ws, num_styles: int = 0) -> None:
    """Blank out style rows (B–L) and the totals row, preserving format for <= 5 styles."""
    if num_styles <= 5:
        # For 5 or fewer styles, clear ALL template style rows (10-18) so past values don't remain
        # Use safe_set_cell_value to preserve formatting/merges
        style_rows = [10, 12, 14, 16, 18]
        for row_idx in style_rows:
            for col_letter in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L']:
                safe_set_cell_value(ws, f"{col_letter}{row_idx}", None)
        
        # Clear totals row (row 20)
        safe_set_cell_value(ws, "B20", None)
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "H20", None)
        safe_set_cell_value(ws, "L20", None)
    else:
        # For more than 5 styles, clear all style rows and totals row
        max_style_rows = num_styles
        for i in range(max_style_rows):
            row_idx = 10 + (i * 2)
            # Only clear if this is a style row (not the totals row)
            totals_row = 20 + (num_styles - 5) * 2
            if row_idx == totals_row:
                continue
            for col_idx in range(2, 13):  # Columns B through L (1-based)
                cell = ws.cell(row=row_idx, column=col_idx)
                # Check if cell is part of a merged range
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left.value = None
                        is_merged = True
                        break
                if not is_merged:
                    cell.value = None
        
        # Clear the totals row
        totals_row = 20 + (num_styles - 5) * 2
        safe_set_cell_value(ws, f"B{totals_row}", None)
        safe_set_cell_value(ws, f"F{totals_row}", None)
        safe_set_cell_value(ws, f"H{totals_row}", None)
        safe_set_cell_value(ws, f"L{totals_row}", None)


def apply_development_package(
    ws,
    *,
    client_name: str,
    client_email: str,
    representative: str,
    style_entries: list[dict],
) -> tuple[float, float]:
    """Write the inputs into the workbook and return totals."""
    # Header metadata
    safe_set_cell_value(ws, "D6", client_email.strip())
    safe_set_cell_value(ws, "J6", (representative or "").strip().upper())
    safe_set_cell_value(ws, "B8", "DEVELOPMENT PACKAGE")

    optional_cells = {
        "H": "wash_dye",
        "I": "design",
        "J": "source",
        "K": "treatment",
    }

    total_development = 0.0
    total_optional = 0.0
    num_styles = len(style_entries)

    # If more than 5 styles, insert rows after row 19 to shift rows 20-32 down
    # For every style beyond 5, we need 2 rows (every other row pattern)
    if num_styles > 5:
        rows_to_insert = (num_styles - 5) * 2
        ws.insert_rows(20, amount=rows_to_insert)

        # Copy formatting from template row 18 to new rows (preserve colors, borders, alignment)
        # New rows start at 20, 22, 24, etc.
        template_row = 18  # Use row 18 as template for formatting
        for i in range(rows_to_insert):
            new_row = 20 + i
            for col_idx in range(2, 13):  # Columns B through L
                source_cell = ws.cell(row=template_row, column=col_idx)
                target_cell = ws.cell(row=new_row, column=col_idx)
                copy_cell_formatting(source_cell, target_cell)

        # Unmerge cells only in newly inserted rows (row 20+) to avoid MergedCell errors
        # Don't touch existing template rows (10-18) to preserve formatting
        first_new_row = 20
        last_new_row = 20 + rows_to_insert - 1
        merged_ranges_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            # Only unmerge if the merged range is entirely within the newly inserted rows
            # and in columns B-L
            if (merged_range.min_row >= first_new_row and merged_range.max_row <= last_new_row) and \
               merged_range.min_col >= 2 and merged_range.max_col <= 12:  # Columns B-L
                merged_ranges_to_unmerge.append(merged_range)
        
        # Unmerge the identified ranges
        for merged_range in merged_ranges_to_unmerge:
            try:
                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
            except Exception:
                pass  # Skip if unmerge fails

    # Clear style rows and totals row (after insertion so row numbers are correct)
    clear_style_rows(ws, num_styles=num_styles)

    # Generate row indices dynamically based on actual number of styles
    # Each style uses 2 rows (merged): 10-11, 12-13, 14-15, 16-17, 18-19, 20-21, etc.
    # We write to the first row of each pair (10, 12, 14, 16, 18, 20, 22, etc.)
    dynamic_row_indices = []
    start_row = 10
    for i in range(num_styles):
        dynamic_row_indices.append(start_row + (i * 2))

    for idx, row_idx in enumerate(dynamic_row_indices):
        entry = style_entries[idx]
        style_name = entry.get("name", "").strip() or "STYLE"
        complexity_pct = float(entry.get("complexity", 0.0))
        is_activewear = entry.get("activewear", False)

        # Calculate base price based on tiered pricing and activewear
        base_price = calculate_base_price(num_styles, is_activewear)

        # Check if this is a new row (row_idx > 18) that needs Arial 20 font
        is_new_row = num_styles > 5 and row_idx > 18
        
        # Each style row spans 2 rows (merged cells)
        row_second = row_idx + 1

        # Merge cells for style row (if new row, merge after writing)
        if is_new_row:
            # Merge columns that need merging (based on template pattern)
            merge_ranges = [
                (3, 3, row_idx, row_second),  # Column C (STYLE)
                (5, 5, row_idx, row_second),  # Column E (COMPLEXITY) - if needed
            ]
            # Only merge complexity if it has a value
            if complexity_pct == 0:
                merge_ranges = [(3, 3, row_idx, row_second)]  # Only merge STYLE column

        # Write style number (#) - merge if new row, left-aligned
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.value = idx + 1
        if is_new_row:
            apply_arial_20_font(cell_b)
            try:
                ws.merge_cells(f"B{row_idx}:B{row_second}")
                cell_b.alignment = Alignment(horizontal="left", vertical="center")
            except Exception:
                pass
        
        # Write style name (merged across 2 rows, left-aligned)
        cell_c = ws.cell(row=row_idx, column=3)
        cell_c.value = style_name.upper()
        if is_new_row:
            apply_arial_20_font(cell_c)
            try:
                ws.merge_cells(f"C{row_idx}:C{row_second}")
                cell_c.alignment = Alignment(horizontal="left", vertical="center")
            except Exception:
                pass
        # For existing rows (10-18), template already has cells merged, don't touch
        
        # Write base price (currency format, integer)
        cell_d = ws.cell(row=row_idx, column=4)
        cell_d.value = int(base_price)
        cell_d.number_format = '$#,##0'  # Currency format
        if is_new_row:
            apply_arial_20_font(cell_d)
            # Merge and center column D
            try:
                ws.merge_cells(f"D{row_idx}:D{row_second}")
                cell_d.alignment = Alignment(horizontal="center", vertical="center")
            except Exception:
                pass
        
        # Set complexity - leave blank if 0, otherwise set the percentage
        cell_e = ws.cell(row=row_idx, column=5)
        if complexity_pct == 0:
            cell_e.value = None
            # When complexity is 0, total = base price
            cell_f = ws.cell(row=row_idx, column=6)
            cell_f.value = f"=D{row_idx}"
            cell_f.number_format = '$#,##0'  # Currency format
            if is_new_row:
                apply_arial_20_font(cell_f)
                # Merge and center column F
                try:
                    ws.merge_cells(f"F{row_idx}:F{row_second}")
                    cell_f.alignment = Alignment(horizontal="center", vertical="center")
                except Exception:
                    pass
        else:
            cell_e.value = complexity_pct / 100.0
            cell_e.number_format = '0%'  # Percentage format
            if is_new_row:
                apply_arial_20_font(cell_e)
                try:
                    ws.merge_cells(f"E{row_idx}:E{row_second}")
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")
                except Exception:
                    pass
            # For existing rows (10-18), template already has cells merged, don't touch
            cell_f = ws.cell(row=row_idx, column=6)
            cell_f.value = f"=D{row_idx}*(1+E{row_idx})"
            cell_f.number_format = '$#,##0'  # Currency format
            if is_new_row:
                apply_arial_20_font(cell_f)
                # Merge and center column F
                try:
                    ws.merge_cells(f"F{row_idx}:F{row_second}")
                    cell_f.alignment = Alignment(horizontal="center", vertical="center")
                except Exception:
                    pass

        # Optional add-ons per row (columns H, I, J, K)
        row_options = entry.get("options", {}) or {}
        row_optional_sum = 0.0
        for col_letter, key in optional_cells.items():
            col_num = ord(col_letter) - 64  # Convert letter to column number
            cell_opt = ws.cell(row=row_idx, column=col_num)
            if row_options.get(key):
                price = int(OPTIONAL_PRICES[key])  # Ensure integer
                cell_opt.value = price
                cell_opt.number_format = '$#,##0'  # Currency format
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                    # Merge and center columns H, I, J, K
                    try:
                        ws.merge_cells(f"{col_letter}{row_idx}:{col_letter}{row_second}")
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        pass
                row_optional_sum += price
            else:
                cell_opt.value = None
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                    # Merge and center even if empty
                    try:
                        ws.merge_cells(f"{col_letter}{row_idx}:{col_letter}{row_second}")
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        pass
        
        cell_l = ws.cell(row=row_idx, column=12)
        cell_l.value = f"=SUM(H{row_idx}:K{row_idx})"
        cell_l.number_format = '$#,##0'  # Currency format
        if is_new_row:
            apply_arial_20_font(cell_l)
            try:
                ws.merge_cells(f"L{row_idx}:L{row_second}")
                cell_l.alignment = Alignment(horizontal="center", vertical="center")
            except Exception:
                pass

        total_development += base_price * (1 + complexity_pct / 100.0)
        total_optional += row_optional_sum

    # Totals section - dynamically calculate totals row and range based on number of styles
    # For 5 or fewer styles: totals at row 20 (original position)
    # For more than 5 styles: totals row shifts down by (num_styles - 5) * 2 rows
    if num_styles > 0:
        first_style_row = dynamic_row_indices[0]  # Should be 10
        last_style_row = dynamic_row_indices[num_styles - 1]
        if num_styles <= 5:
            totals_row = 20  # Original totals row position
        else:
            # Totals row shifts down by the number of rows we inserted
            totals_row = 20 + (num_styles - 5) * 2
        
        # Unmerge any merged cells in the totals row to avoid issues
        merged_ranges_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row):
                merged_ranges_to_unmerge.append(merged_range)
        
        for merged_range in merged_ranges_to_unmerge:
            try:
                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
            except Exception:
                pass
        
        # Set totals row labels
        cell_b_totals = ws.cell(row=totals_row, column=2)
        cell_b_totals.value = "TOTAL DEVELOPMENT"
        if Font is not None:
            cell_b_totals.font = Font(bold=True)
        
        cell_h_totals = ws.cell(row=totals_row, column=8)
        cell_h_totals.value = "TOTAL OPTIONAL ADD-ONS"
        if Font is not None:
            cell_h_totals.font = Font(bold=True)
        
        # Set totals formulas - sum all style rows (dynamic based on actual style rows)
        cell_f_totals = ws.cell(row=totals_row, column=6)
        cell_f_totals.value = f"=SUM(F{first_style_row}:F{last_style_row})"
        cell_f_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_f_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_f_totals.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #709171 to TOTAL DEVELOPMENT
        if PatternFill is not None:
            cell_f_totals.fill = PatternFill(start_color="709171", end_color="709171", fill_type="solid")
        
        cell_l_totals = ws.cell(row=totals_row, column=12)
        cell_l_totals.value = f"=SUM(L{first_style_row}:L{last_style_row})"
        cell_l_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_l_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_l_totals.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #f0cfbb to TOTAL OPTIONAL ADD-ONS
        if PatternFill is not None:
            cell_l_totals.fill = PatternFill(start_color="F0CFBB", end_color="F0CFBB", fill_type="solid")
        
        # Make all cells in totals row bold
        if Font is not None:
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
                else:
                    cell.font = Font(bold=True)
        
        # Apply Arial 20 font to totals row if it's a new row (num_styles > 5)
        if num_styles > 5:
            for col_idx in [2, 6, 8, 12, 14, 16]:  # Columns B, F, H, L, N, P
                cell = ws.cell(row=totals_row, column=col_idx)
                apply_arial_20_font(cell)
                # Ensure bold is maintained
                if Font is not None:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
        else:
            # For <=5 styles (totals_row==20), ensure Arial size 20 bold for key cells
            for col_idx in [2, 6, 8, 12, 14, 16]:
                cell = ws.cell(row=totals_row, column=col_idx)
                if Font is not None:
                    cell.font = Font(name="Arial", size=20, bold=True, color=cell.font.color if cell.font else None)
        # Ensure P (column 16) is also Arial size 20 for <=5 styles
        if num_styles <= 5:
            cell_p_totals = ws.cell(row=totals_row, column=16)
            if Font is not None:
                cell_p_totals.font = Font(name="Arial", size=20, bold=True, color=cell_p_totals.font.color if cell_p_totals.font else None)
        # Apply inferior (bottom) border to entire totals row
        if Border is not None and Side is not None:
            bottom_side = Side(style="thin")
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                # Skip columns A, G, M for bottom borders
                if col_idx in [1, 7, 13]:
                    new_bottom = cell.border.bottom
                else:
                    new_bottom = bottom_side
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=new_bottom,
                )

        # Apply full borders (top/bottom/left/right) to key totals cells (B, F, H, L, N, P)
        if Border is not None and Side is not None:
            full_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for col_idx in [2, 6, 8, 12, 14, 16]:
                cell = ws.cell(row=totals_row, column=col_idx)
                cell.border = full_border
        
        # Update P10 and P12 to use dynamic references (they reference F20 and L20 statically)
        # P10 should reference F{totals_row}, P12 should reference L{totals_row}
        cell_p10 = ws.cell(row=10, column=16)  # Column P, row 10
        if cell_p10.value and isinstance(cell_p10.value, str) and cell_p10.value.startswith("="):
            if "F20" in cell_p10.value:
                cell_p10.value = f"=F{totals_row}"
        
        cell_p12 = ws.cell(row=12, column=16)  # Column P, row 12
        if cell_p12.value and isinstance(cell_p12.value, str) and cell_p12.value.startswith("="):
            if "L20" in cell_p12.value:
                cell_p12.value = f"=L{totals_row}"
        
        # Find and update "TOTAL DUE AT SIGNING" formula
        # The label is in column N (14) and the formula is in column P (16) of the totals row
        cell_n_totals = ws.cell(row=totals_row, column=14)  # Column N
        if cell_n_totals.value and "TOTAL DUE AT SIGNING" in str(cell_n_totals.value).upper():
            # Update the formula in column P to reference the dynamic totals row
            cell_p_totals = ws.cell(row=totals_row, column=16)  # Column P
            cell_p_totals.value = f"=F{totals_row}+L{totals_row}"
            cell_p_totals.number_format = '$#,##0'  # Currency format
            # Apply font size 20 and bold to TOTAL DUE AT SIGNING formula
            if Font is not None:
                cell_p_totals.font = Font(name="Arial", size=20, bold=True, color=cell_p_totals.font.color if cell_p_totals.font else None)
            if Alignment is not None:
                cell_p_totals.alignment = Alignment(horizontal="center", vertical="center")
            # Apply cell color #ffff00 to TOTAL DUE AT SIGNING
            if PatternFill is not None:
                cell_p_totals.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Also make the label bold with size 20
            if Font is not None:
                cell_n_totals.font = Font(name="Arial", size=20, bold=True, color=cell_n_totals.font.color if cell_n_totals.font else None)
        
        # Also update any other formulas in column P that reference F20 or L20 statically
        # Check a few rows around the totals row
        for check_row in range(totals_row - 2, totals_row + 3):
            cell_p = ws.cell(row=check_row, column=16)  # Column P
            if cell_p.value and isinstance(cell_p.value, str) and cell_p.value.startswith("="):
                # Check if it references F20 or L20 (static references)
                if "F20" in cell_p.value or "L20" in cell_p.value:
                    # Replace with dynamic references
                    formula = cell_p.value.replace("F20", f"F{totals_row}").replace("L20", f"L{totals_row}")
                    cell_p.value = formula
                    if Font is not None:
                        existing_font = cell_p.font
                        cell_p.font = Font(
                            name=existing_font.name if existing_font and existing_font.name else "Arial",
                            size=existing_font.size if existing_font and existing_font.size else 20,
                            bold=True,
                            color=existing_font.color if existing_font else None
                        )
                    if Alignment is not None:
                        cell_p.alignment = Alignment(horizontal="center", vertical="center")
    else:
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "L20", None)
    return total_development, total_optional


def build_workbook_bytes(
    *,
    client_name: str,
    client_email: str,
    representative: str,
    style_entries: list[dict],
) -> tuple[bytes, float, float]:
    """Load the template, update it, and return bytes plus totals."""
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. Please add it to the environment first."
        )

    template_path = get_template_path()
    wb = load_workbook(template_path)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Worksheet '{TARGET_SHEET}' is missing from the template."
        )
    ws = wb[TARGET_SHEET]

    # Rename worksheet to "Workbook"
    ws.title = "Workbook"

    # Remove all other worksheets, keeping only "Workbook"
    sheets_to_remove = [name for name in wb.sheetnames if name != "Workbook"]
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])

    # Get reference to the Workbook worksheet after cleanup
    ws = wb["Workbook"]

    update_header_labels(ws, client_name)
    total_dev, total_optional = apply_development_package(
        ws,
        client_name=client_name,
        client_email=client_email,
        representative=representative,
        style_entries=style_entries,
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), total_dev, total_optional


def main() -> None:
    st.set_page_config(
        page_title="Excel Creator",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.title("📊 Excel Creator")

    st.markdown(
        """
<style>

    /* Enlarge checkbox */
    [data-testid="stCheckbox"] label span {
        height: 1.8rem;
        width: 1.8rem;
        transform: translateX(35px);
    }

    /* Hide normal label text (since you use empty label) */
    [data-testid="stCheckbox"] label p {
        font-size: 0;
        margin: 0;
    }

    /* Move right the entire checkbox inside its Streamlit column */
    [data-testid="stCheckbox"] {
        display: flex !important;
        justify-content: right !important;  /* Center horizontally */
        align-items: right !important;       /* Center vertically */
        height: 100%;                          /* Align with text inputs */
        padding: 0 !important;
        margin: 0 !important;
    }

    /* Also move right the internal label wrapper */
    [data-testid="stCheckbox"] > label {
        display: flex !important;
        justify-content: right !important;
        align-items: right !important;
        width: 100%;
        height: 100%;
        padding: 0 !important;
        margin: 0 !important;
    }

</style>
""",
        unsafe_allow_html=True,
    )

    st.caption(
        "Fill in the Development Package inputs and download a formatted workbook "
        "based on the official template."
    )

    client_name = st.text_input("Client Name", placeholder="Enter client name")

    col_a, col_b = st.columns(2)
    with col_a:
        client_email = st.text_input(
            "Client Email", placeholder="client@email.com"
        )
    with col_b:
        representative = st.text_input(
            "Representative", placeholder="Enter representative"
        )

    # Initialize session state for style entries
    if "style_entries" not in st.session_state:
        st.session_state["style_entries"] = []

    st.subheader("**Styles**")
    
    # Column headers
    header_cols = st.columns([2, 0.8, 1.2, 1.2, 1, 1, 1.2, 1])
    with header_cols[0]:
        st.markdown("**Style Name**")
    with header_cols[1]:
        st.markdown("**Activewear?**")
    with header_cols[2]:
        st.markdown("**Complexity (%)**")
    with header_cols[3]:
        st.markdown("**Wash/Dye ($1,330)**")
    with header_cols[4]:
        st.markdown("**Design ($1,330)**")
    with header_cols[5]:
        st.markdown("**Source ($1,330)**")
    with header_cols[6]:
        st.markdown("**Treatment ($760)**")
    with header_cols[7]:
        st.markdown("**Remove**")

    # Display existing style entries in horizontal rows
    if st.session_state["style_entries"]:
        for i, entry in enumerate(st.session_state["style_entries"]):
            with st.container():
                cols = st.columns([2, 0.8, 1.2, 1.2, 1, 1, 1.2, 1])
                with cols[0]:
                    style_name = st.text_input(
                        "Style Name",
                        value=entry.get("name", ""),
                        key=f"style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Dress, Winter Coat",
                    )
                    entry["name"] = style_name
                with cols[1]:
                    activewear = st.checkbox(
                        "",
                        value=entry.get("activewear", False),
                        key=f"activewear_{i}",
                        label_visibility="visible",
                    )
                    entry["activewear"] = activewear
                with cols[2]:
                    complexity = st.number_input(
                        "Complexity (%)",
                        min_value=0,
                        max_value=200,
                        value=int(entry.get("complexity", 100)),
                        step=5,
                        format="%d",
                        key=f"complexity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["complexity"] = float(complexity)
                with cols[3]:
                    wash_dye = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("wash_dye", False),
                        key=f"wash_dye_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["wash_dye"] = wash_dye
                with cols[4]:
                    design = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("design", False),
                        key=f"design_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["design"] = design
                with cols[5]:
                    source = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("source", False),
                        key=f"source_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["source"] = source
                with cols[6]:
                    treatment = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("treatment", False),
                        key=f"treatment_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["treatment"] = treatment
                with cols[7]:
                    if st.button("❌", key=f"remove_{i}", help="Remove this style"):
                        st.session_state["style_entries"].pop(i)
                        st.rerun()

    # Add new style interface
    st.markdown("---")
    st.markdown("**Add New Style**")
    add_cols = st.columns([2, 0.8, 1.2, 1.2, 1, 1, 1.2, 1])
    # Ensure default values exist (placeholders only) without pre-filling
    default_new_style = st.session_state.get("new_style_name", "")
    default_new_activewear = st.session_state.get("new_activewear", False)
    default_new_complexity = st.session_state.get("new_complexity", 0)
    default_new_wash = st.session_state.get("new_wash_dye", False)
    default_new_design = st.session_state.get("new_design", False)
    default_new_source = st.session_state.get("new_source", False)
    default_new_treatment = st.session_state.get("new_treatment", False)

    with add_cols[0]:
        new_style_name = st.text_input(
            "Style Name",
            value=default_new_style,
            key="new_style_name",
            label_visibility="collapsed",
            placeholder="e.g., Dress, Winter Coat",
        )
    with add_cols[1]:
        new_activewear = st.checkbox(
            "",
            value=default_new_activewear,
            key="new_activewear",
            label_visibility="visible",
        )
    with add_cols[2]:
        new_complexity = st.number_input(
            "Complexity (%)",
            min_value=0,
            max_value=200,
            value=default_new_complexity,
            step=5,
            format="%d",
            key="new_complexity",
            label_visibility="collapsed",
        )
    with add_cols[3]:
        new_wash_dye = st.checkbox(
            "",
            value=default_new_wash,
            key="new_wash_dye",
            label_visibility="visible",
        )
    with add_cols[4]:
        new_design = st.checkbox(
            "",
            value=default_new_design,
            key="new_design",
            label_visibility="visible",
        )
    with add_cols[5]:
        new_source = st.checkbox(
            "",
            value=default_new_source,
            key="new_source",
            label_visibility="visible",
        )
    with add_cols[6]:
        new_treatment = st.checkbox(
            "",
            value=default_new_treatment,
            key="new_treatment",
            label_visibility="visible",
        )
    with add_cols[7]:
        if st.button("➕ Add", key="add_style", help="Add this style"):
            if new_style_name.strip():
                st.session_state["style_entries"].append({
                    "name": new_style_name.strip(),
                    "activewear": new_activewear,
                    "complexity": float(new_complexity),
                    "options": {
                        "wash_dye": new_wash_dye,
                        "design": new_design,
                        "source": new_source,
                        "treatment": new_treatment,
                    },
                })
                # Reset add-new-style inputs so the next style starts blank/default
                for key in [
                    "new_style_name",
                    "new_activewear",
                    "new_complexity",
                    "new_wash_dye",
                    "new_design",
                    "new_source",
                    "new_treatment",
                ]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
            else:
                st.warning("Please enter a style name before adding.")

    if not st.session_state["style_entries"]:
        st.info("Add at least one style to enable the generator.")
        return

    try:
        excel_bytes, _, _ = build_workbook_bytes(
            client_name=client_name,
            client_email=client_email,
            representative=representative,
            style_entries=st.session_state["style_entries"],
        )
    except FileNotFoundError as exc:
        st.error(str(exc))
        return
    except Exception as exc:  # pragma: no cover - streamlit runtime feedback
        st.error(f"Failed to build workbook: {exc}")
        return

    safe_client = re.sub(r"[^A-Za-z0-9_-]+", "_", (client_name or "").strip()) or "client"
    download_name = f"workbook_{safe_client.lower()}.xlsx"

    st.success("Workbook is ready.")
    st.download_button(
        label="Generate Workbook",
        data=excel_bytes,
        file_name=download_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


if __name__ == "__main__":
    main()
