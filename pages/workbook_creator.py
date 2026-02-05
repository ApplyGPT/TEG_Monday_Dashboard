"""
Workbook Creator
Builds the Development Package section of the workbook using a template.
"""

from __future__ import annotations

import math
import os
import re
import urllib.parse
from copy import copy
from io import BytesIO

import streamlit as st
import requests

from google_sheets_uploader import (
    GoogleSheetsUploadError,
    upload_workbook_to_google_sheet,
)

# Google API imports for PDF export
try:
    from google.oauth2.service_account import Credentials as SACredentials
    from googleapiclient.discovery import build
    from google.auth.transport.requests import Request as GoogleRequest
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
    from openpyxl.utils import column_index_from_string, get_column_letter
except Exception:  # pragma: no cover - fallback if dependency missing at runtime
    load_workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None
    NamedStyle = None
    column_index_from_string = None
    get_column_letter = None


# Pricing constants - Package rates by # OF STYLES (style rows only)
# 2-4 STYLES, 5-9 STYLES, 10-14 STYLES, +15 STYLES. Activewear includes *2 FITTINGS.
PRICING_TABLE = {
    "Regular": {
        "< 2": 3360.00,   # 1 style → same as 2-4
        "< 5": 3360.00,   # 2-4 STYLES
        "< 10": 2856.00,  # 5-9 STYLES
        "< 15": 2688.00,  # 10-14 STYLES
        "15+": 2520.00,   # +15 STYLES
    },
    "Activewear/Lingerie/Swim": {
        "< 2": 3800.00,
        "< 5": 3800.00,
        "< 10": 3230.00,
        "< 15": 3040.00,
        "15+": 2850.00,
    },
    "Pattern Blocks": {
        "< 2": 2500.00,
        "< 5": 2500.00,
        "< 10": 2125.00,
        "< 15": 2000.00,
        "15+": 1875.00,
    }
}

OPTIONAL_PRICES = {
    "wash_dye": 1500.00,
    "design": 1500.00,
    "treatment": 860.00,
    # Note: "source" removed - no longer an option
}
SUMMARY_LABEL_COL = 14  # Column N
SUMMARY_VALUE_COL = 16  # Column P
SUMMARY_DEV_ROW = 10
SUMMARY_OPT_ROW = 12
SUMMARY_SUBTOTAL_ROW = 14
SUMMARY_DISCOUNT_ROW = 16
SUMMARY_SUM_END_ROW = 13  # Row before discount row
SUMMARY_TOTAL_DUE_BASE_ROW = 20
DELIVERABLE_BLOCK_START = 22
DELIVERABLE_BLOCK_END = 34
DELIVERABLE_BLOCK_HEIGHT = DELIVERABLE_BLOCK_END - DELIVERABLE_BLOCK_START + 1
DELIVERABLE_COL_START = 2  # Column B
DELIVERABLE_COL_END = 16   # Column P
TEMPLATE_FILENAME = "Copy of TEG 2025 WORKBOOK TEMPLATES.xlsx"
TARGET_SHEET = "DEVELOPMENT ONLY"
ROW_INDICES = [10, 12, 14, 16, 18]  # Rows reserved for style entries


@st.cache_data(show_spinner=False)
def get_template_path() -> str:
    """Return the absolute path to the Excel template."""
    base_dir = os.path.dirname(os.path.dirname(__file__))
    template_path = os.path.join(base_dir, "inputs", TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Template '{TEMPLATE_FILENAME}' was not found in the inputs folder."
        )
    return template_path


def calculate_base_price(num_styles: int, style_type: str) -> float:
    """Calculate base price based on number of styles and style type.
    
    Pricing is based on style count brackets and style type:
    - 1 style / 2-4 STYLES: Regular=$3,360, Activewear=$3,800, Pattern Blocks=$2,500
    - 5-9 STYLES: Regular=$2,856, Activewear=$3,230, Pattern Blocks=$2,125
    - 10-14 STYLES: Regular=$2,688, Activewear=$3,040, Pattern Blocks=$2,000
    - +15 STYLES: Regular=$2,520, Activewear=$2,850, Pattern Blocks=$1,875
    
    Args:
        num_styles: Number of style rows (used to determine bracket; custom items are not counted)
        style_type: One of "Regular", "Activewear/Lingerie/Swim", or "Pattern Blocks"
    
    Returns:
        Base price per style for the given bracket and style type
    """
    # Default to "Regular" if style_type is not recognized
    if style_type not in PRICING_TABLE:
        style_type = "Regular"
    
    # Bracket: 1 style, 2-4, 5-9, 10-14, +15
    if num_styles < 2:
        bracket = "< 2"
    elif num_styles < 5:
        bracket = "< 5"
    elif num_styles < 10:
        bracket = "< 10"
    elif num_styles < 15:
        bracket = "< 15"
    else:
        bracket = "15+"
    
    return PRICING_TABLE[style_type][bracket]


def copy_cell_formatting(source_cell, target_cell) -> None:
    """Copy formatting (fill, border, alignment) from source to target cell."""
    if PatternFill is None or Border is None or Alignment is None:
        return
    
    try:
        # Copy fill (background color)
        if source_cell.fill and source_cell.fill.start_color:
            target_cell.fill = PatternFill(
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color,
                fill_type=source_cell.fill.fill_type
            )
        
        # Copy border
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
    except Exception:
        pass  # Skip if copying fails


def apply_arial_20_font(cell) -> None:
    """Apply Arial font size 20 to a cell."""
    if Font is not None:
        try:
            cell.font = Font(name="Arial", size=20)
        except Exception:
            pass


def apply_full_border(cell) -> None:
    """Apply full thin borders around a cell."""
    if Border is None or Side is None:
        return
    try:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    except Exception:
        pass


def apply_full_border_pair(ws, column, start_row: int, end_row: int) -> None:
    """Apply borders to both cells in a vertical pair (before merging)."""
    if column_index_from_string is None:
        return
    col_idx = column if isinstance(column, int) else column_index_from_string(column)
    top_cell = ws.cell(row=start_row, column=col_idx)
    bottom_cell = ws.cell(row=end_row, column=col_idx)
    apply_full_border(top_cell)
    apply_full_border(bottom_cell)


def capture_deliverables_block(ws):
    """Capture the deliverables block (values + formatting) from the template."""
    block_rows = []
    for row in range(DELIVERABLE_BLOCK_START, DELIVERABLE_BLOCK_END + 1):
        row_data = []
        for col in range(DELIVERABLE_COL_START, DELIVERABLE_COL_END + 1):
            # Handle MergedCell - use safe_get_cell_value
            cell_value = safe_get_cell_value(ws, row, col)
            
            # Get cell object for formatting (handle MergedCell)
            cell = None
            try:
                cell = ws.cell(row=row, column=col)
                # Test if we can access attributes (if it's a MergedCell, some attributes might fail)
                _ = cell.number_format
            except Exception:
                # If it's a MergedCell, find the top-left cell of the merged range
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        break
                if cell is None:
                    cell = ws.cell(row=row, column=col)
            
            # Safely get cell attributes (handle MergedCell)
            number_format = None
            font = None
            fill = None
            border = None
            alignment = None
            
            if cell is not None:
                try:
                    number_format = cell.number_format if hasattr(cell, 'number_format') else None
                except Exception:
                    pass
                try:
                    font = copy(cell.font) if cell.font else None
                except Exception:
                    pass
                try:
                    fill = copy(cell.fill) if cell.fill else None
                except Exception:
                    pass
                try:
                    border = copy(cell.border) if cell.border else None
                except Exception:
                    pass
                try:
                    alignment = copy(cell.alignment) if cell.alignment else None
                except Exception:
                    pass
            
            row_data.append(
                {
                    "value": cell_value,
                    "number_format": number_format,
                    "font": font,
                    "fill": fill,
                    "border": border,
                    "alignment": alignment,
                }
            )
        block_rows.append(row_data)
    
    merged_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row >= DELIVERABLE_BLOCK_START
            and merged_range.max_row <= DELIVERABLE_BLOCK_END
            and merged_range.min_col >= DELIVERABLE_COL_START
            and merged_range.max_col <= DELIVERABLE_COL_END
        ):
            merged_ranges.append(
                (
                    merged_range.min_row - DELIVERABLE_BLOCK_START,
                    merged_range.max_row - DELIVERABLE_BLOCK_START,
                    merged_range.min_col,
                    merged_range.max_col,
                )
            )
    
    return {"rows": block_rows, "merges": merged_ranges}


def restore_deliverables_block(ws, template_block: dict, target_start_row: int) -> None:
    """Restore the deliverables block at the specified start row."""
    if not template_block:
        return
    
    rows_data = template_block.get("rows", [])
    block_height = len(rows_data)
    target_end_row = target_start_row + block_height - 1
    
    # Clear existing merges in target area
    to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.max_row < target_start_row
            or merged_range.min_row > target_end_row
            or merged_range.max_col < DELIVERABLE_COL_START
            or merged_range.min_col > DELIVERABLE_COL_END
        ):
            continue
        to_unmerge.append(merged_range)
    
    for merged_range in to_unmerge:
        try:
            ws.unmerge_cells(range_string=str(merged_range))
        except Exception:
            pass
    
    # Write cell data and formatting
    for row_offset, row_cells in enumerate(rows_data):
        target_row = target_start_row + row_offset
        for col_offset, cell_data in enumerate(row_cells):
            target_col = DELIVERABLE_COL_START + col_offset
            coord = None
            if get_column_letter is not None:
                coord = f"{get_column_letter(target_col)}{target_row}"
            value = cell_data.get("value") if cell_data else None
            # Since we already unmerged above, we can set values directly like workbook_creator2.py
            if coord:
                safe_set_cell_value(ws, coord, value)
            else:
                # Direct assignment - cells are already unmerged above
                try:
                    cell = ws.cell(row=target_row, column=target_col)
                    cell.value = value
                except (AttributeError, TypeError):
                    # If it's still a MergedCell (shouldn't happen after unmerge, but handle it)
                    cell = safe_get_writable_cell(ws, target_row, target_col)
                    try:
                        cell.value = value
                    except Exception:
                        # Last resort: use safe_set_cell_value
                        if get_column_letter:
                            from openpyxl.utils import get_column_letter as gcl
                            safe_set_cell_value(ws, f"{gcl(target_col)}{target_row}", value)
            
            # Get cell for formatting - cells should be unmerged by now
            cell = ws.cell(row=target_row, column=target_col)
            
            # Apply formatting (same as workbook_creator2.py)
            cell.number_format = cell_data.get("number_format") or cell.number_format
            if cell_data.get("font"):
                cell.font = copy(cell_data["font"])
            if cell_data.get("fill"):
                cell.fill = copy(cell_data["fill"])
            if cell_data.get("border"):
                cell.border = copy(cell_data["border"])
            if cell_data.get("alignment"):
                cell.alignment = copy(cell_data["alignment"])
    
    # Recreate merged ranges (adjusted for new start row)
    for min_row_offset, max_row_offset, min_col, max_col in template_block.get("merges", []):
        start_row = target_start_row + min_row_offset
        end_row = target_start_row + max_row_offset
        try:
            ws.merge_cells(
                start_row=start_row,
                start_column=min_col,
                end_row=end_row,
                end_column=max_col,
            )
        except Exception:
            pass


def safe_get_writable_cell(ws, row: int, column: int):
    """Get a writable cell, handling MergedCell by returning top-left cell of merged range."""
    # First check if the cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= column <= merged_range.max_col):
            # It's part of a merged range, return top-left cell (always writable)
            # But verify it's actually writable
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            try:
                # Test if writable by trying to read value
                _ = top_left.value
                return top_left
            except (AttributeError, TypeError):
                # Even top-left might be MergedCell in edge cases, find the real top-left
                # This shouldn't happen, but handle it
                return top_left
    
    # Not part of a merged range, try to get the cell
    # But even if not in merged_cells.ranges, ws.cell() might still return MergedCell
    # So we need to test if it's writable
    cell = ws.cell(row=row, column=column)
    # Check if it's a MergedCell by trying to access value (MergedCell raises AttributeError)
    try:
        # Try to read value - if it works, it's a regular cell
        _ = cell.value
        return cell
    except (AttributeError, TypeError):
        # It's a MergedCell even though it's not in merged_cells.ranges
        # This can happen if the merge was just created or there's a timing issue
        # Find the top-left cell by checking all merged ranges again
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= column <= merged_range.max_col):
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        # If we can't find it, the cell might be in a newly merged range
        # Try to get it anyway - caller should handle exceptions
        return cell


def safe_get_cell_value(ws, row: int, column: int):
    """Safely get cell value, handling MergedCell by returning value from top-left cell."""
    # First check if the cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= column <= merged_range.max_col):
            # It's part of a merged range, get value from top-left cell
            try:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                try:
                    return top_left_cell.value
                except (AttributeError, TypeError):
                    # Even the top-left cell might be a MergedCell in some edge cases
                    # Recursively find the actual top-left
                    return safe_get_cell_value(ws, merged_range.min_row, merged_range.min_col)
            except Exception:
                return None
    
    # Not part of a merged range, try to get value directly
    try:
        cell = ws.cell(row=row, column=column)
        # Check if it's a MergedCell by trying to access value
        try:
            return cell.value
        except (AttributeError, TypeError):
            # It's a MergedCell, find the top-left cell
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= column <= merged_range.max_col):
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    try:
                        return top_left_cell.value
                    except (AttributeError, TypeError):
                        # Even the top-left cell might be a MergedCell in some edge cases
                        return safe_get_cell_value(ws, merged_range.min_row, merged_range.min_col)
            return None
    except Exception:
        return None


def safe_merge_cells(ws, range_str: str) -> bool:
    """Safely merge cells, checking if they're already merged first."""
    try:
        # Parse the range string (e.g., "B10:B11")
        parts = range_str.split(':')
        if len(parts) != 2:
            return False
        
        from openpyxl.utils import column_index_from_string
        
        # Parse start cell
        start_col_letter = ''.join([c for c in parts[0] if c.isalpha()])
        start_row = int(''.join([c for c in parts[0] if c.isdigit()]))
        start_col = column_index_from_string(start_col_letter)
        
        # Parse end cell
        end_col_letter = ''.join([c for c in parts[1] if c.isalpha()])
        end_row = int(''.join([c for c in parts[1] if c.isdigit()]))
        end_col = column_index_from_string(end_col_letter)
        
        # Check if any cell in this range is already part of a merged range
        # Do multiple passes to catch all overlapping merges
        max_passes = 3
        for pass_num in range(max_passes):
            ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if ranges overlap (any cell in our range overlaps with merged range)
                if not (merged_range.max_row < start_row or merged_range.min_row > end_row or
                        merged_range.max_col < start_col or merged_range.min_col > end_col):
                    # Ranges overlap, mark for unmerging
                    ranges_to_unmerge.append(merged_range)
            
            if not ranges_to_unmerge:
                break  # No more overlapping merges
            
            # Unmerge overlapping ranges (in reverse to avoid index issues)
            for merged_range in reversed(ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass
        
        # Now merge the cells vertically
        ws.merge_cells(range_str)
        return True
    except Exception:
        return False  # Return False if merge fails


def safe_set_cell_value(ws, cell_ref: str, value) -> None:
    """Safely set cell value, handling merged cells by writing to top-left cell."""
    try:
        # Parse cell reference to get row and column
        from openpyxl.utils import coordinate_from_string, column_index_from_string
        col_letter = ''.join([c for c in cell_ref if c.isalpha()])
        row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
        col_num = column_index_from_string(col_letter)
        
        # Check if cell is part of a merged range
        cell_in_merged = False
        target_range = None
        
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_num <= merged_range.max_row and
                merged_range.min_col <= col_num <= merged_range.max_col):
                cell_in_merged = True
                target_range = merged_range
                break
        
        if cell_in_merged and target_range:
            # Unmerge the range first, then write to the original cell
            try:
                min_cell = ws.cell(row=target_range.min_row, column=target_range.min_col)
                max_cell = ws.cell(row=target_range.max_row, column=target_range.max_col)
                range_str = f"{min_cell.coordinate}:{max_cell.coordinate}"
                ws.unmerge_cells(range_str)
            except Exception:
                pass  # If unmerge fails, try to write anyway
        
        # Write to the cell using cell() method which always returns a writable cell
        target_cell = ws.cell(row=row_num, column=col_num)
        # Check if it's a MergedCell before setting value
        try:
            target_cell.value = value
        except (AttributeError, TypeError):
            # It's a MergedCell, find top-left cell or unmerge
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col_num <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                        target_cell = ws.cell(row=row_num, column=col_num)
                        target_cell.value = value
                        break
                    except Exception:
                        target_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        target_cell.value = value
                        break
        
    except Exception:
        # Fallback: try using the cell reference directly
        try:
            cell = ws[cell_ref]
            # If it's a MergedCell, we need to unmerge first
            if hasattr(cell, 'value'):
                try:
                    cell.value = value
                except (AttributeError, TypeError):
                    # It's a MergedCell, find and unmerge
                    cell_row = cell.row
                    cell_col = cell.column
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= cell_row <= merged_range.max_row and
                            merged_range.min_col <= cell_col <= merged_range.max_col):
                            try:
                                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                                # After unmerging, try to set value with error handling
                                try:
                                    target_cell = safe_get_writable_cell(ws, cell_row, cell_col)
                                    target_cell.value = value
                                except (AttributeError, TypeError):
                                    # If still a MergedCell, use top-left cell
                                    target_cell = safe_get_writable_cell(ws, merged_range.min_row, merged_range.min_col)
                                    target_cell.value = value
                                break
                            except Exception:
                                # Fallback: use top-left cell directly
                                try:
                                    target_cell = safe_get_writable_cell(ws, merged_range.min_row, merged_range.min_col)
                                    target_cell.value = value
                                except (AttributeError, TypeError):
                                    # Last resort: use safe_set_cell_value with coordinate
                                    if get_column_letter:
                                        from openpyxl.utils import get_column_letter as gcl
                                        safe_set_cell_value(ws, f"{gcl(merged_range.min_col)}{merged_range.min_row}", value)
                                    break
                                break
        except Exception:
            pass  # Skip if all methods fail


def update_header_labels(ws, client_name: str) -> None:
    """Ensure headers and client info match the spec."""
    header_map = {
        "H9": "WASH/DYE",
        "I9": "DESIGN",
        "J9": "TREATMENT",
        "K9": "TOTAL",  # K9-L9 will be merged below
    }
    for cell, label in header_map.items():
        safe_set_cell_value(ws, cell, label)
    
    # Merge and center K9-L9 for TOTAL header
    safe_merge_cells(ws, "K9:L9")
    cell_k9 = ws.cell(row=9, column=11)
    if Alignment is not None:
        cell_k9.alignment = Alignment(horizontal="center", vertical="center")

    safe_set_cell_value(ws, "B3", "TEGMADE, JUST FOR")
    client_display = (client_name or "").strip().upper()
    safe_set_cell_value(ws, "J3", client_display)
    if Font is not None:
        ws["J3"].font = Font(
            color="00C9A57A",
            name="Schibsted Grotesk",
            size=48,
            bold=True,
        )
    if Alignment is not None:
        ws["J3"].alignment = Alignment(horizontal="left", vertical="center")

    if Alignment is not None:
        center_cells = ["L20", "P10", "P11", "P12", "P13", "P20"]
        for ref in center_cells:
            ws[ref].alignment = Alignment(horizontal="center", vertical="center")


def clear_style_rows(ws, num_styles: int = 0) -> None:
    """Blank out style rows (B–L) and the totals row, preserving format for <= 5 styles."""
    if num_styles <= 5:
        # For 5 or fewer styles, clear ALL template style rows (10-18) so past values don't remain
        # Use safe_set_cell_value to preserve formatting/merges
        style_rows = [10, 12, 14, 16, 18]
        for row_idx in style_rows:
            for col_letter in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L']:
                safe_set_cell_value(ws, f"{col_letter}{row_idx}", None)
        
        # Clear totals row (row 20)
        safe_set_cell_value(ws, "B20", None)
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "H20", None)
        safe_set_cell_value(ws, "L20", None)
        safe_set_cell_value(ws, f"N{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"N{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, "N19", None)
        safe_set_cell_value(ws, "P19", None)
    else:
        # For more than 5 styles, clear all style rows and totals row
        max_style_rows = num_styles
        totals_row_clear = 20 + (num_styles - 5) * 2
        for i in range(max_style_rows):
            row_idx = 10 + (i * 2)
            # Only clear if this is a style row (not the totals row)
            if row_idx == totals_row_clear:
                continue
            for col_idx in range(2, 13):  # Columns B through L (1-based)
                cell = ws.cell(row=row_idx, column=col_idx)
                # Check if cell is part of a merged range
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        try:
                            top_left.value = None
                        except (AttributeError, TypeError):
                            # If top_left is also a MergedCell, use safe_set_cell_value
                            if get_column_letter:
                                safe_set_cell_value(ws, f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}", None)
                            else:
                                from openpyxl.utils import get_column_letter as gcl
                                safe_set_cell_value(ws, f"{gcl(merged_range.min_col)}{merged_range.min_row}", None)
                        is_merged = True
                        break
                if not is_merged:
                    try:
                        cell.value = None
                    except (AttributeError, TypeError):
                        # If it's a MergedCell, use safe_set_cell_value
                        if get_column_letter:
                            safe_set_cell_value(ws, f"{get_column_letter(col_idx)}{row_idx}", None)
                        else:
                            from openpyxl.utils import get_column_letter as gcl
                            safe_set_cell_value(ws, f"{gcl(col_idx)}{row_idx}", None)
        
        # Clear the totals row and TOTAL DUE AT SIGNING row (dynamic position)
        safe_set_cell_value(ws, f"B{totals_row_clear}", None)
        safe_set_cell_value(ws, f"F{totals_row_clear}", None)
        safe_set_cell_value(ws, f"H{totals_row_clear}", None)
        safe_set_cell_value(ws, f"L{totals_row_clear}", None)
        safe_set_cell_value(ws, f"N{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_SUBTOTAL_ROW}", None)
        safe_set_cell_value(ws, f"N{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"P{SUMMARY_DISCOUNT_ROW}", None)
        safe_set_cell_value(ws, f"N{totals_row_clear - 1}", None)
        safe_set_cell_value(ws, f"P{totals_row_clear - 1}", None)


def apply_development_package(
    ws,
    *,
    client_name: str,
    client_email: str,
    representative: str,
    style_entries: list[dict],
    custom_styles: list[dict],
    discount_percentage: float,
) -> tuple[float, float]:
    """Write the inputs into the workbook and return totals."""
    # Header metadata
    safe_set_cell_value(ws, "D6", client_email.strip())
    safe_set_cell_value(ws, "J6", (representative or "").strip().upper())
    safe_set_cell_value(ws, "B8", "DEVELOPMENT PACKAGE")

    optional_cells = {
        "H": "wash_dye",
        "I": "design",
        "J": "treatment",
        # Note: "source" removed - no longer an option
    }

    total_development = 0.0
    total_optional = 0.0
    num_styles = len(style_entries)
    num_custom_styles = len(custom_styles)
    total_styles_count = num_styles + num_custom_styles  # Total for pricing tier calculation
    discount_percentage = max(0.0, float(discount_percentage or 0))
    deliverables_template = capture_deliverables_block(ws)

    base_capacity = len(ROW_INDICES)
    extra_styles = max(num_styles - base_capacity, 0)
    rows_to_insert_regular = extra_styles * 2
    
    # Calculate how many rows we'll need for Custom Items too (if any)
    # This ensures we insert all rows upfront before writing any styles
    rows_to_insert_custom = 0
    if num_custom_styles > 0:
        # Calculate where Custom Items would start (after all regular styles)
        if num_styles > 5:
            custom_start_row = 20 + (num_styles - 5) * 2
        else:
            custom_start_row = 10 + num_styles * 2
        # Calculate where totals row will be after regular insertions
        if num_styles > 5:
            totals_row_after_regular = 20 + (num_styles - 5) * 2
        else:
            totals_row_after_regular = 20
        # Calculate custom row indices
        custom_row_indices_precalc = []
        for i in range(num_custom_styles):
            custom_row_indices_precalc.append(custom_start_row + (i * 2))
        # Check if we need more rows for Custom Items
        if custom_row_indices_precalc:
            last_custom_row = custom_row_indices_precalc[-1]
            if last_custom_row >= totals_row_after_regular - 2:
                required_totals_row = last_custom_row + 2
                rows_to_insert_custom = required_totals_row - totals_row_after_regular
                if rows_to_insert_custom < 0:
                    rows_to_insert_custom = 0

    # Insert all rows upfront (regular + custom) before writing any styles
    total_rows_to_insert = rows_to_insert_regular + rows_to_insert_custom
    if total_rows_to_insert > 0:
        ws.insert_rows(SUMMARY_TOTAL_DUE_BASE_ROW, amount=total_rows_to_insert)

        # Copy formatting from template row 18 to new rows (preserve colors, borders, alignment)
        # New rows start at 20, 22, 24, etc.
        template_row = 18  # Use row 18 as template for formatting
        for i in range(total_rows_to_insert):
            new_row = 20 + i
            for col_idx in range(2, 13):  # Columns B through L
                source_cell = ws.cell(row=template_row, column=col_idx)
                target_cell = ws.cell(row=new_row, column=col_idx)
                copy_cell_formatting(source_cell, target_cell)

        # Unmerge ALL cells in newly inserted rows to avoid MergedCell errors
        # Excel automatically adjusts merged ranges when rows are inserted, which can cause conflicts
        first_new_row = SUMMARY_TOTAL_DUE_BASE_ROW
        last_new_row = SUMMARY_TOTAL_DUE_BASE_ROW + total_rows_to_insert - 1
        
        # Collect all merged ranges that intersect with the newly inserted rows
        # We need to unmerge both horizontal and vertical merges in the new rows
        # Do this multiple times to catch all merges (some might be created after unmerging others)
        max_iterations = 3
        for iteration in range(max_iterations):
            merged_ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if merged range intersects with newly inserted rows
                # (min_row <= last_new_row and max_row >= first_new_row)
                if (merged_range.min_row <= last_new_row and merged_range.max_row >= first_new_row):
                    merged_ranges_to_unmerge.append(merged_range)
            
            if not merged_ranges_to_unmerge:
                break  # No more merges to unmerge
            
            # Unmerge the identified ranges (do this in reverse to avoid index issues)
            for merged_range in reversed(merged_ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    range_str = f"{min_cell.coordinate}:{max_cell.coordinate}"
                    ws.unmerge_cells(range_str)
                except Exception:
                    pass  # Skip if unmerge fails

    # Clear style rows and totals row (after insertion so row numbers are correct)
    clear_style_rows(ws, num_styles=num_styles)

    # Generate row indices dynamically based on actual number of styles
    # Each style uses 2 rows (merged): 10-11, 12-13, 14-15, 16-17, 18-19, 20-21, etc.
    # We write to the first row of each pair (10, 12, 14, 16, 18, 20, 22, etc.)
    dynamic_row_indices = []
    start_row = 10
    for i in range(num_styles):
        dynamic_row_indices.append(start_row + (i * 2))

    for idx, row_idx in enumerate(dynamic_row_indices):
        entry = style_entries[idx]
        style_name = entry.get("name", "").strip() or "STYLE"
        complexity_pct = float(entry.get("complexity", 0.0))
        style_type = entry.get("style_type", "Regular")
        # Handle migration from old "activewear" boolean
        if "style_type" not in entry and entry.get("activewear", False):
            style_type = "Activewear/Lingerie/Swim"
        row_options = entry.get("options", {})

        # Calculate base price based on tiered pricing and style type
        # Package rate bracket is based on number of styles only (not custom items)
        line_base_price = calculate_base_price(num_styles, style_type)

        # Check if this is a new row (row_idx > 18) that needs Arial 20 font
        is_new_row = num_styles > 5 and row_idx > 18
        
        # Each style row spans 2 rows (merged cells)
        row_second = row_idx + 1

        # For new rows, explicitly unmerge all cells in this row pair before writing
        # This ensures we start with clean, unmerged cells (prevents horizontal merges)
        if is_new_row:
            ranges_to_unmerge = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if merged range intersects with this row pair
                if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                    merged_range.min_col <= 12 and merged_range.max_col >= 2):  # Columns B-L
                    ranges_to_unmerge.append(merged_range)
            
            # Unmerge all overlapping ranges
            for merged_range in reversed(ranges_to_unmerge):
                try:
                    min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                    ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                except Exception:
                    pass

        # Write style number (#) - use style_number from entry (101, 102, 103, etc.)
        style_number = entry.get("style_number", 101 + idx)  # Default to 101, 102, 103... if not set
        cell_b = safe_get_writable_cell(ws, row_idx, 2)
        try:
            cell_b.value = style_number
        except (AttributeError, TypeError):
            # If still MergedCell, use safe_set_cell_value
            safe_set_cell_value(ws, f"B{row_idx}", style_number)
            cell_b = safe_get_writable_cell(ws, row_idx, 2)
        if is_new_row:
            apply_full_border_pair(ws, 2, row_idx, row_second)
            safe_merge_cells(ws, f"B{row_idx}:B{row_second}")
            apply_arial_20_font(cell_b)
            if Alignment is not None:
                cell_b.alignment = Alignment(horizontal="left", vertical="center")
        
        # Write style name (merged across 2 rows, left-aligned)
        cell_c = safe_get_writable_cell(ws, row_idx, 3)
        try:
            cell_c.value = style_name.upper()
        except (AttributeError, TypeError):
            safe_set_cell_value(ws, f"C{row_idx}", style_name.upper())
            cell_c = safe_get_writable_cell(ws, row_idx, 3)
        if is_new_row:
            apply_full_border_pair(ws, 3, row_idx, row_second)
            safe_merge_cells(ws, f"C{row_idx}:C{row_second}")
            apply_arial_20_font(cell_c)
            if Alignment is not None:
                cell_c.alignment = Alignment(horizontal="left", vertical="center")
        # For existing rows (10-18), template already has cells merged, don't touch
        
        # Set complexity - show percentage when provided
        cell_e = ws.cell(row=row_idx, column=5)
        if complexity_pct == 0:
            cell_e.value = None
            # Even when complexity is 0, merge and center for new rows
            if is_new_row:
                apply_full_border_pair(ws, 5, row_idx, row_second)
                safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
                apply_arial_20_font(cell_e)
                if Alignment is not None:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")
        else:
            try:
                cell_e.value = complexity_pct / 100.0
                cell_e.number_format = '0%'  # Percentage format
            except (AttributeError, TypeError):
                safe_set_cell_value(ws, f"E{row_idx}", complexity_pct / 100.0)
                cell_e = safe_get_writable_cell(ws, row_idx, 5)
                try:
                    cell_e.number_format = '0%'
                except Exception:
                    pass
            if is_new_row:
                apply_full_border_pair(ws, 5, row_idx, row_second)
                safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
                apply_arial_20_font(cell_e)
                if Alignment is not None:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")
                elif Alignment is not None and cell_e.alignment:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")

        # Write base price (currency format, integer)
        cell_d = ws.cell(row=row_idx, column=4)
        cell_d.value = int(line_base_price)
        cell_d.number_format = '$#,##0'  # Currency format
        if is_new_row:
            apply_full_border_pair(ws, 4, row_idx, row_second)
            safe_merge_cells(ws, f"D{row_idx}:D{row_second}")
            apply_arial_20_font(cell_d)
            if Alignment is not None:
                cell_d.alignment = Alignment(horizontal="center", vertical="center")
        elif Alignment is not None and cell_d.alignment:
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write total formula (column F)
        cell_f = safe_get_writable_cell(ws, row_idx, 6)
        try:
            if complexity_pct == 0:
                cell_f.value = f"=D{row_idx}"
            else:
                cell_f.value = f"=D{row_idx}*(1+E{row_idx})"
            cell_f.number_format = '$#,##0'  # Currency format
        except (AttributeError, TypeError):
            if complexity_pct == 0:
                safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}")
            else:
                safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}*(1+E{row_idx})")
            cell_f = safe_get_writable_cell(ws, row_idx, 6)
            try:
                cell_f.number_format = '$#,##0'
            except Exception:
                pass
        if is_new_row:
            apply_full_border_pair(ws, 6, row_idx, row_second)
            safe_merge_cells(ws, f"F{row_idx}:F{row_second}")
            apply_arial_20_font(cell_f)
        if Alignment is not None:
            cell_f.alignment = Alignment(horizontal="center", vertical="center")

        # Optional add-ons per row (columns H, I, J, K)
        row_optional_sum = 0.0
        for col_letter, key in optional_cells.items():
            col_num = ord(col_letter) - 64  # Convert letter to column number
            cell_opt = ws.cell(row=row_idx, column=col_num)
            if row_options.get(key):
                price = int(OPTIONAL_PRICES[key])  # Ensure integer
                cell_opt.value = price
                cell_opt.number_format = '$#,##0'  # Currency format
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                    # Merge and center columns H, I, J, K
                    apply_full_border_pair(ws, col_letter, row_idx, row_second)
                    safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                    if Alignment is not None:
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
                row_optional_sum += price
            else:
                try:
                    cell_opt.value = None
                except (AttributeError, TypeError):
                    safe_set_cell_value(ws, f"{col_letter}{row_idx}", None)
                    cell_opt = safe_get_writable_cell(ws, row_idx, col_num)
                if is_new_row:
                    apply_arial_20_font(cell_opt)
                    # Merge and center even if empty
                    apply_full_border_pair(ws, col_letter, row_idx, row_second)
                    safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                    if Alignment is not None:
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
        
        # TOTAL OPTIONAL ADD-ONS now uses columns K and L (merged K10-L11, K12-L13, etc.)
        # Follow the pattern from workbook_creator2.py: set value FIRST, then merge
        cell_k = ws.cell(row=row_idx, column=11)
        cell_k.value = f"=SUM(H{row_idx}:J{row_idx})"
        cell_k.number_format = '$#,##0'  # Currency format
        
        # Always merge and center K-L for each style row (K10-L11, K12-L13, etc.)
        # Unmerge first if needed
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row_second <= merged_range.max_row and
                merged_range.min_row >= row_idx and
                merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        
        # Re-set value after unmerge (unmerge might have cleared it)
        cell_k = ws.cell(row=row_idx, column=11)
        cell_k.value = f"=SUM(H{row_idx}:J{row_idx})"
        cell_k.number_format = '$#,##0'
        
        # Apply borders to individual cells BEFORE merging
        apply_full_border_pair(ws, 11, row_idx, row_second)
        apply_full_border_pair(ws, 12, row_idx, row_second)
        
        # Now merge (value should be preserved in top-left cell K)
        safe_merge_cells(ws, f"K{row_idx}:L{row_second}")
        
        # Get the merged cell and apply formatting
        cell_k = ws.cell(row=row_idx, column=11)
        
        # For more than 5 styles, ensure full borders on merged cell
        # This applies to ALL style rows when total_styles_count > 5
        if total_styles_count > 5 and Border is not None and Side is not None:
            thin = Side(style="thin")
            full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            try:
                cell_k.border = full_border
            except Exception:
                pass
        
        if is_new_row:
            apply_arial_20_font(cell_k)
        if Alignment is not None:
            cell_k.alignment = Alignment(horizontal="center", vertical="center")

        total_development += line_base_price * (1 + complexity_pct / 100.0)
        total_optional += row_optional_sum

    # Process Custom Items (user-defined price, complexity, no add-ons)
    if num_custom_styles > 0:
        # Calculate starting row for Custom Items (after all regular styles)
        custom_start_row = dynamic_row_indices[-1] + 2 if dynamic_row_indices else 10
        custom_row_indices = []
        for i in range(num_custom_styles):
            custom_row_indices.append(custom_start_row + (i * 2))
        
        # Custom row indices are already calculated correctly since we inserted all rows upfront
        # No need to insert more rows here - they were already inserted above
        
        for idx, row_idx in enumerate(custom_row_indices):
            entry = custom_styles[idx]
            style_name = entry.get("name", "").strip() or "STYLE"
            custom_price = float(entry.get("price", 0.0))
            complexity_pct = float(entry.get("complexity", 0.0))
            
            is_new_row = row_idx > 18
            row_second = row_idx + 1
            
            # For new rows, explicitly unmerge all cells in this row pair before writing
            # This ensures we start with clean, unmerged cells (prevents horizontal merges)
            if is_new_row:
                ranges_to_unmerge = []
                for merged_range in list(ws.merged_cells.ranges):
                    # Check if merged range intersects with this row pair
                    if (merged_range.min_row <= row_second and merged_range.max_row >= row_idx and
                        merged_range.min_col <= 12 and merged_range.max_col >= 2):  # Columns B-L
                        ranges_to_unmerge.append(merged_range)
                
                # Unmerge all overlapping ranges
                for merged_range in reversed(ranges_to_unmerge):
                    try:
                        min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                        ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
                    except Exception:
                        pass
            
            # Write style number (#) - use the style_number from entry if available, otherwise default
            # Use style_number from entry if set, otherwise default to 101 + num_styles + idx
            style_number = entry.get("style_number", 101 + num_styles + idx)
            cell_b = safe_get_writable_cell(ws, row_idx, 2)
            try:
                cell_b.value = style_number
            except (AttributeError, TypeError):
                safe_set_cell_value(ws, f"B{row_idx}", style_number)
                cell_b = safe_get_writable_cell(ws, row_idx, 2)
            if is_new_row:
                apply_full_border_pair(ws, 2, row_idx, row_second)
                safe_merge_cells(ws, f"B{row_idx}:B{row_second}")
                apply_arial_20_font(cell_b)
                if Alignment is not None:
                    cell_b.alignment = Alignment(horizontal="left", vertical="center")
            
            # Write style name
            cell_c = safe_get_writable_cell(ws, row_idx, 3)
            try:
                cell_c.value = style_name.upper()
            except (AttributeError, TypeError):
                safe_set_cell_value(ws, f"C{row_idx}", style_name.upper())
                cell_c = safe_get_writable_cell(ws, row_idx, 3)
            if is_new_row:
                apply_full_border_pair(ws, 3, row_idx, row_second)
                safe_merge_cells(ws, f"C{row_idx}:C{row_second}")
                apply_arial_20_font(cell_c)
                if Alignment is not None:
                    cell_c.alignment = Alignment(horizontal="left", vertical="center")
            
            # Write custom price
            cell_d = safe_get_writable_cell(ws, row_idx, 4)
            try:
                cell_d.value = int(custom_price)
                cell_d.number_format = '$#,##0'
            except (AttributeError, TypeError):
                safe_set_cell_value(ws, f"D{row_idx}", int(custom_price))
                cell_d = safe_get_writable_cell(ws, row_idx, 4)
                try:
                    cell_d.number_format = '$#,##0'
                except Exception:
                    pass
            if is_new_row:
                apply_full_border_pair(ws, 4, row_idx, row_second)
                safe_merge_cells(ws, f"D{row_idx}:D{row_second}")
                apply_arial_20_font(cell_d)
                if Alignment is not None:
                    cell_d.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write complexity
            cell_e = ws.cell(row=row_idx, column=5)
            if complexity_pct == 0:
                cell_e.value = None
            else:
                try:
                    cell_e.value = complexity_pct / 100.0
                    cell_e.number_format = '0%'
                except (AttributeError, TypeError):
                    safe_set_cell_value(ws, f"E{row_idx}", complexity_pct / 100.0)
                    cell_e = safe_get_writable_cell(ws, row_idx, 5)
                    try:
                        cell_e.number_format = '0%'
                    except Exception:
                        pass
            if is_new_row:
                apply_full_border_pair(ws, 5, row_idx, row_second)
                safe_merge_cells(ws, f"E{row_idx}:E{row_second}")
                apply_arial_20_font(cell_e)
                if Alignment is not None:
                    cell_e.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write total formula
            cell_f = safe_get_writable_cell(ws, row_idx, 6)
            try:
                if complexity_pct == 0:
                    cell_f.value = f"=D{row_idx}"
                else:
                    cell_f.value = f"=D{row_idx}*(1+E{row_idx})"
                cell_f.number_format = '$#,##0'
            except (AttributeError, TypeError):
                if complexity_pct == 0:
                    safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}")
                else:
                    safe_set_cell_value(ws, f"F{row_idx}", f"=D{row_idx}*(1+E{row_idx})")
                cell_f = safe_get_writable_cell(ws, row_idx, 6)
                try:
                    cell_f.number_format = '$#,##0'
                except Exception:
                    pass
            if is_new_row:
                apply_full_border_pair(ws, 6, row_idx, row_second)
                safe_merge_cells(ws, f"F{row_idx}:F{row_second}")
                apply_arial_20_font(cell_f)
                if Alignment is not None:
                    cell_f.alignment = Alignment(horizontal="center", vertical="center")
            
            # Clear add-ons (Custom Items don't have add-ons) - only H, I, J now (no K)
            for col_letter in ["H", "I", "J"]:
                col_num = ord(col_letter) - 64
                cell_opt = ws.cell(row=row_idx, column=col_num)
                cell_opt.value = None
                if is_new_row:
                    apply_full_border_pair(ws, col_letter, row_idx, row_second)
                    safe_merge_cells(ws, f"{col_letter}{row_idx}:{col_letter}{row_second}")
                    if Alignment is not None:
                        cell_opt.alignment = Alignment(horizontal="center", vertical="center")
            
            # Clear total optional add-on - now uses K-L merged
            cell_k = safe_get_writable_cell(ws, row_idx, 11)
            cell_l = safe_get_writable_cell(ws, row_idx, 12)
            try:
                cell_k.value = None
                cell_l.value = None
            except (AttributeError, TypeError):
                safe_set_cell_value(ws, f"K{row_idx}", None)
                safe_set_cell_value(ws, f"L{row_idx}", None)

            if is_new_row or total_styles_count > 5:
                apply_full_border_pair(ws, 11, row_idx, row_second)
                apply_full_border_pair(ws, 12, row_idx, row_second)
                safe_merge_cells(ws, f"K{row_idx}:L{row_second}")
                # Get merged cell and ensure borders are applied (for more than 5 styles)
                cell_k = safe_get_writable_cell(ws, row_idx, 11)
                if total_styles_count > 5 and Border is not None and Side is not None:
                    thin = Side(style="thin")
                    try:
                        cell_k.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    except Exception:
                        pass
                if Alignment is not None:
                    try:
                        cell_k.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        pass
            
            total_development += custom_price * (1 + complexity_pct / 100.0)
        
        # Update last_style_row to include Custom Items (for totals calculations)
        if custom_row_indices:
            last_style_row = custom_row_indices[-1]
    
    # Determine last_style_row if no Custom Items (for totals calculations)
    if num_custom_styles == 0:
        if dynamic_row_indices:
            last_style_row = dynamic_row_indices[-1]
        else:
            last_style_row = 10

    # Determine last_regular_style_row (only regular + activewear, excluding Custom Items)
    # This is used for deliverables counts (Patterns, First Samples, Final Samples)
    if dynamic_row_indices:
        last_regular_style_row = dynamic_row_indices[-1]
    else:
        last_regular_style_row = 10

    # Count styles by type
    # Handle migration from old "activewear" boolean
    num_activewear = sum(1 for entry in style_entries if (
        entry.get("style_type") == "Activewear/Lingerie/Swim" or 
        (entry.get("style_type") is None and entry.get("activewear", False))
    ))
    num_regular = sum(1 for entry in style_entries if (
        entry.get("style_type") == "Regular" or 
        (entry.get("style_type") is None and not entry.get("activewear", False))
    ))
    num_pattern_blocks = sum(1 for entry in style_entries if entry.get("style_type") == "Pattern Blocks")
    
    total_extra_rows = max(total_styles_count - len(ROW_INDICES), 0) * 2
    deliverables_block_start = DELIVERABLE_BLOCK_START + total_extra_rows
    deliverables_block_end = deliverables_block_start + DELIVERABLE_BLOCK_HEIGHT - 1
    restore_deliverables_block(ws, deliverables_template, deliverables_block_start)
    
    # If there are activewear styles, modify deliverables section
    if num_activewear > 0 and column_index_from_string is not None:
        label_column_idx = column_index_from_string("B")
        col_c_idx = column_index_from_string("C")
        col_d_idx = column_index_from_string("D")
        final_samples_row = None
        row_final_samples_new = None  # Store the row where new FINAL SAMPLES is created
        
        # Find the "FINAL SAMPLES" row (should be around row 31-32)
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = safe_get_cell_value(ws, scan_row, label_column_idx)
            if isinstance(value, str):
                value_lower = value.strip().lower()
                if "final samples" in value_lower and final_samples_row is None:
                    final_samples_row = scan_row
                    break
        
        if final_samples_row:
            # Step 1: Replace FINAL SAMPLES (rows 31-32) with SECOND SAMPLES (only columns B-D)
            # Unmerge cells in columns B-D for these rows first
            for row in [final_samples_row, final_samples_row + 1]:
                for col in [label_column_idx, col_c_idx, col_d_idx]:
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass

            # Clear values in row 32 (second row of pair) for columns B-D
            for col in [label_column_idx, col_c_idx, col_d_idx]:
                # Always use safe_set_cell_value to handle MergedCell
                if get_column_letter:
                    safe_set_cell_value(ws, f"{get_column_letter(col)}{final_samples_row + 1}", None)
                else:
                    # Fallback: use column letter directly
                    from openpyxl.utils import get_column_letter as gcl
                    safe_set_cell_value(ws, f"{gcl(col)}{final_samples_row + 1}", None)

            # Clear column C in row 31 (to ensure no leftover values)
            # Always use safe_set_cell_value to handle MergedCell
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{final_samples_row}", None)
            else:
                safe_set_cell_value(ws, f"C{final_samples_row}", None)

            # Set SECOND SAMPLES in row 31, column B, count in column D
            # Always use safe_set_cell_value to handle MergedCell
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{final_samples_row}", "SECOND SAMPLES")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{final_samples_row}", num_activewear)
            else:
                safe_set_cell_value(ws, f"B{final_samples_row}", "SECOND SAMPLES")
                safe_set_cell_value(ws, f"D{final_samples_row}", num_activewear)
            # Set number format - handle MergedCell
            try:
                count_cell = ws.cell(row=final_samples_row, column=col_d_idx)
                count_cell.number_format = "0"
            except AttributeError:
                # If it's a MergedCell, find top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        count_cell.number_format = "0"
                        break
            
            # Use row 29 (ROUND OF REVISIONS) as reference - it should be around row 29
            # Find ROUND OF REVISIONS row (should be row 29)
            reference_row = None
            for scan_row in range(deliverables_block_start, final_samples_row):
                value = safe_get_cell_value(ws, scan_row, label_column_idx)
                if isinstance(value, str) and "round of revisions" in value.lower():
                    reference_row = scan_row
                    break
            
            # If not found, assume it's row 29 (typical position)
            if reference_row is None:
                reference_row = 29
            
            # Check what merges exist for columns B-D in rows 29-30 (reference_row to reference_row+1)
            b_merged = False
            c_merged = False
            d_merged = False
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row == reference_row and merged_range.max_row == reference_row + 1):
                    if merged_range.min_col == 2 and merged_range.max_col == 2:  # Column B
                        b_merged = True
                    elif merged_range.min_col == 3 and merged_range.max_col == 3:  # Column C
                        c_merged = True
                    elif merged_range.min_col == 4 and merged_range.max_col == 4:  # Column D
                        d_merged = True
            
            # Copy exact formatting from B29-D30 to B31-D32 (SECOND SAMPLES)
            # First, copy cell formatting (but not alignment for column B - we'll center it after merge)
            for col in [label_column_idx, col_c_idx, col_d_idx]:
                source_cell_29 = ws.cell(row=reference_row, column=col)
                source_cell_30 = ws.cell(row=reference_row + 1, column=col)
                target_cell_31 = ws.cell(row=final_samples_row, column=col)
                target_cell_32 = ws.cell(row=final_samples_row + 1, column=col)
                
                # Copy formatting from row 29 to row 31
                if source_cell_29.font:
                    target_cell_31.font = copy(source_cell_29.font)
                if source_cell_29.fill:
                    target_cell_31.fill = copy(source_cell_29.fill)
                if source_cell_29.border:
                    target_cell_31.border = copy(source_cell_29.border)
                # Don't copy alignment for column B - we'll set it to center after merge
                if col != label_column_idx and source_cell_29.alignment:
                    target_cell_31.alignment = copy(source_cell_29.alignment)
                target_cell_31.number_format = source_cell_29.number_format
                
                # Copy formatting from row 30 to row 32
                if source_cell_30.font:
                    target_cell_32.font = copy(source_cell_30.font)
                if source_cell_30.fill:
                    target_cell_32.fill = copy(source_cell_30.fill)
                if source_cell_30.border:
                    target_cell_32.border = copy(source_cell_30.border)
                # Don't copy alignment for column B - we'll set it to center after merge
                if col != label_column_idx and source_cell_30.alignment:
                    target_cell_32.alignment = copy(source_cell_30.alignment)
                target_cell_32.number_format = source_cell_30.number_format
            
            # Merge B:C together and D separately, with center alignment
            if safe_merge_cells:
                # Merge B:C together (columns B and C merged across 2 rows)
                safe_merge_cells(ws, f"B{final_samples_row}:C{final_samples_row + 1}")
                # Set center alignment for merged B:C
                if Alignment:
                    cell_bc = ws.cell(row=final_samples_row, column=label_column_idx)
                    cell_bc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                # Merge column D
                safe_merge_cells(ws, f"D{final_samples_row}:D{final_samples_row + 1}")
                # Set center alignment for merged column D
                if Alignment:
                    cell_d = ws.cell(row=final_samples_row, column=col_d_idx)
                    cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Step 2: Insert 6 rows after row 32 (3 new deliverable entries, each with 2 rows)
            insert_row = final_samples_row + 2  # After row 32
            rows_to_insert = 6
            
            # Insert rows
            ws.insert_rows(insert_row, amount=rows_to_insert)
            
            # Unmerge any cells in the newly inserted rows
            for i in range(rows_to_insert):
                target_row = insert_row + i
                for col in range(DELIVERABLE_COL_START, DELIVERABLE_COL_END + 1):
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= target_row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass
            
            # Step 3: Set values for the new rows and define row variables
            # Row 33-34: 2ND ROUND OF FITTINGS (count in column D)
            row_2nd_fittings = insert_row
            # Clear column C to ensure no leftover values
            # Always use safe_set_cell_value to handle MergedCell
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_2nd_fittings}", None)
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_2nd_fittings}", "2ND ROUND OF FITTINGS")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{row_2nd_fittings}", 1)
            else:
                safe_set_cell_value(ws, f"C{row_2nd_fittings}", None)
                safe_set_cell_value(ws, f"B{row_2nd_fittings}", "2ND ROUND OF FITTINGS")
                safe_set_cell_value(ws, f"D{row_2nd_fittings}", 1)
            # Set number format - handle MergedCell
            try:
                count_cell = ws.cell(row=row_2nd_fittings, column=col_d_idx)
                count_cell.number_format = "0"
            except AttributeError:
                # If it's a MergedCell, find top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row_2nd_fittings <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        count_cell.number_format = "0"
                        break
            row_2nd_fittings_2 = insert_row + 1
            
            # Row 35-36: 2ND ROUND OF REVISIONS (count in column D)
            row_2nd_revisions = insert_row + 2
            # Clear column C to ensure no leftover values
            # Always use safe_set_cell_value to handle MergedCell
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_2nd_revisions}", None)
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_2nd_revisions}", "2ND ROUND OF REVISIONS")
                safe_set_cell_value(ws, f"{get_column_letter(col_d_idx)}{row_2nd_revisions}", 1)
            else:
                safe_set_cell_value(ws, f"C{row_2nd_revisions}", None)
                safe_set_cell_value(ws, f"B{row_2nd_revisions}", "2ND ROUND OF REVISIONS")
                safe_set_cell_value(ws, f"D{row_2nd_revisions}", 1)
            # Set number format - handle MergedCell
            try:
                count_cell = ws.cell(row=row_2nd_revisions, column=col_d_idx)
                count_cell.number_format = "0"
            except AttributeError:
                # If it's a MergedCell, find top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row_2nd_revisions <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        count_cell.number_format = "0"
                        break
            row_2nd_revisions_2 = insert_row + 3
            
            # Row 37-38: FINAL SAMPLES (count of all styles)
            row_final_samples = insert_row + 4
            row_final_samples_new = row_final_samples  # Store for later use
            # Clear column C to ensure no leftover values - always use safe_set_cell_value to handle MergedCell
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(col_c_idx)}{row_final_samples}", None)
                safe_set_cell_value(ws, f"{get_column_letter(label_column_idx)}{row_final_samples}", "FINAL SAMPLES")
            else:
                safe_set_cell_value(ws, f"C{row_final_samples}", None)
                safe_set_cell_value(ws, f"B{row_final_samples}", "FINAL SAMPLES")
            row_final_samples_2 = insert_row + 5
            
            # Step 4: Copy exact formatting from B29-D30 to each new row pair (B33-D34, B35-D36, B37-D38)
            # Use row 29 as source for first row of each pair, row 30 for second row
            # Don't copy alignment for column B - we'll center it after merge
            for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                row_second = base_row + 1
                # Copy formatting from row 29 to first row of pair
                for col in [label_column_idx, col_c_idx, col_d_idx]:
                    source_cell_29 = ws.cell(row=reference_row, column=col)
                    source_cell_30 = ws.cell(row=reference_row + 1, column=col)
                    target_cell_first = ws.cell(row=base_row, column=col)
                    target_cell_second = ws.cell(row=row_second, column=col)
                    
                    # Copy formatting from row 29 to first row
                    if source_cell_29.font:
                        target_cell_first.font = copy(source_cell_29.font)
                    if source_cell_29.fill:
                        target_cell_first.fill = copy(source_cell_29.fill)
                    if source_cell_29.border:
                        target_cell_first.border = copy(source_cell_29.border)
                    # Don't copy alignment for column B - we'll set it to center after merge
                    if col != label_column_idx and source_cell_29.alignment:
                        target_cell_first.alignment = copy(source_cell_29.alignment)
                    target_cell_first.number_format = source_cell_29.number_format
                    
                    # Copy formatting from row 30 to second row
                    if source_cell_30.font:
                        target_cell_second.font = copy(source_cell_30.font)
                    if source_cell_30.fill:
                        target_cell_second.fill = copy(source_cell_30.fill)
                    if source_cell_30.border:
                        target_cell_second.border = copy(source_cell_30.border)
                    # Don't copy alignment for column B - we'll set it to center after merge
                    if col != label_column_idx and source_cell_30.alignment:
                        target_cell_second.alignment = copy(source_cell_30.alignment)
                    target_cell_second.number_format = source_cell_30.number_format
            
            # Step 5: Merge B:C together and D separately, with center alignment for each new row pair
            if safe_merge_cells:
                for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                    row_second = base_row + 1
                    # Merge B:C together (columns B and C merged across 2 rows)
                    safe_merge_cells(ws, f"B{base_row}:C{row_second}")
                    # Set center alignment for merged B:C
                    if Alignment:
                        cell_bc = ws.cell(row=base_row, column=label_column_idx)
                        cell_bc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    # Merge column D
                    safe_merge_cells(ws, f"D{base_row}:D{row_second}")
                    # Set center alignment for merged column D
                    if Alignment:
                        cell_d = ws.cell(row=base_row, column=col_d_idx)
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    
                    # Set FINAL SAMPLES count immediately after merging
                    if base_row == row_final_samples:
                        # Unmerge column D temporarily to set the value
                        try:
                            ws.unmerge_cells(f"D{base_row}:D{row_second}")
                        except Exception:
                            pass
                        # Set the count value (total styles: regular + activewear, excluding custom line items)
                        count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                        # Use safe_set_cell_value to handle MergedCell
                        safe_set_cell_value(ws, f"D{base_row}", None)
                        safe_set_cell_value(ws, f"D{base_row}", count_value)
                        # Set number format on the actual cell (handle MergedCell)
                        count_cell = safe_get_writable_cell(ws, base_row, col_d_idx)
                        try:
                            count_cell.number_format = "0"
                        except Exception:
                            pass
                        # Re-merge column D
                        safe_merge_cells(ws, f"D{base_row}:D{row_second}")
                        # Restore center alignment
                        if Alignment:
                            cell_d = safe_get_writable_cell(ws, base_row, col_d_idx)
                            try:
                                cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                            except Exception:
                                pass
            
            # Restore merged cells for right side columns (H-P) using template merge patterns
            if deliverables_template and "merges" in deliverables_template:
                right_side_merges = []
                for min_row_offset, max_row_offset, min_col, max_col in deliverables_template.get("merges", []):
                    if (min_col >= 8 and max_col <= 16 and 
                        max_row_offset - min_row_offset == 1):
                        right_side_merges.append((min_col, max_col))
                
                # Apply merge patterns to new rows and center align
                for base_row in [row_2nd_fittings, row_2nd_revisions, row_final_samples]:
                    row_second = base_row + 1
                    for min_col, max_col in right_side_merges:
                        try:
                            if get_column_letter:
                                start_col_letter = get_column_letter(min_col)
                                end_col_letter = get_column_letter(max_col)
                                range_str = f"{start_col_letter}{base_row}:{end_col_letter}{row_second}"
                                safe_merge_cells(ws, range_str)
                                # Center align the merged cell
                                if Alignment:
                                    cell = safe_get_writable_cell(ws, base_row, min_col)
                                    try:
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                                    except Exception:
                                        pass
                        except Exception:
                            pass
            
            # Center align all right side columns (H-P) in the entire deliverables section
            if Alignment:
                for row in range(deliverables_block_start, deliverables_block_end + 1):
                    for col in range(8, 17):  # Columns H (8) through P (16)
                        cell = ws.cell(row=row, column=col)
                        if cell.alignment:
                            cell.alignment = Alignment(
                                horizontal='center',
                                vertical=cell.alignment.vertical,
                                wrap_text=cell.alignment.wrap_text,
                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                indent=cell.alignment.indent
                            )
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Update deliverables_block_end to include new rows
            deliverables_block_end += rows_to_insert
        
        # Center align all right side columns (H-P) in the entire deliverables section (including original rows)
        if Alignment:
            for row in range(deliverables_block_start, deliverables_block_end + 1):
                for col in range(8, 17):  # Columns H (8) through P (16)
                    cell = ws.cell(row=row, column=col)
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    else:
        # Even if no activewear, center align right side columns (H-P) in deliverables section
        if Alignment:
            for row in range(deliverables_block_start, deliverables_block_end + 1):
                for col in range(8, 17):  # Columns H (8) through P (16)
                    cell = ws.cell(row=row, column=col)
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Update deliverables table counts (column J) for add-on selections
    if column_index_from_string is not None and total_styles_count > 0:
        label_column_idx = column_index_from_string("H")
        target_col_j = column_index_from_string("J")
        deliverable_addon_map = [
            ("WASH/TREATMENT", "H"),
            ("DESIGN", "I"),
            ("TREATMENT", "J"),  # TREATMENT moved to column J (SOURCING removed)
        ]

        def find_label_row(label_text: str) -> int | None:
            """Locate the row index for a given deliverable label."""
            lowered = label_text.strip().lower()
            partial_match_row = None
            for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
                value = safe_get_cell_value(ws, scan_row, label_column_idx)
                if not isinstance(value, str):
                    continue
                value_clean = value.strip().lower()
                if value_clean == lowered:
                    return scan_row
                if lowered in value_clean and partial_match_row is None:
                    partial_match_row = scan_row
            return partial_match_row

        for label_text, addon_col_letter in deliverable_addon_map:
            row_idx = find_label_row(label_text)
            if row_idx is None:
                continue
            addon_range = f"{addon_col_letter}10:{addon_col_letter}{last_style_row}"
            # Use safe_set_cell_value to handle MergedCell
            if get_column_letter:
                safe_set_cell_value(ws, f"{get_column_letter(target_col_j)}{row_idx}", f"=COUNT({addon_range})")
            else:
                from openpyxl.utils import get_column_letter as gcl
                safe_set_cell_value(ws, f"{gcl(target_col_j)}{row_idx}", f"=COUNT({addon_range})")
            # Set number format - handle MergedCell
            try:
                cell = ws.cell(row=row_idx, column=target_col_j)
                cell.number_format = "0"
            except AttributeError:
                # If it's a MergedCell, find top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row_idx <= merged_range.max_row and
                        merged_range.min_col <= target_col_j <= merged_range.max_col):
                        cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        cell.number_format = "0"
                        break

        # Round of Fittings: always 1
        fittings_row = find_label_row("ROUND OF FITTINGS")
        if fittings_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= fittings_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Use safe_set_cell_value to handle MergedCell
            safe_set_cell_value(ws, f"D{fittings_row}", None)
            safe_set_cell_value(ws, f"D{fittings_row}", 1)
            # Set number format on the actual cell (handle MergedCell)
            try:
                count_cell = ws.cell(row=fittings_row, column=col_d_idx)
                count_cell.number_format = "0"
            except AttributeError:
                # If it's a MergedCell, find top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= fittings_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        count_cell.number_format = "0"
                        break
        
        # Round of Revisions: 
        # - 1 if there's ONLY Regular styles (non-Activewear) OR ONLY Activewear styles
        # - 2 if there's BOTH Regular AND Activewear styles
        # Note: ROUND OF REVISIONS is in column B, not column H, so we need to search column B
        label_col_b = column_index_from_string("B")
        revisions_row = None
        revisions_label_lower = "round of revisions"
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = safe_get_cell_value(ws, scan_row, label_col_b)
            if isinstance(value, str) and revisions_label_lower in value.strip().lower():
                revisions_row = scan_row
                break
        if revisions_row:
            col_d_idx = column_index_from_string("D")
            # Check if this cell is part of a merged range and remember the merge pattern
            was_merged = False
            merge_pattern = None
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= revisions_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    was_merged = True
                    merge_pattern = (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
                    break
            # SIMPLEST APPROACH: Use num_regular and num_activewear already calculated from style_entries
            # These are calculated above from the actual style_entries data - 100% reliable
            # No need to read from Excel cells which might have text/number issues
            
            # Calculate revisions count: 2 if both regular AND activewear exist, 1 otherwise
            # num_regular and num_activewear are calculated at lines 884-885 from style_entries
            revisions_count = 2 if (num_regular > 0 and num_activewear > 0) else 1
            
            # Use safe_set_cell_value to handle MergedCell
            safe_set_cell_value(ws, f"D{revisions_row}", None)
            safe_set_cell_value(ws, f"D{revisions_row}", revisions_count)
            # Set number format on the actual cell (handle MergedCell)
            try:
                count_cell = ws.cell(row=revisions_row, column=col_d_idx)
                count_cell.number_format = "0"
            except AttributeError:
                # If it's a MergedCell, find top-left cell
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= revisions_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        count_cell.number_format = "0"
                        break
            if Alignment is not None:
                count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Re-merge AFTER setting the value
            if was_merged and merge_pattern and safe_merge_cells:
                min_row, max_row, min_col, max_col = merge_pattern
                if get_column_letter:
                    start_col_letter = get_column_letter(min_col)
                    end_col_letter = get_column_letter(max_col)
                    range_str = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"
                    safe_merge_cells(ws, range_str)
                    # Re-apply alignment after merging
                    if Alignment is not None:
                        # Get top-left cell of merged range for alignment
                        try:
                            merged_cell = ws.cell(row=revisions_row, column=col_d_idx)
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        except Exception:
                            # If it's a MergedCell, find top-left cell
                            for merged_range in ws.merged_cells.ranges:
                                if (merged_range.min_row <= revisions_row <= merged_range.max_row and
                                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                                    merged_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                                    break
        
        # If activewear exists, handle SECOND SAMPLES and the new rows
        if num_activewear > 0:
            # SECOND SAMPLES: count of activewear styles (replaced FINAL SAMPLES at rows 31-32)
            # Note: count is already set in column D above, but we ensure it's correct here
            second_sample_row = find_label_row("SECOND SAMPLES")
            if second_sample_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_sample_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Use safe_set_cell_value to handle MergedCell
                safe_set_cell_value(ws, f"D{second_sample_row}", None)
                safe_set_cell_value(ws, f"D{second_sample_row}", num_activewear)
                # Set number format on the actual cell (handle MergedCell)
                try:
                    count_cell = ws.cell(row=second_sample_row, column=col_d_idx)
                    count_cell.number_format = "0"
                except AttributeError:
                    # If it's a MergedCell, find top-left cell
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= second_sample_row <= merged_range.max_row and
                            merged_range.min_col <= col_d_idx <= merged_range.max_col):
                            count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            count_cell.number_format = "0"
                            break
            
            # 2nd Round of Fittings: always 1 (already set in column D above)
            second_fittings_row = find_label_row("2ND ROUND OF FITTINGS")
            if second_fittings_row:
                col_d_idx = column_index_from_string("D")
                # Unmerge if needed and clear any formulas
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= second_fittings_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                # Use safe_set_cell_value to handle MergedCell
                safe_set_cell_value(ws, f"D{second_fittings_row}", None)
                safe_set_cell_value(ws, f"D{second_fittings_row}", 1)
                # Set number format on the actual cell (handle MergedCell)
                try:
                    count_cell = ws.cell(row=second_fittings_row, column=col_d_idx)
                    count_cell.number_format = "0"
                except AttributeError:
                    # If it's a MergedCell, find top-left cell
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= second_fittings_row <= merged_range.max_row and
                            merged_range.min_col <= col_d_idx <= merged_range.max_col):
                            count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            count_cell.number_format = "0"
                            break
            
            # 2nd Round of Revisions: always 1 for Active category (already set in column D above)
            # Merge and alignment will be handled in the unified deliverables section below
            second_revisions_row = find_label_row("2ND ROUND OF REVISIONS")
            if second_revisions_row:
                col_d_idx = column_index_from_string("D")
                # Use safe_set_cell_value to handle MergedCell
                safe_set_cell_value(ws, f"D{second_revisions_row}", 1)
                # Set number format on the actual cell (handle MergedCell)
                try:
                    count_cell = ws.cell(row=second_revisions_row, column=col_d_idx)
                    count_cell.number_format = "0"
                except AttributeError:
                    # If it's a MergedCell, find top-left cell
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= second_revisions_row <= merged_range.max_row and
                            merged_range.min_col <= col_d_idx <= merged_range.max_col):
                            count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            count_cell.number_format = "0"
                            break
            
            # Final Samples: count is already set in the merge section above, but verify it's correct here
            # (This is a backup check - the count should already be set when merging FINAL SAMPLES)
            # Use row_final_samples_new if available (the newly created FINAL SAMPLES row), otherwise search for it
            # Always prefer row_final_samples_new since it's the newly created row
            final_samples_row_to_use = row_final_samples_new if row_final_samples_new else find_label_row("FINAL SAMPLES")
            if final_samples_row_to_use:
                final_samples_row = final_samples_row_to_use
                # Use direct count instead of formula
                count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                col_d_idx = column_index_from_string("D")
                col_b_idx = column_index_from_string("B")
                col_c_idx = column_index_from_string("C")
                # Use safe_set_cell_value to handle MergedCell
                safe_set_cell_value(ws, f"D{final_samples_row}", count_value)
                # Set number format - handle MergedCell
                try:
                    count_cell = ws.cell(row=final_samples_row, column=col_d_idx)
                    count_cell.number_format = "0"
                except AttributeError:
                    # If it's a MergedCell, find top-left cell
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                            merged_range.min_col <= col_d_idx <= merged_range.max_col):
                            count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            count_cell.number_format = "0"
                            break
                # Store row for later merge (we'll do merges at the very end)
                # Merges will be handled separately at the end of deliverables section
        else:
            # When there's no activewear, set FINAL SAMPLES here
            # Final Samples: all styles (regular only, no activewear, excluding custom line items)
            # Note: FINAL SAMPLES is in column B, not column H, so we need to search column B
            final_samples_row = None
            label_col_b = column_index_from_string("B")
            for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
                value = safe_get_cell_value(ws, scan_row, label_col_b)
                if isinstance(value, str) and "final" in value.lower() and "sample" in value.lower():
                    final_samples_row = scan_row
                    break
            if final_samples_row:
                # Use direct count instead of formula
                count_value = num_styles  # num_styles is already regular + activewear (excluding custom)
                col_d_idx = column_index_from_string("D")
                # Use safe_set_cell_value to handle MergedCell
                safe_set_cell_value(ws, f"D{final_samples_row}", count_value)
                # Set number format on the actual cell (handle MergedCell)
                try:
                    count_cell = ws.cell(row=final_samples_row, column=col_d_idx)
                    count_cell.number_format = "0"
                except AttributeError:
                    # If it's a MergedCell, find top-left cell
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= final_samples_row <= merged_range.max_row and
                            merged_range.min_col <= col_d_idx <= merged_range.max_col):
                            count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            count_cell.number_format = "0"
                            break
                # Store row for later merge (we'll do merges at the very end)
                # Merges will be handled separately at the end of deliverables section
        
        # Update PATTERNS and FIRST SAMPLES to have the same value as FINAL SAMPLES (num_styles)
        # This applies to both activewear and non-activewear cases, after find_label_row is defined
        # Note: find_label_row searches column H, but PATTERNS/FIRST SAMPLES are in column B
        # So we need to search column B instead
        patterns_row = None
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = safe_get_cell_value(ws, scan_row, column_index_from_string("B"))
            if isinstance(value, str) and "pattern" in value.lower():
                patterns_row = scan_row
                break
        
        if patterns_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= patterns_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Use safe_set_cell_value to handle MergedCell
            safe_set_cell_value(ws, f"D{patterns_row}", None)
            safe_set_cell_value(ws, f"D{patterns_row}", num_styles)
            # Merge and center (PATTERNS typically spans 2 rows)
            patterns_row_second = patterns_row + 1
            col_b_idx = column_index_from_string("B")
            col_c_idx = column_index_from_string("C")
            # Merge B:C for label
            if safe_merge_cells:
                safe_merge_cells(ws, f"B{patterns_row}:C{patterns_row_second}")
                safe_merge_cells(ws, f"D{patterns_row}:D{patterns_row_second}")
            # Set number format and alignment - handle MergedCell
            count_cell = safe_get_writable_cell(ws, patterns_row, col_d_idx)
            label_cell = safe_get_writable_cell(ws, patterns_row, col_b_idx)
            try:
                count_cell.number_format = "0"
                if Alignment:
                    count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            except Exception:
                pass
        
        first_samples_row = None
        for scan_row in range(deliverables_block_start, deliverables_block_end + 1):
            value = safe_get_cell_value(ws, scan_row, column_index_from_string("B"))
            if isinstance(value, str) and "first" in value.lower() and "sample" in value.lower():
                first_samples_row = scan_row
                break
        
        if first_samples_row:
            col_d_idx = column_index_from_string("D")
            # Unmerge if needed and clear any formulas
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= first_samples_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Use safe_set_cell_value to handle MergedCell
            safe_set_cell_value(ws, f"D{first_samples_row}", None)
            safe_set_cell_value(ws, f"D{first_samples_row}", num_styles)
            # Merge and center (FIRST SAMPLES typically spans 2 rows)
            first_samples_row_second = first_samples_row + 1
            col_b_idx = column_index_from_string("B")
            col_c_idx = column_index_from_string("C")
            # Merge B:C for label
            if safe_merge_cells:
                safe_merge_cells(ws, f"B{first_samples_row}:C{first_samples_row_second}")
                safe_merge_cells(ws, f"D{first_samples_row}:D{first_samples_row_second}")
            # Set number format and alignment - handle MergedCell
            count_cell = safe_get_writable_cell(ws, first_samples_row, col_d_idx)
            label_cell = safe_get_writable_cell(ws, first_samples_row, col_b_idx)
            try:
                count_cell.number_format = "0"
                if Alignment:
                    count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            except Exception:
                pass
        
        # Set TSU TechPak and Costing Workbook to number of styles
        # These are typically at rows 33 and 35 (column D), but may shift with row insertions
        teg_tech_pack_row = None
        costing_workbook_row = None
        col_b_idx = column_index_from_string("B")
        # Scan more broadly to find them, including checking merged cells
        scan_end_row = min(deliverables_block_end + 5, ws.max_row)
        for scan_row in range(deliverables_block_start, scan_end_row + 1):
            value = safe_get_cell_value(ws, scan_row, col_b_idx)
            # Also check if it's in a merged cell
            if not value or not isinstance(value, str):
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= scan_row <= merged_range.max_row and
                        merged_range.min_col <= col_b_idx <= merged_range.max_col):
                        value = safe_get_cell_value(ws, merged_range.min_row, merged_range.min_col)
                        scan_row = merged_range.min_row
                        break
            
            if isinstance(value, str):
                value_lower = value.lower().strip()
                if "teg" in value_lower and "tech" in value_lower and "pack" in value_lower:
                    if teg_tech_pack_row is None:
                        teg_tech_pack_row = scan_row
                elif "costing" in value_lower and "workbook" in value_lower:
                    if costing_workbook_row is None:
                        costing_workbook_row = scan_row
        
        col_d_idx = column_index_from_string("D")
        col_b_idx = column_index_from_string("B")
        col_c_idx = column_index_from_string("C")
        
        # Set TEG TECH PACK - merge and center like other deliverables
        if teg_tech_pack_row:
            # Ensure TEG TECH PACK label exists in column B
            teg_label = safe_get_cell_value(ws, teg_tech_pack_row, col_b_idx)
            if not teg_label or not isinstance(teg_label, str) or "teg" not in teg_label.lower():
                # Label is missing, set it
                safe_set_cell_value(ws, f"B{teg_tech_pack_row}", "TEG TECH PACK")
            
            # Unmerge B:C and D first to set value
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= teg_tech_pack_row <= merged_range.max_row and
                    (merged_range.min_col <= col_b_idx <= merged_range.max_col or
                     merged_range.min_col <= col_d_idx <= merged_range.max_col)):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear and set value to num_styles (should be 4)
            safe_set_cell_value(ws, f"D{teg_tech_pack_row}", None)
            safe_set_cell_value(ws, f"D{teg_tech_pack_row}", num_styles)
            # Merge B:C and D, then center
            teg_tech_pack_row_2 = teg_tech_pack_row + 1
            if safe_merge_cells:
                safe_merge_cells(ws, f"B{teg_tech_pack_row}:C{teg_tech_pack_row_2}")
                safe_merge_cells(ws, f"D{teg_tech_pack_row}:D{teg_tech_pack_row_2}")
            # Set number format and alignment
            count_cell = safe_get_writable_cell(ws, teg_tech_pack_row, col_d_idx)
            label_cell = safe_get_writable_cell(ws, teg_tech_pack_row, col_b_idx)
            try:
                count_cell.number_format = "0"
                if Alignment:
                    count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            except Exception:
                pass
        
        if costing_workbook_row and teg_tech_pack_row:
            # Ensure COSTING WORKBOOK label exists in column B
            costing_label = safe_get_cell_value(ws, costing_workbook_row, col_b_idx)
            if not costing_label or not isinstance(costing_label, str) or "costing" not in costing_label.lower():
                # Label is missing, set it
                safe_set_cell_value(ws, f"B{costing_workbook_row}", "COSTING WORKBOOK")
            
            # Read the actual value from TEG TECH PACK cell to ensure they match
            teg_tech_pack_value = safe_get_cell_value(ws, teg_tech_pack_row, col_d_idx)
            # If TEG TECH PACK value is wrong, use num_styles directly
            if teg_tech_pack_value != num_styles:
                teg_tech_pack_value = num_styles
            
            # Unmerge B:C and D first to set value
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= costing_workbook_row <= merged_range.max_row and
                    (merged_range.min_col <= col_b_idx <= merged_range.max_col or
                     merged_range.min_col <= col_d_idx <= merged_range.max_col)):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Set COSTING WORKBOOK value
            if get_column_letter:
                costing_workbook_ref = f"{get_column_letter(col_d_idx)}{costing_workbook_row}"
                safe_set_cell_value(ws, costing_workbook_ref, teg_tech_pack_value)
            else:
                safe_set_cell_value(ws, f"D{costing_workbook_row}", teg_tech_pack_value)
            # Merge B:C and D, then center
            costing_workbook_row_2 = costing_workbook_row + 1
            if safe_merge_cells:
                safe_merge_cells(ws, f"B{costing_workbook_row}:C{costing_workbook_row_2}")
                safe_merge_cells(ws, f"D{costing_workbook_row}:D{costing_workbook_row_2}")
            # Set number format and alignment
            count_cell = safe_get_writable_cell(ws, costing_workbook_row, col_d_idx)
            label_cell = safe_get_writable_cell(ws, costing_workbook_row, col_b_idx)
            try:
                count_cell.number_format = "0"
                if Alignment:
                    count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            except Exception:
                pass
        elif costing_workbook_row:
            # Fallback if TEG TECH PACK not found
            if get_column_letter:
                costing_workbook_ref = f"{get_column_letter(col_d_idx)}{costing_workbook_row}"
                safe_set_cell_value(ws, costing_workbook_ref, num_styles)
                # Set number format - use safe_get_writable_cell
                count_cell = safe_get_writable_cell(ws, costing_workbook_row, col_d_idx)
                try:
                    count_cell.number_format = "0"
                except Exception:
                    pass
            else:
                # Fallback - handle MergedCell
                try:
                    count_cell = ws.cell(row=costing_workbook_row, column=col_d_idx)
                    count_cell.value = num_styles
                    count_cell.number_format = "0"
                except Exception:
                    # Handle MergedCell by finding top-left cell or unmerging
                    for merged_range in list(ws.merged_cells.ranges):
                        if (merged_range.min_row <= costing_workbook_row <= merged_range.max_row and
                            merged_range.min_col <= col_d_idx <= merged_range.max_col):
                            try:
                                # Use safe_set_cell_value to handle MergedCell
                                if get_column_letter:
                                    safe_set_cell_value(ws, f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}", num_styles)
                                else:
                                    from openpyxl.utils import get_column_letter as gcl
                                    safe_set_cell_value(ws, f"{gcl(merged_range.min_col)}{merged_range.min_row}", num_styles)
                                # Set number format - handle MergedCell
                                try:
                                    count_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                    count_cell.number_format = "0"
                                except AttributeError:
                                    pass
                                break
                            except Exception:
                                pass
        
        # Add right borders to column D (values) for all deliverables in the block
        # Find all deliverable rows and add right border to their D column values
        if Border is not None and Side is not None:
            thin = Side(style="thin")
            col_b_idx = column_index_from_string("B")
            col_d_idx = column_index_from_string("D")
            
            # Scan for all deliverables and add right border to their D column
            for scan_row in range(deliverables_block_start, deliverables_block_end + 5):
                value = safe_get_cell_value(ws, scan_row, col_b_idx)
                if isinstance(value, str) and value.strip():
                    value_lower = value.lower().strip()
                    # Check if it's a deliverable label
                    deliverable_keywords = ["pattern", "sample", "fitting", "revision", "tech pack", "costing", "workbook"]
                    if any(keyword in value_lower for keyword in deliverable_keywords):
                        # Found a deliverable, add right border to its D column
                        count_cell = safe_get_writable_cell(ws, scan_row, col_d_idx)
                        try:
                            existing_border = count_cell.border
                            if existing_border:
                                count_cell.border = Border(
                                    left=existing_border.left,
                                    right=thin,
                                    top=existing_border.top,
                                    bottom=existing_border.bottom
                                )
                            else:
                                count_cell.border = Border(right=thin)
                        except Exception:
                            pass
        
        # Final check: Ensure TEG TECH PACK and COSTING WORKBOOK are correct
        # Re-scan to make sure we have the right rows
        if not teg_tech_pack_row or not costing_workbook_row:
            for scan_row in range(deliverables_block_start, min(deliverables_block_end + 5, ws.max_row + 1)):
                value = safe_get_cell_value(ws, scan_row, col_b_idx)
                if isinstance(value, str):
                    value_lower = value.lower().strip()
                    if "teg" in value_lower and "tech" in value_lower and "pack" in value_lower:
                        if not teg_tech_pack_row:
                            teg_tech_pack_row = scan_row
                    elif "costing" in value_lower and "workbook" in value_lower:
                        if not costing_workbook_row:
                            costing_workbook_row = scan_row
        
        # Ensure COSTING WORKBOOK has the same value as TEG TECH PACK (final check after all processing)
        if teg_tech_pack_row and costing_workbook_row:
            # Ensure labels exist
            teg_label = safe_get_cell_value(ws, teg_tech_pack_row, col_b_idx)
            if not teg_label or not isinstance(teg_label, str) or "teg" not in teg_label.lower():
                safe_set_cell_value(ws, f"B{teg_tech_pack_row}", "TEG TECH PACK")
            
            costing_label = safe_get_cell_value(ws, costing_workbook_row, col_b_idx)
            if not costing_label or not isinstance(costing_label, str) or "costing" not in costing_label.lower():
                safe_set_cell_value(ws, f"B{costing_workbook_row}", "COSTING WORKBOOK")
            
            # Read TEG TECH PACK value, but use num_styles if it's wrong
            teg_tech_pack_value = safe_get_cell_value(ws, teg_tech_pack_row, col_d_idx)
            if teg_tech_pack_value != num_styles:
                # TEG TECH PACK has wrong value, fix it
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_row <= teg_tech_pack_row <= merged_range.max_row and
                        merged_range.min_col <= col_d_idx <= merged_range.max_col):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass
                safe_set_cell_value(ws, f"D{teg_tech_pack_row}", None)
                safe_set_cell_value(ws, f"D{teg_tech_pack_row}", num_styles)
                teg_tech_pack_value = num_styles
                # Re-merge
                if safe_merge_cells:
                    safe_merge_cells(ws, f"D{teg_tech_pack_row}:D{teg_tech_pack_row + 1}")
            
            # Set COSTING WORKBOOK to match TEG TECH PACK
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= costing_workbook_row <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            safe_set_cell_value(ws, f"D{costing_workbook_row}", None)
            safe_set_cell_value(ws, f"D{costing_workbook_row}", teg_tech_pack_value)
            # Re-merge
            if safe_merge_cells:
                safe_merge_cells(ws, f"B{costing_workbook_row}:C{costing_workbook_row + 1}")
                safe_merge_cells(ws, f"D{costing_workbook_row}:D{costing_workbook_row + 1}")
            # Set alignment
            count_cell = safe_get_writable_cell(ws, costing_workbook_row, col_d_idx)
            label_cell = safe_get_writable_cell(ws, costing_workbook_row, col_b_idx)
            try:
                count_cell.number_format = "0"
                if Alignment:
                    count_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            except Exception:
                pass

    # Totals section - dynamically calculate totals row and range based on number of styles
    # For 5 or fewer styles: totals at row 20 (original position)
    # For more than 5 styles: totals row shifts down by (total_styles_count - 5) * 2 rows
    if total_styles_count > 0:
        first_style_row = dynamic_row_indices[0] if dynamic_row_indices else 10
        if total_styles_count <= 5:
            totals_row = 20  # Original totals row position
        else:
            # Totals row shifts down by the number of rows we inserted
            totals_row = 20 + (total_styles_count - 5) * 2
        
        # Unmerge any merged cells in the totals row to avoid issues
        merged_ranges_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row):
                merged_ranges_to_unmerge.append(merged_range)
        
        for merged_range in merged_ranges_to_unmerge:
            try:
                min_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                max_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                ws.unmerge_cells(f"{min_cell.coordinate}:{max_cell.coordinate}")
            except Exception:
                pass
        
        # Set totals row labels
        cell_b_totals = ws.cell(row=totals_row, column=2)
        cell_b_totals.value = "TOTAL DEVELOPMENT"
        if Font is not None:
            cell_b_totals.font = Font(bold=True)
        
        # Merge and center "TOTAL OPTIONAL ADD-ONS" H20-J20
        cell_h_totals = ws.cell(row=totals_row, column=8)
        cell_h_totals.value = "TOTAL OPTIONAL ADD-ONS"
        if Font is not None:
            cell_h_totals.font = Font(bold=True)
        # Unmerge any existing merges in H-J for totals row
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row and
                merged_range.min_col <= 8 <= merged_range.max_col <= 10):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        # Merge H-J
        safe_merge_cells(ws, f"H{totals_row}:J{totals_row}")
        # Get writable cell after merge for alignment and borders
        cell_h_totals = safe_get_writable_cell(ws, totals_row, 8)
        if Alignment is not None:
            try:
                cell_h_totals.alignment = Alignment(horizontal="left", vertical="center")
            except Exception:
                pass
        # Apply full borders to H-J merged cell (for more than 5 styles)
        if total_styles_count > 5 and Border is not None and Side is not None:
            thin = Side(style="thin")
            try:
                cell_h_totals.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            except Exception:
                pass
        
        # Set totals formulas - sum all style rows (dynamic based on actual style rows)
        cell_f_totals = ws.cell(row=totals_row, column=6)
        cell_f_totals.value = f"=SUM(F10:F{last_style_row})"
        cell_f_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_f_totals.font = Font(bold=True)
        if Alignment is not None:
            cell_f_totals.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #709171 to TOTAL DEVELOPMENT
        if PatternFill is not None:
            cell_f_totals.fill = PatternFill(start_color="709171", end_color="709171", fill_type="solid")
        
        # TOTAL OPTIONAL ADD-ONS now uses columns K-L (merged K20-L20)
        cell_k_totals = ws.cell(row=totals_row, column=11)
        cell_l_totals = ws.cell(row=totals_row, column=12)
        # Formula now sums K column (which contains SUM(H:J) for each row)
        cell_k_totals.value = f"=SUM(K10:K{last_style_row})"
        cell_k_totals.number_format = '$#,##0'  # Currency format
        if Font is not None:
            cell_k_totals.font = Font(bold=True)
        # Unmerge any existing merges in K-L for totals row
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= totals_row <= merged_range.max_row and
                merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        # Merge and center K-L
        safe_merge_cells(ws, f"K{totals_row}:L{totals_row}")
        # Get writable cell after merge for alignment, font, fill, and borders
        cell_k_totals = safe_get_writable_cell(ws, totals_row, 11)
        if Alignment is not None:
            try:
                cell_k_totals.alignment = Alignment(horizontal="center", vertical="center")
            except Exception:
                pass
        # Apply font size 20 to TOTAL OPTIONAL ADD-ONS
        if Font is not None:
            try:
                cell_k_totals.font = Font(name="Arial", size=20, bold=True)
            except Exception:
                pass
        # Apply cell color #f0cfbb to TOTAL OPTIONAL ADD-ONS
        if PatternFill is not None:
            try:
                cell_k_totals.fill = PatternFill(start_color="F0CFBB", end_color="F0CFBB", fill_type="solid")
            except Exception:
                pass
        # Clear L cell since it's merged with K
        try:
            cell_l_totals.value = None
        except Exception:
            pass
        
        # Apply full borders to K-L merged cell (ALWAYS, not just for more than 5 styles)
        # For merged cells, we need to ensure the border is applied correctly
        # Get the merged cell and apply borders
        cell_k_totals = safe_get_writable_cell(ws, totals_row, 11)
        if Border is not None and Side is not None:
            thin = Side(style="thin")
            full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            try:
                # Apply border to the merged cell (top-left cell K)
                cell_k_totals.border = full_border
            except Exception:
                try:
                    # Fallback: get cell directly
                    cell_k_totals = ws.cell(row=totals_row, column=11)
                    cell_k_totals.border = full_border
                except Exception:
                    pass
            
            # Also ensure column L (rightmost of merged range) has the right border
            # This is important for merged cells - sometimes the right border needs to be on the rightmost column
            try:
                cell_l_totals = ws.cell(row=totals_row, column=12)
                # Get existing border or create new one
                existing_border = cell_l_totals.border if cell_l_totals.border else Border()
                # Ensure right border is set
                cell_l_totals.border = Border(
                    left=existing_border.left if existing_border.left else thin,
                    right=thin,  # Ensure right border is set
                    top=existing_border.top if existing_border.top else thin,
                    bottom=existing_border.bottom if existing_border.bottom else thin
                )
            except Exception:
                pass
        
        # For less than 5 styles, merge, center and apply all borders to empty OPTIONAL ADD-ONS totals
        if total_styles_count <= 5 and Border is not None and Side is not None:
            # Find which style rows are empty (not used)
            used_rows = set()
            if dynamic_row_indices:
                for row_idx in dynamic_row_indices:
                    used_rows.add(row_idx)
            
            # Process each row in ROW_INDICES that wasn't used
            for row_idx in ROW_INDICES:
                if row_idx not in used_rows:
                    row_second = row_idx + 1
                    # Check if K-L cells are empty (no formula/value)
                    cell_k_value = safe_get_cell_value(ws, row_idx, 11)
                    if cell_k_value is None or (isinstance(cell_k_value, str) and cell_k_value.strip() == ""):
                        # Unmerge any existing merges in K-L for this row pair
                        for merged_range in list(ws.merged_cells.ranges):
                            if (merged_range.min_row <= row_idx <= merged_range.max_row <= row_second and
                                merged_range.min_col <= 11 <= merged_range.max_col <= 12):
                                try:
                                    ws.unmerge_cells(range_string=str(merged_range))
                                except Exception:
                                    pass
                        
                        # Apply full borders to K and L columns
                        apply_full_border_pair(ws, 11, row_idx, row_second)
                        apply_full_border_pair(ws, 12, row_idx, row_second)
                        
                        # Merge and center K-L
                        safe_merge_cells(ws, f"K{row_idx}:L{row_second}")
                        if Alignment is not None:
                            # Get writable cell for alignment
                            cell_k = safe_get_writable_cell(ws, row_idx, 11)
                            try:
                                cell_k.alignment = Alignment(horizontal="center", vertical="center")
                            except Exception:
                                pass
                        # Clear L cell since it's merged with K (handle MergedCell)
                        try:
                            cell_l = safe_get_writable_cell(ws, row_idx, 12)
                            cell_l.value = None
                        except Exception:
                            # If it's a MergedCell, the value is already None or handled by merge
                            pass

        try:
            first_note_row = None
            second_note_row = None

            # Search a reasonable band of rows where the notes live
            for row in range(20, 90):
                for col in (5, 6):  # Columns E (5) and F (6)
                    value = safe_get_cell_value(ws, row, col)
                    if not isinstance(value, str):
                        continue
                    upper = value.upper()
                    if ("DESIGNS ARE REVIEWED" in upper and
                        "SURCHARGE APPLIES" in upper and
                        first_note_row is None):
                        first_note_row = row
                    if ("DEVELOPMENT DOES NOT INCLUDE" in upper and
                        "BULK PRODUCTION INVENTORY" in upper and
                        second_note_row is not None) is False:
                        # Only set once when found
                        if ("DEVELOPMENT DOES NOT INCLUDE" in upper and
                            "BULK PRODUCTION INVENTORY" in upper and
                            second_note_row is None):
                            second_note_row = row
                if first_note_row is not None and second_note_row is not None:
                    break

            # Helper to (re)merge and fully border a note block in E-F
            def _format_note_block(start_row: int, row_span: int) -> None:
                end_row = start_row + row_span - 1
                # Unmerge any existing merges in E-F intersecting this vertical span
                for merged_range in list(ws.merged_cells.ranges):
                    if (merged_range.min_col <= 6 and merged_range.max_col >= 5 and
                        merged_range.min_row <= end_row and merged_range.max_row >= start_row):
                        try:
                            ws.unmerge_cells(range_string=str(merged_range))
                        except Exception:
                            pass

                # Merge E-F across the span
                safe_merge_cells(ws, f"E{start_row}:F{end_row}")

                # Center-align the merged cell
                if Alignment is not None:
                    top_left = safe_get_writable_cell(ws, start_row, 5)
                    try:
                        top_left.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=True,
                        )
                    except Exception:
                        pass

                # Apply full borders to all cells in the E-F block
                if Border is not None and Side is not None:
                    thin = Side(style="thin")
                    for r in range(start_row, end_row + 1):
                        for c in (5, 6):
                            cell = ws.cell(row=r, column=c)
                            try:
                                cell.border = Border(
                                    left=thin,
                                    right=thin,
                                    top=thin,
                                    bottom=thin,
                                )
                            except Exception:
                                pass

            # First note: 6-row span
            if first_note_row is not None:
                _format_note_block(first_note_row, 6)

            # Second note: 8-row span (position will already include any extra rows
            # introduced by Activewear logic, so we just span 8 rows from its text row)
            if second_note_row is not None:
                _format_note_block(second_note_row, 8)

                # After we format the main 8-row block, remove any *other* instances
                for row in range(20, 90):
                    # Skip inside the formatted 8-row block
                    if second_note_row <= row <= second_note_row + 7:
                        continue
                    for col in (5, 6):  # E/F
                        value = safe_get_cell_value(ws, row, col)
                        if not isinstance(value, str):
                            continue
                        upper = value.upper()
                        if ("DEVELOPMENT DOES NOT INCLUDE" in upper and
                            "BULK PRODUCTION INVENTORY" in upper):
                            try:
                                safe_set_cell_value(
                                    ws,
                                    f"{'E' if col == 5 else 'F'}{row}",
                                    None,
                                )
                            except Exception:
                                pass

        except Exception:
            # Notes formatting should never break workbook creation
            pass

        try:
            # Only build the tall merged E–F \"box\" when there is at least one Activewear style. For regular-only projects we keep the original template behavior untouched.
            if (
                num_activewear > 0
                and Border is not None
                and Side is not None
                and column_index_from_string is not None
            ):
                label_col_b = column_index_from_string("B")
                final_samples_row_scan = None
                costing_workbook_row_scan = None

                # Scan a reasonable band where these labels live
                for row in range(DELIVERABLE_BLOCK_START, DELIVERABLE_BLOCK_END + 15):
                    value = safe_get_cell_value(ws, row, label_col_b)
                    if not isinstance(value, str):
                        continue
                    lower = value.strip().lower()
                    if "final samples" in lower and final_samples_row_scan is None:
                        final_samples_row_scan = row
                    elif ("costing" in lower and "workbook" in lower and
                          costing_workbook_row_scan is None):
                        costing_workbook_row_scan = row

                if final_samples_row_scan is not None and costing_workbook_row_scan is not None:
                    start_row = final_samples_row_scan
                    end_row = costing_workbook_row_scan
                    # Include one extra row below COSTING WORKBOOK to match the visual box
                    box_end_row = end_row + 1

                    thin = Side(style="thin")
                    # First, apply full borders to E/F for the whole vertical span
                    for r in range(start_row, box_end_row + 1):
                        for c in (5, 6):  # E and F
                            cell = ws.cell(row=r, column=c)
                            try:
                                cell.border = Border(
                                    left=thin,
                                    right=thin,
                                    top=thin,
                                    bottom=thin,
                                )
                            except Exception:
                                pass

                    # Then merge this entire E–F block into a single tall cell and center its content.
                    for merged_range in list(ws.merged_cells.ranges):
                        if (
                            merged_range.min_col <= 6
                            and merged_range.max_col >= 5
                            and merged_range.min_row <= box_end_row
                            and merged_range.max_row >= start_row
                        ):
                            try:
                                ws.unmerge_cells(range_string=str(merged_range))
                            except Exception:
                                pass

                    # Merge E-F from FINAL SAMPLES row down through COSTING WORKBOOK box
                    safe_merge_cells(ws, f"E{start_row}:F{box_end_row}")

                    if Alignment is not None:
                        top_left = safe_get_writable_cell(ws, start_row, 5)
                        try:
                            top_left.alignment = Alignment(
                                horizontal="center",
                                vertical="center",
                                wrap_text=True,
                            )
                        except Exception:
                            pass
        except Exception:
            # Border \"box\" improvement should not break workbook creation
            pass

        # SUB-TOTAL row (new) and Discount row (moved down)
        subtotal_row = SUMMARY_SUBTOTAL_ROW
        discount_row = SUMMARY_DISCOUNT_ROW
        discount_decimal = discount_percentage / 100.0 if discount_percentage else 0.0
        cell_n_subtotal = ws.cell(row=subtotal_row, column=SUMMARY_LABEL_COL)
        cell_p_subtotal = ws.cell(row=subtotal_row, column=SUMMARY_VALUE_COL)
        cell_n_discount = ws.cell(row=discount_row, column=SUMMARY_LABEL_COL)
        cell_p_discount = ws.cell(row=discount_row, column=SUMMARY_VALUE_COL)
        
        if discount_percentage > 0:
            # SUB-TOTAL = TOTAL DEVELOPMENT + TOTAL OPTIONAL ADD-ONS
            cell_n_subtotal.value = "SUB-TOTAL"
            if Font is not None:
                cell_n_subtotal.font = Font(name="Arial", size=20, bold=True, color=cell_n_subtotal.font.color if cell_n_subtotal.font else None)
            if Alignment is not None:
                cell_n_subtotal.alignment = Alignment(horizontal="center", vertical="center")
            cell_p_subtotal.value = "=SUM(P10:P13)"
            cell_p_subtotal.number_format = '$#,##0'
            if Font is not None:
                cell_p_subtotal.font = Font(name="Arial", size=20, bold=True, color=cell_p_subtotal.font.color if cell_p_subtotal.font else None)
            if Alignment is not None:
                cell_p_subtotal.alignment = Alignment(horizontal="center", vertical="center")
            
            # DISCOUNT uses SUB-TOTAL as base
            cell_n_discount.value = f"DISCOUNT ({discount_percentage:.0f}%)"
            if Font is not None:
                cell_n_discount.font = Font(name="Arial", size=20, bold=True, color=cell_n_discount.font.color if cell_n_discount.font else None)
            if Alignment is not None:
                cell_n_discount.alignment = Alignment(horizontal="center", vertical="center")
            
            cell_p_discount.value = f"=P{subtotal_row}*{discount_decimal}"
            cell_p_discount.number_format = '$#,##0'
            if Font is not None:
                cell_p_discount.font = Font(name="Arial", size=20, bold=True, color=cell_p_discount.font.color if cell_p_discount.font else None)
            if Alignment is not None:
                cell_p_discount.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Clear SUB-TOTAL and DISCOUNT rows when no discount
            for clear_row in [subtotal_row, subtotal_row + 1, discount_row, discount_row + 1]:
                safe_set_cell_value(ws, f"N{clear_row}", None)
                safe_set_cell_value(ws, f"P{clear_row}", None)
            
            # Merge and center cleared ranges for SUB-TOTAL and DISCOUNT
            if safe_merge_cells:
                safe_merge_cells(ws, f"N{subtotal_row}:P{subtotal_row + 1}")
                safe_merge_cells(ws, f"N{discount_row}:P{discount_row + 1}")
                if Alignment is not None:
                    merged_cell_subtotal = ws.cell(row=subtotal_row, column=SUMMARY_LABEL_COL)
                    merged_cell_subtotal.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    merged_cell_discount = ws.cell(row=discount_row, column=SUMMARY_LABEL_COL)
                    merged_cell_discount.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            
            # Ensure N12 and P12 have inferior (bottom) borders
            if Border is not None and Side is not None:
                bottom_side = Side(style="thin")
                row_12 = SUMMARY_OPT_ROW  # Row 12
                cell_n12 = ws.cell(row=row_12, column=SUMMARY_LABEL_COL)
                cell_p12 = ws.cell(row=row_12, column=SUMMARY_VALUE_COL)
                
                existing_n12_border = cell_n12.border if cell_n12.border else Border()
                existing_p12_border = cell_p12.border if cell_p12.border else Border()
                
                cell_n12.border = Border(
                    left=existing_n12_border.left,
                    right=existing_n12_border.right,
                    top=existing_n12_border.top,
                    bottom=bottom_side
                )
                cell_p12.border = Border(
                    left=existing_p12_border.left,
                    right=existing_p12_border.right,
                    top=existing_p12_border.top,
                    bottom=bottom_side
                )
        
        # Clear N23 if it contains discount percentage (remove duplicate)
        # N23 should remain empty or contain "TOTAL DUE AT SIGNING" if that's what the template has
        cell_n23_check = ws.cell(row=23, column=SUMMARY_LABEL_COL)
        cell_n23_check_value = safe_get_cell_value(ws, 23, SUMMARY_LABEL_COL)
        if cell_n23_check_value and "%" in str(cell_n23_check_value):
            # Clear the duplicate discount percentage from N23
            cell_n23_check.value = None
        
        # Make all cells in totals row bold
        if Font is not None:
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
                else:
                    cell.font = Font(bold=True)
        
        # Apply Arial 20 font to totals row if it's a new row (num_styles > 5)
        if num_styles > 5:
            for col_idx in [2, 6, 8, 12, 14, 16]:  # Columns B, F, H, L, N, P
                cell = ws.cell(row=totals_row, column=col_idx)
                apply_arial_20_font(cell)
                # Ensure bold is maintained
                if Font is not None:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        color=cell.font.color
                    )
        else:
            # For <=5 styles (totals_row==20), ensure Arial size 20 bold for key cells
            for col_idx in [2, 6, 8, 12, 14, 16]:
                cell = ws.cell(row=totals_row, column=col_idx)
                if Font is not None:
                    cell.font = Font(name="Arial", size=20, bold=True, color=cell.font.color if cell.font else None)
        # Ensure P (column 16) is also Arial size 20 for <=5 styles
        if num_styles <= 5:
            cell_p_totals = ws.cell(row=totals_row, column=16)
            if Font is not None:
                cell_p_totals.font = Font(name="Arial", size=20, bold=True, color=cell_p_totals.font.color if cell_p_totals.font else None)
        # Apply inferior (bottom) border to entire totals row
        if Border is not None and Side is not None:
            bottom_side = Side(style="thin")
            for col_idx in range(1, 17):  # Columns A through P
                cell = ws.cell(row=totals_row, column=col_idx)
                # Skip columns A, G, M for bottom borders
                if col_idx in [1, 7, 13]:
                    new_bottom = cell.border.bottom
                else:
                    new_bottom = bottom_side
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=new_bottom,
                )

        # Apply full borders (top/bottom/left/right) to key totals cells (B, F, H-J merged, K-L merged, N, P)
        # Note: H-J and K-L are merged, so we need to handle them separately
        if Border is not None and Side is not None:
            full_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            # Apply borders to individual cells (B, F, N, P)
            for col_idx in [2, 6, 14, 16]:
                cell = safe_get_writable_cell(ws, totals_row, col_idx)
                try:
                    cell.border = full_border
                except Exception:
                    pass
            # Apply borders to merged cells H-J and K-L (for more than 5 styles)
            if total_styles_count > 5:
                # H-J merged cell
                cell_h_totals = safe_get_writable_cell(ws, totals_row, 8)
                try:
                    cell_h_totals.border = full_border
                except Exception:
                    pass
                # K-L merged cell - ensure right border is on column L
                cell_k_totals = safe_get_writable_cell(ws, totals_row, 11)
                try:
                    cell_k_totals.border = full_border
                except Exception:
                    pass
                # Also set right border on column L (rightmost of merged range)
                try:
                    cell_l_totals = ws.cell(row=totals_row, column=12)
                    existing_border = cell_l_totals.border if cell_l_totals.border else Border()
                    cell_l_totals.border = Border(
                        left=existing_border.left if existing_border.left else Side(style="thin"),
                        right=Side(style="thin"),  # Ensure right border
                        top=existing_border.top if existing_border.top else Side(style="thin"),
                        bottom=existing_border.bottom if existing_border.bottom else Side(style="thin")
                    )
                except Exception:
                    pass
            else:
                # For <=5 styles, also ensure K-L merged cell has full borders including right border
                cell_k_totals = safe_get_writable_cell(ws, totals_row, 11)
                try:
                    cell_k_totals.border = full_border
                except Exception:
                    pass
                # Also set right border on column L
                try:
                    cell_l_totals = ws.cell(row=totals_row, column=12)
                    existing_border = cell_l_totals.border if cell_l_totals.border else Border()
                    cell_l_totals.border = Border(
                        left=existing_border.left if existing_border.left else Side(style="thin"),
                        right=Side(style="thin"),  # Ensure right border
                        top=existing_border.top if existing_border.top else Side(style="thin"),
                        bottom=existing_border.bottom if existing_border.bottom else Side(style="thin")
                    )
                except Exception:
                    pass
                # Also apply to individual H cell (H-J is merged, but H still needs border)
                cell_h_totals = safe_get_writable_cell(ws, totals_row, 8)
                try:
                    cell_h_totals.border = full_border
                except Exception:
                    pass
        
        # Update P10 and P12 to use dynamic references (they reference F20 and L20 statically)
        # P10 should reference F{totals_row}, P12 should reference L{totals_row}
        cell_p10 = safe_get_writable_cell(ws, 10, 16)  # Column P, row 10
        cell_p10_value = safe_get_cell_value(ws, 10, 16)
        if cell_p10_value and isinstance(cell_p10_value, str) and cell_p10_value.startswith("="):
            if "F20" in cell_p10_value:
                try:
                    cell_p10.value = f"=F{totals_row}"
                except Exception:
                    safe_set_cell_value(ws, "P10", f"=F{totals_row}")
        
        cell_p12 = ws.cell(row=12, column=16)  # Column P, row 12
        cell_p12 = safe_get_writable_cell(ws, 12, 16)
        cell_p12_value = safe_get_cell_value(ws, 12, 16)
        if cell_p12_value and isinstance(cell_p12_value, str) and cell_p12_value.startswith("="):
            # Update to reference K{totals_row} (the merged K-L cell where TOTAL OPTIONAL ADD-ONS value is)
            # Check for both L20 and K20 references, and also check if it's referencing a style row instead of totals
            if "L20" in cell_p12_value or "K20" in cell_p12_value:
                cell_p12.value = f"=K{totals_row}"
            elif "L10" in cell_p12_value or "L12" in cell_p12_value or "L14" in cell_p12_value or "L16" in cell_p12_value or "L18" in cell_p12_value:
                # If it's referencing a style row (L10, L12, etc.), update to totals row
                cell_p12.value = f"=K{totals_row}"
            elif "K10" in cell_p12_value or "K12" in cell_p12_value or "K14" in cell_p12_value or "K16" in cell_p12_value or "K18" in cell_p12_value:
                # If it's referencing a style row in K, update to totals row
                cell_p12.value = f"=K{totals_row}"
        
        # Place "TOTAL DUE AT SIGNING" in the 2 rows just above totals row: N(totals_row-1):O(totals_row) and P(totals_row-1):P(totals_row).
        # For <=5 styles: totals_row=20 → N19:O20, P19:P20. For 6 styles: totals_row=22 → N21:O22, P21:P22. Always dynamic.
        total_due_start = totals_row - 1
        total_due_end = totals_row
        cell_n_label = ws.cell(row=total_due_start, column=14)
        # Always place in dynamic position when we have a totals row (so 6+ styles get N21:O22, not N19:O20)
        # When totals row > 20, unmerge and clear old position (N19:O20, P19:P20) to avoid duplicate
        if total_due_start > 19:
            for merged_range in list(ws.merged_cells.ranges):
                if (merged_range.min_row <= 20 and merged_range.max_row >= 19 and
                    merged_range.min_col >= 14 and merged_range.max_col <= 16):
                    try:
                        ws.unmerge_cells(range_string=str(merged_range))
                    except Exception:
                        pass
            # Clear value, borders, and fill (yellow) from original N19:O20, P19:P20 so no leftover formatting
            for r in (19, 20):
                for c in (14, 15, 16):
                    try:
                        cell = ws.cell(row=r, column=c)
                        cell.value = None
                        if Border is not None:
                            cell.border = Border()
                        if PatternFill is not None:
                            cell.fill = PatternFill(fill_type="none")
                    except Exception:
                        pass
        # Unmerge any existing merge in N-P for the target two rows (total_due_start:total_due_end)
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= total_due_end and merged_range.max_row >= total_due_start and
                merged_range.min_col >= 14 and merged_range.max_col <= 16):
                try:
                    ws.unmerge_cells(range_string=str(merged_range))
                except Exception:
                    pass
        # Apply borders to every cell in both ranges BEFORE merging so the full perimeter is bordered
        if Border is not None and Side is not None:
            thin_side = Side(style="thin")
            no_side = Side(style=None)
            for r in range(total_due_start, total_due_end + 1):
                top_side = thin_side if r == total_due_start else no_side
                bottom_side = thin_side if r == total_due_end else no_side
                for c in range(14, 16 + 1):  # N=14, O=15, P=16
                    try:
                        cell = ws.cell(row=r, column=c)
                        if c == 14:
                            cell.border = Border(left=thin_side, right=no_side, top=top_side, bottom=bottom_side)
                        elif c == 15:
                            cell.border = Border(left=no_side, right=thin_side, top=top_side, bottom=bottom_side)
                        else:
                            cell.border = Border(left=thin_side, right=thin_side, top=top_side, bottom=bottom_side)
                    except Exception:
                        pass
        # Merge N(total_due_start):O(total_due_end) (label spanning 2 rows)
        cell_n_label.value = "TOTAL DUE AT SIGNING"
        safe_merge_cells(ws, f"N{total_due_start}:O{total_due_end}")
        cell_n_label = safe_get_writable_cell(ws, total_due_start, 14)
        # Merge P(total_due_start):P(total_due_end) (value spanning 2 rows)
        cell_p_val = ws.cell(row=total_due_start, column=16)  # Column P
        if discount_percentage > 0:
            cell_p_val.value = f"=P{SUMMARY_SUBTOTAL_ROW}-P{SUMMARY_DISCOUNT_ROW}"
        else:
            cell_p_val.value = f"=SUM(P10:P13)"
        safe_merge_cells(ws, f"P{total_due_start}:P{total_due_end}")
        cell_p_val = safe_get_writable_cell(ws, total_due_start, 16)
        cell_p_val.number_format = '$#,##0'  # Currency format
        # Apply font size 20 and bold to TOTAL DUE AT SIGNING formula
        if Font is not None:
            cell_p_val.font = Font(name="Arial", size=20, bold=True, color=cell_p_val.font.color if cell_p_val.font else None)
        if Alignment is not None:
            cell_p_val.alignment = Alignment(horizontal="center", vertical="center")
        # Apply cell color #ffff00 to TOTAL DUE AT SIGNING
        if PatternFill is not None:
            cell_p_val.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Label (merged N:O) bold size 20
        if Font is not None:
            cell_n_label.font = Font(name="Arial", size=20, bold=True, color=cell_n_label.font.color if cell_n_label.font else None)
        if Alignment is not None:
            cell_n_label.alignment = Alignment(horizontal="left", vertical="center")
        
        # Also update any other formulas in column P that reference F20 or L20 statically
        # Check a few rows around the totals row
        for check_row in range(totals_row - 2, totals_row + 3):
            cell_p = safe_get_writable_cell(ws, check_row, 16)  # Column P
            if cell_p.value and isinstance(cell_p.value, str) and cell_p.value.startswith("="):
                # Check if it references F20 or L20 (static references)
                if "F20" in cell_p.value or "L20" in cell_p.value:
                    # Replace with dynamic references
                    formula = cell_p.value.replace("F20", f"F{totals_row}").replace("L20", f"L{totals_row}")
                    cell_p.value = formula
                    if Font is not None:
                        existing_font = cell_p.font
                        cell_p.font = Font(
                            name=existing_font.name if existing_font and existing_font.name else "Arial",
                            size=existing_font.size if existing_font and existing_font.size else 20,
                            bold=True,
                            color=existing_font.color if existing_font else None
                        )
                    if Alignment is not None:
                        cell_p.alignment = Alignment(horizontal="center", vertical="center")
    else:
        safe_set_cell_value(ws, "F20", None)
        safe_set_cell_value(ws, "L20", None)
    
    # Final absolute check: Ensure COSTING WORKBOOK matches TEG TECH PACK (right before return)
    col_d_idx = column_index_from_string("D")
    teg_tech_pack_row_final = None
    costing_workbook_row_final = None
    for scan_row in range(20, 50):  # Scan a reasonable range
        # Handle MergedCell when reading value
        value = None
        try:
            cell = ws.cell(row=scan_row, column=column_index_from_string("B"))
            value = cell.value
        except Exception:
            # If it's a MergedCell, find top-left cell
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= scan_row <= merged_range.max_row and
                    merged_range.min_col <= column_index_from_string("B") <= merged_range.max_col):
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    value = cell.value
                    break
        
        if isinstance(value, str):
            value_lower = value.lower()
            if "teg" in value_lower and "tech" in value_lower and "pack" in value_lower:
                teg_tech_pack_row_final = scan_row
            elif "costing" in value_lower and "workbook" in value_lower:
                costing_workbook_row_final = scan_row
    
    if teg_tech_pack_row_final and costing_workbook_row_final:
        # Read TEG TECH PACK value (handle merged cells)
        teg_tech_pack_value_final = safe_get_cell_value(ws, teg_tech_pack_row_final, col_d_idx)
        
        if teg_tech_pack_value_final is not None:
            # Find top-left cell of merged range if COSTING WORKBOOK is merged
            costing_workbook_top_row = costing_workbook_row_final
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= costing_workbook_row_final <= merged_range.max_row and
                    merged_range.min_col <= col_d_idx <= merged_range.max_col):
                    costing_workbook_top_row = merged_range.min_row
                    break
            
            # Use safe_set_cell_value to handle merged cells properly
            if get_column_letter:
                costing_workbook_ref = f"{get_column_letter(col_d_idx)}{costing_workbook_top_row}"
                safe_set_cell_value(ws, costing_workbook_ref, teg_tech_pack_value_final)
                # Set number format on the actual cell - use safe_get_writable_cell
                costing_workbook_cell_final = safe_get_writable_cell(ws, costing_workbook_top_row, col_d_idx)
                try:
                    costing_workbook_cell_final.number_format = "0"
                except Exception:
                    pass
            else:
                # Fallback if get_column_letter not available - use safe functions
                safe_set_cell_value(ws, f"D{costing_workbook_top_row}", teg_tech_pack_value_final)
                costing_workbook_cell_final = safe_get_writable_cell(ws, costing_workbook_top_row, col_d_idx)
                try:
                    costing_workbook_cell_final.number_format = "0"
                except Exception:
                    pass
    
    return total_development, total_optional


def build_workbook_bytes(
    *,
    client_name: str,
    client_email: str,
    representative: str,
    style_entries: list[dict],
    custom_styles: list[dict],
    discount_percentage: float,
    notes: list[str] = None,
) -> tuple[bytes, float, float]:
    """Load the template, update it, and return bytes plus totals."""
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. Please add it to the environment first."
        )

    template_path = get_template_path()
    wb = load_workbook(template_path)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Worksheet '{TARGET_SHEET}' is missing from the template."
        )
    ws = wb[TARGET_SHEET]

    # Rename worksheet to "Workbook"
    ws.title = "Workbook"

    # Remove all other worksheets, keeping only "Workbook"
    sheets_to_remove = [name for name in wb.sheetnames if name != "Workbook"]
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])

    # Get reference to the Workbook worksheet after cleanup
    ws = wb["Workbook"]

    update_header_labels(ws, client_name)
    total_dev, total_optional = apply_development_package(
        ws,
        client_name=client_name,
        client_email=client_email,
        representative=representative,
        style_entries=style_entries,
        custom_styles=custom_styles,
        discount_percentage=discount_percentage,
    )
    
    # Write notes to column N starting below "PROJECT NOTES"
    # Find "PROJECT NOTES" dynamically (it moves when more styles are added)
    if notes:
        project_notes_row = None
        # Search for "PROJECT NOTES" in column N (column 14)
        for search_row in range(20, 50):  # Search from row 20 to 50
            cell_value = safe_get_cell_value(ws, search_row, SUMMARY_LABEL_COL)
            if cell_value and isinstance(cell_value, str):
                if "PROJECT" in cell_value.upper() and "NOTES" in cell_value.upper():
                    project_notes_row = search_row
                    break
        
        if project_notes_row:
            # Place notes starting one row below "PROJECT NOTES" (every other row: N27, N29, etc.)
            notes_start_row = project_notes_row + 1
            note_index = 0
            for note in notes:
                if note and note.strip():
                    # Place notes at every other row (skip one row between notes)
                    cell_row = notes_start_row + (note_index * 2)
                    cell = ws.cell(row=cell_row, column=SUMMARY_LABEL_COL)
                    # Uppercase and center-align the notes
                    cell.value = note.strip().upper()
                    if Font is not None:
                        cell.font = Font(name="Arial", size=20)
                    if Alignment is not None:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    note_index += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), total_dev, total_optional


@st.cache_data(ttl=300)  # Cache for 5 minutes
def get_sales_records():
    """Get sales records from monday.com for dropdown selection."""
    try:
        from database_utils import get_sales_data
        
        sales_data = get_sales_data()
        items = sales_data.get("data", {}).get("boards", [{}])[0].get("items_page", {}).get("items", [])
        
        # Return list of (item_id, item_name) tuples
        records = [(item.get("id"), item.get("name", "")) for item in items if item.get("name")]
        return sorted(records, key=lambda x: x[1])  # Sort by name
    except Exception as e:
        st.warning(f"Could not load sales records: {e}")
        return []


def update_monday_item_workbook_url(item_id: str, workbook_url: str) -> bool:
    """Update a monday.com item with the workbook URL in a 'Workbook Link' field."""
    try:
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            st.error("Monday.com API token not found in secrets.")
            return False
        
        # First, we need to find the column ID for "Workbook Link" or create it
        # For now, we'll use a URL column type. The column ID needs to be found or created in monday.com
        # This is a placeholder - the actual column ID needs to be configured
        
        # Query to get board columns to find the "Workbook Link" column
        query = f"""
        query {{
            items(ids: [{item_id}]) {{
                board {{
                    id
                    columns {{
                        id
                        title
                        type
                    }}
                }}
            }}
        }}
        """
        
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            st.error(f"Error fetching monday.com columns: {data['errors']}")
            return False
        
        items = data.get("data", {}).get("items", [])
        if not items:
            st.error("Item not found in monday.com")
            return False
        
        board = items[0].get("board", {})
        board_id = board.get("id")
        columns = board.get("columns", [])
        
        if not board_id:
            st.error("Could not determine board ID from monday.com item")
            return False
        
        # Find "Workbook Link" or "Workbook URL Link" column
        workbook_column = None
        for col in columns:
            title_lower = col.get("title", "").lower()
            if "workbook" in title_lower and ("link" in title_lower or "url" in title_lower):
                workbook_column = col
                break
        
        if not workbook_column:
            st.warning("⚠️ 'Workbook Link' column not found in monday.com. Please create a URL column named 'Workbook Link' or 'Workbook URL Link' in the Sales board.")
            return False
        
        column_id = workbook_column.get("id")
        column_type = workbook_column.get("type")
        
        # Update the item with the workbook URL
        # For URL columns, the value format is: {"url": "https://...", "text": "Link Text"}
        if column_type == "link":
            mutation = f"""
            mutation {{
                change_column_value(
                    board_id: {board_id},
                    item_id: {item_id},
                    column_id: "{column_id}",
                    value: "{{\\"url\\": \\"{workbook_url}\\", \\"text\\": \\"View Workbook\\"}}"
                ) {{
                    id
                }}
            }}
            """
        else:
            # For text columns, just use the URL as text
            mutation = f"""
            mutation {{
                change_column_value(
                    board_id: {board_id},
                    item_id: {item_id},
                    column_id: "{column_id}",
                    value: "{workbook_url}"
                ) {{
                    id
                }}
            }}
            """
        
        response = requests.post(url, json={"query": mutation}, headers=headers, timeout=30)
        result = response.json()
        
        if "errors" in result:
            st.error(f"Error updating monday.com: {result['errors']}")
            return False
        
        return True
        
    except Exception as e:
        st.error(f"Failed to update monday.com: {e}")
        return False


def extract_spreadsheet_id_from_url(url: str) -> str | None:
    """Extract spreadsheet ID from Google Sheets or Google Drive URL.
    
    Args:
        url: Google Sheets URL (e.g., https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit)
             or Google Drive URL (e.g., https://drive.google.com/file/d/FILE_ID/view)
    
    Returns:
        Spreadsheet/file ID or None if not found
    """
    import re
    
    # Pattern for Google Sheets URL: /spreadsheets/d/{id}/
    sheets_pattern = r'/spreadsheets/d/([a-zA-Z0-9-_]+)'
    match = re.search(sheets_pattern, url)
    if match:
        return match.group(1)
    
    # Pattern for Google Drive URL: /file/d/{id}/
    drive_pattern = r'/file/d/([a-zA-Z0-9-_]+)'
    match = re.search(drive_pattern, url)
    if match:
        return match.group(1)
    
    return None


def export_google_sheet_as_pdf(sheet_url: str) -> bytes:
    """Export Google Sheet as PDF using Google Sheets export API.
    
    Args:
        sheet_url: The Google Sheet URL
        
    Returns:
        PDF file as bytes
        
    Raises:
        RuntimeError: If spreadsheet ID cannot be extracted or export fails
    """
    # Extract spreadsheet ID from URL
    spreadsheet_id = extract_spreadsheet_id_from_url(sheet_url)
    if not spreadsheet_id:
        raise RuntimeError(f"Could not extract spreadsheet ID from URL: {sheet_url}")
    
    # Build export URL with PDF format
    # Using landscape orientation and removing gridlines for better readability
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=pdf&portrait=false&gridlines=false"
    
    # Get Google credentials for authentication
    if not GOOGLE_API_AVAILABLE:
        raise RuntimeError("Google API libraries not available. Please install google-auth and google-api-python-client.")
    
    try:
        info = st.secrets.get("google_service_account")
        if not info:
            raise RuntimeError("Google service account credentials not found in secrets")
        
        credentials = SACredentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        
        # Refresh credentials to get access token
        if not credentials.valid:
            credentials.refresh(GoogleRequest())
        
        # Fetch the PDF using authenticated request
        import urllib.request
        
        req = urllib.request.Request(export_url)
        req.add_header('Authorization', f'Bearer {credentials.token}')
        
        with urllib.request.urlopen(req) as response:
            pdf_bytes = response.read()
            return pdf_bytes
            
    except Exception as e:
        raise RuntimeError(f"Failed to export Google Sheet as PDF: {e}")


def get_board_id_from_item(item_id: str) -> str | None:
    """Get board ID from a Monday.com item ID.
    
    Args:
        item_id: The Monday.com item ID
        
    Returns:
        Board ID or None if not found
    """
    try:
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            return None
        
        query = f"""
        query {{
            items(ids: [{item_id}]) {{
                board {{
                    id
                }}
            }}
        }}
        """
        
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            return None
        
        items = data.get("data", {}).get("items", [])
        if not items:
            return None
        
        board = items[0].get("board", {})
        return board.get("id")
        
    except Exception:
        return None


def upload_file_to_monday_item(item_id: str, board_id: str, file_bytes: bytes, filename: str) -> bool:
    """Upload a file to a monday.com item using the GraphQL file upload API."""
    try:
        import json
        
        monday_config = st.secrets.get("monday", {})
        api_token = monday_config.get("api_token")
        
        if not api_token:
            st.error("Monday.com API token not found in secrets.")
            return False
        
        # Step 1: Get the files column ID for this board
        url = "https://api.monday.com/v2"
        headers = {
            "Authorization": api_token,
            "Content-Type": "application/json",
        }
        
        # Query to find the files column
        query = f"""
        query {{
            boards(ids: [{board_id}]) {{
                columns {{
                    id
                    title
                    type
                }}
            }}
        }}
        """
        
        response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
        data = response.json()
        
        if "errors" in data:
            st.error(f"Error fetching board columns: {data['errors']}")
            return False
        
        boards = data.get("data", {}).get("boards", [])
        if not boards:
            st.error("Board not found")
            return False
        
        columns = boards[0].get("columns", [])
        
        # Find files column (type is "file" or title contains "file")
        files_column = None
        for col in columns:
            if col.get("type") == "file" or "file" in col.get("title", "").lower():
                files_column = col
                break
        
        if not files_column:
            st.warning("⚠️ No files column found on the board. File upload requires a files column.")
            return False
        
        column_id = files_column.get("id")
        
        # Step 2: Upload file using GraphQL file upload API
        # Use the /v2/file endpoint for file uploads with multipart/form-data
        file_url = "https://api.monday.com/v2/file"
        
        # GraphQL mutation for adding file to column
        mutation = """
        mutation addFile($file: File!) {
            add_file_to_column(
                file: $file,
                item_id: %s,
                column_id: "%s"
            ) {
                id
            }
        }
        """ % (item_id, column_id)
        
        # Prepare multipart form data for GraphQL file upload
        # The format is: query, variables, map, and file
        variables = {
            "file": None
        }
        
        file_map = {
            "file": ["variables.file"]
        }
        
        # Determine MIME type based on file extension
        if filename.lower().endswith('.pdf'):
            mime_type = 'application/pdf'
        elif filename.lower().endswith('.xlsx'):
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mime_type = 'application/octet-stream'
        
        # Create multipart form data
        files_data = {
            'query': (None, mutation),
            'variables': (None, json.dumps(variables)),
            'map': (None, json.dumps(file_map)),
            'file': (filename, file_bytes, mime_type)
        }
        
        # Upload the file
        upload_response = requests.post(
            file_url,
            headers={"Authorization": api_token},
            files=files_data,
            timeout=60
        )
        
        if upload_response.status_code not in [200, 201]:
            error_text = upload_response.text
            try:
                error_json = upload_response.json()
                if "errors" in error_json:
                    st.error(f"Error uploading file: {error_json['errors']}")
                else:
                    st.error(f"Error uploading file: {error_text}")
            except:
                st.error(f"Error uploading file: {upload_response.status_code} - {error_text}")
            return False
        
        result = upload_response.json()
        
        if "errors" in result:
            st.error(f"Error uploading file: {result['errors']}")
            return False
        
        # Check if file was added successfully
        if result.get("data", {}).get("add_file_to_column", {}).get("id"):
            return True
        else:
            st.warning("File upload may have succeeded but no file ID returned")
            return True  # Assume success if no errors
        
    except Exception as e:
        st.error(f"Failed to upload file to monday.com: {e}")
        return False


def main() -> None:
    st.set_page_config(
        page_title="Workbook Creator",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.title("📊 Workbook Creator")

    st.markdown(
        """
<style>

    /* Enlarge checkbox */
    [data-testid="stCheckbox"] label span {
        height: 1.8rem;
        width: 1.8rem;
        transform: translateX(35px);
    }

    /* Hide normal label text (since you use empty label) */
    [data-testid="stCheckbox"] label p {
        font-size: 0;
        margin: 0;
    }

    /* Move right the entire checkbox inside its Streamlit column */
    [data-testid="stCheckbox"] {
        display: flex !important;
        justify-content: right !important;  /* Center horizontally */
        align-items: right !important;       /* Center vertically */
        height: 100%;                          /* Align with text inputs */
        padding: 0 !important;
        margin: 0 !important;
    }

    /* Also move right the internal label wrapper */
    [data-testid="stCheckbox"] > label {
        display: flex !important;
        justify-content: right !important;
        align-items: right !important;
        width: 100%;
        height: 100%;
        padding: 0 !important;
        margin: 0 !important;
    }

    /* Hide ALL sidebar list items by default */
    [data-testid="stSidebarNav"] li {
        display: none !important;
    }

    /* Show list items that contain allowed tool pages using :has() selector */
    [data-testid="stSidebarNav"] li:has(a[href*="signnow"]),
    [data-testid="stSidebarNav"] li:has(a[href*="/tools"]),
    [data-testid="stSidebarNav"] li:has(a[href*="workbook"]),
    [data-testid="stSidebarNav"] li:has(a[href*="deck_creator"]),
    [data-testid="stSidebarNav"] li:has(a[href*="a_la_carte"]) {
        display: block !important;
    }

</style>
<script>
// JavaScript to show only tool pages and hide everything else
(function() {
    function showToolPagesOnly() {
        const navItems = document.querySelectorAll('[data-testid="stSidebarNav"] li');
        const allowedPages = ['signnow', 'tools', 'workbook', 'deck_creator', 'a_la_carte'];
        
        // Check if we're currently on an ads dashboard page
        const currentUrl = window.location.href.toLowerCase();
        const currentPath = window.location.pathname.toLowerCase();
        const isOnAdsDashboard = currentUrl.includes('ads') && currentUrl.includes('dashboard') ||
                                 currentPath.includes('ads') && currentPath.includes('dashboard');
        
        navItems.forEach(item => {
            const link = item.querySelector('a');
            if (!link) {
                item.style.setProperty('display', 'none', 'important');
                return;
            }
            
            const href = (link.getAttribute('href') || '').toLowerCase();
            const text = link.textContent.trim().toLowerCase();
            
            // Check if this is an allowed tool page
            const isToolPage = allowedPages.some(page => {
                return href.includes(page) || text.includes(page.toLowerCase());
            });
            
            // Make sure it's not ads dashboard or other dashboards
            const isDashboard = (text.includes('ads') && text.includes('dashboard')) || 
                              (href.includes('ads') && href.includes('dashboard'));
            
            // Hide a_la_carte if we're on an ads dashboard page
            const isDevInspection = href.includes('a_la_carte') || text.includes('a_la_carte');
            if (isOnAdsDashboard && isDevInspection) {
                item.style.setProperty('display', 'none', 'important');
                return;
            }
            
            if (isToolPage && !isDashboard) {
                item.style.setProperty('display', 'block', 'important');
                link.style.setProperty('display', 'block', 'important');
            } else {
                item.style.setProperty('display', 'none', 'important');
            }
        });
    }
    
    // Run immediately and on load
    showToolPagesOnly();
    window.addEventListener('load', function() {
        setTimeout(showToolPagesOnly, 50);
        setTimeout(showToolPagesOnly, 200);
        setTimeout(showToolPagesOnly, 500);
    });
    
    // Watch for DOM changes
    const observer = new MutationObserver(function() {
        showToolPagesOnly();
    });
    
    setTimeout(function() {
        const sidebar = document.querySelector('[data-testid="stSidebarNav"]');
        if (sidebar) {
            observer.observe(sidebar, { 
                childList: true, 
                subtree: true,
                attributes: true
            });
        }
    }, 100);
})();
</script>
""",
        unsafe_allow_html=True,
    )

    st.caption(
        "Fill in the Development Package inputs and download a formatted workbook "
        "based on the official template."
    )

    # Get query parameters from Monday.com link
    query_params = st.query_params
    first_name = query_params.get("first_name", "").strip()
    last_name = query_params.get("last_name", "").strip()
    email = query_params.get("email", "").strip()
    representative = query_params.get("representative", "").strip()
    num_styles_param = query_params.get("num_styles", "").strip()
    item_id = query_params.get("item_id", "").strip()
    
    # Clean up representative (remove "$" prefix if present from Monday.com formula)
    if representative.startswith("$"):
        representative = representative[1:].strip()
    
    # Combine first_name and last_name for client_name
    client_name_default = ""
    if first_name or last_name:
        client_name_default = f"{first_name} {last_name}".strip()
    
    # Use query params as default values if available
    client_name = st.text_input(
        "Client Name", 
        value=client_name_default,
        placeholder="Enter client name"
    )

    col_a, col_b = st.columns(2)
    with col_a:
        client_email = st.text_input(
            "Client Email", 
            value=email,
            placeholder="client@email.com"
        )
    with col_b:
        representative = st.text_input(
            "Representative", 
            value=representative,
            placeholder="Enter representative"
        )

    # Initialize session state for style entries
    if "style_entries" not in st.session_state:
        st.session_state["style_entries"] = []
    
    # Initialize Custom Items
    if "custom_styles" not in st.session_state:
        st.session_state["custom_styles"] = []
    
    st.subheader("**Styles**")
    
    # Number of Styles field (auto-fills from query param, but editable) - in first column, half width
    col_num_styles, col_spacer = st.columns([1, 3])
    with col_num_styles:
        num_styles_default = len(st.session_state["style_entries"])
        if num_styles_param:
            try:
                num_styles_default = int(num_styles_param)
            except ValueError:
                num_styles_default = len(st.session_state["style_entries"])
        
        num_styles_input = st.number_input(
            "Number of Styles",
            min_value=0,
            value=num_styles_default,
            step=1,
            key="num_styles_input",
            help="Number of styles/items. Auto-filled from Monday.com if available."
        )
        
        # Update style entries list to match num_styles_input
        current_count = len(st.session_state["style_entries"])
        if num_styles_input > current_count:
            # Add new styles with blank names
            for i in range(current_count, num_styles_input):
                st.session_state["style_entries"].append({
                    "name": "",  # Leave blank, don't default to "Style 1", etc.
                    "style_type": "Regular",  # Default to Regular
                    "complexity": 0.0,
                    "style_number": 101 + i,  # Default style numbers: 101, 102, 103...
                    "options": {
                        "wash_dye": False,
                        "design": False,
                        "treatment": False,
                        # Note: "source" removed - no longer an option
                    },
                })
        elif num_styles_input < current_count:
            # Remove excess styles
            st.session_state["style_entries"] = st.session_state["style_entries"][:num_styles_input]
    
    # Column headers (removed Source column)
    header_cols = st.columns([1.1, 1.8, 1.2, 1.2, 1.2, 1, 1.1])
    with header_cols[0]:
        st.markdown("**Style Number**")
    with header_cols[1]:
        st.markdown("**Style Name**")
    with header_cols[2]:
        st.markdown("**Style Type**")
    with header_cols[3]:
        st.markdown("**Complexity (%)**")
    with header_cols[4]:
        st.markdown("**Wash/Dye ($1,500)**")
    with header_cols[5]:
        st.markdown("**Design ($1,500)**")
    with header_cols[6]:
        st.markdown("**Treatment ($860)**")
    
    # Display existing style entries in horizontal rows
    if st.session_state["style_entries"]:
        for i, entry in enumerate(st.session_state["style_entries"]):
            with st.container():
                cols = st.columns([1.2, 1.8, 1.2, 1.2, 1.2, 1, 1.1])  # Removed source column
                with cols[0]:
                    # Style Number field with default value (101, 102, 103...)
                    default_style_number = entry.get("style_number", 101 + i)
                    style_number = st.number_input(
                        "Style Number",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with cols[1]:
                    style_name = st.text_input(
                        "Custom Item Name",
                        value=entry.get("name", ""),
                        key=f"style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Dress, Winter Coat",
                    )
                    entry["name"] = style_name
                with cols[2]:
                    # Style Type dropdown instead of checkbox
                    style_type_options = ["Regular", "Activewear/Lingerie/Swim", "Pattern Blocks"]
                    current_style_type = entry.get("style_type", "Regular")
                    # Handle migration from old "activewear" boolean
                    if "style_type" not in entry and entry.get("activewear", False):
                        current_style_type = "Activewear/Lingerie/Swim"
                    style_type = st.selectbox(
                        "Style Type",
                        options=style_type_options,
                        index=style_type_options.index(current_style_type) if current_style_type in style_type_options else 0,
                        key=f"style_type_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_type"] = style_type
                    # Remove old activewear key if it exists
                    if "activewear" in entry:
                        del entry["activewear"]
                with cols[3]:
                    complexity = st.number_input(
                        "Complexity (%)",
                        min_value=0,
                        max_value=200,
                        value=int(entry.get("complexity", 100)),
                        step=5,
                        format="%d",
                        key=f"complexity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["complexity"] = float(complexity)
                with cols[4]:
                    wash_dye = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("wash_dye", False),
                        key=f"wash_dye_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["wash_dye"] = wash_dye
                with cols[5]:
                    design = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("design", False),
                        key=f"design_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["design"] = design
                # Removed source column - no longer an option
                with cols[6]:
                    treatment = st.checkbox(
                        "",
                        value=entry.get("options", {}).get("treatment", False),
                        key=f"treatment_{i}",
                        label_visibility="visible",
                    )
                    entry.setdefault("options", {})["treatment"] = treatment

    # Custom Item section
    st.markdown("---")
    st.subheader("**Custom Item**")
    
    # Calculate number of regular styles for custom item numbering
    num_regular_styles = len(st.session_state.get("style_entries", []))
    
    # QuickSelect buttons for common custom items
    quick_select_cols = st.columns([1, 1, 2])
    with quick_select_cols[0]:
        if st.button("➕ Sourcing ($2,050)", key="quick_sourcing", help="Add Sourcing — $2,050"):
            next_num = 101 + num_regular_styles + len(st.session_state["custom_styles"])
            st.session_state["custom_styles"].append({
                "name": "Sourcing",
                "price": 2050.0,
                "complexity": 0.0,
                "style_number": next_num,
            })
            st.rerun()
    with quick_select_cols[1]:
        if st.button("➕ Sourcing Consult ($860)", key="quick_sourcing_consult", help="Add Sourcing Consult — $860"):
            next_num = 101 + num_regular_styles + len(st.session_state["custom_styles"])
            st.session_state["custom_styles"].append({
                "name": "Sourcing Consult",
                "price": 860.0,
                "complexity": 0.0,
                "style_number": next_num,
            })
            st.rerun()
    
    # Column headers for Custom Items
    custom_header_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
    with custom_header_cols[0]:
        st.markdown("**Style Number**")
    with custom_header_cols[1]:
        st.markdown("**Custom Item Name**")
    with custom_header_cols[2]:
        st.markdown("**Price ($)**")
    with custom_header_cols[3]:
        st.markdown("**Complexity (%)**")
    with custom_header_cols[4]:
        st.markdown("**Remove**")
    
    # Display existing Custom Items
    if st.session_state["custom_styles"]:
        for i, entry in enumerate(st.session_state["custom_styles"]):
            with st.container():
                custom_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
                with custom_cols[0]:
                    # Style Number field with default value (101 + num_regular_styles + i)
                    default_style_number = entry.get("style_number", 101 + num_regular_styles + i)
                    style_number = st.number_input(
                        "Style Number",
                        min_value=1,
                        value=int(default_style_number),
                        step=1,
                        key=f"custom_style_number_{i}",
                        label_visibility="collapsed",
                    )
                    entry["style_number"] = style_number
                with custom_cols[1]:
                    custom_style_name = st.text_input(
                        "Custom Item Name",
                        value=entry.get("name", ""),
                        key=f"custom_style_name_{i}",
                        label_visibility="collapsed",
                        placeholder="e.g., Custom Item",
                    )
                    entry["name"] = custom_style_name
                with custom_cols[2]:
                    custom_price = st.number_input(
                        "Price",
                        min_value=0.0,
                        value=float(entry.get("price", 0.0)),
                        step=100.0,
                        format="%.2f",
                        key=f"custom_price_{i}",
                        label_visibility="collapsed",
                    )
                    entry["price"] = float(custom_price)
                with custom_cols[3]:
                    custom_complexity = st.number_input(
                        "Complexity (%)",
                        min_value=0,
                        max_value=200,
                        value=int(entry.get("complexity", 0)),
                        step=5,
                        format="%d",
                        key=f"custom_complexity_{i}",
                        label_visibility="collapsed",
                    )
                    entry["complexity"] = float(custom_complexity)
                with custom_cols[4]:
                    if st.button("❌", key=f"remove_custom_{i}", help="Remove this Custom Item"):
                        st.session_state["custom_styles"].pop(i)
                        st.rerun()
    
    # Add new Custom Item interface
    st.markdown("**Add New Custom Item**")
    add_custom_cols = st.columns([1.2, 2, 1.5, 1.5, 0.8])
    default_new_custom_name = st.session_state.get("new_custom_style_name", "")
    default_new_custom_price = st.session_state.get("new_custom_price", 0.0)
    default_new_custom_complexity = st.session_state.get("new_custom_complexity", 0)
    default_new_custom_style_number = st.session_state.get("new_custom_style_number", 101 + num_regular_styles + len(st.session_state.get("custom_styles", [])))
    
    with add_custom_cols[0]:
        new_custom_style_number = st.number_input(
            "Style Number",
            min_value=1,
            value=int(default_new_custom_style_number),
            step=1,
            key="new_custom_style_number",
            label_visibility="collapsed",
        )
    with add_custom_cols[1]:
        new_custom_style_name = st.text_input(
            "Custom Item Name",
            value=default_new_custom_name,
            key="new_custom_style_name",
            label_visibility="collapsed",
            placeholder="e.g., Custom Item",
        )
    with add_custom_cols[2]:
        new_custom_price = st.number_input(
            "Price",
            min_value=0.0,
            value=default_new_custom_price,
            step=100.0,
            format="%.2f",
            key="new_custom_price",
            label_visibility="collapsed",
        )
    with add_custom_cols[3]:
        new_custom_complexity = st.number_input(
            "Complexity (%)",
            min_value=0,
            max_value=200,
            value=default_new_custom_complexity,
            step=5,
            format="%d",
            key="new_custom_complexity",
            label_visibility="collapsed",
        )
    with add_custom_cols[4]:
        if st.button("➕ Add", key="add_custom_style", help="Add this Custom Item"):
            if new_custom_style_name.strip() and new_custom_price > 0:
                st.session_state["custom_styles"].append({
                    "name": new_custom_style_name.strip(),
                    "price": float(new_custom_price),
                    "complexity": float(new_custom_complexity),
                    "style_number": int(new_custom_style_number),
                })
                # Reset add-new-custom-style inputs
                for key in [
                    "new_custom_style_name",
                    "new_custom_price",
                    "new_custom_complexity",
                    "new_custom_style_number",
                ]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
            else:
                st.warning("Please enter a Custom Item name and price before adding.")

    if not st.session_state["style_entries"] and not st.session_state["custom_styles"]:
        st.info("Add at least one style or Custom Item to enable the generator.")
        return

    st.markdown("---")
    st.subheader("Discount")
    discount_cols = st.columns([1.8, 1, 1, 1.2, 1, 0.8, 0.8, 1, 0.8])
    with discount_cols[0]:
        discount_percentage = st.number_input(
            "Discount (%)",
            min_value=0,
            max_value=100,
            value=st.session_state.get("discount_percentage_value", 0),
            step=1,
        )
    st.session_state["discount_percentage_value"] = discount_percentage
    
    # Notes Section
    st.markdown("---")
    st.subheader("Notes")
    
    # Initialize notes list in session state
    if "notes_list" not in st.session_state:
        st.session_state["notes_list"] = [""]
    
    # Display existing notes
    for i, note in enumerate(st.session_state["notes_list"]):
        note_cols = st.columns([10, 1])
        with note_cols[0]:
            updated_note = st.text_input(
                "Note",
                value=note,
                key=f"note_{i}",
                placeholder="Enter a note...",
                label_visibility="collapsed"
            )
            st.session_state["notes_list"][i] = updated_note
        with note_cols[1]:
            if st.button("❌", key=f"remove_note_{i}", help="Remove this note"):
                st.session_state["notes_list"].pop(i)
                st.rerun()
    
    # Add new note button
    if st.button("➕ Add Note", key="add_note"):
        st.session_state["notes_list"].append("")
        st.rerun()
    
    # Filter out empty notes for workbook generation
    notes = [note for note in st.session_state["notes_list"] if note and note.strip()]

    try:
        excel_bytes, _, _ = build_workbook_bytes(
            client_name=client_name,
            client_email=client_email,
            representative=representative,
            style_entries=st.session_state["style_entries"],
            custom_styles=st.session_state.get("custom_styles", []),
            discount_percentage=discount_percentage,
            notes=notes if notes else [],
        )
    except FileNotFoundError as exc:
        st.error(str(exc))
        return
    except Exception as exc:  # pragma: no cover - streamlit runtime feedback
        st.error(f"Failed to build workbook: {exc}")
        return

    safe_client = re.sub(r"[^A-Za-z0-9_-]+", "_", (client_name or "").strip()) or "client"
    download_name = f"workbook_{safe_client.lower()}.xlsx"

    st.success("Workbook is ready.")
    st.download_button(
        label="Generate Workbook",
        data=excel_bytes,
        file_name=download_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    # Google Drive upload section (service account only)
    st.markdown("---")
    st.subheader("Google Sheets --> Monday.com Upload")

    sheet_title = (client_name or "Workbook").strip() or "Workbook"
    sheet_title = f"{sheet_title} - Development Package"

    st.caption("Uploads will use the shared Google Drive folder configured for the service account.")

    if st.button("Upload to Monday.com", type="primary"):
        with st.spinner("Uploading workbook to Google Sheets and updating Monday.com..."):
            try:
                # Upload to Google Sheets
                sheet_url, converted = upload_workbook_to_google_sheet(excel_bytes, sheet_title)
                st.session_state["google_sheet_url"] = sheet_url
                
                if converted:
                    st.success(f"✅ Google Sheet created: [Open Sheet]({sheet_url})")
                else:
                    st.success(
                        f"✅ Workbook uploaded as XLSX: [Download / Open]({sheet_url})  \n"
                        "Google denied automatic conversion because service accounts report zero Drive quota. "
                        "Open the file in Google Drive and use **File → Save as Google Sheets** if you need an editable sheet."
                    )
                
                # Update Monday.com with the Google Sheet URL if item_id is provided
                if item_id:
                    # Update Monday.com item with the Google Sheet link
                    if update_monday_item_workbook_url(item_id, sheet_url):
                        st.success(f"✅ Monday.com item updated with Google Sheet link!")
                    else:
                        st.warning("⚠️ Google Sheet uploaded, but failed to update Monday.com item. Please update manually.")
                    
                    # Also upload PDF version to Files column (only if converted to Google Sheet)
                    if converted:
                        try:
                            # Get board_id from item_id
                            board_id = get_board_id_from_item(item_id)
                            if board_id:
                                # Export Google Sheet as PDF using Google Sheets export API
                                pdf_bytes = export_google_sheet_as_pdf(sheet_url)
                                pdf_filename = f"{safe_client}_workbook.pdf"
                                
                                # Upload PDF to Monday.com Files column
                                if upload_file_to_monday_item(item_id, board_id, pdf_bytes, pdf_filename):
                                    st.success(f"✅ PDF version uploaded to Monday.com Files column!")
                                else:
                                    st.warning("⚠️ PDF upload to Monday.com Files column failed. Link was updated successfully.")
                            else:
                                st.warning("⚠️ Could not retrieve board ID. PDF upload skipped.")
                        except Exception as pdf_error:
                            st.warning(f"⚠️ PDF export/upload failed: {pdf_error}. Link was updated successfully.")
                    else:
                        st.info("ℹ️ PDF export skipped (workbook not converted to Google Sheet format).")
                else:
                    st.info("ℹ️ No Monday.com item ID provided. Workbook uploaded to Google Sheets only.")

            except GoogleSheetsUploadError as exc:
                message = str(exc)
                st.error(f"❌ Google Sheets upload failed: {message}")
                if "storage quota" in message.lower():
                    _show_drive_quota_help()
            except Exception as exc:  # pragma: no cover - runtime failures
                st.error(f"❌ Unexpected error: {exc}")


if __name__ == "__main__":
    main()
